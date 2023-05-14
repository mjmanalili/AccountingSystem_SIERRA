from tkinter import *
from tkinter import messagebox
import tkinter.ttk as tk
from tkcalendar import DateEntry
from calendar import monthrange
from PIL import ImageTk, Image
import datetime
import mysql.connector as mysql
import bcrypt
from tkinter import filedialog
from os import listdir, startfile, system
from tkinter.filedialog import askopenfilename
from openpyxl import load_workbook
# from itertools import groupby
import pandas as pd
from num2words import num2words
from pypdf import PdfReader
import pandas._libs.tslibs.base
import babel.numbers

### MY SQL CONNECTION ###
db = mysql.connect(
    host = "192.168.1.248",
    user = "ACCTG",
    passwd = "ACCTGasd123!",
    database = "dbpsc")
cursor = db.cursor()

### KEY VARIABLES ###

PATH_TEMPLATE = "//192.168.1.248/sierra/TEMPLATE/" #D:/SIERRA/TEMPLATE
PATH_SAVE = "C:/SIERRA/SAVED/" #D:/SIERRA/SAVED/
PATH_ICON = "//192.168.1.248/sierra/ICON_DARK/" #D:/SIERRA/ICON_DARK/
PATH_UPDATE = "C:/SIERRA/APP/"

APP_NAME = "SIERRA"
APP_VERSION = "2.1"
APP_SIZE = "1195x695+100+0"
APP_BG = "#2E5894" #000000 #gray25 #gray35 #51A0D5 #78C5EF #2C528C
APP_DASHBOARD_BG = "#072F5F" #"gray20" #gray15 #2C528C #51A0D5
APP_FG = "#FFFFFF"
APP_FG2 = "#000000"
APP_DASHBOARD_FG = "#FFFFFF"
APP_FONT = ("calibri", 10)
APP_FONT_LARGE = ("calibri", 12)
APP_FONT_SMALL = ("calibri", 8)
# APP_TOPBAR_SIZE = 3

# WORKSPACE VARIABLES
CLICK_BG = "#78C5EF"
BUTTON_BG = "#FFFFFF"
BUTTON_FG = "#2C528C"
BUTTON_FG_RED = "#FF0000"
BUTTON_FG_GRN = "#228C22"
LABEL_BG = "#FFFFFF"
TREE_TAG_EVENROW = "#D1E7FF"
TREE_TAG_VOID = "#FFB5B7"
BUTTON_FONT = ("calibri", 8, "bold")
BUTTON_FONT2 = ("calibri", 7, "bold")
BUTTON_FONT3 = ("calibri", 10, "underline")
TOP_BUTTON_WIDTH = 8
TOP_PADX = 3
TOP_PADY = 3

MENU_PADX = 5
MENU_PADY = 5
MORPH = False
MORPH_SPEED = 1
MORPH_DROPSPEED = 15
SUBMENU_PADY = 2

class Main:
    def __init__(self, master):
        self.master = master
        self.USERNAME = StringVar()
        self.PASSWORD = StringVar()
        
        global FRAME_LOGIN
        FRAME_LOGIN = Frame(self.master, bg = APP_BG)
        FRAME_LOGIN.grid(column = 0, row = 0)
        
        self.showLogin()
        
        master.protocol("WM_DELETE_WINDOW", self.exitApplication)
        self.master.grid_rowconfigure(0, weight = 1)
        self.master.grid_columnconfigure(0, weight = 1)

    def showLogin(self):
        ENTRY_USERNAME = Entry(FRAME_LOGIN, textvariable = self.USERNAME, font = ("calibri", 12))
        ENTRY_USERNAME.grid(column = 0, row = 0, pady = MENU_PADY)
        ENTRY_USERNAME.bind("<FocusOut>", lambda e: self.USERNAME.set(self.USERNAME.get().upper()))

        ENTRY_PASSWORD = Entry(FRAME_LOGIN, textvariable = self.PASSWORD, font = ("calibri", 12), show = "●")
        ENTRY_PASSWORD.grid(column = 0, row = 1, pady = MENU_PADY)
        ENTRY_PASSWORD.bind("<Return>", self.loginUser)

        BUTTON_LOGIN = Button(FRAME_LOGIN, text = "Login", font = APP_FONT, bg = APP_BG, fg = APP_FG, bd = 0, cursor = "hand2", command = self.loginUser)
        BUTTON_LOGIN.grid(column = 0, row = 2)
        BUTTON_LOGIN.bind("<Return>", self.loginUser)

        BUTTON_CHANGE = Button(FRAME_LOGIN, text = "Change Password", font = APP_FONT, bg = APP_BG, fg = APP_FG, bd = 0, cursor = "hand2", command = self.showChangePassword)
        BUTTON_CHANGE.grid(column = 0, row = 3)

        BUTTON_EXIT = Button(FRAME_LOGIN, text = "Exit", font = APP_FONT, bg = APP_BG, fg = APP_FG, bd = 0, cursor = "hand2", command = self.exitApplication)
        BUTTON_EXIT.grid(column = 0, row = 4)
        
        ENTRY_USERNAME.focus()

    def loginUser(self, *args):
        global USER
        find = "SELECT ID, password FROM tblusers WHERE ID = %s LIMIT 1"
        cursor.execute(find, [self.USERNAME.get()])
        result = cursor.fetchone()
        password = self.PASSWORD.get()
        if result:
            if bcrypt.checkpw(password.encode("utf-8"), result[1].encode("utf-8")):
                USER = self.USERNAME.get()
                self.showDashboard()
            else:
                messagebox.showerror("SIERRA!", "Invalid username and password!")
        else:
            messagebox.showerror("SIERRA!", "Invalid ID and/or password!")

    def showChangePassword(self, *args):
        for i in FRAME_LOGIN.winfo_children():
            i.destroy()
            
        LABEL_SEARCH = Label(FRAME_LOGIN, text = "ID", width = 7, font = APP_FONT, anchor = W)
        LABEL_SEARCH.grid(column = 0, row = 0, pady = SUBMENU_PADY, sticky = W)
        
        LABEL_SEARCH = Label(FRAME_LOGIN, text = "Current", width = 7, font = APP_FONT, anchor = W)
        LABEL_SEARCH.grid(column = 0, row = 1, pady = SUBMENU_PADY, sticky = W)
        
        LABEL_SEARCH = Label(FRAME_LOGIN, text = "New", width = 7, font = APP_FONT, anchor = W)
        LABEL_SEARCH.grid(column = 0, row = 2, pady = SUBMENU_PADY, sticky = W)
        
        LABEL_SEARCH = Label(FRAME_LOGIN, text = "Confirm New", width = 7, font = APP_FONT, anchor = W)
        LABEL_SEARCH.grid(column = 0, row = 3, pady = SUBMENU_PADY, sticky = W)
        
        global ENTRY_ID
        ENTRY_ID = Entry(FRAME_LOGIN, font = APP_FONT, bg = APP_BG, fg = APP_FG)
        ENTRY_ID.grid(column = 1, row = 0, pady = MENU_PADY)
        
        global ENTRY_CURRENT
        ENTRY_CURRENT = Entry(FRAME_LOGIN, font = APP_FONT, bg = APP_BG, fg = APP_FG, show = "●")
        ENTRY_CURRENT.grid(column = 1, row = 1, pady = MENU_PADY)
        
        global ENTRY_NEW
        ENTRY_NEW = Entry(FRAME_LOGIN, font = APP_FONT, bg = APP_BG, fg = APP_FG, show = "●")
        ENTRY_NEW.grid(column = 1, row = 2, pady = MENU_PADY)
        
        global ENTRY_CONFIRM
        ENTRY_CONFIRM = Entry(FRAME_LOGIN, font = APP_FONT, bg = APP_BG, fg = APP_FG, show = "●")
        ENTRY_CONFIRM.grid(column = 1, row = 3, pady = MENU_PADY)
        
        BUTTON_CHANGE = Button(FRAME_LOGIN, text = "Submit", font = APP_FONT, bg = APP_BG, fg = APP_FG, bd = 0, cursor = "hand2", command = self.savePassword)
        BUTTON_CHANGE.grid(column = 1, row = 4)

        BUTTON_BACK = Button(FRAME_LOGIN, text = "Back", font = APP_FONT, bg = APP_BG, fg = APP_FG, bd = 0, cursor = "hand2", command = self.backToLogin)
        BUTTON_BACK.grid(column = 1, row = 5)

    def exitApplication(self, *args):
        db.close()
        self.master.destroy()
        
    def savePassword(self):
        if ENTRY_NEW.get() != "" and ENTRY_CONFIRM.get() != "":
            find = "SELECT ID, password FROM tblusers WHERE ID = %s LIMIT 1"
            cursor.execute(find, [ENTRY_ID.get()])
            result = cursor.fetchone()
            password = ENTRY_CURRENT.get()
            if result:
                if bcrypt.checkpw(password.encode("utf-8"), result[1].encode("utf-8")):
                    if ENTRY_NEW.get() == ENTRY_CONFIRM.get():
                        ask = messagebox.askyesno("Change Password", "Are you sure?")
                        if ask:
                            update = "UPDATE tblusers SET password = %s WHERE ID = %s"
                            cursor.execute(update, [self.hashPassword(ENTRY_CONFIRM.get()), ENTRY_ID.get()])
                            db.commit()
                            messagebox.showinfo("Change Password", "Your password has been updated!")
                            self.backToLogin()
                    else:
                        messagebox.showerror("Change Password", "Passwords do not match!")
                else:
                    messagebox.showerror("Change Password", "Invalid ID and/or password!")
        else:
            messagebox.showerror("Change Password", "Empty fields detected!")
    
    def backToLogin(self):
        for i in FRAME_LOGIN.winfo_children():
            i.destroy()
        self.showLogin()

### FRAME_MANAGER ###
    def showDashboard(self, *args):
        FRAME_LOGIN.destroy()
        
        self.master.grid_rowconfigure(0, weight = 0)

        # FRAME 1 top dashboard frame
        FRAME_1 = Frame(self.master, bg = APP_DASHBOARD_BG)
        FRAME_1.grid(column = 0, row = 0, pady = 2, sticky = NW)

        # logo
        LABEL_LOGO = Label(FRAME_1, image = ICON_LOGO, width = self.master.winfo_screenwidth(), bg = APP_DASHBOARD_BG, anchor = CENTER)
        # LABEL_LOGO.grid(column = 0, row = 0, sticky = N) width = 1190
        LABEL_LOGO.pack(fill = "both")

        # date
        LABEL_DATE = Label(FRAME_1, text = str(datetime.date.today()) + " v" + APP_VERSION, font = APP_FONT, bg = APP_DASHBOARD_BG, fg = APP_DASHBOARD_FG, anchor = W)
        LABEL_DATE.place(x = 5, y = 5)

        # buttons
        SUB1_FRAME1 = Frame(FRAME_1, bg = APP_DASHBOARD_BG)
        SUB1_FRAME1.place(x = self.master.winfo_screenwidth()-200)
        
        LABEL_NAME = Label(SUB1_FRAME1, text = USER, font = APP_FONT, bg = APP_DASHBOARD_BG, fg = APP_DASHBOARD_FG, justify = RIGHT)
        LABEL_NAME.grid(column = 0, row = 0, padx = MENU_PADX)
        
        BUTTON_BELL = Button(SUB1_FRAME1, bd = 0, bg = APP_DASHBOARD_BG, cursor = "hand2", image = ICON_PROFILE, command = self.showForApprovals)
        BUTTON_BELL.grid(column = 1, row = 0, padx = MENU_PADX)

        BUTTON_PROFILE = Button(SUB1_FRAME1, bd = 0, bg = APP_DASHBOARD_BG, cursor = "hand2", image = ICON_PROFILE, command = None)
        BUTTON_PROFILE.grid(column = 1, row = 0, padx = MENU_PADX)

        BUTTON_SETTINGS = Button(SUB1_FRAME1, bd = 0, bg = APP_DASHBOARD_BG, cursor = "hand2", image = ICON_SETTINGS, command = None)
        BUTTON_SETTINGS.grid(column = 2, row = 0, padx = MENU_PADX)

        # FRAME 2 side dashboard frame
        global FRAME_2
        FRAME_2 = Frame(self.master, bg = APP_BG)
        FRAME_2.grid(column = 0, row = 1, sticky = NW)
        
        FRAME_3 = Frame(FRAME_2, bg = APP_BG)
        FRAME_3.grid(column = 0, row = 0, sticky = NW)

        global SUB1_FRAME3
        SUB1_FRAME3 = Frame(FRAME_3, bg = APP_DASHBOARD_BG)
        SUB1_FRAME3.grid(column = 0, row = 0, ipadx = 50, pady = 1, sticky = W)

        global SUB2_FRAME3
        SUB2_FRAME3 = Frame(FRAME_3, bg = APP_BG)
        SUB2_FRAME3.grid(column = 0, row = 1, ipadx = 50, sticky = E)

        global SUB3_FRAME3
        SUB3_FRAME3 = Frame(FRAME_3, bg = APP_DASHBOARD_BG)
        SUB3_FRAME3.grid(column = 0, row = 2, ipadx = 50, pady = 1, sticky = W)

        global SUB4_FRAME3
        SUB4_FRAME3 = Frame(FRAME_3, bg = APP_BG)
        SUB4_FRAME3.grid(column = 0, row = 3, ipadx = 50, sticky = E)

        global SUB5_FRAME3
        SUB5_FRAME3 = Frame(FRAME_3, bg = APP_DASHBOARD_BG)
        SUB5_FRAME3.grid(column = 0, row = 4, ipadx = 50, pady = 1, sticky = W)

        global SUB6_FRAME3
        SUB6_FRAME3 = Frame(FRAME_3, bg = APP_BG)
        SUB6_FRAME3.grid(column = 0, row = 5, ipadx = 50, sticky = E)

        global SUB7_FRAME3
        SUB7_FRAME3 = Frame(FRAME_3, bg = APP_DASHBOARD_BG)
        SUB7_FRAME3.grid(column = 0, row = 6, ipadx = 50, pady = 1, sticky = W)

        global SUB8_FRAME3
        SUB8_FRAME3 = Frame(FRAME_3, bg = APP_BG)
        SUB8_FRAME3.grid(column = 0, row = 7, ipadx = 50, sticky = E)
        
        global SUB9_FRAME3
        SUB9_FRAME3 = Frame(FRAME_3, bg = APP_DASHBOARD_BG)
        SUB9_FRAME3.grid(column = 0, row = 8, ipadx = 50, pady = 1, sticky = W)
        
        global SUB10_FRAME3
        SUB10_FRAME3 = Frame(FRAME_3, bg = APP_BG)
        SUB10_FRAME3.grid(column = 0, row = 9, ipadx = 50, sticky = E)

        global SUBMENU_FRAMES
        SUBMENU_FRAMES = [SUB2_FRAME3, SUB4_FRAME3, SUB6_FRAME3, SUB8_FRAME3, SUB10_FRAME3]

        # menu
        global BUTTON_ACCOUNTING
        BUTTON_ACCOUNTING = Button(SUB1_FRAME3, text = " Accounting", width = 107, font = APP_FONT_LARGE, bg = APP_DASHBOARD_BG, fg = APP_FG, bd = 0, anchor = W, image = ICON_ACCOUNTING, compound = LEFT, cursor = "hand2", command = self.showAccountingMenu)
        BUTTON_ACCOUNTING.grid(column = 0, row = 0, pady = MENU_PADY, sticky = W)

        global BUTTON_FINANCE
        BUTTON_FINANCE = Button(SUB3_FRAME3, text = " Finance", width = 107, font = APP_FONT_LARGE, bg = APP_DASHBOARD_BG, fg = APP_FG, bd = 0, anchor = W, image = ICON_FINANCE, compound = LEFT, cursor = "hand2", command = self.showFinanceMenu)
        BUTTON_FINANCE.grid(column = 0, row = 0, pady = MENU_PADY, sticky = W)

        global BUTTON_BCD
        BUTTON_BCD = Button(SUB5_FRAME3, text = " Billing", width = 107, font = APP_FONT_LARGE, bg = APP_DASHBOARD_BG, fg = APP_FG, bd = 0, anchor = W, image = ICON_BILLING, compound = LEFT, cursor = "hand2", command = self.showBCDMenu)
        BUTTON_BCD.grid(column = 0, row = 0, pady = MENU_PADY, sticky = W)

        global BUTTON_GSAD
        BUTTON_GSAD = Button(SUB7_FRAME3, text = " GSAD", width = 107, font = APP_FONT_LARGE, bg = APP_DASHBOARD_BG, fg = APP_FG, bd = 0, anchor = W, image = ICON_GSAD, compound = LEFT, cursor = "hand2", command = self.showGeneralServicesMenu)
        BUTTON_GSAD.grid(column = 0, row = 0, pady = MENU_PADY, sticky = W)

        global BUTTON_ADMINISTRATOR
        BUTTON_ADMINISTRATOR = Button(SUB9_FRAME3, text = " Administrator", width = 107, font = APP_FONT_LARGE, bg = APP_DASHBOARD_BG, fg = APP_FG, bd = 0, anchor = W, image = ICON_ADMIN, compound = LEFT, cursor = "hand2", command = self.showAdministratorMenu)
        BUTTON_ADMINISTRATOR.grid(column = 0, row = 0, pady = MENU_PADY, sticky = W)

        global MENU_BUTTONS
        MENU_BUTTONS = [BUTTON_ACCOUNTING, BUTTON_FINANCE, BUTTON_BCD, BUTTON_GSAD, BUTTON_ADMINISTRATOR]

        # workspace
        global FRAME_4
        FRAME_4 = Frame(FRAME_2, bg = APP_BG)
        FRAME_4.grid(column = 1, row = 0, pady = SUBMENU_PADY, padx = MENU_PADX, sticky = NW)

    def showForApprovals(self):
        pass

### MENU ###
    def showAccountingMenu(self, *args):
        self.morphMenuButton(BUTTON_ACCOUNTING)
        SUB2_FRAME3.grid(column = 0, row = 1, ipadx = 50, sticky = E)
        BUTTON_GENERAL = Button(SUB2_FRAME3, text = "General Journal", width = 15, font = APP_FONT, bg = APP_BG, fg = APP_FG, bd = 0, anchor = E, state = DISABLED, cursor = "arrow", command = self.showGeneralJournal)
        BUTTON_GENERAL.grid(column = 0, row = 0, pady = SUBMENU_PADY, sticky = E)
        SUB2_FRAME3.after(MORPH_DROPSPEED)
        SUB2_FRAME3.update()

        BUTTON_IMPORTER = Button(SUB2_FRAME3, text = "Post Master", width = 15, font = APP_FONT, bg = APP_BG, fg = APP_FG, bd = 0, anchor = E, state = DISABLED, cursor = "arrow", command = self.showPostMaster)
        BUTTON_IMPORTER.grid(column = 0, row = 5, pady = SUBMENU_PADY, sticky = E)
        SUB2_FRAME3.after(MORPH_DROPSPEED)
        SUB2_FRAME3.update()

        BUTTON_PAYROLL = Button(SUB2_FRAME3, text = "Payroll Importer", width = 15, font = APP_FONT, bg = APP_BG, fg = APP_FG, bd = 0, anchor = E, state = DISABLED, cursor = "arrow", command = self.showPayrollImporter)
        BUTTON_PAYROLL.grid(column = 0, row = 6, pady = SUBMENU_PADY, sticky = E)
        SUB2_FRAME3.after(MORPH_DROPSPEED)
        SUB2_FRAME3.update()

        BUTTON_LOCKER = Button(SUB2_FRAME3, text = "Period Locker", width = 15, font = APP_FONT, bg = APP_BG, fg = APP_FG, bd = 0, anchor = E, state = DISABLED, cursor = "arrow", command = self.showPeriodLocker)
        BUTTON_LOCKER.grid(column = 0, row = 8, pady = SUBMENU_PADY, sticky = E)
        SUB2_FRAME3.after(MORPH_DROPSPEED)
        SUB2_FRAME3.update()

        BUTTON_REPORTS = Button(SUB2_FRAME3, text = "Reports", width = 15, font = APP_FONT, bg = APP_BG, fg = APP_FG, bd = 0, anchor = E, state = DISABLED, cursor = "arrow", command = self.showAccountingReports)
        BUTTON_REPORTS.grid(column = 0, row = 9, pady = SUBMENU_PADY, sticky = E)
        
        if self.returnUserName(USER, 3) == "ASD":
            for i in SUB2_FRAME3.winfo_children():
                i.config(state = NORMAL, cursor = "hand2")

    def showFinanceMenu(self, *args):
        self.morphMenuButton(BUTTON_FINANCE)
        SUB4_FRAME3.grid(column = 0, row = 3, ipadx = 50, sticky = E)
        BUTTON_DISBURSEMENTS = Button(SUB4_FRAME3, text = "Disbursements", width = 15, font = APP_FONT, bg = APP_BG, fg = APP_FG, bd = 0, anchor = E, state = DISABLED, cursor = "arrow", command = self.showDisbursements)
        BUTTON_DISBURSEMENTS.grid(column = 0, row = 0, pady = SUBMENU_PADY, sticky = E)
        SUB4_FRAME3.after(MORPH_DROPSPEED)
        SUB4_FRAME3.update()
        
        BUTTON_PAYABLE = Button(SUB4_FRAME3, text = "Accounts Payable", width = 15, font = APP_FONT, bg = APP_BG, fg = APP_FG, bd = 0, anchor = E, state = DISABLED, cursor = "arrow", command = self.showAccountsPayable)
        BUTTON_PAYABLE.grid(column = 0, row = 1, pady = SUBMENU_PADY, sticky = E)

        BUTTON_TEXTMASTER = Button(SUB4_FRAME3, text = "Text Master", width = 15, font = APP_FONT, bg = APP_BG, fg = APP_FG, bd = 0, anchor = E, state = DISABLED, cursor = "arrow", command = self.showTextMaster)
        BUTTON_TEXTMASTER.grid(column = 0, row = 2, pady = SUBMENU_PADY, sticky = E)
        SUB4_FRAME3.after(MORPH_DROPSPEED)
        SUB4_FRAME3.update()

        BUTTON_TRACKER = Button(SUB4_FRAME3, text = "Tracker", width = 15, font = APP_FONT, bg = APP_BG, fg = APP_FG, bd = 0, anchor = E, state = DISABLED, cursor = "arrow", command = None)
        # BUTTON_TRACKER.grid(column = 0, row = 2, pady = SUBMENU_PADY, sticky = E)
        SUB4_FRAME3.after(MORPH_DROPSPEED)
        SUB4_FRAME3.update()
        
        BUTTON_REPORTS = Button(SUB4_FRAME3, text = "Reports", width = 15, font = APP_FONT, bg = APP_BG, fg = APP_FG, bd = 0, anchor = E, state = DISABLED, cursor = "arrow", command = self.showFinanceReports)
        BUTTON_REPORTS.grid(column = 0, row = 4, pady = SUBMENU_PADY, sticky = E)
        
        if self.returnUserName(USER, 3) == "FMD" or self.returnUserName(USER, 4) == "administrator":
            for i in SUB4_FRAME3.winfo_children():
                i.config(state = NORMAL, cursor = "hand2")

    def showBCDMenu(self, *args):
        self.morphMenuButton(BUTTON_BCD)
        SUB6_FRAME3.grid(column = 0, row = 5, ipadx = 50, sticky = E)
        BUTTON_RECEIVABLES = Button(SUB6_FRAME3, text = "Receivables", width = 15, font = APP_FONT, bg = APP_BG, fg = APP_FG, bd = 0, anchor = E, state = DISABLED, cursor = "arrow", command = self.showReceivables)
        BUTTON_RECEIVABLES.grid(column = 0, row = 0, pady = SUBMENU_PADY, sticky = E)
        SUB6_FRAME3.after(MORPH_DROPSPEED)
        SUB6_FRAME3.update()

        BUTTON_COLLECTION = Button(SUB6_FRAME3, text = "Collection", width = 15, font = APP_FONT, bg = APP_BG, fg = APP_FG, bd = 0, anchor = E, state = DISABLED, cursor = "arrow", command = self.showCollections)
        BUTTON_COLLECTION.grid(column = 0, row = 1, pady = SUBMENU_PADY, sticky = E)
        SUB6_FRAME3.after(MORPH_DROPSPEED)
        SUB6_FRAME3.update()
        
        if self.returnUserName(USER, 3) == "BCD" or self.returnUserName(USER, 4) == "administrator":
            # for i in SUB6_FRAME3.winfo_children():
            #     i.config(state = NORMAL, cursor = "hand2")
            BUTTON_RECEIVABLES.config(state = NORMAL, cursor = "hand2")
        if self.returnUserName(USER, 4) == "administrator":
            BUTTON_COLLECTION.config(state = NORMAL, cursor = "hand2")

    def showGeneralServicesMenu(self, *args):
        self.morphMenuButton(BUTTON_GSAD)
        SUB8_FRAME3.grid(column = 0, row = 7, ipadx = 50, sticky = E)
        BUTTON_PAYMENT = Button(SUB8_FRAME3, text = "Payment Request", width = 15, font = APP_FONT, bg = APP_BG, fg = APP_FG, bd = 0, anchor = E, state = DISABLED, cursor = "arrow", command = None)
        # BUTTON_PAYMENT.grid(column = 0, row = 0, pady = SUBMENU_PADY, sticky = E)
        SUB8_FRAME3.after(MORPH_DROPSPEED)
        SUB8_FRAME3.update()

        BUTTON_ORDER = Button(SUB8_FRAME3, text = "Purchase Order", width = 15, font = APP_FONT, bg = APP_BG, fg = APP_FG, bd = 0, anchor = E, state = DISABLED, cursor = "arrow", command = self.showPurchaseOrder)
        BUTTON_ORDER.grid(column = 0, row = 1, pady = SUBMENU_PADY, sticky = E)
        SUB8_FRAME3.after(MORPH_DROPSPEED)
        SUB8_FRAME3.update()

        BUTTON_RECEIVE = Button(SUB8_FRAME3, text = "Receiving", width = 15, font = APP_FONT, bg = APP_BG, fg = APP_FG, bd = 0, anchor = E, state = DISABLED, cursor = "arrow", command = self.showReceivingReport)
        BUTTON_RECEIVE.grid(column = 0, row = 2, pady = SUBMENU_PADY, sticky = E)
        SUB8_FRAME3.after(MORPH_DROPSPEED)
        SUB8_FRAME3.update()

        BUTTON_ISSUANCE = Button(SUB8_FRAME3, text = "Issuance", width = 15, font = APP_FONT, bg = APP_BG, fg = APP_FG, bd = 0, anchor = E, state = DISABLED, cursor = "arrow", command = None)
        # BUTTON_ISSUANCE.grid(column = 0, row = 3, pady = SUBMENU_PADY, sticky = E)
        SUB8_FRAME3.after(MORPH_DROPSPEED)
        SUB8_FRAME3.update()

        BUTTON_INVENTORY = Button(SUB8_FRAME3, text = "Inventory", width = 15, font = APP_FONT, bg = APP_BG, fg = APP_FG, bd = 0, anchor = E, state = DISABLED, cursor = "arrow", command = self.showInventory)
        BUTTON_INVENTORY.grid(column = 0, row = 4, pady = SUBMENU_PADY, sticky = E)
        SUB8_FRAME3.after(MORPH_DROPSPEED)
        SUB8_FRAME3.update()

        BUTTON_REPORTS = Button(SUB8_FRAME3, text = "Reports", width = 15, font = APP_FONT, bg = APP_BG, fg = APP_FG, bd = 0, anchor = E, state = DISABLED, cursor = "arrow", command = None)
        # BUTTON_REPORTS.grid(column = 0, row = 5, pady = SUBMENU_PADY, sticky = E)
        
        if self.returnUserName(USER, 3) == "GSAD" or self.returnUserName(USER, 4) == "administrator":
            for i in SUB8_FRAME3.winfo_children():
                i.config(state = NORMAL, cursor = "hand2")

    def showAdministratorMenu(self, *args):
        self.morphMenuButton(BUTTON_ADMINISTRATOR)
        SUB10_FRAME3.grid(column = 0, row = 9, ipadx = 50, sticky = E)
        BUTTON_BANKS = Button(SUB10_FRAME3, text = "Banks", width = 15, font = APP_FONT, bg = APP_BG, fg = APP_FG, bd = 0, anchor = E, state = DISABLED, cursor = "arrow", command = self.showBanks)
        BUTTON_BANKS.grid(column = 0, row = 0, pady = SUBMENU_PADY, sticky = E)
        SUB10_FRAME3.after(MORPH_DROPSPEED)
        SUB10_FRAME3.update()

        BUTTON_CENTERS = Button(SUB10_FRAME3, text = "Cost Centers", width = 15, font = APP_FONT, bg = APP_BG, fg = APP_FG, bd = 0, anchor = E, state = DISABLED, cursor = "arrow", command = self.showCostCenters)
        BUTTON_CENTERS.grid(column = 0, row = 1, pady = SUBMENU_PADY, sticky = E)
        SUB10_FRAME3.after(MORPH_DROPSPEED)
        SUB10_FRAME3.update()

        BUTTON_CHART = Button(SUB10_FRAME3, text = "Chart of Accounts", width = 15, font = APP_FONT, bg = APP_BG, fg = APP_FG, bd = 0, anchor = E, state = DISABLED, cursor = "arrow", command = self.showChartofAccounts)
        BUTTON_CHART.grid(column = 0, row = 2, pady = SUBMENU_PADY, sticky = E)
        SUB10_FRAME3.after(MORPH_DROPSPEED)
        SUB10_FRAME3.update()

        BUTTON_CLIENTS = Button(SUB10_FRAME3, text = "Clients", width = 15, font = APP_FONT, bg = APP_BG, fg = APP_FG, bd = 0, anchor = E, state = DISABLED, cursor = "arrow", command = self.showClients)
        BUTTON_CLIENTS.grid(column = 0, row = 3, pady = SUBMENU_PADY, sticky = E)
        SUB10_FRAME3.after(MORPH_DROPSPEED)
        SUB10_FRAME3.update()

        BUTTON_LOCKER = Button(SUB10_FRAME3, text = "Period Locker", width = 15, font = APP_FONT, bg = APP_BG, fg = APP_FG, bd = 0, anchor = E, state = DISABLED, cursor = "arrow", command = self.clearWorkspace)
        BUTTON_LOCKER.grid(column = 0, row = 4, pady = SUBMENU_PADY, sticky = E)
        SUB10_FRAME3.after(MORPH_DROPSPEED)
        SUB10_FRAME3.update()

        BUTTON_SETTINGS = Button(SUB10_FRAME3, text = "Settings", width = 15, font = APP_FONT, bg = APP_BG, fg = APP_FG, bd = 0, anchor = E, state = DISABLED, cursor = "arrow", command = self.clearWorkspace)
        BUTTON_SETTINGS.grid(column = 0, row = 5, pady = SUBMENU_PADY, sticky = E)
        SUB10_FRAME3.after(MORPH_DROPSPEED)
        SUB10_FRAME3.update()

        BUTTON_SUPPLIERS = Button(SUB10_FRAME3, text = "Suppliers", width = 15, font = APP_FONT, bg = APP_BG, fg = APP_FG, bd = 0, anchor = E, state = DISABLED, cursor = "arrow", command = self.showSuppliers)
        BUTTON_SUPPLIERS.grid(column = 0, row = 6, pady = SUBMENU_PADY, sticky = E)
        SUB10_FRAME3.after(MORPH_DROPSPEED)
        SUB10_FRAME3.update()

        BUTTON_TAXES = Button(SUB10_FRAME3, text = "Tax Codes", width = 15, font = APP_FONT, bg = APP_BG, fg = APP_FG, bd = 0, anchor = E, state = DISABLED, cursor = "arrow", command = self.showTaxCodes)
        BUTTON_TAXES.grid(column = 0, row = 7, pady = SUBMENU_PADY, sticky = E)
        SUB10_FRAME3.after(MORPH_DROPSPEED)
        SUB10_FRAME3.update()

        BUTTON_TYPES = Button(SUB10_FRAME3, text = "Transactions Types", width = 15, font = APP_FONT, bg = APP_BG, fg = APP_FG, bd = 0, anchor = E, state = DISABLED, cursor = "arrow", command = self.showTransactionTypes)
        BUTTON_TYPES.grid(column = 0, row = 8, pady = SUBMENU_PADY, sticky = E)
        SUB10_FRAME3.after(MORPH_DROPSPEED)
        SUB10_FRAME3.update()

        BUTTON_USERS = Button(SUB10_FRAME3, text = "Users", width = 15, font = APP_FONT, bg = APP_BG, fg = APP_FG, bd = 0, anchor = E, state = DISABLED, cursor = "arrow", command = self.showUsers)
        BUTTON_USERS.grid(column = 0, row = 9, pady = SUBMENU_PADY, sticky = E)
        
        if self.returnUserName(USER, 3) == "ASD":
            for i in SUB10_FRAME3.winfo_children():
                i.config(state = NORMAL, cursor = "hand2")
        if self.returnUserName(USER, 3) == "GSAD":
            BUTTON_SUPPLIERS.config(state = NORMAL, cursor = "hand2")

### MENU_ACCOUNTING ###
    def showGeneralJournal(self, *args):
        self.clearWorkspace()
        FRAME_JOURNAL = LabelFrame(FRAME_4, text = "General Journal", font = APP_FONT)
        FRAME_JOURNAL.grid(column = 1, row = 0)

        global SUB1_JOURNAL
        SUB1_JOURNAL = Frame(FRAME_JOURNAL)
        SUB1_JOURNAL.grid(column = 0, row = 0, sticky = W, pady = SUBMENU_PADY)

        SUB2_JOURNAL = Frame(FRAME_JOURNAL)
        SUB2_JOURNAL.grid(column = 0, row = 1)

        LABEL_SEARCH = Label(SUB1_JOURNAL, text = "Search", width = 5, font = APP_FONT, anchor = W)
        LABEL_SEARCH.grid(column = 0, row = 0, pady = SUBMENU_PADY, sticky = W)

        TEXTVAR_SEARCH = StringVar()
        ENTRY_SEARCH = Entry(SUB1_JOURNAL, textvariable = TEXTVAR_SEARCH, width = 25, font = APP_FONT)
        ENTRY_SEARCH.grid(column = 1, row = 0, padx = MENU_PADX, sticky = W)
        ENTRY_SEARCH.bind("<Return>", lambda e: self.searchJournal(TEXTVAR_SEARCH.get()))

        BUTTON_GO = Button(SUB1_JOURNAL, text = "GO", width = 5, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = lambda: self.searchJournal(TEXTVAR_SEARCH.get()))
        BUTTON_GO.grid(column = 2, row = 0, padx = MENU_PADX)

        BUTTON_REFRESH = Button(SUB1_JOURNAL, text = "REFRESH", width = 10, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = self.refreshJournal)
        # BUTTON_REFRESH.grid(column = 3, row = 0, padx = MENU_PADX)

        BUTTON_ADD = Button(SUB1_JOURNAL, text = "ADD", width = 5, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = self.showAddEditJournal)
        BUTTON_ADD.grid(column = 4, row = 0, padx = MENU_PADX)
        
        if self.returnAccess(USER, 18) == 0:
            BUTTON_ADD.config(state = DISABLED, cursor = "arrow")
        
        global CALENDAR_START
        CALENDAR_START = DateEntry(SUB1_JOURNAL, firstweekday = "sunday", date_pattern = "yyyy-mm-dd", state = "readonly")
        CALENDAR_START.grid(column = 5, row = 0, sticky = W, padx = TOP_PADX + 10)
        # CALENDAR_START.set_date(self.returnFirstDayOfMonth(str(CALENDAR_START.get_date())))
        
        global CALENDAR_END
        CALENDAR_END = DateEntry(SUB1_JOURNAL, firstweekday = "sunday", date_pattern = "yyyy-mm-dd", state = "readonly")
        CALENDAR_END.grid(column = 6, row = 0, sticky = W, padx = TOP_PADX + 10)

        global PROGRESS_BAR
        PROGRESS_BAR = tk.Progressbar(SUB1_JOURNAL, orient = HORIZONTAL, length = 200, mode = "determinate")
        PROGRESS_BAR.grid(column = 7, row = 0, padx = TOP_PADX + 10, sticky = E)

        global TREE_JOURNAL
        TREE_JOURNAL = tk.Treeview(SUB2_JOURNAL, height = 28, selectmode = "browse")
        TREE_JOURNAL["columns"] = ("GJ No.", "Doc Date", "GL Date", "Particulars", "Source", "Reference", "Amount", "isPosted", "isVoid", "Encoder", "Encoded")
        TREE_JOURNAL.column("#0", width = 0, minwidth = 0, stretch = True)
        TREE_JOURNAL.column("GJ No.", anchor = W, minwidth = 75, width = 75)
        TREE_JOURNAL.column("Doc Date", anchor = W, minwidth = 75, width = 75)
        TREE_JOURNAL.column("GL Date", anchor = W, minwidth = 75, width = 75)
        TREE_JOURNAL.column("Particulars", anchor = W, minwidth = 165, width = 300)
        TREE_JOURNAL.column("Source", anchor = W, minwidth = 100, width = 75)
        TREE_JOURNAL.column("Reference", anchor = W, minwidth = 165, width = 100)
        TREE_JOURNAL.column("Amount", anchor = E, minwidth = 100, width = 100)
        TREE_JOURNAL.column("isPosted", anchor = W, minwidth = 75, width = 75)
        TREE_JOURNAL.column("isVoid", anchor = W, minwidth = 75, width = 75)
        TREE_JOURNAL.column("Encoder", anchor = W, minwidth = 75, width = 75)
        TREE_JOURNAL.column("Encoded", anchor = W, minwidth = 75, width = 75)
        
        TREE_JOURNAL.heading("#0", text = "", anchor = W)
        TREE_JOURNAL.heading("GJ No.", text = "GJ No.", anchor = N)
        TREE_JOURNAL.heading("Doc Date", text = "Doc Date", anchor = N)
        TREE_JOURNAL.heading("GL Date", text = "GL Date", anchor = N)
        TREE_JOURNAL.heading("Particulars", text = "Particulars", anchor = N)
        TREE_JOURNAL.heading("Source", text = "Source", anchor = N)
        TREE_JOURNAL.heading("Reference", text = "Reference", anchor = N)
        TREE_JOURNAL.heading("Amount", text = "Amount", anchor = N)
        TREE_JOURNAL.heading("isPosted", text = "isPosted", anchor = N)
        TREE_JOURNAL.heading("isVoid", text = "isVoid", anchor = N)
        TREE_JOURNAL.heading("Encoder", text = "Encoder", anchor = N)
        TREE_JOURNAL.heading("Encoded", text = "Encoded", anchor = N)

        global POPUP_JOURNAL
        POPUP_JOURNAL = Menu(TREE_JOURNAL, tearoff = 0)
        POPUP_JOURNAL.add_command(command = self.editJournal, label = "Edit", state = DISABLED)
        TREE_JOURNAL.bind("<Button-3>", lambda e: self.popupMenu(TREE_JOURNAL, POPUP_JOURNAL, e))
        TREE_JOURNAL.bind("<Double-1>", self.editJournal)

        global STYLE_JOURNAL
        STYLE_JOURNAL = tk.Style()
        STYLE_JOURNAL.map("Treeview", foreground = self.fixedMap("foreground", STYLE_JOURNAL), background = self.fixedMap("background", STYLE_JOURNAL))

        TREE_JOURNAL.tag_configure("oddrow", background = None)
        TREE_JOURNAL.tag_configure("evenrow", background = TREE_TAG_EVENROW)
        db.commit()
        cursor.execute(f"SELECT gjNumber, docDate, glDate, particulars, source, reference, isPosted, isVoid, encoder, DATE(encoded) FROM tblgeneraljournal WHERE docDate BETWEEN '{CALENDAR_START.get_date()}' AND '{CALENDAR_END.get_date()}' ORDER BY gjNumber DESC")
        result = cursor.fetchall()

        YSCROLLBAR = tk.Scrollbar(SUB2_JOURNAL, orient = "vertical", command = TREE_JOURNAL.yview)
        YSCROLLBAR.pack(side = "right", fill = "y")

        XSCROLLBAR = tk.Scrollbar(SUB2_JOURNAL, orient = "horizontal", command = TREE_JOURNAL.xview)
        
        TREE_JOURNAL.configure(yscrollcommand = YSCROLLBAR.set, xscrollcommand = XSCROLLBAR.set) 
        TREE_JOURNAL.pack()
        XSCROLLBAR.pack(fill ="x")

        count = 0
        gjnumbers, skipped = [], []
        for i in result:
            if i[0] not in gjnumbers:
                if count % 2 == 0:
                    TREE_JOURNAL.insert("", "end", values = (str(i[0]).zfill(8),i[1],i[2],i[3],i[4],i[5],self.validateAmount2(self.returnTotalJournalAmount(i[0])),i[6],i[7],self.returnUserName(i[8], 0),i[9]), tags = "evenrow")
                else:
                    TREE_JOURNAL.insert("", "end", values = (str(i[0]).zfill(8),i[1],i[2],i[3],i[4],i[5],self.validateAmount2(self.returnTotalJournalAmount(i[0])),i[6],i[7],self.returnUserName(i[8], 0),i[9]), tags = "oddrow")
                count += 1
                gjnumbers.append(i[0])
                PROGRESS_BAR["value"] = round((count/(len(result)-len(skipped)))*100, 0)
                SUB1_JOURNAL.update()
            else:
                skipped.append(i[0])
        PROGRESS_BAR.grid_remove()

        try:
            TOP_JOURNAL.destroy()
        except:
            pass

    def showPostMaster(self, *args):
        self.clearWorkspace()
        FRAME_POST = LabelFrame(FRAME_4, text = "Post Master", font = APP_FONT)
        FRAME_POST.grid(column = 1, row = 0)
        
        global SUB_FRAME1
        SUB_FRAME1 = Frame(FRAME_POST)
        SUB_FRAME1.grid(column = 0, row = 1)
        
        LABEL_BOOK = Label(SUB_FRAME1, text = "Book", width = 10, font = APP_FONT, anchor = W)
        LABEL_BOOK.grid(column = 0, row = 0, pady = SUBMENU_PADY, sticky = W)
        
        global TEXTVAR_BOOK
        TEXTVAR_BOOK = StringVar()
        COMBO_BOOK = tk.Combobox(SUB_FRAME1, values = ["General Journal", "Purchase Book", "Cash Disbursement Book", "Sales Book", "Cash Receipt Book"], textvariable = TEXTVAR_BOOK, font = APP_FONT, width = 30, state = "readonly")
        COMBO_BOOK.grid(column = 1, row = 0, sticky = W, ipadx = 2)
        
        LABEL_FROM = Label(SUB_FRAME1, text = "From", width = 10, font = APP_FONT, anchor = W)
        LABEL_FROM.grid(column = 0, row = 1, pady = SUBMENU_PADY, sticky = W)
        
        LABEL_TO = Label(SUB_FRAME1, text = "To", width = 10, font = APP_FONT, anchor = W)
        LABEL_TO.grid(column = 0, row = 2, pady = SUBMENU_PADY, sticky = W)
        
        global CALENDAR_FROM
        CALENDAR_FROM = DateEntry(SUB_FRAME1, firstweekday = "sunday", date_pattern = "yyyy-mm-dd", state = "readonly")
        CALENDAR_FROM.grid(column = 1, row = 1, sticky = W, pady = TOP_PADY)
        CALENDAR_FROM.set_date(self.returnFirstDayOfMonth(str(CALENDAR_FROM.get_date())))
        
        global CALENDAR_TO
        CALENDAR_TO = DateEntry(SUB_FRAME1, firstweekday = "sunday", date_pattern = "yyyy-mm-dd", state = "readonly")
        CALENDAR_TO.grid(column = 1, row = 2, sticky = W, pady = TOP_PADY)
        CALENDAR_TO.set_date(self.returnLastDayOfMonth(str(CALENDAR_TO.get_date())))
        
        BUTTON_POST = Button(SUB_FRAME1, text = "POST", width = 5, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, state = DISABLED, cursor = "hand2", command = self.postBook)
        BUTTON_POST.grid(column = 1, row = 4, pady = MENU_PADY)

        LABEL_PROGRESS = Label(SUB_FRAME1, text = "Progress", width = 10, font = APP_FONT, anchor = W)
        LABEL_PROGRESS.grid(column = 0, row = 5, pady = SUBMENU_PADY, sticky = W)

        global PROGRESS_BAR
        PROGRESS_BAR = tk.Progressbar(SUB_FRAME1, orient = HORIZONTAL, length = 500, mode = "determinate")
        PROGRESS_BAR.grid(column = 1, row = 5)
        
        if self.returnUserName(USER, 4) == "administrator":
            BUTTON_POST.config(state = NORMAL, cursor = "hand2")

    def showPayrollImporter(self, *args):
        self.clearWorkspace()
        FRAME_PAYROLL = LabelFrame(FRAME_4, text = "Payroll Importer", font = APP_FONT)
        FRAME_PAYROLL.grid(column = 1, row = 0)

        SUB1_PAYROLL = Frame(FRAME_PAYROLL)
        SUB1_PAYROLL.grid(column = 0, row = 0, sticky = W, pady = SUBMENU_PADY)

        SUB2_PAYROLL = Frame(FRAME_PAYROLL)
        SUB2_PAYROLL.grid(column = 0, row = 1)

        global SUB3_PAYROLL
        SUB3_PAYROLL = Frame(FRAME_PAYROLL)
        SUB3_PAYROLL.grid(column = 0, row = 2)

        LABEL_FOLDER = Label(SUB1_PAYROLL, text = "Folder", font = APP_FONT)
        LABEL_FOLDER.grid(column = 0, row = 0, pady = TOP_PADY)

        LABEL_DOC = Label(SUB1_PAYROLL, text = "Doc Date", font = APP_FONT)
        LABEL_DOC.grid(column = 0, row = 1, pady = TOP_PADY, sticky = E)

        LABEL_GL = Label(SUB1_PAYROLL, text = "GL Date", font = APP_FONT)
        LABEL_GL.grid(column = 0, row = 2, pady = TOP_PADY, sticky = E)
        
        global TEXTVAR_FOLDER
        TEXTVAR_FOLDER = StringVar()
        ENTRY_FOLDER = Entry(SUB1_PAYROLL, textvariable = TEXTVAR_FOLDER, font = APP_FONT, width = 40, state = "readonly")
        ENTRY_FOLDER.grid(column = 1, row = 0, pady = TOP_PADY)

        BUTTON_FOLDER = Button(SUB1_PAYROLL, text = "...", font = APP_FONT, command = lambda: self.getFolderAddress(TEXTVAR_FOLDER)) #self.getFolderAddress(TEXTVAR_FOLDER)) #self.getFileAddress2("Excel files", "*.xlsx")) 
        BUTTON_FOLDER.grid(column = 1, row = 0, sticky = E, pady = TOP_PADY)

        global CALENDAR_DOC
        CALENDAR_DOC = DateEntry(SUB1_PAYROLL, firstweekday = "sunday", date_pattern = "yyyy-mm-dd", state = "readonly")
        CALENDAR_DOC.grid(column = 1, row = 1, sticky = W)
        
        global CALENDAR_GL, TEXTVAR_GL
        TEXTVAR_GL= StringVar()
        CALENDAR_GL = DateEntry(SUB1_PAYROLL, textvariable = TEXTVAR_GL, firstweekday = "sunday", date_pattern = "yyyy-mm-dd", state = "readonly")
        CALENDAR_GL.grid(column = 1, row = 2, sticky = W)
        CALENDAR_GL.set_date(self.returnLastDayOfMonth(str(datetime.datetime.today())))
        CALENDAR_GL.bind("<FocusOut>", lambda e: self.formatDate(CALENDAR_GL, TEXTVAR_GL))

        BUTTON_IMPORT = Button(SUB2_PAYROLL, text = "Import", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, command = self.importPayrollSummary)
        BUTTON_IMPORT.grid(column = 0, row = 0, sticky = E, pady = TOP_PADY)

        global PROGRESS_BAR
        PROGRESS_BAR = tk.Progressbar(SUB3_PAYROLL, orient = HORIZONTAL, length = 200, mode = "determinate")

    def showPeriodLocker(self, *args):
        self.clearWorkspace()
        FRAME_LOCKER = LabelFrame(FRAME_4, text = "Period Locker", font = APP_FONT)
        FRAME_LOCKER.grid(column = 1, row = 0)

        SUB1_LOCKER = Frame(FRAME_LOCKER)
        SUB1_LOCKER.grid(column = 0, row = 0, sticky = W, pady = SUBMENU_PADY)

        SUB2_LOCKER = Frame(FRAME_LOCKER)
        SUB2_LOCKER.grid(column = 0, row = 1)
        
        global CALENDAR_START
        CALENDAR_START = DateEntry(SUB1_LOCKER, firstweekday = "sunday", date_pattern = "yyyy-mm-dd", state = "readonly")
        CALENDAR_START.grid(column = 0, row = 0, sticky = W, padx = TOP_PADX + 10)
        CALENDAR_START.set_date(self.returnFirstDayOfMonth(str(CALENDAR_START.get_date())))
        
        BUTTON_GO = Button(SUB1_LOCKER, text = "GO", width = 5, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = None)
        BUTTON_GO.grid(column = 2, row = 0, padx = MENU_PADX)
        
        global TREE_LOCKER
        TREE_LOCKER = tk.Treeview(SUB2_LOCKER, height = 28, selectmode = "browse")
        TREE_LOCKER["columns"] = ("Year", "Month", "Status")
        TREE_LOCKER.column("#0", width = 0, minwidth = 0, stretch = True)
        TREE_LOCKER.column("Year", anchor = W, minwidth = 75, width = 75)
        TREE_LOCKER.column("Month", anchor = W, minwidth = 75, width = 75)
        TREE_LOCKER.column("Status", anchor = W, minwidth = 75, width = 75)
        
        TREE_LOCKER.heading("#0", text = "", anchor = W)
        TREE_LOCKER.heading("Year", text = "Year", anchor = N)
        TREE_LOCKER.heading("Month", text = "Month", anchor = N)
        TREE_LOCKER.heading("Status", text = "Status", anchor = N)

        global POPUP_LOCKER
        POPUP_LOCKER = Menu(TREE_LOCKER, tearoff = 0)
        POPUP_LOCKER.add_command(command = self.editLocker, label = "Edit")
        TREE_LOCKER.bind("<Button-3>", lambda e: self.popupMenu(TREE_LOCKER, POPUP_LOCKER, e))
        TREE_LOCKER.bind("<Double-1>", self.editLocker)

        global STYLE_LOCKER
        STYLE_LOCKER = tk.Style()
        STYLE_LOCKER.map("Treeview", foreground = self.fixedMap("foreground", STYLE_LOCKER), background = self.fixedMap("background", STYLE_LOCKER))

        TREE_LOCKER.tag_configure("oddrow", background = None)
        TREE_LOCKER.tag_configure("evenrow", background = TREE_TAG_EVENROW)
        
        yearvar = int(str(CALENDAR_START.get_date()).split("-")[0])
        db.commit()
        cursor.execute(f"SELECT year, month, status FROM tblperiodlocker WHERE year = {yearvar} ORDER BY month")
        result = cursor.fetchall()

        count = 0
        for i in result:
            if count % 2 == 0:
                TREE_LOCKER.insert("", "end", values = (i[0],i[1],i[2]), tags = "evenrow")
            else:
                TREE_LOCKER.insert("", "end", values = (i[0],i[1],i[2]), tags = "oddrow")
            count += 1

        YSCROLLBAR = tk.Scrollbar(SUB2_LOCKER, orient = "vertical", command = TREE_LOCKER.yview)
        YSCROLLBAR.pack(side = "right", fill = "y")

        XSCROLLBAR = tk.Scrollbar(SUB2_LOCKER, orient = "horizontal", command = TREE_LOCKER.xview)
        
        TREE_LOCKER.configure(yscrollcommand = YSCROLLBAR.set, xscrollcommand = XSCROLLBAR.set) 
        TREE_LOCKER.pack()
        XSCROLLBAR.pack(fill ="x")

        try:
            TOP_LOCKER.destroy()
        except:
            pass

    def showAccountingReports(self, *args):
        self.clearWorkspace()
        global FRAME_ACCOUNTING
        FRAME_ACCOUNTING = LabelFrame(FRAME_4, text = "Reports", font = APP_FONT)
        FRAME_ACCOUNTING.grid(column = 1, row = 0)
        
        global SUB_FRAME1
        SUB_FRAME1 = Frame(FRAME_ACCOUNTING)
        SUB_FRAME1.grid(column = 0, row = 5)
        
        global SUB_FRAME2
        SUB_FRAME2 = Frame(FRAME_ACCOUNTING) #treeview
        SUB_FRAME2.grid(column = 0, row = 6)
        
        LABEL_REPORT = Label(SUB_FRAME1, text = "Report Type", font = APP_FONT, anchor = W)
        LABEL_REPORT.grid(column = 0, row = 0, pady = SUBMENU_PADY, sticky = W)

        LABEL_STATUS = Label(SUB_FRAME1, text = "Posting Status", font = APP_FONT, anchor = W)
        LABEL_STATUS.grid(column = 0, row = 1, pady = SUBMENU_PADY, sticky = W)

        global LABEL_BOOK
        LABEL_BOOK = Label(SUB_FRAME1, text = "Book Type", font = APP_FONT, anchor = W)
        global LABEL_CODE
        LABEL_CODE = Label(SUB_FRAME1, text = "Chart Code", font = APP_FONT, anchor = W)
        global LABEL_TITLE
        LABEL_TITLE = Label(SUB_FRAME1, text = "Chart Title", font = APP_FONT, anchor = W)
        
        global TEXTVAR_REPORTS
        TEXTVAR_REPORTS = StringVar()
        COMBO_REPORTS = tk.Combobox(SUB_FRAME1, values = ["General Ledger", "Trial Balance", "Income Statement", "Statement of Financial Position", "End of Month", "Account Analysis", "Books"], textvariable = TEXTVAR_REPORTS, font = APP_FONT, width = 30, state = "readonly")
        COMBO_REPORTS.grid(column = 1, row = 0, sticky = W, ipadx = 2, pady = TOP_PADY)
        COMBO_REPORTS.bind("<<ComboboxSelected>>", self.showSelectedAccountingReport)
        
        global TEXTVAR_OPTION, COMBO_OPTION
        TEXTVAR_OPTION = StringVar()
        COMBO_OPTION = tk.Combobox(SUB_FRAME1, values = ["Posted", "Posted and Unposted"], textvariable = TEXTVAR_OPTION, font = APP_FONT, width = 30, state = "readonly")
        COMBO_OPTION.grid(column = 1, row = 1, sticky = W, ipadx = 2)
        TEXTVAR_OPTION.set("Posted")
        
        entryCode, entryTitle = [], []
        global TEXTVAR_CHARTCODE, ENTRY_CHARTCODE
        TEXTVAR_CHARTCODE = StringVar()
        ENTRY_CHARTCODE = Entry(SUB_FRAME1, textvariable = TEXTVAR_CHARTCODE, font = APP_FONT, width = 30)
        ENTRY_CHARTCODE.bind("<FocusOut>", lambda e: self.populateChartFields(entryCode[0], entryCode[0]))
        entryCode.append(TEXTVAR_CHARTCODE)
        
        global BUTTON_CHARTCODE
        BUTTON_CHARTCODE = Button(SUB_FRAME1, text = "...", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, cursor = "hand2", command = lambda: self.showChartSelection(0, entryCode, entryTitle))
        
        global TEXTVAR_BOOKS, COMBO_BOOKS
        TEXTVAR_BOOKS = StringVar()
        COMBO_BOOKS = tk.Combobox(SUB_FRAME1, values = ["GJ", "CDB", "PB", "SB"], textvariable = TEXTVAR_BOOKS, font = APP_FONT, width = 30, state = "readonly")

        global TEXTVAR_CHARTTITLE, ENTRY_CHARTTITLE
        TEXTVAR_CHARTTITLE = StringVar()
        ENTRY_CHARTTITLE = Entry(SUB_FRAME1, textvariable = TEXTVAR_CHARTTITLE, font = APP_FONT, width = 30, state = "readonly")
        entryTitle.append(TEXTVAR_CHARTTITLE)

        LABEL_FROM = Label(SUB_FRAME1, text = "From", width = 10, font = APP_FONT, anchor = W)
        LABEL_FROM.grid(column = 0, row = 5, pady = SUBMENU_PADY, sticky = W)
        
        LABEL_TO = Label(SUB_FRAME1, text = "To", width = 10, font = APP_FONT, anchor = W)
        LABEL_TO.grid(column = 0, row = 6, pady = SUBMENU_PADY, sticky = W)
        
        global CALENDAR_FROM
        CALENDAR_FROM = DateEntry(SUB_FRAME1, firstweekday = "sunday", date_pattern = "yyyy-mm-dd", state = "readonly")
        CALENDAR_FROM.grid(column = 1, row = 5, sticky = W, pady = TOP_PADY)
        CALENDAR_FROM.set_date(self.returnFirstDayOfMonth(str(CALENDAR_FROM.get_date())))
        
        global CALENDAR_TO
        CALENDAR_TO = DateEntry(SUB_FRAME1, firstweekday = "sunday", date_pattern = "yyyy-mm-dd", state = "readonly")
        CALENDAR_TO.grid(column = 1, row = 6, sticky = W, pady = TOP_PADY)
        
        global BUTTON_EXPORT
        BUTTON_EXPORT = Button(SUB_FRAME1, text = "EXPORT", width = 5, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "arrow", state = DISABLED, command = self.exportAccountingReport)
        BUTTON_EXPORT.grid(column = 1, row = 7, pady = MENU_PADY, sticky = W)

        global LABEL_PROGRESS
        LABEL_PROGRESS = Label(SUB_FRAME1, text = "Progress", font = APP_FONT, anchor = W)
        LABEL_PROGRESS.grid(column = 0, row = 8, sticky = W)

        global PROGRESS_BAR
        PROGRESS_BAR = tk.Progressbar(SUB_FRAME1, orient = HORIZONTAL, length = 400, mode = "determinate")
        PROGRESS_BAR.grid(column = 1, row = 8)

### MENU_ACCOUNTING_JOURNAL ###
    def showAddEditJournal(self, *args):
        global TOP_JOURNAL
        TOP_JOURNAL = Toplevel()
        TOP_JOURNAL.title("Create - General Journal")
        TOP_JOURNAL.iconbitmap(PATH_ICON + "icon.ico")
        TOP_JOURNAL.geometry("970x500+100+20")
        TOP_JOURNAL.resizable(height = False, width = False)
        TOP_JOURNAL.grab_set()
        TOP_JOURNAL.focus()
        TOP_JOURNAL.protocol("WM_DELETE_WINDOW", lambda: self.closeTopLevel(TOP_JOURNAL))

        SUB_FRAME1 = Frame(TOP_JOURNAL)
        SUB_FRAME1.grid(column = 0, row = 0, sticky = W, pady = TOP_PADY)
        
        SUB_FRAME2 = Frame(TOP_JOURNAL) #header
        SUB_FRAME2.grid(column = 0, row = 1, sticky = W)
        
        SUB_FRAME3 = Frame(TOP_JOURNAL) #scroll
        SUB_FRAME3.grid(column = 0, row = 2, sticky = W)
        
        SUB_FRAME4 = Frame(TOP_JOURNAL) #buttons
        SUB_FRAME4.grid(column = 0, row = 3, sticky = E, pady = TOP_PADY + 10)
        
        LABEL_DESC = Label(SUB_FRAME1, text = "Particulars", font = APP_FONT)
        LABEL_DESC.grid(column = 0, row = 0, pady = TOP_PADY, sticky = NE)

        LABEL_REF = Label(SUB_FRAME1, text = "Reference", font = APP_FONT)
        LABEL_REF.grid(column = 0, row = 1, pady = TOP_PADY, sticky = E)

        LABEL_APV = Label(SUB_FRAME1, text = "GJ No.", font = APP_FONT)
        LABEL_APV.grid(column = 2, row = 0, pady = TOP_PADY, sticky = E)
        
        LABEL_DOC = Label(SUB_FRAME1, text = "Doc Date", font = APP_FONT)
        LABEL_DOC.grid(column = 2, row = 1, pady = TOP_PADY, sticky = E)

        LABEL_GL = Label(SUB_FRAME1, text = "GL Date", font = APP_FONT)
        LABEL_GL.grid(column = 2, row = 2, pady = TOP_PADY, sticky = E)
        
        global TEXTVAR_PARTICULARS, ENTRY_PARTICULARS
        TEXTVAR_PARTICULARS = StringVar()
        ENTRY_PARTICULARS = Entry(SUB_FRAME1, textvariable = TEXTVAR_PARTICULARS, font = APP_FONT, width = 60)
        ENTRY_PARTICULARS.grid(column = 1, row = 0, sticky = W, padx = TOP_PADX + 20)
        
        global TEXTVAR_REFERENCE, ENTRY_REFERENCE
        TEXTVAR_REFERENCE = StringVar()
        ENTRY_REFERENCE = Entry(SUB_FRAME1, textvariable = TEXTVAR_REFERENCE, font = APP_FONT, width = 30)
        ENTRY_REFERENCE.grid(column = 1, row = 1, sticky = W, padx = TOP_PADX + 20)
        
        global TEXTVAR_GJ
        TEXTVAR_GJ = StringVar()
        ENTRY_GJ = Entry(SUB_FRAME1, textvariable = TEXTVAR_GJ, font = APP_FONT, width = 15, justify = RIGHT, state = "readonly")
        ENTRY_GJ.grid(column = 3, row = 0, sticky = W)
        
        global CALENDAR_DOC
        CALENDAR_DOC = DateEntry(SUB_FRAME1, firstweekday = "sunday", date_pattern = "yyyy-mm-dd", state = "readonly")
        CALENDAR_DOC.grid(column = 3, row = 1, sticky = W)
        
        global CALENDAR_GL, TEXTVAR_GL
        TEXTVAR_GL= StringVar()
        CALENDAR_GL = DateEntry(SUB_FRAME1, textvariable = TEXTVAR_GL, firstweekday = "sunday", date_pattern = "yyyy-mm-dd", state = "readonly")
        CALENDAR_GL.grid(column = 3, row = 2, sticky = W)
        CALENDAR_GL.set_date(self.returnLastDayOfMonth(str(datetime.datetime.today())))
        CALENDAR_GL.bind("<FocusOut>", lambda e: self.formatDate(CALENDAR_GL, TEXTVAR_GL))
        
        LABEL_CODE = Label(SUB_FRAME2, text = "Code", width = 12, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_CODE.grid(column = 0, row = 0)

        LABEL_TITLE = Label(SUB_FRAME2, text = "Title", width = 30, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_TITLE.grid(column = 1, row = 0)

        LABEL_CENTER = Label(SUB_FRAME2, text = "Amount", width = 15, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_CENTER.grid(column = 2, row = 0)

        LABEL_GROSS = Label(SUB_FRAME2, text = "Dr/Cr", width = 10, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_GROSS.grid(column = 3, row = 0)

        LABEL_EXPENSE = Label(SUB_FRAME2, text = "Remarks", width = 31, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_EXPENSE.grid(column = 4, row = 0)

        LABEL_X = Label(SUB_FRAME2, text = "[x]", width = 2, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_X.grid(column = 5, row = 0)
        
        self.createScrollFrame(SUB_FRAME3, 350, 740, 20, 0, 0)
        
        global entryFrame, entryCode, entryTitle, entryAmount, entryDrCr, entryRemarks, entryClear
        entryFrame, entryCode, entryTitle, entryAmount, entryDrCr, entryRemarks, entryClear = [], [], [], [], [], [], []
        
        for i in range(40):
            self.showSOAEntryLines(SCROLLABLE_FRAME, i)

        SUB_FRAME7 = Frame(SUB_FRAME3) #entry total frame
        SUB_FRAME7.grid(column = 1, row = 0, sticky = NW, padx = TOP_PADX + 10)
        
        LABEL_DEBIT = Label(SUB_FRAME7, text = "Debit", width = 12, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_DEBIT.grid(column = 0, row = 0)

        LABEL_CREDIT = Label(SUB_FRAME7, text = "Credit", width = 12, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_CREDIT.grid(column = 0, row = 1)

        LABEL_VARIANCE = Label(SUB_FRAME7, text = "Variance", width = 12, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_VARIANCE.grid(column = 0, row = 2)
        
        global TEXTVAR_DEBIT
        TEXTVAR_DEBIT = StringVar()
        ENTRY_DEBIT = Entry(SUB_FRAME7, textvariable = TEXTVAR_DEBIT, font = APP_FONT, width = 12, justify = RIGHT, state = "readonly")
        ENTRY_DEBIT.grid(column = 1, row = 0, sticky = W)
        
        global TEXTVAR_CREDIT
        TEXTVAR_CREDIT = StringVar()
        ENTRY_CREDIT = Entry(SUB_FRAME7, textvariable = TEXTVAR_CREDIT, font = APP_FONT, width = 12, justify = RIGHT, state = "readonly")
        ENTRY_CREDIT.grid(column = 1, row = 1, sticky = W)
        
        global TEXTVAR_VARIANCE
        TEXTVAR_VARIANCE = StringVar()
        ENTRY_VARIANCE = Entry(SUB_FRAME7, textvariable = TEXTVAR_VARIANCE, font = APP_FONT, width = 12, justify = RIGHT, state = "readonly")
        ENTRY_VARIANCE.grid(column = 1, row = 2, sticky = W)
        
        global BUTTON_SUBMIT
        BUTTON_SUBMIT = Button(SUB_FRAME4, text = "SUBMIT", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "hand2", state = NORMAL, command = self.saveJournal)
        BUTTON_SUBMIT.grid(column = 1, row = 0, padx = TOP_PADX)

        global BUTTON_VOID
        BUTTON_VOID = Button(SUB_FRAME4, text = "VOID", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG_RED, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "arrow", state = DISABLED, command = self.voidJournal)
        BUTTON_VOID.grid(column = 3, row = 0, padx = TOP_PADX)

        global BUTTON_PRINT
        BUTTON_PRINT = Button(SUB_FRAME4, text = "PRINT", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "arrow", state = DISABLED, command = self.printJournal)
        BUTTON_PRINT.grid(column = 4, row = 0, padx = TOP_PADX)

        BUTTON_CLOSE = Button(SUB_FRAME4, text = "CLOSE", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "hand2", command = lambda: self.closeTopLevel(TOP_JOURNAL))
        BUTTON_CLOSE.grid(column = 5, row = 0, padx = TOP_PADX)
    
    def saveJournal(self, *args):
        insert = """INSERT INTO tblgeneraljournal (
            gjNumber, docDate, glDate, particulars, reference,
            source, chartcode, amount, side, remarks,
            isPosted, isVoid, encoder, encoded) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
            
        delete = "DELETE FROM tblgeneraljournal WHERE gjNumber = %s"

        if self.returnPeriodStatus(CALENDAR_GL.get_date()) == "open":
            ask = messagebox.askyesno("General Journal", "Are you sure?")
            if ask:
                if TEXTVAR_GJ.get() != "":
                    cursor.execute(delete, [int(TEXTVAR_GJ.get())])
                    db.commit()
                    num = TEXTVAR_GJ.get()
                else:
                    num = self.generateGJNumber()
                if TEXTVAR_PARTICULARS.get() != "" and self.returnFloatAmount(TEXTVAR_VARIANCE.get()) == 0:
                    validlines = []
                    for i in range(len(entryFrame)):
                        if entryCode[i].get() != "":
                            validlines.append([
                                int(num), CALENDAR_DOC.get_date(), CALENDAR_GL.get_date(), TEXTVAR_PARTICULARS.get(), TEXTVAR_REFERENCE.get(),
                                "GJ", int(entryCode[i].get()), self.returnFloatAmount(entryAmount[i].get()), entryDrCr[i].get(), entryRemarks[i].get(),
                                "No", "No", USER, datetime.datetime.now()
                            ])
                
                    if len(validlines) > 0:
                        for i in validlines:
                            cursor.execute(insert, i)
                        db.commit()
                        messagebox.showinfo("General Journal", "General Journal has been saved!")
                        TEXTVAR_GJ.set(num)
                        self.disableJournalWidgets()
                        BUTTON_SUBMIT.config(state = DISABLED, cursor = "arrow")
                        BUTTON_VOID.config(state = NORMAL, cursor = "hand2")
                        BUTTON_PRINT.config(state = NORMAL, cursor = "hand2")
                        TOP_JOURNAL.focus()
                    else:
                        messagebox.showerror("General Journal", "Zero valid lines!")
                        TOP_JOURNAL.focus()
                else:
                    messagebox.showerror("General Journal", "Check particulars if empty or entry amounts!")
                    TOP_JOURNAL.focus()
        else:
            messagebox.showerror("General Journal", "GL Date is locked!")
            TOP_JOURNAL.focus()
    
    def voidJournal(self):
        insertbook = """INSERT INTO tblgeneraljournal (
            gjNumber, docDate, glDate, particulars, reference,
            source, chartcode, amount, side, remarks,
            isPosted, isVoid, encoder, encoded) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
        
        validbook = []
        for i in range(len(entryFrame)):
            if entryCode[i].get() != "" and self.returnFloatAmount(entryAmount[i].get()) > 0:
                if entryDrCr[i].get() == "Debit":
                    side = "Credit"
                else:
                    side = "Debit"
                validbook.append([
                    int(self.generateGJNumber()), datetime.date.today(), self.returnLastDayOfMonth(str(datetime.date.today())), "REVERSAL OF GJ#" + TEXTVAR_GJ.get() + " " + TEXTVAR_PARTICULARS.get(), TEXTVAR_GJ.get(),
                    "GJ", int(entryCode[i].get()), self.returnFloatAmount(entryAmount[i].get()), side, entryRemarks[i].get(),
                    "No", "No", USER, datetime.datetime.now()
                ])
                
        voidGJ = "UPDATE tblgeneraljournal SET isVoid = %s WHERE gjNumber = %s"
        ask = messagebox.askyesno("VOID General Journal", "Are you sure?")
        if ask:
            cursor.execute(voidGJ, ["Yes", int(TEXTVAR_GJ.get())])
            db.commit()
            
            if len(validbook) != 0:
                for i in validbook:
                    cursor.execute(insertbook, i)
                db.commit()

            messagebox.showinfo("VOID General Journal", "GJ has been voided!")
            TOP_JOURNAL.focus()
            self.disableJournalWidgets()
            BUTTON_SUBMIT.config(state = DISABLED, cursor = "arrow")
            BUTTON_VOID.config(state = DISABLED, cursor = "arrow")
            BUTTON_PRINT.config(state = NORMAL, cursor = "hand2")
    
    def printJournal(self):
        pass
    
    def editJournal(self, *args):
        self.copySelection(TREE_JOURNAL)
        self.showAddEditJournal()
        TOP_JOURNAL.title("Edit - GJ")
        
        select = """SELECT 
            gjNumber, docDate, glDate, particulars, reference,
            source, chartcode, amount, side, remarks,
            isPosted, isVoid
                FROM tblgeneraljournal WHERE gjNUmber = %s"""
        
        db.commit()
        cursor.execute(select, [int(content[0])])
        result = cursor.fetchall()
        if result:
            TEXTVAR_GJ.set(str(result[0][0]).zfill(8))
            TEXTVAR_PARTICULARS.set(result[0][3])
            TEXTVAR_REFERENCE.set(result[0][4])
            CALENDAR_GL.set_date(result[0][2])
            CALENDAR_DOC.set_date(result[0][1])
        
            count = 0
            for i in result:
                entryCode[count].set(i[6])
                self.populateChartFields(entryCode[count], entryTitle[count])
                entryAmount[count].set(self.validateAmount2(i[7]))
                entryDrCr[count].set(i[8])
                entryRemarks[count].set(i[9])
                count += 1
            self.updateSOAEntriesTotals()
            
            if self.returnJournalStatus(int(content[0]), 0) == "Yes" and self.returnJournalStatus(int(content[0]), 1) == "No":
                self.disableJournalWidgets()
                BUTTON_SUBMIT.config(state = DISABLED, cursor = "arrow")
                BUTTON_VOID.config(state = NORMAL, cursor = "hand2")
            if self.returnJournalStatus(int(content[0]), 0) == "Yes" and self.returnJournalStatus(int(content[0]), 1) == "Yes":
                self.disableJournalWidgets()
                BUTTON_SUBMIT.config(state = DISABLED, cursor = "arrow")
            BUTTON_PRINT.config(state = NORMAL, cursor = "hand2")
    
    def searchJournal(self, var, *args):
        db.commit()
        find = "SELECT gjNumber, docDate, glDate, particulars, source, reference, isPosted, isVoid, encoder, DATE(encoded) FROM tblgeneraljournal WHERE (gjNumber LIKE %s OR reference LIKE %s OR particulars LIKE %s) AND docDate BETWEEN %s AND %s ORDER BY gjNumber DESC"
        cursor.execute(find, [f"%{var}%", f"%{var}%", f"%{var}%", CALENDAR_START.get_date(), CALENDAR_END.get_date()])
        result = cursor.fetchall()
        for i in TREE_JOURNAL.get_children():
            TREE_JOURNAL.delete(i)
        if result:
            PROGRESS_BAR.grid(column = 7, row = 0, padx = TOP_PADX + 10, sticky = E)
            count = 0
            gjnumbers, skipped = [], []
            for i in result:
                if i[0] not in gjnumbers:
                    if count % 2 == 0:
                        TREE_JOURNAL.insert("", "end", values = (str(i[0]).zfill(8),i[1],i[2],i[3],i[4],i[5],self.validateAmount2(self.returnTotalJournalAmount(i[0])),i[6],i[7],self.returnUserName(i[8], 0),i[9]), tags = "evenrow")
                    else:
                        TREE_JOURNAL.insert("", "end", values = (str(i[0]).zfill(8),i[1],i[2],i[3],i[4],i[5],self.validateAmount2(self.returnTotalJournalAmount(i[0])),i[6],i[7],self.returnUserName(i[8], 0),i[9]), tags = "oddrow")
                    count += 1
                    gjnumbers.append(i[0])
                    PROGRESS_BAR["value"] = round((count/(len(result)-len(skipped)))*100, 0)
                    SUB1_JOURNAL.update()
                else:
                    skipped.append(i[0])
            PROGRESS_BAR.grid_remove()
        else:
            messagebox.showerror("General Journal", "No record found!")

    def refreshJournal(self, *args):
        for i in TREE_JOURNAL.get_children():
            TREE_JOURNAL.delete(i)
        db.commit()
        cursor.execute(f"SELECT gjNumber, docDate, glDate, particulars, source, reference, isPosted, isVoid, encoder, DATE(encoded) FROM tblgeneraljournal WHERE docDate BETWEEN '{CALENDAR_START.get_date()}' AND '{CALENDAR_END.get_date()}' ORDER BY gjNumber DESC")
        result = cursor.fetchall()
        PROGRESS_BAR.grid(column = 7, row = 0, padx = TOP_PADX + 10, sticky = E)
        count = 0
        gjnumbers, skipped = [], []
        for i in result:
            if i[0] not in gjnumbers:
                if count % 2 == 0:
                    TREE_JOURNAL.insert("", "end", values = (str(i[0]).zfill(8),i[1],i[2],i[3],i[4],i[5],self.validateAmount2(self.returnTotalJournalAmount(i[0])),i[6],i[7],self.returnUserName(i[8], 0),i[9]), tags = "evenrow")
                else:
                    TREE_JOURNAL.insert("", "end", values = (str(i[0]).zfill(8),i[1],i[2],i[3],i[4],i[5],self.validateAmount2(self.returnTotalJournalAmount(i[0])),i[6],i[7],self.returnUserName(i[8], 0),i[9]), tags = "oddrow")
                count += 1
                gjnumbers.append(i[0])
                PROGRESS_BAR["value"] = round((count/(len(result)-len(skipped)))*100, 0)
                SUB1_JOURNAL.update()
            else:
                skipped.append(i[0])
        PROGRESS_BAR.grid_remove()
    
    def disableJournalWidgets(self):
        ENTRY_PARTICULARS.config(state = DISABLED)
        ENTRY_REFERENCE.config(state = DISABLED)
        CALENDAR_DOC.config(state = DISABLED)
        CALENDAR_GL.config(state = DISABLED)
        for i in entryFrame:
            for x in i.winfo_children():
                x.config(state = DISABLED)
    
    def generateGJNumber(self):
        cursor.execute("SELECT MAX(gjNumber) FROM tblgeneralJournal")
        result = cursor.fetchone()
        if result[0] == None:
            return str(783).zfill(8)
        else:
            return str(int(result[0])+1).zfill(8)

    def returnTotalJournalAmount(self, var):
        cursor.execute(f"SELECT SUM(amount) FROM tblgeneraljournal WHERE gjNumber = '{var}' AND side = 'Debit' LIMIT 1")
        result = cursor.fetchone()
        try:
            return result[0]
        except:
            return 0

    def returnJournalStatus(self, var, i):
        select = "SELECT isPosted, isVoid FROM tblgeneraljournal WHERE gjNumber = %s LIMIT 1"
        cursor.execute(select, [var])
        result = cursor.fetchone()
        if result:
            return result[i]

### MENU_ACCOUNTING_POST MASTER ###
    def postBook(self, *args):
        insert = """INSERT INTO tblgeneralledger (
            glDate, reference, source, chartCode, amount,
            side, remarks, poster, posted
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        if TEXTVAR_BOOK.get() == "General Journal":
            select = "SELECT glDate, gjNumber, chartCode, amount, side, remarks FROM tblgeneraljournal WHERE isPosted = 'No' AND glDate BETWEEN %s AND %s"
            update = "UPDATE tblgeneraljournal SET isPosted = 'Yes' WHERE isPosted = 'No' AND glDate BETWEEN %s AND %s"
            book = "GJ"
        elif TEXTVAR_BOOK.get() == "Purchase Book":
            select = "SELECT glDate, apvNumber, chartCode, amount, side, remarks FROM tblpurchasebook WHERE isPosted = 'No' AND glDate BETWEEN %s AND %s"
            update = "UPDATE tblpurchasebook SET isPosted = 'Yes' WHERE isPosted = 'No' AND glDate BETWEEN %s AND %s"
            book = "PB"
        elif TEXTVAR_BOOK.get() == "Cash Disbursement Book":
            select = "SELECT glDate, dvNumber, chartCode, amount, side, remarks FROM tblcashdisbursementbook WHERE isPosted = 'No' AND glDate BETWEEN %s AND %s"
            update = "UPDATE tblcashdisbursementbook SET isPosted = 'Yes' WHERE isPosted = 'No' AND glDate BETWEEN %s AND %s"
            book = "CDB"
        elif TEXTVAR_BOOK.get() == "Sales Book":
            select = "SELECT glDate, soaNumber, chartCode, amount, side, remarks FROM tblsalesbook WHERE isPosted = 'No' AND glDate BETWEEN %s AND %s"
            update = "UPDATE tblsalesbook SET isPosted = 'Yes' WHERE isPosted = 'No' AND glDate BETWEEN %s AND %s"
            book = "SB"
        elif TEXTVAR_BOOK.get() == "Cash Receipt Book":
            select = "SELECT glDate, orNumber, chartCode, amount, side, remarks FROM tblcashreceiptbook WHERE isPosted = 'No' AND glDate BETWEEN %s AND %s"
            update = "UPDATE tblcashreceiptbook SET isPosted = 'Yes' WHERE isPosted = 'No' AND glDate BETWEEN %s AND %s"
            book = "CRB"
        
        cursor.execute(select, [str(CALENDAR_FROM.get_date()), str(CALENDAR_TO.get_date())])
        result = cursor.fetchall()
        if result:
            ask = messagebox.askyesno("Post Master", "Are you sure?")
            if ask:
                listcount = 0
                for i in result:
                    cursor.execute(insert, [
                        i[0], i[1], book, i[2], i[3],
                        i[4], i[5], USER, datetime.datetime.now()
                        ])
                    listcount += 1
                    PROGRESS_BAR["value"] = round((listcount/len(result))*100, 0)
                    SUB_FRAME1.update()
                cursor.execute(update, [str(CALENDAR_FROM.get_date()), str(CALENDAR_TO.get_date())])
                db.commit()
                messagebox.showinfo("Post Master", "Posting of " + book +" successful!")
        else:
            messagebox.showerror("Post Master", "Nothing to post. Check the date range!")

### MENU_ACCOUNTING_POST MASTER ###
    def importPayrollSummary(self):
        global pclient,pperiod,pdate,basicpay,overtimepay,holidaypay,ssspremium,hdmfpremium,philhealthpremium,sssmpf,sssecc,coverup,ecola,primetime,lwp,nightdifferential,sea,ctpa,nthmonthpay,hazardpay,refund,monetization,communication,deminimis,family,gas,maintenance,meal,motorcycle,travel,uniform,adjbasicpay,adjovertimepay,adjlwp,apeccpremium,apwithholdingtax,aphdmfpremium,apphilhealthpremium,apsssmpf,apssspremium,apssscalamity,apssssalary,apsssinvestment,apssshousing,aphdmfcalamity,aphdmfmp2,aphdmfmpl,aphdmfpivol,apcashbond,apewt,aptriplehgc,aptriplehgrocery,aptriplehprepaid,aptriplehhmo,aptriplehmfl,aptriplehabuloy,aptriplehothers,arcashadvance,artelephonebill,arrental,arrepairs,idorsupplies,otherdeductions,aratmcharges,sssclaims,adjbasicpay2,adjovertimepay2,adjlwp2,bdo,cash,rcbc,payroll,lbp,dbp,gcash,aphdmfhousing
        files = listdir(TEXTVAR_FOLDER.get())
        insert = """INSERT INTO tblgeneraljournal (
                    gjNumber, docDate, glDate, particulars, reference,
                    source, chartcode, amount, side, remarks,
                    isPosted, isVoid, encoder, encoded) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
        posted, duplicates, errors = [], [], []
        pclient, pdate = "", ""
        ask = messagebox.askyesno("Payroll Importer", "Are you sure?")
        if ask:
            count = 0
            PROGRESS_BAR.grid(column = 0, row = 0, padx = TOP_PADX + 10, sticky = E)
            for i in files:
                validlines = []
                wb = load_workbook(TEXTVAR_FOLDER.get() + "/" + i)
                st = wb.active
                try:
                    if (st['A5'].value.split("Name: ")[1] != pclient and st['A6'].value.split("Date: ")[1] != pdate) or (st['A5'].value.split("Name: ")[1] != "" and st['A6'].value.split("Date: ")[1] != ""):
                        pclient = st['A5'].value.split("Name: ")[1]
                        pperiod = st['A6'].value.split("Period : ")[1]
                        pdate = st['A6'].value.split("Date: ")[1]

                        ### debit ###
                        basicpay = st['H29'].value
                        overtimepay = st['H30'].value
                        holidaypay = st['H31'].value
                        ssspremium = st['H32'].value
                        hdmfpremium = st['H33'].value
                        philhealthpremium = st['H34'].value
                        sssmpf = st['H35'].value
                        sssecc = st['H36'].value
                        coverup = st['H37'].value
                        ecola = st['H38'].value
                        primetime = st['H39'].value
                        lwp = st['H40'].value
                        nightdifferential = st['H41'].value
                        if nightdifferential == None:
                            nightdifferential = 0
                        sea = st['H42'].value 
                        ctpa = st['H43'].value 
                        nthmonthpay = st['H44'].value
                        hazardpay = st['H45'].value
                        refund = st['H46'].value
                        monetization = st['H47'].value
                        communication = st['H49'].value
                        deminimis = st['H50'].value
                        family = st['H51'].value
                        gas = st['H52'].value
                        maintenance = st['H53'].value
                        meal = st['H54'].value
                        motorcycle = st['H55'].value
                        travel = st['H56'].value
                        uniform = st['H57'].value
                        adjbasicpay = st['H60'].value
                        adjovertimepay = st['H61'].value
                        adjlwp = st['H62'].value

                        ### credit ###
                        apeccpremium = st['P30'].value
                        apwithholdingtax = st['P31'].value
                        apssspremium = st['P32'].value
                        aphdmfpremium = st['P33'].value
                        apphilhealthpremium = st['P34'].value
                        apsssmpf = st['P35'].value
                        apssscalamity = st['P37'].value
                        apssssalary = st['P38'].value
                        apsssinvestment = st['P39'].value
                        apssshousing = st['P40'].value
                        aphdmfhousing = st['P42'].value
                        aphdmfcalamity = st['P43'].value
                        aphdmfmp2 = st['P44'].value
                        aphdmfmpl = st['P45'].value
                        aphdmfpivol = st['P46'].value
                        apcashbond = st['P47'].value
                        apewt = st['P48'].value
                        aptriplehgc = st['O19'].value
                        aptriplehgrocery = st['O20'].value
                        aptriplehprepaid = st['O21'].value
                        aptriplehhmo = st['P19'].value
                        aptriplehmfl = st['P20'].value
                        aptriplehabuloy = st['P21'].value
                        aptriplehothers = st['O22'].value
                        arcashadvance = st['K19'].value
                        artelephonebill = st['P51'].value
                        arrental = st['P52'].value
                        arrepairs = st['P53'].value
                        idorsupplies = st['P55'].value
                        otherdeductions = st['P56'].value
                        if otherdeductions == None:
                            otherdeductions = 0
                        aratmcharges = st['P57'].value
                        sssclaims = st['P58'].value
                        adjbasicpay2 = st['P60'].value
                        adjovertimepay2 = st['P61'].value
                        adjlwp2 = st['P62'].value
                        bdo = st['T30'].value
                        cash = st['T31'].value
                        rcbc = st['T32'].value
                        lbp = st['T33'].value
                        dbp = st['T34'].value
                        gcash = st['T35'].value
        
                        clericalSum = st['T39'].value
                        janitorialSum = st['T41'].value
                        skilledMSum = st['T45'].value
                        skilledTSum = st['T47'].value

                        self.payrollItemsGrouper()
                        ddtotalsum, cctotalsum = [], []
                        for d in dd:
                            ddtotalsum.append(d[0])
                        for c in cc:
                            cctotalsum.append(c[0])

                        if round(sum(ddtotalsum), 2) == round(sum(cctotalsum), 2):
                            if (clericalSum + skilledTSum) > (janitorialSum + skilledMSum):
                                employeeCategory = "Administrative"
                            else:
                                employeeCategory = "Maintenance"
                            
                            gjnumber = self.generateGJNumber()
                            for x in dd:
                                if x[0] > 0:
                                    if employeeCategory == "Administrative":
                                        validlines.append([
                                            int(gjnumber), CALENDAR_DOC.get_date(), CALENDAR_GL.get_date(), pclient + "_Payroll Register Summary for the period_" + pperiod, "EIS PayDate_" + pdate.split("   ")[0],
                                            "GJ", x[1], x[0], "Debit", "",
                                            "No", "No", USER, datetime.datetime.now()
                                        ])
                                    else:
                                        validlines.append([
                                            int(gjnumber), CALENDAR_DOC.get_date(), CALENDAR_GL.get_date(), pclient + "_Payroll Register Summary for the period_" + pperiod, "EIS PayDate_" + pdate.split("   ")[0],
                                            "GJ", x[2], x[0], "Debit", "",
                                            "No", "No", USER, datetime.datetime.now()
                                        ])
                            for x in cc:
                                if x[0] > 0:
                                    validlines.append([
                                            int(gjnumber), CALENDAR_DOC.get_date(), CALENDAR_GL.get_date(), pclient + "_Payroll Register Summary for the period_" + pperiod, "EIS PayDate_" + pdate.split("   ")[0],
                                            "GJ", x[1], x[0], "Credit", x[2],
                                            "No", "No", USER, datetime.datetime.now()
                                    ])
                        else:
                            if st['A5'].value.split("Name: ")[1] not in errors:
                                errors.append(st['A5'].value.split("Name: ")[1])
                    else:
                        if st['A5'].value.split("Name: ")[1] not in duplicates:
                            duplicates.append(st['A5'].value.split("Name: ")[1])
                except:
                    pass
                wb.close()
            
                if len(validlines) != 0:
                    for y in validlines:
                        cursor.execute(insert, y)
                    db.commit()
                    posted.append(pclient)

                count += 1
                PROGRESS_BAR["value"] = round((count/(len(files)))*100, 0)
                SUB3_PAYROLL.update()
            PROGRESS_BAR.grid_remove()

            if len(posted) != 0:
                messagebox.showinfo("Payroll Importer", "Successfully posted " + str(len(posted)) + "payroll registers! \nDuplicates are: " + str(duplicates) + "\nError files are: " + str(errors))
            else:
                messagebox.showerror("Payroll Importer", "No register has been posted!")

    def payrollItemsGrouper(self):
        global dd, cc
        salary = [sum([basicpay, coverup, nightdifferential, hazardpay, adjbasicpay]), 611010, 611020]
        overtime = [sum([overtimepay, primetime, adjovertimepay, ctpa]), 611050, 611060]
        holiday = [holidaypay, 611070, 611080]
        sssp = [ssspremium, 611220, 611230]
        hdmfp = [hdmfpremium, 611200, 611210]
        phicp = [philhealthpremium, 611240, 611250]
        sssmp = [sssmpf, 611280, 611290]
        ssse = [sssecc, 611260, 611270]
        ecol = [ecola, 611030, 611040]
        lwpay = [sum([lwp, adjlwp, monetization]), 611110, 611120]
        allow = [deminimis, 611190, 611300]
        travels = [sum([communication, gas, maintenance, meal, motorcycle, travel, uniform]), 611310, 611310]
        fallow = [family, 611190, 611300]
        mp13 = [nthmonthpay, 611130, 611140]
        refunds = [refund, 215020, 215020]
        
        dd = [salary,overtime,holiday,sssp,hdmfp,phicp,sssmp,ssse,ecol,lwpay,allow,travels,fallow,mp13,refunds]
       
        aphl = [apssshousing, 213020, "HOUSING LOAN"]
        aphhl = [aphdmfhousing, 213020, "HOUSING LOAN"]
        apsssp = [apssspremium, 213110, "SSS PREMIUMS"]
        aphdmfpr = [aphdmfpremium, 213090, "HDMF PREMIUMS"]
        aphdmfm2 = [aphdmfmp2, 213090, "HDMF MP2"]
        aphdmfpvl = [aphdmfpivol, 213090, "HDMF PIVOL"]
        apphicp = [apphilhealthpremium, 213030, "PHIC PREMIUM"]
        apsssmp = [apsssmpf, 213120, "SSS MPF"]
        apsssl = [apssssalary, 213100, "SSS SALARY"]
        apssslc = [apssscalamity, 213100, "SSS CALAMITY"]
        apsssli = [apsssinvestment, 213110, "SSS INVESTMENT"]
        aphdmfl = [aphdmfmpl, 213070, "HDMF MPL"]
        apssse = [apeccpremium, 213010, "ECC PREMIUM"]
        apwtx = [apwithholdingtax, 213130, "WTX"]
        apewtx = [apewt, 213050, "EWT"]
        apabu = [aptriplehabuloy, 215020, "ABULOY"]
        ap3h = [sum([aptriplehgc, aptriplehgrocery, aptriplehprepaid, aptriplehhmo, aptriplehothers]), 214010, "TRIPLE H"]
        apcb = [apcashbond, 215020, "CASH BOND"]
        arca = [sum([arcashadvance, idorsupplies]), 116020, "CA CTRLS"]
        arcarl = [arrepairs, 116020, "REPAIRS"]
        arcatb = [artelephonebill, 116020, "TELEPHONE BILL"]
        arcar = [arrental, 116020, "RENTAL"]
        arcaatm = [aratmcharges, 116020, "ATM CHARGES"]
        aphdmfl2 = [0, "", ""]
        aphdmfcl = [aphdmfcalamity, 213080, "HDMF CALAMITY"]
        ap3hmfl = [aptriplehmfl, 214020, "MFL"]
        arothrs = [otherdeductions, 116030, "OTHERS"]
        ssscr = [sssclaims, 116030, ""]
        adjsalary = [adjbasicpay2, 611010, ""]
        adjot = [adjovertimepay2, 611050, ""]
        adjleave = [adjlwp2, 611110, ""]
        apoebdo = [bdo, 215020, "BDO"]
        apoewo = [cash, 215020, "CASH"]
        apoercbc = [rcbc, 215020, "RCBC"]
        aplbp = [lbp, 215020, "LBP"]
        apdbp = [dbp, 215020, "DBP"]
        apgc = [gcash, 215020, "GCASH"]
        
        cc = [aphl,aphhl,apsssp,aphdmfpr,aphdmfm2,aphdmfpvl,apphicp,apsssmp,apsssl,apssslc,apsssli,aphdmfl,apssse,apwtx,apewtx,apabu,ap3h,apcb,arca,arcarl,arcatb,arcar,arcaatm,aphdmfl2,aphdmfcl,ap3hmfl,arothrs,ssscr,adjsalary,adjot,adjleave,apoebdo,apoewo,apoercbc,aplbp,apdbp,apgc]
            
### MENU_ACCOUNTING_LOCKER ###
    def showAddEditLocker(self, *args):
        global TOP_LOCKER
        TOP_LOCKER = Toplevel()
        TOP_LOCKER.title("Create - General Journal")
        TOP_LOCKER.iconbitmap(PATH_ICON + "icon.ico")
        TOP_LOCKER.geometry("200x200+100+20")
        TOP_LOCKER.resizable(height = False, width = False)
        TOP_LOCKER.grab_set()
        TOP_LOCKER.focus()
        TOP_LOCKER.protocol("WM_DELETE_WINDOW", self.showPeriodLocker)
        
        LABEL_YEAR = Label(TOP_LOCKER, text = "Year", font = APP_FONT)
        LABEL_YEAR.grid(column = 0, row = 0, pady = TOP_PADY, sticky = W)
        
        LABEL_MONTH = Label(TOP_LOCKER, text = "Month", font = APP_FONT)
        LABEL_MONTH.grid(column = 0, row = 1, pady = TOP_PADY, sticky = W)
        
        LABEL_STATUS = Label(TOP_LOCKER, text = "Status", font = APP_FONT)
        LABEL_STATUS.grid(column = 0, row = 2, pady = TOP_PADY, sticky = W)
        
        global TEXTVAR_YEAR
        TEXTVAR_YEAR = StringVar()
        ENTRY_YEAR = Entry(TOP_LOCKER, textvariable = TEXTVAR_YEAR, font = APP_FONT, width = 10)
        ENTRY_YEAR.grid(column = 1, row = 0, sticky = W)
        
        global TEXTVAR_MONTH
        TEXTVAR_MONTH = StringVar()
        ENTRY_MONTH = Entry(TOP_LOCKER, textvariable = TEXTVAR_MONTH, font = APP_FONT, width = 10)
        ENTRY_MONTH.grid(column = 1, row = 1, sticky = W)
        
        global TEXTVAR_STATUS
        TEXTVAR_STATUS = StringVar()
        COMBO_STATUS = tk.Combobox(TOP_LOCKER, values = ["open", "locked"], textvariable = TEXTVAR_STATUS, font = APP_FONT, width = 10, state = "readonly")
        COMBO_STATUS.grid(column = 1, row = 2, sticky = W)
        
        global BUTTON_SUBMIT
        BUTTON_SUBMIT = Button(TOP_LOCKER, text = "SAVE", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "hand2", command = self.saveLocker)
        BUTTON_SUBMIT.grid(column = 1, row = 3, pady = TOP_PADY)

        global BUTTON_CLOSE
        BUTTON_CLOSE = Button(TOP_LOCKER, text = "CLOSE", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "hand2", command = self.showPeriodLocker)
        BUTTON_CLOSE.grid(column = 2, row = 3)

    def saveLocker(self):
        update = "UPDATE tblperiodlocker SET status = %s WHERE year = %s AND month = %s"
        ask = messagebox.askyesno("Period Locker", "Are you sure?")
        if ask:
            cursor.execute(update, [TEXTVAR_STATUS.get(), int(TEXTVAR_YEAR.get()), int(TEXTVAR_MONTH.get())])
            db.commit()
            messagebox.showinfo("Period Locker", "Period has been updated!")
            TOP_LOCKER.focus()

    def editLocker(self, *args):
        self.copySelection(TREE_LOCKER)
        self.showAddEditLocker()
        TOP_LOCKER.title("Edit - Locker")
        
        select = """SELECT 
            year, month, status
                FROM tblperiodlocker WHERE year = %s AND month = %s LIMIT 1"""
        
        cursor.execute(select, [int(content[0]), int(content[1])])
        result = cursor.fetchone()
        if result:
            TEXTVAR_YEAR.set(result[0])
            TEXTVAR_MONTH.set(result[1])
            TEXTVAR_STATUS.set(result[2])

    def returnPeriodStatus(self, var):
        select = "SELECT status FROM tblperiodlocker WHERE year = %s AND month = %s LIMIT 1"
        cursor.execute(select, [int(str(var).split("-")[0]), int(str(var).split("-")[1])])
        result = cursor.fetchone()
        try:
            return result[0]
        except:
            return "locked"

### MENU_ACCOUNTING_REPORTS ###
    def showSelectedAccountingReport(self, *args):
        if TEXTVAR_REPORTS.get() == "General Ledger":
            CALENDAR_FROM.config(state = NORMAL)
            BUTTON_EXPORT.config(state = NORMAL, cursor = "hand2")
            COMBO_OPTION.config(state = "readonly")
            LABEL_CODE.grid_remove()
            LABEL_TITLE.grid_remove()
            ENTRY_CHARTCODE.grid_remove()
            BUTTON_CHARTCODE.grid_remove()
            ENTRY_CHARTTITLE.grid_remove()
            COMBO_BOOKS.grid_remove()
            LABEL_BOOK.grid_remove()
        elif TEXTVAR_REPORTS.get() == "Trial Balance":
            CALENDAR_FROM.config(state = DISABLED)
            COMBO_OPTION.config(state = DISABLED)
            BUTTON_EXPORT.config(state = NORMAL, cursor = "hand2")
            LABEL_CODE.grid_remove()
            LABEL_TITLE.grid_remove()
            ENTRY_CHARTCODE.grid_remove()
            BUTTON_CHARTCODE.grid_remove()
            ENTRY_CHARTTITLE.grid_remove()
            COMBO_BOOKS.grid_remove()
            LABEL_BOOK.grid_remove()
        elif TEXTVAR_REPORTS.get() == "Income Statement":
            CALENDAR_FROM.config(state = DISABLED)
            COMBO_OPTION.config(state = DISABLED)
            BUTTON_EXPORT.config(state = NORMAL, cursor = "hand2")
            LABEL_CODE.grid_remove()
            LABEL_TITLE.grid_remove()
            ENTRY_CHARTCODE.grid_remove()
            BUTTON_CHARTCODE.grid_remove()
            ENTRY_CHARTTITLE.grid_remove()
            COMBO_BOOKS.grid_remove()
            LABEL_BOOK.grid_remove()
        elif TEXTVAR_REPORTS.get() == "Statement of Financial Position":
            CALENDAR_FROM.config(state = DISABLED)
            COMBO_OPTION.config(state = DISABLED)
            BUTTON_EXPORT.config(state = NORMAL, cursor = "hand2")
            LABEL_CODE.grid_remove()
            LABEL_TITLE.grid_remove()
            ENTRY_CHARTCODE.grid_remove()
            BUTTON_CHARTCODE.grid_remove()
            ENTRY_CHARTTITLE.grid_remove()
            COMBO_BOOKS.grid_remove()
            LABEL_BOOK.grid_remove()
        elif TEXTVAR_REPORTS.get() == "End of Month":
            CALENDAR_FROM.config(state = NORMAL)
            COMBO_OPTION.config(state = DISABLED)
            BUTTON_EXPORT.config(state = NORMAL, cursor = "hand2")
            LABEL_CODE.grid_remove()
            LABEL_TITLE.grid_remove()
            ENTRY_CHARTCODE.grid_remove()
            BUTTON_CHARTCODE.grid_remove()
            ENTRY_CHARTTITLE.grid_remove()
            COMBO_BOOKS.grid_remove()
            LABEL_BOOK.grid_remove()
        elif TEXTVAR_REPORTS.get() == "Account Analysis":
            CALENDAR_FROM.config(state = NORMAL)
            BUTTON_EXPORT.config(state = NORMAL, cursor = "hand2")
            COMBO_OPTION.config(state = "readonly")
            LABEL_CODE.grid(column = 0, row = 3, sticky = W)
            LABEL_TITLE.grid(column = 0, row = 4, sticky = W)
            ENTRY_CHARTCODE.grid(column = 1, row = 3, sticky = W)
            BUTTON_CHARTCODE.grid(column = 1, row = 3, sticky = W)
            ENTRY_CHARTTITLE.grid(column = 1, row = 4, sticky = W, ipadx = 1)
            COMBO_BOOKS.grid_remove()
            LABEL_BOOK.grid_remove()
        elif TEXTVAR_REPORTS.get() == "Books":
            CALENDAR_FROM.config(state = NORMAL)
            LABEL_BOOK.grid(column = 0, row = 3, sticky = W)
            COMBO_OPTION.config(state = "readonly")
            BUTTON_EXPORT.config(state = NORMAL, cursor = "hand2")
            LABEL_CODE.grid_remove()
            LABEL_TITLE.grid_remove()
            ENTRY_CHARTCODE.grid_remove()
            BUTTON_CHARTCODE.grid_remove()
            ENTRY_CHARTTITLE.grid_remove()
            COMBO_BOOKS.grid(column = 1, row = 3, sticky = W)
            
    def showAccountingReportTreeView(self, frame):
        global TREE_ACCOUNTING
        TREE_ACCOUNTING = tk.Treeview(frame, height = 25, selectmode = "browse")
        TREE_ACCOUNTING["columns"] = ("GL Date", "Code", "Title", "Debit", "Credit", "Remarks", "Source", "Reference", "Poster", "Posted")
        TREE_ACCOUNTING.column("#0", width = 0, minwidth = 0, stretch = True)
        TREE_ACCOUNTING.column("GL Date", anchor = W, minwidth = 75, width = 75)
        TREE_ACCOUNTING.column("Code", anchor = W, minwidth = 75, width = 75)
        TREE_ACCOUNTING.column("Title", anchor = W, minwidth = 100, width = 200)
        TREE_ACCOUNTING.column("Debit", anchor = E, minwidth = 100, width = 100)
        TREE_ACCOUNTING.column("Credit", anchor = E, minwidth = 100, width = 100)
        TREE_ACCOUNTING.column("Remarks", anchor = W, minwidth = 75, width = 150)
        TREE_ACCOUNTING.column("Source", anchor = W, minwidth = 75, width = 75)
        TREE_ACCOUNTING.column("Reference", anchor = W, minwidth = 75, width = 100)
        TREE_ACCOUNTING.column("Poster", anchor = W, minwidth = 100, width = 75)
        TREE_ACCOUNTING.column("Posted", anchor = W, minwidth = 100, width = 75)

        TREE_ACCOUNTING.heading("#0", text = "", anchor = W)
        TREE_ACCOUNTING.heading("GL Date", text = "GL Date", anchor = N)
        TREE_ACCOUNTING.heading("Code", text = "Code", anchor = N)
        TREE_ACCOUNTING.heading("Title", text = "Title", anchor = N)
        TREE_ACCOUNTING.heading("Debit", text = "Debit", anchor = N)
        TREE_ACCOUNTING.heading("Credit", text = "Credit", anchor = N)
        TREE_ACCOUNTING.heading("Remarks", text = "Remarks", anchor = N)
        TREE_ACCOUNTING.heading("Source", text = "Source", anchor = N)
        TREE_ACCOUNTING.heading("Reference", text = "Reference", anchor = N)
        TREE_ACCOUNTING.heading("Poster", text = "Poster", anchor = N)
        TREE_ACCOUNTING.heading("Posted", text = "Posted", anchor = N)

        global STYLE_ACCOUNTING
        STYLE_ACCOUNTING = tk.Style()
        STYLE_ACCOUNTING.map("Treeview", foreground = self.fixedMap("foreground", STYLE_ACCOUNTING), background = self.fixedMap("background", STYLE_ACCOUNTING))

        TREE_ACCOUNTING.tag_configure("oddrow", background = None)
        TREE_ACCOUNTING.tag_configure("evenrow", background = TREE_TAG_EVENROW)

        YSCROLLBAR = tk.Scrollbar(frame, orient = "vertical", command = TREE_ACCOUNTING.yview)
        YSCROLLBAR.pack(side = "right", fill = "y")

        XSCROLLBAR = tk.Scrollbar(frame, orient = "horizontal", command = TREE_ACCOUNTING.xview)
        
        TREE_ACCOUNTING.configure(yscrollcommand = YSCROLLBAR.set, xscrollcommand = XSCROLLBAR.set) 
        TREE_ACCOUNTING.pack()
        XSCROLLBAR.pack(fill ="x")

    def fetchGeneralLedger(self):
        books = []
        db.commit()
        cursor.execute(f"SELECT glDate, chartCode, amount, side, remarks, source, reference, poster, DATE(posted) FROM tblgeneralledger WHERE glDate BETWEEN '{CALENDAR_FROM.get_date()}' AND'{CALENDAR_TO.get_date()}' ORDER BY chartCode")
        result = cursor.fetchall()
        if result:
            books.append(result)
        if TEXTVAR_OPTION.get() != "Posted":
            db.commit()
            cursor.execute(f"SELECT glDate, chartCode, amount, side, remarks, source, gjNumber, encoder, DATE(encoded) FROM tblgeneraljournal WHERE isPosted = 'No' AND glDate BETWEEN '{CALENDAR_FROM.get_date()}' AND'{CALENDAR_TO.get_date()}' ORDER BY chartCode")
            resultgj = cursor.fetchall()
            if resultgj:
                books.append(resultgj)
            
            db.commit()
            cursor.execute(f"SELECT glDate, chartCode, amount, side, remarks, source, apvNumber, poster, DATE(posted) FROM tblpurchasebook WHERE isPosted = 'No' AND glDate BETWEEN '{CALENDAR_FROM.get_date()}' AND'{CALENDAR_TO.get_date()}' ORDER BY chartCode")
            resultpb = cursor.fetchall()
            if resultpb:
                books.append(resultpb)
            
            db.commit()
            cursor.execute(f"SELECT glDate, chartCode, amount, side, remarks, source, dvNumber, poster, DATE(posted) FROM tblcashdisbursementbook WHERE isPosted = 'No' AND glDate BETWEEN '{CALENDAR_FROM.get_date()}' AND'{CALENDAR_TO.get_date()}' ORDER BY chartCode")
            resultcdb = cursor.fetchall()
            if resultcdb:
                books.append(resultcdb)
            
            db.commit()
            cursor.execute(f"SELECT glDate, chartCode, amount, side, remarks, source, soaNumber, poster, DATE(posted) FROM tblsalesbook WHERE isPosted = 'No' AND glDate BETWEEN '{CALENDAR_FROM.get_date()}' AND'{CALENDAR_TO.get_date()}' ORDER BY chartCode")
            resultsb = cursor.fetchall()
            if resultsb:
                books.append(resultsb)
            
            # db.commit()
            # cursor.execute(f"SELECT glDate, chartCode, amount, side, remarks, source, orNumber, poster, DATE(posted) FROM tblcashreceiptbook WHERE isPosted = 'No' AND glDate BETWEEN '{CALENDAR_FROM.get_date()}' AND'{CALENDAR_TO.get_date()}' ORDER BY chartCode")
            # resultcrb = cursor.fetchall()
            # if resultcrb:
                # books.append(resultcrb)
            
        global finalresults
        allresults, finalresults = [], []
        if len(books) != 0:
            for i in books:
                for x in i:
                    allresults.append(x)
        count = 0 #glDate, chartCode, amount, side, remarks, source, reference, poster, DATE(posted)
        for i in allresults:
            if i[3] == "Debit":
                finalresults.append([i[0],i[1],self.returnChartTitle(i[1])[0],i[2],0,i[4],i[5],i[6],self.returnUserName(i[7], 0),i[8]])
            else:
                finalresults.append([i[0],i[1],self.returnChartTitle(i[1])[0],0,i[2],i[4],i[5],i[6],self.returnUserName(i[7], 0),i[8]])
            count += 1
            PROGRESS_BAR["value"] = round((count/len(allresults))*100, 0)
            SUB_FRAME1.update()

    def exportAccountingReport(self):
        if TEXTVAR_REPORTS.get() == "General Ledger":
            self.fetchGeneralLedger()
            wb = load_workbook(PATH_TEMPLATE + "GLSUMMARY.xlsx")
            sheet = wb.active
            sheet["B5"] = f"{CALENDAR_FROM.get_date()} to {CALENDAR_TO.get_date()}"
            count = 8
            listcount = 0
            resultcount = len(finalresults)
            for i in finalresults:
                sheet["A" + str(count)] = i[0]
                sheet["B" + str(count)] = i[1]
                sheet["C" + str(count)] = i[2]
                sheet["D" + str(count)] = i[3]
                sheet["E" + str(count)] = i[4]
                sheet["F" + str(count)] = i[5]
                sheet["G" + str(count)] = i[6]
                sheet["H" + str(count)] = i[7]
                sheet["I" + str(count)] = i[8]
                sheet["J" + str(count)] = i[9]
                count += 1
                listcount += 1
                PROGRESS_BAR["value"] = round((listcount/resultcount)*100, 0)
                SUB_FRAME1.update()
            sheet["A" + str(count)] = "/" + self.returnUserName(USER, 0) + str(datetime.date.today())
            wb.save(PATH_SAVE + "GLSUMMARY.xlsx")
            startfile(PATH_SAVE + "GLSUMMARY.xlsx", "open")
        elif TEXTVAR_REPORTS.get() == "Trial Balance":
            db.commit()
            cursor.execute(f"SELECT chartCode, amount, side FROM tblgeneralledger WHERE glDate BETWEEN '2022-12-31' AND'{CALENDAR_TO.get_date()}' ORDER BY chartCode")
            result = cursor.fetchall()
            if result:
                codes = []
                listcount = 0
                for i in result:
                    if i[2] == "Debit":
                        codes.append([i[0], float(i[1])])
                    else:
                        codes.append([i[0], -float(i[1])])
                    listcount += 1
                    PROGRESS_BAR["value"] = round((listcount/len(result))*100, 0)
                    SUB_FRAME1.update()

                df = pd.DataFrame(codes)
                codebalance = df.groupby([0], as_index = False).sum()
                codebalance = codebalance.values.tolist()

                wb = load_workbook(PATH_TEMPLATE + "BOOK.xlsx")
                sheet = wb.active
                sheet["A3"] = TEXTVAR_REPORTS.get()
                sheet["B5"] = f"{CALENDAR_FROM.get_date()} to {CALENDAR_TO.get_date()}"
                count = 8
                listcount = 0
                for i in codebalance:
                    sheet["C" + str(count)] = int(str(i[0])[:2]) * 10000
                    sheet["D" + str(count)] = self.returnChartTitle(int(str(i[0])[:2]) * 10000)[0]
                    sheet["E" + str(count)] = int(str(i[0])[:3]) * 1000
                    sheet["F" + str(count)] = self.returnChartTitle(int(str(i[0])[:3]) * 1000)[0]
                    sheet["G" + str(count)] = i[0]
                    sheet["H" + str(count)] = self.returnChartTitle(i[0])[0]
                    sheet["I" + str(count)] = i[1]
                    count += 1
                    listcount += 1
                    PROGRESS_BAR["value"] = round((listcount/len(codebalance))*100, 0)
                    SUB_FRAME1.update()
                sheet["A" + str(count)] = "/" + self.returnUserName(USER, 0) + str(datetime.date.today())
                wb.save(PATH_SAVE + TEXTVAR_REPORTS.get() + ".xlsx")
                startfile(PATH_SAVE + TEXTVAR_REPORTS.get() + ".xlsx", "open")
        elif TEXTVAR_REPORTS.get() == "Income Statement":
            db.commit()
            cursor.execute(f"SELECT chartCode, amount, side FROM tblgeneralledger WHERE glDate BETWEEN '2022-12-31' AND'{CALENDAR_TO.get_date()}' ORDER BY chartCode")
            result = cursor.fetchall()
            if result:
                codes = []
                listcount = 0
                for i in result:
                    if i[2] == "Debit":
                        codes.append([int(str(i[0])[:2]) * 10000, float(i[1])])
                    else:
                        codes.append([int(str(i[0])[:2]) * 10000, -float(i[1])])
                    listcount += 1
                    PROGRESS_BAR["value"] = round((listcount/len(result))*100, 0)
                    SUB_FRAME1.update()

                df = pd.DataFrame(codes)
                codebalance = df.groupby([0], as_index = False).sum()
                codebalance = codebalance.values.tolist()

                wb = load_workbook(PATH_TEMPLATE + "IS.xlsx")
                sheet = wb.active
                sheet["B7"] = f"2023-01-01 to {CALENDAR_TO.get_date()}"

                for i in codebalance:
                    if i[0] == 510000:
                        sheet["C" + str(10)] = i[1]
                    elif i[0] == 610000:
                        sheet["C" + str(11)] = i[1]
                    elif i[0] == 620000:
                        sheet["C" + str(14)] = i[1]
                    elif i[0] == 630000:
                        sheet["C" + str(15)] = i[1]
                    elif i[0] == 640000:
                        sheet["C" + str(16)] = i[1]
                    elif i[0] == 520000:
                        sheet["C" + str(17)] = i[1]
                
                sheet["A" + str(25  )] = "/" + self.returnUserName(USER, 0) + str(datetime.date.today())
                wb.save(PATH_SAVE + TEXTVAR_REPORTS.get() + ".xlsx")
                wb.close()
                startfile(PATH_SAVE + TEXTVAR_REPORTS.get() + ".xlsx", "open")
        elif TEXTVAR_REPORTS.get() == "Statement of Financial Position":
            db.commit()
            cursor.execute(f"SELECT chartCode, amount, side FROM tblgeneralledger WHERE glDate BETWEEN '2022-12-31' AND'{CALENDAR_TO.get_date()}' ORDER BY chartCode")
            result = cursor.fetchall()
            if result:
                codes = []
                listcount = 0
                for i in result:
                    if i[2] == "Debit":
                        codes.append([int(str(i[0])[:2]) * 10000, float(i[1])])
                    else:
                        codes.append([int(str(i[0])[:2]) * 10000, -float(i[1])])
                    listcount += 1
                    PROGRESS_BAR["value"] = round((listcount/len(result))*100, 0)
                    SUB_FRAME1.update()

                df = pd.DataFrame(codes)
                codebalance = df.groupby([0], as_index = False).sum()
                codebalance = codebalance.values.tolist()

                wb = load_workbook(PATH_TEMPLATE + "BS.xlsx")
                sheet = wb.active
                sheet["B7"] = f"2023-01-01 to {CALENDAR_TO.get_date()}"

                netincome = []
                for i in codebalance:
                    if str(i[0])[0] == '5' or str(i[0])[0] == '6':
                        netincome.append(i[1])

                for i in codebalance:
                    if i[0] == 110000:
                        sheet["C" + str(10)] = i[1]
                    elif i[0] == 120000:
                        sheet["C" + str(11)] = i[1]
                    elif i[0] == 210000:
                        sheet["C" + str(14)] = i[1]
                    elif i[0] == 220000:
                        sheet["C" + str(15)] = i[1]
                    elif i[0] == 310000:
                        sheet["C" + str(18)] = i[1]
                    elif i[0] == 320000:
                        sheet["C" + str(19)] = i[1]
                    elif i[0] == 340000:
                        sheet["C" + str(20)] = sum((i[1], sum(netincome)))
                    elif i[0] == 330000:
                        sheet["C" + str(21)] = i[1]
                    elif i[0] == 350000:
                        sheet["C" + str(22)] = i[1]

                sheet["A" + str(25)] = "/" + self.returnUserName(USER, 0) + str(datetime.date.today())
                wb.save(PATH_SAVE + TEXTVAR_REPORTS.get() + ".xlsx")
                wb.close()
                startfile(PATH_SAVE + TEXTVAR_REPORTS.get() + ".xlsx", "open")
        elif TEXTVAR_REPORTS.get() == "End of Month":
            db.commit()
            cursor.execute(f"SELECT chartCode, amount, side FROM tblgeneralledger WHERE glDate BETWEEN '{CALENDAR_FROM.get_date()}' AND '{CALENDAR_TO.get_date()}' ORDER BY chartCode")
            result = cursor.fetchall()
            if result:
                codes = []
                for i in result:
                    if i[2] == "Debit":
                        codes.append([i[0], float(i[1])])
                    else:
                        codes.append([i[0], -float(i[1])])

                df = pd.DataFrame(codes)
                codebalance = df.groupby([0], as_index = False).sum()
                codebalance = codebalance.values.tolist()

                wb = load_workbook(PATH_TEMPLATE + "BOOK.xlsx")
                sheet = wb.active
                sheet["A3"] = TEXTVAR_BOOKS.get()
                sheet["B5"] = f"{CALENDAR_FROM.get_date()} to {CALENDAR_TO.get_date()}"
                count = 8
                listcount = 0
                for i in codebalance:
                    sheet["G" + str(count)] = i[0]
                    sheet["H" + str(count)] = self.returnChartTitle(i[0])[0]
                    sheet["I" + str(count)] = i[1]
                    count += 1
                    listcount += 1
                    PROGRESS_BAR["value"] = round((listcount/len(codebalance))*100, 0)
                    SUB_FRAME1.update()
                sheet["A" + str(count)] = "/" + self.returnUserName(USER, 0) + str(datetime.date.today())
                wb.save(PATH_SAVE + TEXTVAR_REPORTS.get() + f"{CALENDAR_FROM.get_date()} to {CALENDAR_TO.get_date()}" + ".xlsx")
                startfile(PATH_SAVE + TEXTVAR_REPORTS.get() + f"{CALENDAR_FROM.get_date()} to {CALENDAR_TO.get_date()}" + ".xlsx", "open")
            else:
                messagebox.showerror("Reports", "No results found!")
        elif TEXTVAR_REPORTS.get() == "Account Analysis":
            db.commit()
            cursor.execute(f"SELECT glDate, chartCode, amount, side, remarks, source, reference, poster, DATE(posted) FROM tblgeneralledger WHERE chartCode = {ENTRY_CHARTCODE.get()} AND glDate BETWEEN '{CALENDAR_FROM.get_date()}' AND'{CALENDAR_TO.get_date()}' ORDER BY chartCode")
            result = cursor.fetchall()
            wb = load_workbook(PATH_TEMPLATE + "GLSUMMARY.xlsx")
            sheet = wb.active
            sheet["B5"] = f"{CALENDAR_FROM.get_date()} to {CALENDAR_TO.get_date()}"
            count = 8
            listcount = 0
            resultcount = len(result)
            for i in result:
                sheet["A" + str(count)] = i[0]
                sheet["B" + str(count)] = i[1]
                sheet["C" + str(count)] = self.returnChartTitle(i[1])[0]
                if i[3] == "Debit":
                    sheet["D" + str(count)] = i[2]
                    sheet["E" + str(count)] = 0
                else:
                    sheet["D" + str(count)] = 0
                    sheet["E" + str(count)] = i[2]
                sheet["F" + str(count)] = i[4]
                sheet["G" + str(count)] = i[5]
                sheet["H" + str(count)] = i[6]
                sheet["I" + str(count)] = i[7]
                sheet["J" + str(count)] = i[8]
                count += 1
                listcount += 1
                PROGRESS_BAR["value"] = round((listcount/resultcount)*100, 0)
                SUB_FRAME1.update()
            sheet["A" + str(count)] = "/" + self.returnUserName(USER, 0) + str(datetime.date.today())
            wb.save(PATH_SAVE + "GLSUMMARY.xlsx")
            startfile(PATH_SAVE + "GLSUMMARY.xlsx", "open")
        elif TEXTVAR_REPORTS.get() == "Books":
            db.commit()
            if TEXTVAR_BOOKS.get() == "CDB":
                if TEXTVAR_OPTION.get() == "Posted":
                    cursor.execute(f"SELECT DISTINCT source, tblcashdisbursementbook.dvNumber, tbldisbursements.dvDate, tblcashdisbursementbook.glDate, tbldisbursements.particulars, tbldisbursements.reference, chartCode, amount, side, remarks, poster, DATE(posted) FROM tblcashdisbursementbook INNER JOIN tbldisbursements ON tblcashdisbursementbook.dvNumber = tbldisbursements.dvNumber WHERE isPosted = 'Yes' AND tblcashdisbursementbook.glDate BETWEEN '{CALENDAR_FROM.get_date()}' AND'{CALENDAR_TO.get_date()}' ORDER BY tblcashdisbursementbook.dvNumber")
                else:
                    cursor.execute(f"SELECT DISTINCT source, tblcashdisbursementbook.dvNumber, tbldisbursements.dvDate, tblcashdisbursementbook.glDate, tbldisbursements.particulars, tbldisbursements.reference, chartCode, amount, side, remarks, poster, DATE(posted) FROM tblcashdisbursementbook INNER JOIN tbldisbursements ON tblcashdisbursementbook.dvNumber = tbldisbursements.dvNumber WHERE tblcashdisbursementbook.glDate BETWEEN '{CALENDAR_FROM.get_date()}' AND'{CALENDAR_TO.get_date()}' ORDER BY tblcashdisbursementbook.dvNumber")
            elif TEXTVAR_BOOKS.get() == "PB":
                if TEXTVAR_OPTION.get() == "Posted":
                    cursor.execute(f"SELECT DISTINCT source, tblpurchasebook.apvNumber, tblpayables.apvDate, tblpurchasebook.glDate, tblpayables.particulars, tblpayables.reference, tblpurchasebook.chartCode, tblpurchasebook.amount, side, tblpurchasebook.remarks, poster, DATE(posted) FROM tblpurchasebook INNER JOIN tblpayables ON tblpurchasebook.apvNumber = tblpayables.apvNumber WHERE isPosted = 'Yes' AND tblpurchasebook.glDate BETWEEN '{CALENDAR_FROM.get_date()}' AND'{CALENDAR_TO.get_date()}' ORDER BY tblpurchasebook.apvNumber")
                else:
                    cursor.execute(f"SELECT DISTINCT source, tblpurchasebook.apvNumber, tblpayables.apvDate, tblpurchasebook.glDate, tblpayables.particulars, tblpayables.reference, tblpurchasebook.chartCode, tblpurchasebook.amount, side, tblpurchasebook.remarks, poster, DATE(posted) FROM tblpurchasebook INNER JOIN tblpayables ON tblpurchasebook.apvNumber = tblpayables.apvNumber WHERE tblpurchasebook.glDate BETWEEN '{CALENDAR_FROM.get_date()}' AND'{CALENDAR_TO.get_date()}' ORDER BY tblpurchasebook.apvNumber")
            elif TEXTVAR_BOOKS.get() == "SB":
                if TEXTVAR_OPTION.get() == "Posted":
                    cursor.execute(f"SELECT DISTINCT source, tblsalesbook.soaNumber, tblreceivables.soaDate, tblsalesbook.glDate, tblreceivables.particulars, tblreceivables.reference, tblsalesbook.chartCode, tblsalesbook.amount, side, tblsalesbook.remarks, poster, DATE(posted) FROM tblsalesbook INNER JOIN tblreceivables ON tblsalesbook.soaNumber = tblreceivables.soaNumber WHERE isPosted = 'Yes' AND tblsalesbook.glDate BETWEEN '{CALENDAR_FROM.get_date()}' AND'{CALENDAR_TO.get_date()}' ORDER BY tblsalesbook.soaNumber")
                else:
                    cursor.execute(f"SELECT DISTINCT source, tblsalesbook.soaNumber, tblreceivables.soaDate, tblsalesbook.glDate, tblreceivables.particulars, tblreceivables.reference, tblsalesbook.chartCode, tblsalesbook.amount, side, tblsalesbook.remarks, poster, DATE(posted) FROM tblsalesbook INNER JOIN tblreceivables ON tblsalesbook.soaNumber = tblreceivables.soaNumber WHERE tblsalesbook.glDate BETWEEN '{CALENDAR_FROM.get_date()}' AND'{CALENDAR_TO.get_date()}' ORDER BY tblsalesbook.soaNumber")
            elif TEXTVAR_BOOKS.get() == "GJ":
                if TEXTVAR_OPTION.get() == "Posted":
                    cursor.execute(f"SELECT source, gjNumber, docDate, glDate, particulars, reference, chartCode, amount, side, remarks, encoder, DATE(encoded) FROM tblgeneraljournal WHERE isPosted = 'Yes' AND glDate BETWEEN '{CALENDAR_FROM.get_date()}' AND'{CALENDAR_TO.get_date()}' ORDER BY gjNumber")
                else:
                    cursor.execute(f"SELECT source, gjNumber, docDate, glDate, particulars, reference, chartCode, amount, side, remarks, encoder, DATE(encoded) FROM tblgeneraljournal WHERE glDate BETWEEN '{CALENDAR_FROM.get_date()}' AND'{CALENDAR_TO.get_date()}' ORDER BY gjNumber")

            result = cursor.fetchall()
            if result:
                wb = load_workbook(PATH_TEMPLATE + "BOOK.xlsx")
                sheet = wb.active
                sheet["A3"] = TEXTVAR_BOOKS.get()
                sheet["B5"] = f"{CALENDAR_FROM.get_date()} to {CALENDAR_TO.get_date()}"
                count = 8
                listcount = 0
                for i in result:
                    sheet["A" + str(count)] = i[0] #source
                    sheet["B" + str(count)] = i[1] #number
                    sheet["C" + str(count)] = i[2] #docdate
                    sheet["D" + str(count)] = i[3] #gldate
                    sheet["E" + str(count)] = i[4] #partic
                    sheet["F" + str(count)] = i[5] #ref
                    sheet["G" + str(count)] = i[6] #chart
                    sheet["H" + str(count)] = self.returnChartTitle(i[6])[0] #title
                    sheet["I" + str(count)] = i[7] #amount
                    sheet["J" + str(count)] = i[8] #side
                    sheet["K" + str(count)] = i[9] #remarks
                    sheet["L" + str(count)] = self.returnUserName(i[10],0) #poster
                    sheet["M" + str(count)] = i[11] #posted
                    count += 1
                    listcount += 1
                    PROGRESS_BAR["value"] = round((listcount/len(result))*100, 0)
                    FRAME_ACCOUNTING.update()
                sheet["A" + str(count)] = "/" + self.returnUserName(USER, 0) + str(datetime.date.today())
                wb.save(PATH_SAVE + TEXTVAR_BOOKS.get() + ".xlsx")
                startfile(PATH_SAVE + TEXTVAR_BOOKS.get() + ".xlsx", "open")
            else:
                messagebox.showerror("Reports", "No results found!")

### MENU_FINANCE ###
    def showAccountsPayable(self, *args):
        self.clearWorkspace()
        FRAME_PAYABLE = LabelFrame(FRAME_4, text = "Accounts Payable", font = APP_FONT)
        FRAME_PAYABLE.grid(column = 1, row = 0)

        global SUB1_PAYABLE
        SUB1_PAYABLE = Frame(FRAME_PAYABLE)
        SUB1_PAYABLE.grid(column = 0, row = 0, sticky = W, pady = SUBMENU_PADY)

        SUB2_PAYABLE = Frame(FRAME_PAYABLE)
        SUB2_PAYABLE.grid(column = 0, row = 1)

        LABEL_SEARCH = Label(SUB1_PAYABLE, text = "Search", width = 5, font = APP_FONT, anchor = W)
        LABEL_SEARCH.grid(column = 0, row = 0, pady = SUBMENU_PADY, sticky = W)

        TEXTVAR_SEARCH = StringVar()
        ENTRY_SEARCH = Entry(SUB1_PAYABLE, textvariable = TEXTVAR_SEARCH, width = 25, font = APP_FONT)
        ENTRY_SEARCH.grid(column = 1, row = 0, padx = MENU_PADX, sticky = W)
        ENTRY_SEARCH.bind("<Return>", lambda e: self.searchAPV(TEXTVAR_SEARCH.get()))

        BUTTON_GO = Button(SUB1_PAYABLE, text = "GO", width = 5, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = lambda: self.searchAPV(TEXTVAR_SEARCH.get()))
        BUTTON_GO.grid(column = 2, row = 0, padx = MENU_PADX)

        BUTTON_REFRESH = Button(SUB1_PAYABLE, text = "REFRESH", width = 10, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = self.refreshAPV)
        BUTTON_REFRESH.grid(column = 3, row = 0, padx = MENU_PADX)

        BUTTON_ADD = Button(SUB1_PAYABLE, text = "ADD", width = 5, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = self.showAddEditAccountsPayable)
        BUTTON_ADD.grid(column = 4, row = 0, padx = MENU_PADX)
        
        if self.returnAccess(USER, 6) == 0:
            BUTTON_ADD.config(state = DISABLED, cursor = "arrow")
        
        global CALENDAR_START
        CALENDAR_START = DateEntry(SUB1_PAYABLE, firstweekday = "sunday", date_pattern = "yyyy-mm-dd", state = "readonly")
        CALENDAR_START.grid(column = 5, row = 0, sticky = W, padx = TOP_PADX + 10)
        # CALENDAR_START.set_date(self.returnFirstDayOfMonth(str(CALENDAR_START.get_date())))
        
        global CALENDAR_END
        CALENDAR_END = DateEntry(SUB1_PAYABLE, firstweekday = "sunday", date_pattern = "yyyy-mm-dd", state = "readonly")
        CALENDAR_END.grid(column = 6, row = 0, sticky = W, padx = TOP_PADX + 10)

        global PROGRESS_BAR
        PROGRESS_BAR = tk.Progressbar(SUB1_PAYABLE, orient = HORIZONTAL, length = 200, mode = "determinate")
        PROGRESS_BAR.grid(column = 7, row = 0, padx = TOP_PADX + 10, sticky = E)

        global TREE_PAYABLE
        TREE_PAYABLE = tk.Treeview(SUB2_PAYABLE, height = 28, selectmode = "browse")
        TREE_PAYABLE["columns"] = ("APV No.", "Doc Date", "GL Date", "Payee", "Particulars", "Amount", "Balance", "isApproved", "Encoder", "Encoded", "Approver", "Approved")
        TREE_PAYABLE.column("#0", width = 0, minwidth = 0, stretch = True)
        TREE_PAYABLE.column("APV No.", anchor = W, minwidth = 75, width = 75)
        TREE_PAYABLE.column("Doc Date", anchor = W, minwidth = 75, width = 75)
        TREE_PAYABLE.column("GL Date", anchor = W, minwidth = 75, width = 75)
        TREE_PAYABLE.column("Payee", anchor = W, minwidth = 165, width = 150)
        TREE_PAYABLE.column("Particulars", anchor = W, minwidth = 165, width = 300)
        TREE_PAYABLE.column("Amount", anchor = E, minwidth = 100, width = 75)
        TREE_PAYABLE.column("Balance", anchor = E, minwidth = 100, width = 75)
        TREE_PAYABLE.column("isApproved", anchor = W, minwidth = 50, width = 70)
        TREE_PAYABLE.column("Encoder", anchor = W, minwidth = 75, width = 75)
        TREE_PAYABLE.column("Encoded", anchor = W, minwidth = 75, width = 75)
        TREE_PAYABLE.column("Approver", anchor = W, minwidth = 75, width = 75)
        TREE_PAYABLE.column("Approved", anchor = W, minwidth = 75, width = 75)
        
        TREE_PAYABLE.heading("#0", text = "", anchor = W)
        TREE_PAYABLE.heading("APV No.", text = "APV No.", anchor = N)
        TREE_PAYABLE.heading("Doc Date", text = "Doc Date", anchor = N)
        TREE_PAYABLE.heading("GL Date", text = "GL Date", anchor = N)
        TREE_PAYABLE.heading("Payee", text = "Payee", anchor = N)
        TREE_PAYABLE.heading("Particulars", text = "Particulars", anchor = N)
        TREE_PAYABLE.heading("Amount", text = "Amount", anchor = N)
        TREE_PAYABLE.heading("Balance", text = "Balance", anchor = N)
        TREE_PAYABLE.heading("isApproved", text = "isApproved", anchor = N)
        TREE_PAYABLE.heading("Encoder", text = "Encoder", anchor = N)
        TREE_PAYABLE.heading("Encoded", text = "Encoded", anchor = N)
        TREE_PAYABLE.heading("Approver", text = "Approver", anchor = N)
        TREE_PAYABLE.heading("Approved", text = "Approved", anchor = N)

        global POPUP_PAYABLE
        POPUP_PAYABLE = Menu(TREE_PAYABLE, tearoff = 0)
        POPUP_PAYABLE.add_command(command = self.editPayable, label = "Edit", state = DISABLED)
        POPUP_PAYABLE.add_command(command = self.createDV, label = "Create Disbursement", state = DISABLED)
        POPUP_PAYABLE.add_command(command = self.viewDV, label = "View Disbursement", state = DISABLED)
        TREE_PAYABLE.bind("<Button-3>", lambda e: self.modifyAPVPopupMenu(TREE_PAYABLE, POPUP_PAYABLE, e))
        TREE_PAYABLE.bind("<Double-1>", self.editPayable)

        global STYLE_PAYABLE
        STYLE_PAYABLE = tk.Style()
        STYLE_PAYABLE.map("Treeview", foreground = self.fixedMap("foreground", STYLE_PAYABLE), background = self.fixedMap("background", STYLE_PAYABLE))

        TREE_PAYABLE.tag_configure("oddrow", background = None)
        TREE_PAYABLE.tag_configure("evenrow", background = TREE_TAG_EVENROW)
        db.commit()
        cursor.execute(f"SELECT apvNumber, apvDate, glDate, supplierCode, particulars, amount, isApproved, encoder, DATE(encoded), approver, DATE(approved) FROM tblpayables WHERE apvDate BETWEEN '{CALENDAR_START.get_date()}' AND '{CALENDAR_END.get_date()}' ORDER BY apvNumber DESC")
        result = cursor.fetchall()

        YSCROLLBAR = tk.Scrollbar(SUB2_PAYABLE, orient = "vertical", command = TREE_PAYABLE.yview)
        YSCROLLBAR.pack(side = "right", fill = "y")

        XSCROLLBAR = tk.Scrollbar(SUB2_PAYABLE, orient = "horizontal", command = TREE_PAYABLE.xview)
        
        TREE_PAYABLE.configure(yscrollcommand = YSCROLLBAR.set, xscrollcommand = XSCROLLBAR.set) 
        TREE_PAYABLE.pack()
        XSCROLLBAR.pack(fill ="x")

        count = 0
        apvnumbers, skipped = [], []
        for i in result:
            if i[0] not in apvnumbers:
                if count % 2 == 0:
                    TREE_PAYABLE.insert("", "end", values = (str(i[0]).zfill(8),i[1],i[2],self.returnSupplierName(i[3]),i[4],self.returnTotalAPVNet(int(i[0])),self.returnTotalAPVBalance(int(i[0])),i[6],self.returnUserName(i[7], 0),i[8],self.returnUserName(i[9], 0),i[10]), tags = "evenrow")
                else:
                    TREE_PAYABLE.insert("", "end", values = (str(i[0]).zfill(8),i[1],i[2],self.returnSupplierName(i[3]),i[4],self.returnTotalAPVNet(int(i[0])),self.returnTotalAPVBalance(int(i[0])),i[6],self.returnUserName(i[7], 0),i[8],self.returnUserName(i[9], 0),i[10]), tags = "oddrow")
                count += 1
                apvnumbers.append(i[0])
                PROGRESS_BAR["value"] = round((count/(len(result)-len(skipped)))*100, 0)
                SUB1_PAYABLE.update()
            else:
                skipped.append(i[0])
        PROGRESS_BAR.grid_remove()

        try:
            TOP_PAYABLE.destroy()
        except:
            pass

    def showDisbursements(self, *args):
        self.clearWorkspace()
        FRAME_DISBURSEMENT = LabelFrame(FRAME_4, text = "Disbursements", font = APP_FONT)
        FRAME_DISBURSEMENT.grid(column = 1, row = 0)

        global SUB1_DISBURSEMENT
        SUB1_DISBURSEMENT = Frame(FRAME_DISBURSEMENT)
        SUB1_DISBURSEMENT.grid(column = 0, row = 0, sticky = W, pady = SUBMENU_PADY)

        SUB2_DISBURSEMENT = Frame(FRAME_DISBURSEMENT)
        SUB2_DISBURSEMENT.grid(column = 0, row = 1)

        LABEL_SEARCH = Label(SUB1_DISBURSEMENT, text = "Search", width = 5, font = APP_FONT, anchor = W)
        LABEL_SEARCH.grid(column = 0, row = 0, pady = SUBMENU_PADY, sticky = W)

        TEXTVAR_SEARCH = StringVar()
        ENTRY_SEARCH = Entry(SUB1_DISBURSEMENT, textvariable = TEXTVAR_SEARCH, width = 25, font = APP_FONT)
        ENTRY_SEARCH.grid(column = 1, row = 0, padx = MENU_PADX, sticky = W)
        ENTRY_SEARCH.bind("<Return>", lambda e: self.searchDV(TEXTVAR_SEARCH.get()))

        BUTTON_GO = Button(SUB1_DISBURSEMENT, text = "GO", width = 5, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = lambda: self.searchDV(TEXTVAR_SEARCH.get()))
        BUTTON_GO.grid(column = 2, row = 0, padx = MENU_PADX)
        
        BUTTON_REFRESH = Button(SUB1_DISBURSEMENT, text = "REFRESH", width = 10, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = self.refreshDV)
        BUTTON_REFRESH.grid(column = 3, row = 0, padx = MENU_PADX)

        BUTTON_ADD = Button(SUB1_DISBURSEMENT, text = "ADD", width = 5, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = self.showAddEditDisbursement)
        BUTTON_ADD.grid(column = 4, row = 0, padx = MENU_PADX)
        
        if self.returnAccess(USER, 9) == 0:
            BUTTON_ADD.config(state = DISABLED, cursor = "arrow")
        
        global CALENDAR_START
        CALENDAR_START = DateEntry(SUB1_DISBURSEMENT, firstweekday = "sunday", date_pattern = "yyyy-mm-dd", state = "readonly")
        CALENDAR_START.grid(column = 5, row = 0, sticky = W, padx = TOP_PADX + 10)
        # CALENDAR_START.set_date(self.returnFirstDayOfMonth(str(CALENDAR_START.get_date())))
        
        global CALENDAR_END
        CALENDAR_END = DateEntry(SUB1_DISBURSEMENT, firstweekday = "sunday", date_pattern = "yyyy-mm-dd", state = "readonly")
        CALENDAR_END.grid(column = 6, row = 0, sticky = W, padx = TOP_PADX + 10)

        global PROGRESS_BAR
        PROGRESS_BAR = tk.Progressbar(SUB1_DISBURSEMENT, orient = HORIZONTAL, length = 200, mode = "determinate")
        PROGRESS_BAR.grid(column = 7, row = 0, padx = TOP_PADX + 10, sticky = E)

        global TREE_DISBURSEMENT
        TREE_DISBURSEMENT = tk.Treeview(SUB2_DISBURSEMENT, height = 28, selectmode = "browse")
        TREE_DISBURSEMENT["columns"] = ("DV No.", "Doc Date", "GL Date", "Payee", "Description", "Amount", "Status", "Encoder", "Encoded", "Approver", "Approved")
        TREE_DISBURSEMENT.column("#0", width = 0, minwidth = 0, stretch = True)
        TREE_DISBURSEMENT.column("DV No.", anchor = W, minwidth = 75, width = 75)
        TREE_DISBURSEMENT.column("Doc Date", anchor = W, minwidth = 75, width = 75)
        TREE_DISBURSEMENT.column("GL Date", anchor = W, minwidth = 75, width = 75)
        TREE_DISBURSEMENT.column("Payee", anchor = W, minwidth = 100, width = 150)
        TREE_DISBURSEMENT.column("Description", anchor = W, minwidth = 100, width = 300)
        TREE_DISBURSEMENT.column("Amount", anchor = E, minwidth = 100, width = 75)
        TREE_DISBURSEMENT.column("Status", anchor = W, minwidth = 100, width = 75)
        TREE_DISBURSEMENT.column("Encoder", anchor = W, minwidth = 100, width = 75)
        TREE_DISBURSEMENT.column("Encoded", anchor = W, minwidth = 75, width = 50)
        TREE_DISBURSEMENT.column("Approver", anchor = W, minwidth = 75, width = 75)
        TREE_DISBURSEMENT.column("Approved", anchor = W, minwidth = 75, width = 75)
        
        TREE_DISBURSEMENT.heading("#0", text = "", anchor = W)
        TREE_DISBURSEMENT.heading("DV No.", text = "DV No.", anchor = N)
        TREE_DISBURSEMENT.heading("Doc Date", text = "Doc Date", anchor = N)
        TREE_DISBURSEMENT.heading("GL Date", text = "GL Date", anchor = N)
        TREE_DISBURSEMENT.heading("Payee", text = "Payee", anchor = N)
        TREE_DISBURSEMENT.heading("Description", text = "Description", anchor = N)
        TREE_DISBURSEMENT.heading("Amount", text = "Amount", anchor = N)
        TREE_DISBURSEMENT.heading("Status", text = "Status", anchor = N)
        TREE_DISBURSEMENT.heading("Encoder", text = "Encoder", anchor = N)
        TREE_DISBURSEMENT.heading("Encoded", text = "Encoded", anchor = N)
        TREE_DISBURSEMENT.heading("Approver", text = "Approver", anchor = N)
        TREE_DISBURSEMENT.heading("Approved", text = "Approved", anchor = N)

        POPUP_DISBURSEMENT = Menu(TREE_DISBURSEMENT, tearoff = 0)
        POPUP_DISBURSEMENT.add_command(command = self.editDV, label = "Edit")
        TREE_DISBURSEMENT.bind("<Button-3>", lambda e: self.popupMenu(TREE_DISBURSEMENT, POPUP_DISBURSEMENT, e))
        TREE_DISBURSEMENT.bind("<Double-1>", self.editDV)

        global STYLE_DISBURSEMENT
        STYLE_DISBURSEMENT = tk.Style()
        STYLE_DISBURSEMENT.map("Treeview", foreground = self.fixedMap("foreground", STYLE_DISBURSEMENT), background = self.fixedMap("background", STYLE_DISBURSEMENT))

        TREE_DISBURSEMENT.tag_configure("oddrow", background = None)
        TREE_DISBURSEMENT.tag_configure("evenrow", background = TREE_TAG_EVENROW)
        TREE_DISBURSEMENT.tag_configure("void", background = TREE_TAG_VOID)

        db.commit()
        cursor.execute(f"SELECT dvNumber, dvDate, glDate, payeeCode, payeeName, particulars, disbNet, isApproved, isVoid, encoder, DATE(encoded), approver, DATE(approved) FROM tbldisbursements WHERE dvDate BETWEEN '{CALENDAR_START.get_date()}' AND'{CALENDAR_END.get_date()}' ORDER BY dvNumber DESC")
        result = cursor.fetchall()

        YSCROLLBAR = tk.Scrollbar(SUB2_DISBURSEMENT, orient = "vertical", command = TREE_DISBURSEMENT.yview)
        YSCROLLBAR.pack(side = "right", fill = "y")

        XSCROLLBAR = tk.Scrollbar(SUB2_DISBURSEMENT, orient = "horizontal", command = TREE_DISBURSEMENT.xview)
        
        TREE_DISBURSEMENT.configure(yscrollcommand = YSCROLLBAR.set, xscrollcommand = XSCROLLBAR.set) 
        TREE_DISBURSEMENT.pack()
        XSCROLLBAR.pack(fill ="x")

        count = 0
        dvnumbers, skipped = [], []
        for i in result:
            if i[0] not in dvnumbers:
                if i[8] != "Yes":
                    if count % 2 == 0:
                        TREE_DISBURSEMENT.insert("", "end", values = (str(i[0]).zfill(8),i[1],i[2],self.returnPayeeName(i[3],i[4]),i[5],self.validateAmount2(self.returnDVNetAmount(i[0])),self.returnDVStatus(i[7],i[8]),self.returnUserName(i[9], 0),i[10],self.returnUserName(i[11], 0),i[12]), tags = "evenrow")
                    else:
                        TREE_DISBURSEMENT.insert("", "end", values = (str(i[0]).zfill(8),i[1],i[2],self.returnPayeeName(i[3],i[4]),i[5],self.validateAmount2(self.returnDVNetAmount(i[0])),self.returnDVStatus(i[7],i[8]),self.returnUserName(i[9], 0),i[10],self.returnUserName(i[11], 0),i[12]), tags = "oddrow")
                else:
                    TREE_DISBURSEMENT.insert("", "end", values = (str(i[0]).zfill(8),i[1],i[2],self.returnPayeeName(i[3],i[4]),i[5],self.validateAmount2(self.returnDVNetAmount(i[0])),self.returnDVStatus(i[7],i[8]),self.returnUserName(i[9], 0),i[10],self.returnUserName(i[11], 0),i[12]), tags = "void")
                count += 1
                dvnumbers.append(i[0])
                PROGRESS_BAR["value"] = round((count/(len(result)-len(skipped)))*100, 0)
                SUB1_DISBURSEMENT.update()
            else:
                skipped.append(i[0])
        PROGRESS_BAR.grid_remove()

        try:
            TOP_DISBURSEMENTS.destroy()
        except:
            pass

    def showFinanceReports(self, *args):
        self.clearWorkspace()
        FRAME_DISBURSEMENT = LabelFrame(FRAME_4, text = "Reports", font = APP_FONT)
        FRAME_DISBURSEMENT.grid(column = 1, row = 0)
        
        global SUB_FRAME1
        SUB_FRAME1 = Frame(FRAME_DISBURSEMENT)
        SUB_FRAME1.grid(column = 0, row = 1)
        
        global SUB_FRAME2
        SUB_FRAME2 = Frame(FRAME_DISBURSEMENT) #treeview
        SUB_FRAME2.grid(column = 0, row = 2)
        
        global TEXTVAR_REPORTS
        TEXTVAR_REPORTS = StringVar()
        COMBO_REPORTS = tk.Combobox(FRAME_DISBURSEMENT, values = ["Disbursement Summary", "Transaction Type Summary", "ISO Report"], textvariable = TEXTVAR_REPORTS, font = APP_FONT, width = 30, state = "readonly")
        COMBO_REPORTS.grid(column = 0, row = 0, sticky = W, ipadx = 2)
        COMBO_REPORTS.bind("<<ComboboxSelected>>", self.showSelectedFinanceReport)

    def showTextMaster(self, *args):
        self.clearWorkspace()
        FRAME_TEXT = LabelFrame(FRAME_4, text = "Text Master", font = APP_FONT)
        FRAME_TEXT.grid(column = 1, row = 0)

        SUB_FRAME1 = Frame(FRAME_TEXT)
        SUB_FRAME1.grid(column = 0, row = 0) #submenu

        global SUB_FRAME2
        SUB_FRAME2 = Frame(FRAME_TEXT)
        SUB_FRAME2.grid(column = 0, row = 1, sticky = W) #options
        
        global SUB_FRAME3
        SUB_FRAME3 = Frame(FRAME_TEXT)
        SUB_FRAME3.grid(column = 0, row = 2, sticky = W, pady = TOP_PADY) #tree
        
        global SUB_FRAME4
        SUB_FRAME4 = Frame(FRAME_TEXT)
        SUB_FRAME4.grid(column = 0, row = 3, sticky = E, pady = TOP_PADY + 5) #buttons
        
        global SUB_FRAME5
        SUB_FRAME5 = Frame(FRAME_TEXT)
        SUB_FRAME5.grid(column = 0, row = 4, pady = TOP_PADY + 5) #progress
        
        BUTTON_AUTO = Button(SUB_FRAME1, text = "AUTO", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "hand2", command = self.showTextAuto)
        BUTTON_AUTO.grid(column = 0, row = 0, padx = TOP_PADX + 20)
        
        BUTTON_MANUAL = Button(SUB_FRAME1, text = "MANUAL", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, state = DISABLED, cursor = "hand2", command = None)
        BUTTON_MANUAL.grid(column = 1, row = 0, padx = TOP_PADX + 20)
        
        BUTTON_SENT = Button(SUB_FRAME1, text = "SENT", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, state = DISABLED, cursor = "hand2", command = None)
        BUTTON_SENT.grid(column = 2, row = 0, padx = TOP_PADX + 20)
        
        BUTTON_SENT = Button(SUB_FRAME1, text = "EXTRACT", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, state = NORMAL, cursor = "arrow", command = self.showMLExtractor)
        BUTTON_SENT.grid(column = 3, row = 0, padx = TOP_PADX + 20)

### MENU_FINANCE_TEXT MASTER ###
    def showTextAuto(self):
        self.clearTextMasterFields()
        
        global TEXTVAR_CONNECTION, LABEL_CONNECTION
        TEXTVAR_CONNECTION = StringVar()
        LABEL_CONNECTION = Label(SUB_FRAME2, textvariable = TEXTVAR_CONNECTION, font = APP_FONT)
        LABEL_CONNECTION.grid(column = 0, row = 0, pady = TOP_PADY, sticky = E)
        TEXTVAR_CONNECTION.set("Offline")
        
        BUTTON_CONNECT = Button(SUB_FRAME2, text = "Connect", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "hand2", command = self.connectADB)
        BUTTON_CONNECT.grid(column = 1, row = 0, padx = TOP_PADX, sticky = W)
        BUTTON_CONNECT.bind("<Return>", self.connectADB)
        
        LABEL_PICKER = Label(SUB_FRAME2, text = "Mode", font = APP_FONT)
        LABEL_PICKER.grid(column = 0, row = 2, pady = TOP_PADY, sticky = E)
        
        global TEXTVAR_PICKER
        TEXTVAR_PICKER = StringVar()
        COMBO_PICKER = tk.Combobox(SUB_FRAME2, values = ["Cebuana Ref. No.", "MLhuillier KPTN"], textvariable = TEXTVAR_PICKER, font = APP_FONT, width = 20, state = "readonly")
        COMBO_PICKER.grid(column = 1, row = 2, sticky = W, ipadx = 2)
        
        LABEL_CONNECTION = Label(SUB_FRAME2, text = "Excel File", font = APP_FONT)
        LABEL_CONNECTION.grid(column = 0, row = 1, pady = TOP_PADY, sticky = E)
        
        global TEXTVAR_FILE
        TEXTVAR_FILE = StringVar()
        ENTRY_FILE = Entry(SUB_FRAME2, textvariable = TEXTVAR_FILE, font = APP_FONT, width = 60, state = "readonly")
        ENTRY_FILE.grid(column = 1, row = 1, sticky = W)
        
        BUTTON_FILE = Button(SUB_FRAME2, text = "...", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "hand2", command = self.getFileAddress)
        BUTTON_FILE.grid(column = 1, row = 1, padx = TOP_PADX, sticky = E)
        BUTTON_FILE.bind("<Return>", self.getFileAddress)
        
        BUTTON_TEMPLATE = Button(SUB_FRAME2, text = "TEMPLATE", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "hand2", state = DISABLED, command = lambda: startfile(PATH_TEMPLATE + "TXT.XLSX", "open"))
        BUTTON_TEMPLATE.grid(column = 2, row = 1, padx = TOP_PADX)
        BUTTON_TEMPLATE.bind("<Return>", lambda: startfile(PATH_TEMPLATE + "TXT.XLSX", "open"))
        
        BUTTON_SEND = Button(SUB_FRAME4, text = "SEND", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "hand2", command = self.sendTextFile)
        BUTTON_SEND.grid(column = 0, row = 0, padx = TOP_PADX)
        BUTTON_SEND.bind("<Return>", self.sendTextFile)
        
        global TEXTVAR_COUNT
        TEXTVAR_COUNT = StringVar()
        LABEL_COUNT = Label(SUB_FRAME5, textvariable = TEXTVAR_COUNT, font = APP_FONT)
        LABEL_COUNT.grid(column = 0, row = 1, sticky = E)

        global PROGRESS_BAR
        PROGRESS_BAR = tk.Progressbar(SUB_FRAME5, orient = HORIZONTAL, length = 885, mode = "determinate")
        PROGRESS_BAR.grid(column = 0, row = 0)
    
    def sendTextFile(self, *args):
        self.connectADB()
        remitter = TEXTVAR_PICKER.get()
        tab = "cmd /c adb shell input keyevent 22"
        send = "cmd /c adb shell input keyevent 66"
        sender = "INSERT INTO tblmessaging (sender, particulars, receiver, client, number, reference, amount, date, status, user) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
        sent = []
        now = str(datetime.date.today())
        if TEXTVAR_CONNECTION.get() == "Online":
            ask = messagebox.askyesno("Send", "Are you sure?")
            if ask:
                row = 2
                listcount = 0
                sheet["H1"] = "status"
                sheet["I1"] = "date sent"
                for i in sendlist:
                    if i[4] != "invalid":
                        message = f"{i[1]}, Receiver: {i[2].split(',')[0]}, Sender: {i[0]}, {remitter}: {i[5]}, Amount: P{i[6]}"
                        system(f"cmd /c adb shell am start -a android.intent.action.SENDTO -d sms:{i[4]} --es sms_body '{message}' --ez exit_on_sent true")
                        system(tab)
                        system(send)
                        sheet["H" + str(row)] = "sent"
                        sheet["I" + str(row)] = now
                        sent.append(1)
                        cursor.execute(sender, [i[0], i[1], i[2], i[3], i[4], i[5], self.returnFloatAmount(i[6]), now, "sent", USER])
                        db.commit()
                    row += 1
                    listcount += 1
                    PROGRESS_BAR["value"] = round((listcount/len(sendlist))*100, 0)
                    SUB_FRAME5.update()
                messagebox.showinfo("Send", f"Sending completed!\nSent: {len(sent)}\nTotal Records: {len(sent)}")
                wb.save(TEXTVAR_FILE.get())
                startfile(TEXTVAR_FILE.get(), "open")
        else:
            messagebox.showerror("Send", "Sending failed! No connection!")
    
    def getFileAddress(self):
        filename = askopenfilename(filetypes = (("Excel files", "*.xlsx"),("All files", "*.*")))
        if filename:
            TEXTVAR_FILE.set(filename)
            self.showTextTreeView(SUB_FRAME3)
            self.loadTextTreeView(TREE_TEXT)
        else:
            TEXTVAR_FILE.set("")
            try:
                for i in TREE_TEXT.get_children():
                    TREE_TEXT.delete(i)
            except:
                pass
            
    def getFileAddress2(self, filetype1, filetype2):
        filename = askopenfilename(filetypes = ((filetype1, filetype2),("All files", "*.*")))  #"PDF files", "*.pdf"
        if filename:
            TEXTVAR_FILE.set(filename)
        else:
            TEXTVAR_FILE.set("")
    
    def loadTextTreeView(self, tree):
        global sendlist, wb, sheet
        wb = load_workbook(TEXTVAR_FILE.get())
        sheet = wb.active
        sendlist = []
        count = 0
        for i in range(sheet.max_row-1):
            sender = sheet["A" + str(i+2)].value
            particulars = sheet["B" + str(i+2)].value
            receiver = sheet["C" + str(i+2)].value
            client = sheet["D" + str(i+2)].value
            number = sheet["E" + str(i+2)].value
            reference = sheet["F" + str(i+2)].value
            amount = sheet["G" + str(i+2)].value
            try:
                if len(self.validatePhoneNumber(number)) == 11:
                    sendlist.append([sender.upper(), particulars, receiver.upper(), client.upper(), self.validatePhoneNumber(number), self.formatQuickSendReferenceNumber(reference), self.formatAmount(amount)])
                    format = "valid"
                else:
                    sendlist.append([sender.upper(), particulars, receiver.upper(), client.upper(), "invalid", self.formatQuickSendReferenceNumber(reference), self.formatAmount(amount)])
                    format = "invalid"
                if count % 2 == 0:
                    tree.insert(parent = "", index = "end", values = (sender.upper(), particulars, receiver.upper(), client.upper(), self.validatePhoneNumber(number), self.formatQuickSendReferenceNumber(reference), self.validateAmount2(amount), format), tags = "evenrow")
                else:
                    tree.insert(parent = "", index = "end", values = (sender.upper(), particulars, receiver.upper(), client.upper(), self.validatePhoneNumber(number), self.formatQuickSendReferenceNumber(reference), self.validateAmount2(amount), format), tags = "oddrow")
            except:
                pass
            count += 1
    
    def formatQuickSendReferenceNumber(self, ref):
        kptn = str(ref).replace(" ","")
        if len(kptn) == 18:
            return f"{kptn[0:4]} {kptn[4:7]} {kptn[7:11]} {kptn[11:14]} {kptn[14:18]}"
        else:
            return kptn
    
    def validatePhoneNumber(self, num):
        try:
            if num > 0:
                return "0" + str(num)
        except:
            return num
    
    def showTextTreeView(self, frame):
        global TREE_TEXT
        TREE_TEXT = tk.Treeview(frame, height = 20, selectmode = "browse")
        TREE_TEXT["columns"] = ("Sender", "Particulars", "Receiver", "Client", "Number", "Reference", "Amount")
        TREE_TEXT.column("#0", width = 0, minwidth = 0, stretch = True)
        TREE_TEXT.column("Sender", anchor = W, minwidth = 150, width = 200)
        TREE_TEXT.column("Particulars", anchor = W, minwidth = 250, width = 280)
        TREE_TEXT.column("Receiver", anchor = W, minwidth = 150, width = 290)
        TREE_TEXT.column("Client", anchor = W, minwidth = 165, width = 150)
        TREE_TEXT.column("Number", anchor = W, minwidth = 165, width = 110)
        TREE_TEXT.column("Reference", anchor = W, minwidth = 100, width = 75)
        TREE_TEXT.column("Amount", anchor = E, minwidth = 100, width = 75)
        
        TREE_TEXT.heading("#0", text = "", anchor = W)
        TREE_TEXT.heading("Sender", text = "Sender", anchor = N)
        TREE_TEXT.heading("Particulars", text = "Particulars", anchor = N)
        TREE_TEXT.heading("Receiver", text = "Receiver", anchor = N)
        TREE_TEXT.heading("Client", text = "Client", anchor = N)
        TREE_TEXT.heading("Number", text = "Number", anchor = N)
        TREE_TEXT.heading("Reference", text = "Reference", anchor = N)
        TREE_TEXT.heading("Amount", text = "Amount", anchor = N)

        global STYLE_TEXT
        STYLE_TEXT = tk.Style()
        STYLE_TEXT.map("Treeview", foreground = self.fixedMap("foreground", STYLE_TEXT), background = self.fixedMap("background", STYLE_TEXT))

        TREE_TEXT.tag_configure("oddrow", background = None)
        TREE_TEXT.tag_configure("evenrow", background = TREE_TAG_EVENROW)

        YSCROLLBAR = tk.Scrollbar(frame, orient = "vertical", command = TREE_TEXT.yview)
        YSCROLLBAR.pack(side = "right", fill = "y")

        XSCROLLBAR = tk.Scrollbar(frame, orient = "horizontal", command = TREE_TEXT.xview)
        
        TREE_TEXT.configure(yscrollcommand = YSCROLLBAR.set, xscrollcommand = XSCROLLBAR.set) 
        TREE_TEXT.pack()
        XSCROLLBAR.pack(fill ="x")
        
    def connectADB(self):
        if system(f"cmd /c adb shell am start") == 1:
            TEXTVAR_CONNECTION.set("Offline")
            LABEL_CONNECTION.config(bg = "red2")
        else:
            TEXTVAR_CONNECTION.set("Online")
            LABEL_CONNECTION.config(bg = "lawn green")
    
    def clearTextMasterFields(self):
        for i in SUB_FRAME2.winfo_children():
            i.destroy()
        for i in SUB_FRAME3.winfo_children():
            i.destroy()
        for i in SUB_FRAME4.winfo_children():
            i.destroy()

    def showMLExtractor(self):
        self.clearTextMasterFields()
        
        LABEL_CONNECTION = Label(SUB_FRAME2, text = "PDF File", font = APP_FONT)
        LABEL_CONNECTION.grid(column = 0, row = 0, pady = TOP_PADY, sticky = E)
        
        global TEXTVAR_FILE
        TEXTVAR_FILE = StringVar()
        ENTRY_FILE = Entry(SUB_FRAME2, textvariable = TEXTVAR_FILE, font = APP_FONT, width = 60, state = "readonly")
        ENTRY_FILE.grid(column = 1, row = 1, sticky = W)
        
        BUTTON_FILE = Button(SUB_FRAME2, text = "...", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "hand2", command = self.getFileAddress2)
        BUTTON_FILE.grid(column = 1, row = 1, padx = TOP_PADX, sticky = E)
        BUTTON_FILE.bind("<Return>", self.getFileAddress2("PDF files", "*.pdf"))
        
        BUTTON_SEND = Button(SUB_FRAME4, text = "Extract", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "hand2", command = self.extractMLDetails)
        BUTTON_SEND.grid(column = 0, row = 0, padx = TOP_PADX)
        BUTTON_SEND.bind("<Return>", self.extractMLDetails)
    
    def extractMLDetails(self):
        wb = load_workbook(PATH_TEMPLATE + "EXTRACT.xlsx")
        sheet = wb.active
        reader = PdfReader(TEXTVAR_FILE.get())
        text = []
        for page in reader.pages:
            # sheet["A" + str(count)] = page.extract_text() + "\n"
            text.append([page.extract_text() + "\n"])
            # if page.extract_text().find("CLEMENTE, HANZEL") > 0:
            # text += page.extract_text() + "\n"
        
        textlist = []
        for i in text:
            for x in i:
                for y in x.split("\n"):
                    textlist.append(y)
        
        processors = ["GEERLEEN MIRA ", "ANDREA KEITH ABAD ", "ROSEMARIE BALASTA "]
        processor = ""
        receiver = ""
        count = 2
        for i in textlist:
            if i.find(processors[0]) >= 0:
                receiver = i.replace(processors[0],"")
                processor = processors[0]
            if i.find(processors[1]) >= 0:
                receiver = i.replace(processors[1],"")
                processor = processors[1]
            if i.find(processors[2]) >= 0:
                receiver = i.replace(processors[2],"")
                processor = processors[2]
            newreceiver = receiver.replace(" CLEMENTE, HANZEL ","")
            allnewreceiver = newreceiver.replace(processor, "")
            if i.find("CLEMENTE, HANZEL ") > 0:
                receiver += i.replace("CLEMENTE, HANZEL ","")
            try:
                if i.find("ROSARIO") >= 0 and int(i[7]) >= 0:
                    sheet["A" + str(count)] = i.split(" ")[1] #control#
                    sheet["B" + str(count)] = "HANZEL ROSARIO CLEMENTE" #sender
                    sheet["C" + str(count)] = allnewreceiver #receiver
                    sheet["D" + str(count)] = i.split(" ")[2] #kptn
                    sheet["E" + str(count)] = i.split(" ")[4] #or#
                    sheet["F" + str(count)] = i.split(" ")[6] #amount
                    sheet["G" + str(count)] = i.split("  ")[2].split(" ")[0] #charge
                    sheet["H" + str(count)] = processor #processor
                    count += 1
            except:
                pass
        wb.save(PATH_SAVE + "EXTRACT.xlsx")
        startfile(PATH_SAVE + "EXTRACT.xlsx", "open")

### MENU_FINANCE_ACCOUNTS PAYABLE ###
    def showAddEditAccountsPayable(self, *args):
        global TOP_PAYABLE
        TOP_PAYABLE = Toplevel()
        TOP_PAYABLE.title("Create - Accounts Payable")
        TOP_PAYABLE.iconbitmap(PATH_ICON + "icon.ico")
        TOP_PAYABLE.geometry("1195x650+100+20")
        TOP_PAYABLE.resizable(height = False, width = False)
        TOP_PAYABLE.grab_set()
        TOP_PAYABLE.focus()
        # TOP_PAYABLE.protocol("WM_DELETE_WINDOW", self.showAccountsPayable)

        global required
        required = []

        SUB_FRAME1 = Frame(TOP_PAYABLE)
        SUB_FRAME1.grid(column = 0, row = 0, sticky = W)

        LABEL_APV = Label(SUB_FRAME1, text = "APV No.", font = APP_FONT)
        LABEL_APV.grid(column = 0, row = 0, pady = TOP_PADY, sticky = E)

        LABEL_GL = Label(SUB_FRAME1, text = "GL Date", font = APP_FONT)
        LABEL_GL.grid(column = 3, row = 0, pady = TOP_PADY, sticky = E)

        LABEL_DOC = Label(SUB_FRAME1, text = "Doc Date", font = APP_FONT)
        LABEL_DOC.grid(column = 3, row = 1, pady = TOP_PADY, sticky = E)

        LABEL_DUE = Label(SUB_FRAME1, text = "Due Date", font = APP_FONT)
        LABEL_DUE.grid(column = 3, row = 2, pady = TOP_PADY, sticky = NE)

        LABEL_PAYEE = Label(SUB_FRAME1, text = "Payee", font = APP_FONT)
        LABEL_PAYEE.grid(column = 0, row = 1, pady = TOP_PADY, sticky = E)

        LABEL_DESC = Label(SUB_FRAME1, text = "Description", font = APP_FONT)
        LABEL_DESC.grid(column = 0, row = 2, pady = TOP_PADY, sticky = NE)

        LABEL_DESC = Label(SUB_FRAME1, text = "Reference", font = APP_FONT)
        LABEL_DESC.grid(column = 0, row = 3, pady = TOP_PADY, sticky = E)

        global TEXTVAR_APV
        TEXTVAR_APV = StringVar()
        ENTRY_APV = Entry(SUB_FRAME1, textvariable = TEXTVAR_APV, font = APP_FONT, width = 15, justify = RIGHT, state = "readonly")
        ENTRY_APV.grid(column = 1, row = 0, sticky = W)

        SUB_1_1 = Frame(SUB_FRAME1)
        SUB_1_1.grid(column = 1, row = 1, sticky = W)

        global TEXTVAR_PAYEECODE, ENTRY_PAYEECODE
        TEXTVAR_PAYEECODE = StringVar()
        ENTRY_PAYEECODE = Entry(SUB_1_1, textvariable = TEXTVAR_PAYEECODE, font = APP_FONT, width = 15)
        ENTRY_PAYEECODE.grid(column = 0, row = 0, sticky = W)
        ENTRY_PAYEECODE.bind("<FocusOut>", lambda e: self.populateAPVFields(TEXTVAR_PAYEECODE.get()))
        required.append(TEXTVAR_PAYEECODE)

        global TEXTVAR_PAYEENAME
        TEXTVAR_PAYEENAME = StringVar()
        ENTRY_PAYEENAME = Entry(SUB_1_1, textvariable = TEXTVAR_PAYEENAME, font = APP_FONT, width = 55, state = "readonly")
        ENTRY_PAYEENAME.grid(column = 1, row = 0, sticky = W)

        global BUTTON_PAYEE
        BUTTON_PAYEE = Button(SUB_1_1, text = "...", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, cursor = "hand2", command = self.showSupplierSelection)
        BUTTON_PAYEE.grid(column = 0, row = 0, sticky = E)

        global ENTRY_DESC
        ENTRY_DESC = Text(SUB_FRAME1, font = APP_FONT, width = 70, height = 3)
        ENTRY_DESC.grid(column = 1, row = 2, sticky = W)
        required.append(ENTRY_DESC)

        SUB_1_2 = Frame(SUB_FRAME1)
        SUB_1_2.grid(column = 1, row = 3, sticky = W)

        global TEXTVAR_REF, ENTRY_REF
        TEXTVAR_REF = StringVar()
        ENTRY_REF = Entry(SUB_1_2, textvariable = TEXTVAR_REF, font = APP_FONT, width = 30)
        ENTRY_REF.grid(column = 0, row = 0, sticky = W)
        required.append(TEXTVAR_REF)
        
        LABEL_DESC = Label(SUB_1_2, text = "RR #", font = APP_FONT)
        LABEL_DESC.grid(column = 1, row = 0, pady = TOP_PADY, sticky = E)
        
        global TEXTVAR_RR
        TEXTVAR_RR = StringVar()
        ENTRY_RR = Entry(SUB_1_2, textvariable = TEXTVAR_RR, font = APP_FONT, width = 10, state = DISABLED)
        ENTRY_RR.grid(column = 2, row = 0, sticky = W, padx = TOP_PADX)
        
        LABEL_DESC = Label(SUB_1_2, text = "PO #", font = APP_FONT)
        LABEL_DESC.grid(column = 3, row = 0, pady = TOP_PADY, sticky = E)
        
        global TEXTVAR_PO
        TEXTVAR_PO = StringVar()
        ENTRY_PO = Entry(SUB_1_2, textvariable = TEXTVAR_PO, font = APP_FONT, width = 10, state = DISABLED)
        ENTRY_PO.grid(column = 4, row = 0, sticky = W, padx = TOP_PADX)

        LABEL_DIVIDER = Label(SUB_FRAME1, text = "", font = APP_FONT)
        LABEL_DIVIDER.grid(column = 2, row = 0, padx = TOP_PADX + 20, sticky = N)

        global CALENDAR_GL, TEXTVAR_GL
        TEXTVAR_GL= StringVar()
        CALENDAR_GL = DateEntry(SUB_FRAME1, textvariable = TEXTVAR_GL, firstweekday = "sunday", date_pattern = "yyyy-mm-dd")
        CALENDAR_GL.grid(column = 4, row = 0, sticky = W)
        CALENDAR_GL.set_date(self.returnLastDayOfMonth(str(datetime.datetime.today())))
        CALENDAR_GL.bind("<FocusOut>", lambda e: self.formatDate(CALENDAR_GL, TEXTVAR_GL))

        global CALENDAR_DOC
        CALENDAR_DOC = DateEntry(SUB_FRAME1, firstweekday = "sunday", date_pattern = "yyyy-mm-dd")
        CALENDAR_DOC.grid(column = 4, row = 1, sticky = W)

        global CALENDAR_DUE
        CALENDAR_DUE = DateEntry(SUB_FRAME1, firstweekday = "sunday", date_pattern = "yyyy-mm-dd")
        CALENDAR_DUE.grid(column = 4, row = 2, sticky = NW)

        LABEL_DIVIDER = Label(SUB_FRAME1, text = "", font = APP_FONT)
        LABEL_DIVIDER.grid(column = 5, row = 0, padx = TOP_PADX + 20, sticky = N)

        LABEL_GROSS = Label(SUB_FRAME1, text = "Gross", font = APP_FONT)
        LABEL_GROSS.grid(column = 6, row = 0, pady = TOP_PADY, sticky = E)

        LABEL_EXPENSE = Label(SUB_FRAME1, text = "Expense", font = APP_FONT)
        LABEL_EXPENSE.grid(column = 6, row = 1, pady = TOP_PADY, sticky = E)

        LABEL_VAT = Label(SUB_FRAME1, text = "VAT", font = APP_FONT)
        LABEL_VAT.grid(column = 6, row = 2, pady = TOP_PADY, sticky = NE)

        LABEL_EWT = Label(SUB_FRAME1, text = "EWT", font = APP_FONT, width = 10, anchor = E)
        LABEL_EWT.grid(column = 8, row = 0, pady = TOP_PADY, sticky = E)

        LABEL_NET = Label(SUB_FRAME1, text = "Net", font = APP_FONT, width = 10, anchor = E)
        LABEL_NET.grid(column = 8, row = 1, pady = TOP_PADY, sticky = E)

        global TEXTVAR_TGROSS
        TEXTVAR_TGROSS = StringVar()
        ENTRY_GROSSTOTAL = Entry(SUB_FRAME1, textvariable = TEXTVAR_TGROSS, font = APP_FONT, width = 12, justify = RIGHT, state = "readonly")
        ENTRY_GROSSTOTAL.grid(column = 7, row = 0, sticky = W)

        global TEXTVAR_TEXPENSE
        TEXTVAR_TEXPENSE = StringVar()
        ENTRY_EXPENSETOTAL = Entry(SUB_FRAME1, textvariable = TEXTVAR_TEXPENSE, font = APP_FONT, width = 12, justify = RIGHT, state = "readonly")
        ENTRY_EXPENSETOTAL.grid(column = 7, row = 1, sticky = W)

        global TEXTVAR_TVAT
        TEXTVAR_TVAT = StringVar()
        ENTRY_VATTOTAL = Entry(SUB_FRAME1, textvariable = TEXTVAR_TVAT, font = APP_FONT, width = 12, justify = RIGHT, state = "readonly")
        ENTRY_VATTOTAL.grid(column = 7, row = 2, sticky = NW)

        global TEXTVAR_TEWT
        TEXTVAR_TEWT = StringVar()
        ENTRY_EWTTOTAL = Entry(SUB_FRAME1, textvariable = TEXTVAR_TEWT, font = APP_FONT, width = 12, justify = RIGHT, state = "readonly")
        ENTRY_EWTTOTAL.grid(column = 9, row = 0, sticky = W)

        global TEXTVAR_TNET
        TEXTVAR_TNET = StringVar()
        ENTRY_NETTOTAL = Entry(SUB_FRAME1, textvariable = TEXTVAR_TNET, font = APP_FONT, width = 12, justify = RIGHT, state = "readonly")
        ENTRY_NETTOTAL.grid(column = 9, row = 1, sticky = W)

        LABEL_DIVIDER = Label(TOP_PAYABLE, text = "", font = APP_FONT)
        LABEL_DIVIDER.grid(column = 0, row = 1, pady = TOP_PADY, sticky = N)

        SUB_FRAME2 = Frame(TOP_PAYABLE)
        SUB_FRAME2.grid(column = 0, row = 2, sticky = W)

        SUB_FRAME3 = Frame(TOP_PAYABLE)
        SUB_FRAME3.grid(column = 0, row = 3, sticky = W)

        LABEL_CODE = Label(SUB_FRAME2, text = "Code", width = 19, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_CODE.grid(column = 0, row = 0)

        LABEL_TITLE = Label(SUB_FRAME2, text = "Title", width = 25, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_TITLE.grid(column = 1, row = 0)

        LABEL_CENTER = Label(SUB_FRAME2, text = "Cost Center", width = 20, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_CENTER.grid(column = 2, row = 0)

        LABEL_GROSS = Label(SUB_FRAME2, text = "Gross", width = 12, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_GROSS.grid(column = 3, row = 0)

        LABEL_TAX = Label(SUB_FRAME2, text = "Tax", width = 9, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_TAX.grid(column = 4, row = 0)

        LABEL_EXPENSE = Label(SUB_FRAME2, text = "Expense", width = 12, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_EXPENSE.grid(column = 5, row = 0)

        LABEL_VAT = Label(SUB_FRAME2, text = "VAT", width = 12, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_VAT.grid(column = 6, row = 0)

        LABEL_EWT = Label(SUB_FRAME2, text = "EWT", width = 12, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_EWT.grid(column = 7, row = 0)

        LABEL_NET = Label(SUB_FRAME2, text = "Net", width = 12, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_NET.grid(column = 8, row = 0)

        LABEL_DESC = Label(SUB_FRAME2, text = "Description", width = 22, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_DESC.grid(column = 9, row = 0)

        LABEL_CANCEL = Label(SUB_FRAME2, text = "[x]", width = 2, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_CANCEL.grid(column = 10, row = 0)

        self.createScrollFrame(SUB_FRAME3, 200, 1165, 20, 0, 0)
        global apvFrame, apvCode, apvTitle, apvCenterCode, apvCenterName, apvGross, apvTax, apvExpense, apvVat, apvEwt, apvNet, apvDesc, apvClear
        apvFrame, apvCode, apvTitle, apvCenterCode, apvCenterName, apvGross, apvTax, apvExpense, apvVat, apvEwt, apvNet, apvDesc, apvClear  = [], [], [], [], [], [], [], [], [], [], [], [], []

        for i in range(8):
            self.showPayableLines(SCROLLABLE_FRAME, i)

        FRAME_ENTRY = Frame(TOP_PAYABLE)
        FRAME_ENTRY.grid(column = 0, row = 4, sticky = W, pady = TOP_PADY + 10)
        
        global SUB_FRAME6
        SUB_FRAME6 = Frame(FRAME_ENTRY) #entry label frame
        SUB_FRAME6.grid(column = 0, row = 3, sticky = NW)
        
        global SUB_FRAME5
        SUB_FRAME5 = Frame(FRAME_ENTRY) #entry lines frame
        SUB_FRAME5.grid(column = 0, row = 4, sticky = NW)
        
        LABEL_CODE = Label(SUB_FRAME6, text = "Code", width = 12, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_CODE.grid(column = 0, row = 0)

        LABEL_TITLE = Label(SUB_FRAME6, text = "Title", width = 30, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_TITLE.grid(column = 1, row = 0)

        LABEL_CENTER = Label(SUB_FRAME6, text = "Amount", width = 15, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_CENTER.grid(column = 2, row = 0)

        LABEL_GROSS = Label(SUB_FRAME6, text = "Dr/Cr", width = 10, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_GROSS.grid(column = 3, row = 0)

        LABEL_EXPENSE = Label(SUB_FRAME6, text = "Remarks", width = 31, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_EXPENSE.grid(column = 4, row = 0)

        LABEL_X = Label(SUB_FRAME6, text = "[x]", width = 2, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_X.grid(column = 5, row = 0)
        
        self.createScrollFrame(SUB_FRAME5, 210, 735, 20, 0, 0)
        
        global entryFrame, entryCode, entryTitle, entryAmount, entryDrCr, entryRemarks, entryClear
        entryFrame, entryCode, entryTitle, entryAmount, entryDrCr, entryRemarks, entryClear = [], [], [], [], [], [], []
        
        for i in range(15):
            self.showSOAEntryLines(SCROLLABLE_FRAME, i)

        SUB_FRAME7 = Frame(SUB_FRAME5) #entry total frame
        SUB_FRAME7.grid(column = 1, row = 0, sticky = NW)
        
        LABEL_DEBIT = Label(SUB_FRAME7, text = "Debit", width = 12, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_DEBIT.grid(column = 0, row = 0)

        LABEL_CREDIT = Label(SUB_FRAME7, text = "Credit", width = 12, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_CREDIT.grid(column = 0, row = 1)

        LABEL_VARIANCE = Label(SUB_FRAME7, text = "Variance", width = 12, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_VARIANCE.grid(column = 0, row = 2)
        
        global TEXTVAR_DEBIT
        TEXTVAR_DEBIT = StringVar()
        ENTRY_DEBIT = Entry(SUB_FRAME7, textvariable = TEXTVAR_DEBIT, font = APP_FONT, width = 12, justify = RIGHT, state = "readonly")
        ENTRY_DEBIT.grid(column = 1, row = 0, sticky = W)
        
        global TEXTVAR_CREDIT
        TEXTVAR_CREDIT = StringVar()
        ENTRY_CREDIT = Entry(SUB_FRAME7, textvariable = TEXTVAR_CREDIT, font = APP_FONT, width = 12, justify = RIGHT, state = "readonly")
        ENTRY_CREDIT.grid(column = 1, row = 1, sticky = W)
        
        global TEXTVAR_VARIANCE
        TEXTVAR_VARIANCE = StringVar()
        ENTRY_VARIANCE = Entry(SUB_FRAME7, textvariable = TEXTVAR_VARIANCE, font = APP_FONT, width = 12, justify = RIGHT, state = "readonly")
        ENTRY_VARIANCE.grid(column = 1, row = 2, sticky = W)
        
        FRAME_BUTTON = Frame(TOP_PAYABLE)
        FRAME_BUTTON.grid(column = 0, row = 4, sticky = SE, pady = TOP_PADY + 10)
        
        global BUTTON_SUBMIT
        BUTTON_SUBMIT = Button(FRAME_BUTTON, text = "SUBMIT", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "hand2", command = self.savePayable)
        BUTTON_SUBMIT.grid(column = 0, row = 0, padx = TOP_PADX)

        global BUTTON_APPROVE
        BUTTON_APPROVE = Button(FRAME_BUTTON, text = "APPROVE", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG_GRN, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, state = DISABLED, cursor = "hand2", command = self.approvePayable)
        BUTTON_APPROVE.grid(column = 1, row = 0, padx = TOP_PADX)
        
        global BUTTON_VOID
        BUTTON_VOID = Button(FRAME_BUTTON, text = "VOID", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG_RED, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, state = DISABLED, cursor = "hand2", command = self.voidPayable)
        BUTTON_VOID.grid(column = 2, row = 0, padx = TOP_PADX)

        global BUTTON_PRINT
        BUTTON_PRINT = Button(FRAME_BUTTON, text = "PRINT", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, state = DISABLED, cursor = "hand2", command = self.printPayable)
        BUTTON_PRINT.grid(column = 3, row = 0, padx = TOP_PADX)

        global BUTTON_CLOSE
        BUTTON_CLOSE = Button(FRAME_BUTTON, text = "CLOSE", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "hand2", command = lambda: self.closeTopLevel(TOP_PAYABLE))
        BUTTON_CLOSE.grid(column = 5, row = 0, padx = TOP_PADX)

    def showPayableLines(self, frame, rown):
        FRAME_LINE = Frame(frame)
        FRAME_LINE.grid(column = 0, row = rown, sticky = W)
        apvFrame.append(FRAME_LINE)

        FRAME_CODE = Frame(FRAME_LINE)
        FRAME_CODE.grid(column = 0, row = 0, sticky = W, ipadx = 1)

        global TEXTVAR_CODE
        TEXTVAR_CODE = StringVar()
        ENTRY_CODE = Entry(FRAME_CODE, textvariable = TEXTVAR_CODE, font = APP_FONT, width = 17)
        ENTRY_CODE.grid(column = 0, row = 0, sticky = W)
        ENTRY_CODE.bind("<FocusOut>", lambda e: self.populateChartFields(apvCode[rown], apvTitle[rown]))
        apvCode.append(TEXTVAR_CODE)

        BUTTON_CODE = Button(FRAME_CODE, text = "...", font = BUTTON_FONT2, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, cursor = "hand2", command = lambda: self.showChartSelection(rown, apvCode, apvTitle))
        BUTTON_CODE.grid(column = 1, row = 0, sticky = W)

        global TEXTVAR_TITLE
        TEXTVAR_TITLE = StringVar()
        ENTRY_TITLE = Entry(FRAME_LINE, textvariable = TEXTVAR_TITLE, font = APP_FONT, width = 25, state = "readonly")
        ENTRY_TITLE.grid(column = 1, row = 0, sticky = W, ipadx = 1)
        apvTitle.append(TEXTVAR_TITLE)

        FRAME_CENTER = Frame(FRAME_LINE)
        FRAME_CENTER.grid(column = 2, row = 0, sticky = W)

        global TEXTVAR_CENTERCODE
        TEXTVAR_CENTERCODE = StringVar()
        ENTRY_CENTERCODE = Entry(FRAME_CENTER, textvariable = TEXTVAR_CENTERCODE, font = APP_FONT, width = 9)
        ENTRY_CENTERCODE.grid(column = 0, row = 0, sticky = W, ipadx = 1)
        ENTRY_CENTERCODE.bind("<FocusOut>", lambda e: self.returnCenterName(apvCenterCode[rown], apvCenterName[rown]))
        apvCenterCode.append(TEXTVAR_CENTERCODE)
        
        global TEXTVAR_CENTERNAME
        TEXTVAR_CENTERNAME = StringVar()
        ENTRY_CENTERNAME = Entry(FRAME_CENTER, textvariable = TEXTVAR_CENTERNAME, font = APP_FONT, width = 10, state = DISABLED)
        ENTRY_CENTERNAME.grid(column = 1, row = 0, sticky = W, ipadx = 1)
        apvCenterName.append(TEXTVAR_CENTERNAME)

        BUTTON_CENTER = Button(FRAME_CENTER, text = "...", font = BUTTON_FONT2, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, cursor = "hand2", command = lambda: self.showCenterSelection(rown, apvCenterCode, apvCenterName))
        BUTTON_CENTER.grid(column = 1, row = 0, sticky = E)

        global TEXTVAR_GROSS
        TEXTVAR_GROSS = StringVar()
        ENTRY_GROSS = Entry(FRAME_LINE, textvariable = TEXTVAR_GROSS, font = APP_FONT, width = 12, justify = RIGHT)
        ENTRY_GROSS.grid(column = 3, row = 0, sticky = W, ipadx = 1)
        ENTRY_GROSS.bind("<FocusOut>", lambda e: self.updateAPVLine(apvGross[rown], apvTax[rown], rown))
        apvGross.append(TEXTVAR_GROSS)

        global TEXTVAR_TAX
        TEXTVAR_TAX = StringVar()
        COMBO_TAX = tk.Combobox(FRAME_LINE, values = ["NV-00", "NV-01", "NV-02", "NV-05", "NV-10", "NV-15", "WV-01", "WV-02", "WV-05", "WV-10", "WV-15", "WV-00"], textvariable = TEXTVAR_TAX, font = APP_FONT, width = 6, state = "readonly")
        COMBO_TAX.grid(column = 4, row = 0, sticky = W, ipadx = 2)
        COMBO_TAX.bind("<<ComboboxSelected>>", lambda e: self.updateAPVLine(apvGross[rown], apvTax[rown], rown))
        apvTax.append(TEXTVAR_TAX)

        global TEXTVAR_EXPENSE
        TEXTVAR_EXPENSE = StringVar()
        ENTRY_EXPENSE = Entry(FRAME_LINE, textvariable = TEXTVAR_EXPENSE, font = APP_FONT, width = 12, justify = RIGHT, state = "readonly")
        ENTRY_EXPENSE.grid(column = 5, row = 0, sticky = W, ipadx = 1)
        apvExpense.append(TEXTVAR_EXPENSE)

        global TEXTVAR_VAT
        TEXTVAR_VAT = StringVar()
        ENTRY_VAT = Entry(FRAME_LINE, textvariable = TEXTVAR_VAT, font = APP_FONT, width = 12, justify = RIGHT, state = "readonly")
        ENTRY_VAT.grid(column = 6, row = 0, sticky = W, ipadx = 1)
        apvVat.append(TEXTVAR_VAT)

        global TEXTVAR_EWT
        TEXTVAR_EWT = StringVar()
        ENTRY_EWT = Entry(FRAME_LINE, textvariable = TEXTVAR_EWT, font = APP_FONT, width = 12, justify = RIGHT, state = "readonly")
        ENTRY_EWT.grid(column = 7, row = 0, sticky = W, ipadx = 1)
        apvEwt.append(TEXTVAR_EWT)

        global TEXTVAR_NET
        TEXTVAR_NET = StringVar()
        ENTRY_NET = Entry(FRAME_LINE, textvariable = TEXTVAR_NET, font = APP_FONT, width = 12, justify = RIGHT, state = "readonly")
        ENTRY_NET.grid(column = 8, row = 0, sticky = W, ipadx = 1)
        apvNet.append(TEXTVAR_NET)

        global TEXTVAR_DESC
        TEXTVAR_DESC = StringVar()
        ENTRY_DESC = Entry(FRAME_LINE, textvariable = TEXTVAR_DESC, font = APP_FONT, width = 22)
        ENTRY_DESC.grid(column = 9, row = 0, sticky = W, ipadx = 1)
        apvDesc.append(TEXTVAR_DESC)

        global BUTTON_CLEAR
        BUTTON_CLEAR = Button(FRAME_LINE, text = "[X]", font = BUTTON_FONT2, bg = BUTTON_BG, fg = BUTTON_FG_RED, activebackground = CLICK_BG, cursor = "hand2", command = None)
        BUTTON_CLEAR.grid(column = 10, row = 0, sticky = W)
        apvClear.append(BUTTON_CLEAR)

    def editPayable(self, *args):
        self.copySelection(TREE_PAYABLE)
        self.showAddEditAccountsPayable()
        TOP_PAYABLE.title("Edit - APV")
        
        select = """SELECT 
            apvNumber, apvDate, glDate, dueDate, supplierCode,
            particulars, reference, rrNumber, chartCode, centerCode,
            amount, taxType, remarks, encoder, encoded,
            isApproved, isVoid
                FROM tblpayables WHERE apvNumber = %s"""
                
        selectbook = """SELECT 
            chartCode, amount, side, remarks
                FROM tblpurchasebook WHERE apvNumber = %s"""
        
        db.commit()
        cursor.execute(selectbook, [int(content[0])])
        result2 = cursor.fetchall()
        db.commit()
        cursor.execute(select, [int(content[0])])
        result = cursor.fetchall()
        if result:
            TEXTVAR_APV.set(str(result[0][0]).zfill(8))
            TEXTVAR_PAYEECODE.set(result[0][4])
            TEXTVAR_PAYEENAME.set(self.returnSupplierName(result[0][4]))
            ENTRY_DESC.insert("1.0", result[0][5])
            TEXTVAR_REF.set(result[0][6])
            if int(result[0][7]) == 0:
                TEXTVAR_RR.set(0)
                TEXTVAR_PO.set(0)
            else:
                TEXTVAR_RR.set(str(result[0][7]).zfill(8))
                TEXTVAR_PO.set(str(str(self.returnPONumber(int(result[0][7])))).zfill(8))
            CALENDAR_GL.set_date(result[0][2])
            CALENDAR_DOC.set_date(result[0][1])
            CALENDAR_DUE.set_date(result[0][3])
            
            count = 0
            for i in result:
                apvCode[count].set(i[8])
                self.populateChartFields(apvCode[count], apvTitle[count])
                apvCenterCode[count].set(i[9])
                self.returnCenterName(apvCenterCode[count], apvCenterName[count])
                apvGross[count].set(self.validateAmount2(i[10]))
                apvTax[count].set(i[11])
                self.updateAPVLine(apvGross[count], apvTax[count], count)
                apvDesc[count].set(i[12])
                count += 1
                
            if result2:
                self.clearSOAEntryLines()
                count = 0
                for i in result2:
                    entryCode[count].set(i[0])
                    self.populateChartFields(entryCode[count], entryTitle[count])
                    entryAmount[count].set(self.validateAmount2(i[1]))
                    entryDrCr[count].set(i[2])
                    entryRemarks[count].set(i[3])
                    count += 1
                self.updateSOAEntriesTotals()
            
            if result[0][15] == "No":
                BUTTON_SUBMIT.config(state = NORMAL, cursor = "hand2")
                BUTTON_APPROVE.config(state = NORMAL, cursor = "hand2")
                BUTTON_VOID.config(state = DISABLED, cursor = "arrow")
                BUTTON_PRINT.config(state = NORMAL, cursor = "hand2")
            else:
                self.disableAPVWidgets()
                if result[0][16] == "Yes":
                    BUTTON_SUBMIT.config(state = DISABLED, cursor = "arrow")
                    BUTTON_APPROVE.config(state = DISABLED, cursor = "arrow")
                    BUTTON_VOID.config(state = DISABLED, cursor = "arrow")
                    BUTTON_PRINT.config(state = NORMAL, cursor = "hand2")
                else:
                    BUTTON_SUBMIT.config(state = DISABLED, cursor = "arrow")
                    BUTTON_APPROVE.config(state = DISABLED, cursor = "arrow")
                    BUTTON_VOID.config(state = NORMAL, cursor = "hand2")
                    BUTTON_PRINT.config(state = NORMAL, cursor = "hand2")
                    
            if self.returnAccess(USER, 8) == 0:
                BUTTON_VOID.config(state = DISABLED, cursor = "arrow")
                
            if self.returnAccess(USER, 7) == 0:
                BUTTON_APPROVE.config(state = DISABLED, cursor = "arrow")

    def savePayable(self):
        insertAPV = """INSERT INTO tblpayables (
            apvNumber, apvDate, glDate, dueDate, supplierCode,
            particulars, reference, rrNumber, chartCode, centerCode,
            amount, taxType, remarks, encoder, encoded,
            isApproved, isVoid
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
            
        deleteAPV = "DELETE FROM tblpayables WHERE apvNumber = %s"
        
        deleteBook = "DELETE FROM tblpurchasebook WHERE apvNumber = %s"
        insertBook = """INSERT INTO tblpurchasebook (
                glDate, apvNumber, chartCode, amount, side,
                remarks, isPosted, isVoid, source, poster, 
                posted
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
            
        wrong = []
        for i in required:
            try:
                if i.get() == "":
                    wrong.append("Check APV details!")
            except:
                if i.get("1.0", END) == "":
                    wrong.append("Please fill up the particulars field!")
        for i in range(len(apvFrame)):
            if self.returnFloatAmount(apvNet[i].get()) > 0 and apvCode[i].get() == "":
                wrong.append("Check entry codes!")
        if self.returnFloatAmount(TEXTVAR_TGROSS.get()) == 0:
            wrong.append("Total Payable is zero!")
        if self.returnFloatAmount(TEXTVAR_VARIANCE.get()) != 0:
            wrong.append("Check variance in entries")
        if self.returnPeriodStatus(CALENDAR_GL.get_date()) == "locked":
            wrong.append("GL date is locked")
        if len(wrong) > 0:
            messagebox.showerror("Accounts Payable Voucher", wrong)
        else:
            ask = messagebox.askyesno("Accounts Payable Voucher", "Are you sure?")
            if ask:
                if TEXTVAR_APV.get() == "":
                    APVnumber = int(self.generateAPVNumber())
                else:
                    cursor.execute(deleteAPV, [int(TEXTVAR_APV.get())])
                    db.commit()
                    cursor.execute(deleteBook, [int(TEXTVAR_APV.get())])
                    db.commit()
                    APVnumber = int(TEXTVAR_APV.get())
                validlines = []
                try:
                    if int(TEXTVAR_RR.get()) != 0:
                        rr = int(TEXTVAR_RR.get())
                    else:
                        rr = 0
                except:
                    rr = 0
                for i in range(len(apvFrame)):
                    if self.returnFloatAmount(apvNet[i].get()) != 0 and apvCode[i].get() != "":
                        validlines.append([
                            APVnumber, CALENDAR_DOC.get_date(), CALENDAR_GL.get_date(), CALENDAR_DUE.get_date(), TEXTVAR_PAYEECODE.get(), 
                            ENTRY_DESC.get("1.0", END).replace("\n", ""), TEXTVAR_REF.get(), rr, int(apvCode[i].get()), int(apvCenterCode[i].get()),
                            self.returnFloatAmount(apvGross[i].get()), apvTax[i].get(), apvDesc[i].get(), USER, datetime.datetime.now(),
                            "No", "No"
                        ])
                
                validbook = []
                for i in range(len(entryFrame)):
                    if entryCode[i].get() != "" and self.returnFloatAmount(entryAmount[i].get()) > 0:
                        validbook.append([
                            CALENDAR_GL.get_date(), APVnumber, int(entryCode[i].get()), self.returnFloatAmount(entryAmount[i].get()), entryDrCr[i].get(),
                            entryRemarks[i].get(), "No", "No", "APV", USER,
                            datetime.datetime.now()
                        ])
                
                if len(validlines) > 0 and len(validbook) > 0:
                    for i in validlines:
                        cursor.execute(insertAPV, i)
                    db.commit()

                    for i in validbook:
                        cursor.execute(insertBook, i)
                    db.commit()

                    messagebox.showinfo("Accounts Payable Voucher", "APV has been saved!")
                    TEXTVAR_APV.set(str(APVnumber).zfill(8))
                    TOP_PAYABLE.focus()
                    BUTTON_APPROVE.config(state = NORMAL, cursor = "hand2")
                    BUTTON_VOID.config(state = DISABLED, cursor = "arrow")
                    BUTTON_PRINT.config(state = NORMAL, cursor = "hand2")
                    
                    if self.returnAccess(USER, 7) == 0:
                        BUTTON_APPROVE.config(state = DISABLED, cursor = "arrow")
                else:
                    messagebox.showerror("Accounts Payable Voucher", "No valid lines detected!")
                    TOP_PAYABLE.focus()

    def approvePayable(self):
        insertBook = """INSERT INTO tblpurchasebook (
                glDate, apvNumber, chartCode, amount, side,
                remarks, isPosted, isVoid, source, poster, 
                posted
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
                
        deleteBook = "DELETE FROM tblpurchasebook WHERE apvNumber = %s"
        validbook = []
        for i in range(len(entryFrame)):
            if entryCode[i].get() != "" and self.returnFloatAmount(entryAmount[i].get()) > 0:
                validbook.append([
                    CALENDAR_GL.get_date(), int(TEXTVAR_APV.get()), int(entryCode[i].get()), self.returnFloatAmount(entryAmount[i].get()), entryDrCr[i].get(),
                    entryRemarks[i].get(), "No", "No", "APV", USER,
                    datetime.datetime.now()
                ])
        
        approveAPV = "UPDATE tblpayables SET approver = %s, approved = %s, isApproved = %s WHERE apvNumber = %s"
        ask = messagebox.askyesno("APPROVE Accounts Payable Voucher", "Are you sure?")
        if ask:
            cursor.execute(deleteBook, [int(TEXTVAR_APV.get())])
            db.commit()
            
            if len(validbook) > 0:
                for i in validbook:
                    cursor.execute(insertBook, i)
                db.commit()
            
            cursor.execute(approveAPV, [USER, datetime.datetime.now(), "Yes", int(TEXTVAR_APV.get())])
            db.commit()
            
            messagebox.showinfo("APPROVE Accounts Payable Voucher", "APV has been approved!")
            TOP_PAYABLE.focus()
            BUTTON_SUBMIT.config(state = DISABLED, cursor = "arrow")
            BUTTON_APPROVE.config(state = DISABLED, cursor = "arrow")
            BUTTON_VOID.config(state = NORMAL, cursor = "hand2")
            BUTTON_PRINT.config(state = NORMAL, cursor = "hand2")
            
            if self.returnAccess(USER, 8) == 0:
                BUTTON_VOID.config(state = DISABLED, cursor = "arrow")
            self.disableAPVWidgets()
    
    def printPayable(self):
        wb = load_workbook(PATH_TEMPLATE + "APV.xlsx")
        sheet = wb.active
        sheet["A1"] = f"Printed Date: {datetime.datetime.now()} {self.returnUserName(USER, 0)}"
        sheet["D3"] = f"APV{TEXTVAR_APV.get()}"
        sheet["A9"] = TEXTVAR_PAYEENAME.get()
        sheet["C9"] = datetime.datetime.strptime(str(CALENDAR_DOC.get_date()), '%Y-%m-%d').strftime('%B %d, %Y')
        sheet["D9"] = datetime.datetime.strptime(str(CALENDAR_DUE.get_date()), '%Y-%m-%d').strftime('%B %d, %Y')
        sheet["A11"] = self.convertNumberToWords(TEXTVAR_TNET.get())
        sheet["D11"] = self.returnFloatAmount(TEXTVAR_TNET.get())
        try:
            if int(TEXTVAR_RR.get()) == 0:
                sheet["A13"] = ENTRY_DESC.get("1.0", END).replace("\n", "")
            else:
                sheet["A13"] = ENTRY_DESC.get("1.0", END).replace("\n", "") + " RR#" + TEXTVAR_RR.get() + " PO#" + TEXTVAR_PO.get()
        except:
            sheet["A13"] = ENTRY_DESC.get("1.0", END).replace("\n", "")
        sheet["A34"] = f"{self.returnUserName(USER, 1).upper()} {self.returnUserName(USER, 2).upper()}"
        sheet["A35"] = self.returnUserName(USER, 3)
        sheet["D34"] = TEXTVAR_REF.get()

        count = 15
        for i in range(len(entryFrame)):
            if entryCode[i].get() != "":
                sheet["A" + str(count)] = entryCode[i].get()
                sheet["B" + str(count)] = entryTitle[i].get()
                if entryDrCr[i].get() == "Debit":
                    sheet["C" + str(count)] = self.returnFloatAmount(entryAmount[i].get())
                else:
                    sheet["D" + str(count)] = self.returnFloatAmount(entryAmount[i].get())
                count += 1
            
        for i in range(15,35):
            cell = sheet["A" + str(i)].value
            if cell == None:
                sheet.row_dimensions[i].hidden = True
                    
        wb.save(PATH_SAVE + "APV.xlsx")
        startfile(PATH_SAVE + "APV.xlsx", "open")
    
    def voidPayable(self):
        insertbook = """INSERT INTO tblgeneraljournal (
            gjNumber, docDate, glDate, particulars, reference,
            source, chartcode, amount, side, remarks,
            isPosted, isVoid, encoder, encoded) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
        
        validbook = []
        for i in range(len(entryFrame)):
            if entryCode[i].get() != "" and self.returnFloatAmount(entryAmount[i].get()) > 0:
                if entryDrCr[i].get() == "Debit":
                    side = "Credit"
                else:
                    side = "Debit"
                validbook.append([
                    int(self.generateGJNumber()), datetime.date.today(), self.returnLastDayOfMonth(str(datetime.date.today())), "REVERSAL OF APV#" + TEXTVAR_APV.get() + " " + ENTRY_DESC.get("1.0", END).replace("\n",""), TEXTVAR_APV.get(),
                    "APV", int(entryCode[i].get()), self.returnFloatAmount(entryAmount[i].get()), side, entryRemarks[i].get(),
                    "No", "No", USER, datetime.datetime.now()
                ])
                
        voidAPV = "UPDATE tblpayables SET isVoid = %s WHERE apvNumber = %s"
        ask = messagebox.askyesno("VOID Accounts Payable Voucher", "Are you sure?")
        if ask:
            cursor.execute(voidAPV, ["Yes", int(TEXTVAR_APV.get())])
            db.commit()
            
            if len(validbook) != 0:
                for i in validbook:
                    cursor.execute(insertbook, i)
                db.commit()

            messagebox.showinfo("VOID Accounts Payable Voucher", "APV has been voided!")
            TOP_PAYABLE.focus()
            self.disableAPVWidgets()
            BUTTON_SUBMIT.config(state = DISABLED, cursor = "arrow")
            BUTTON_APPROVE.config(state = DISABLED, cursor = "arrow")
            BUTTON_VOID.config(state = DISABLED, cursor = "arrow")
            BUTTON_PRINT.config(state = NORMAL, cursor = "hand2")
    
    def returnAPVStatus(self, var):
        select = "SELECT isApproved, isVoid FROM tblpayables WHERE apvNumber = %s LIMIT 1"
        cursor.execute(select, [var])
        result = cursor.fetchone()
        if result:
            return result
    
    def disableAPVWidgets(self):
        ENTRY_PAYEECODE.config(state = DISABLED)
        BUTTON_PAYEE.config(state = DISABLED, cursor = "arrow")
        ENTRY_DESC.config(state = DISABLED)
        ENTRY_REF.config(state = DISABLED)
        CALENDAR_GL.config(state = DISABLED)
        CALENDAR_DOC.config(state = DISABLED)
        CALENDAR_DUE.config(state = DISABLED)
        for i in apvFrame:
            for x in i.winfo_children():
                try:
                    x.config(state = DISABLED)
                except:
                    for y in x.winfo_children():
                        y.config(state = DISABLED)
        for i in entryFrame:
            for x in i.winfo_children():
                x.config(state = DISABLED)
        
    def updateAPVLine(self, gross, tax, rown):
        self.validateAmount(gross)
        if self.returnFloatAmount(gross.get()) != 0 and tax.get() != "":
            if tax.get()[:2] == "WV":
                expense = self.returnFloatAmount(gross.get())/1.12
                vat = (self.returnFloatAmount(gross.get())/1.12)*.12
                ewt = (self.returnFloatAmount(gross.get())/1.12)*float("0." + tax.get()[3:])
                net = self.returnFloatAmount(gross.get()) - ewt
            else:
                expense = self.returnFloatAmount(gross.get())
                vat = 0
                if tax.get()[3:] == "00":
                    ewt = 0
                    net = self.returnFloatAmount(gross.get())
                else:
                    ewt = self.returnFloatAmount(gross.get())*float("0." + tax.get()[3:])
                    net = self.returnFloatAmount(gross.get()) - ewt
        else:
            expense = 0
            vat = 0
            ewt = 0
            net = 0
            
        apvExpense[rown].set(self.validateAmount2(expense))
        apvVat[rown].set(self.validateAmount2(vat))
        apvEwt[rown].set(self.validateAmount2(ewt))
        apvNet[rown].set(self.validateAmount2(net))
        
        self.updateAPVTotals()
        
    def updateAPVTotals(self):
        tgross, texpense, tvat, tewt, tnet = [], [], [], [], []
        for i in range(len(apvFrame)):
            try:
                if self.returnFloatAmount(apvNet[i].get()) != 0:
                    tgross.append(self.returnFloatAmount(apvGross[i].get()))
                    texpense.append(self.returnFloatAmount(apvExpense[i].get()))
                    tvat.append(self.returnFloatAmount(apvVat[i].get()))
                    tewt.append(self.returnFloatAmount(apvEwt[i].get()))
                    tnet.append(self.returnFloatAmount(apvNet[i].get()))
            except:
                pass
            TEXTVAR_TGROSS.set(self.validateAmount2(sum(tgross)))
            TEXTVAR_TEXPENSE.set(self.validateAmount2(sum(texpense)))
            TEXTVAR_TVAT.set(self.validateAmount2(sum(tvat)))
            TEXTVAR_TEWT.set(self.validateAmount2(sum(tewt)))
            TEXTVAR_TNET.set(self.validateAmount2(sum(tnet)))
            
        self.updateAPVEntries()
            
    def updateAPVEntries(self):
        self.clearSOAEntryLines()
        if self.returnFloatAmount(TEXTVAR_TNET.get()) > 0:
            count = 0
            for i in range(len(apvFrame)):
                if self.returnFloatAmount(apvNet[i].get()) > 0 and apvCode[i].get() != "":
                    entryCode[count].set(apvCode[i].get())
                    self.populateChartFields(entryCode[count], entryTitle[count])
                    entryAmount[count].set(self.validateAmount2(self.returnFloatAmount(apvExpense[i].get())))
                    entryDrCr[count].set("Debit")
                    count += 1
            if self.returnFloatAmount(TEXTVAR_TVAT.get()) > 0:
                entryCode[count].set("117050")
                self.populateChartFields(entryCode[count], entryTitle[count])
                entryAmount[count].set(self.validateAmount2(self.returnFloatAmount(TEXTVAR_TVAT.get())))
                entryDrCr[count].set("Debit")
                count += 1
            if self.returnFloatAmount(TEXTVAR_TEWT.get()) > 0:
                entryCode[count].set("213050")
                self.populateChartFields(entryCode[count], entryTitle[count])
                entryAmount[count].set(self.validateAmount2(self.returnFloatAmount(TEXTVAR_TEWT.get())))
                entryDrCr[count].set("Credit")
                count += 1
            if self.returnFloatAmount(TEXTVAR_TNET.get()) > 0:
                entryCode[count].set("211010")
                self.populateChartFields(entryCode[count], entryTitle[count])
                entryAmount[count].set(self.validateAmount2(self.returnFloatAmount(TEXTVAR_TNET.get())))
                entryDrCr[count].set("Credit")
                count += 1
        self.updateSOAEntriesTotals()
    
    def populateAPVFields(self, var):
        self.capitalLetters(var)
        select = "SELECT code, name FROM tblsuppliers WHERE code = %s LIMIT 1"
        cursor.execute(select, [var])
        result = cursor.fetchone()
        if result:
            TEXTVAR_PAYEECODE.set(result[0])
            TEXTVAR_PAYEENAME.set(result[1])
        else:
            TEXTVAR_PAYEECODE.set("")
            TEXTVAR_PAYEENAME.set("")

    def returnCenterName(self, var, var2):
        select = "SELECT name FROM tblcenters WHERE code = %s LIMIT 1"
        cursor.execute(select, [var.get()])
        result = cursor.fetchone()
        if result:
            var2.set(result)
        else:
            var.set("")
            var2.set("")

    def computeAPVNetAmount(self, var, tax):
        if tax[:2] == "WV":
            ewt = (float(var)/1.12)*float("0." + tax[3:])
            net = float(var) - ewt
        else:
            if tax[3:] == "00":
                ewt = 0
                net = float(var)
            else:
                ewt = float(var)*float("0." + tax[3:])
                net = float(var) - ewt

        return net

    def generateAPVNumber(self):
        cursor.execute("SELECT MAX(apvNumber) FROM tblpayables")
        result = cursor.fetchone()
        if result[0] == None:
            return str(795).zfill(8)
        else:
            return str(int(result[0])+1).zfill(8)
    
    def returnTotalAPVNet(self, var):
        select = "SELECT amount, taxType FROM tblpayables WHERE apvNumber = %s"
        db.commit()
        cursor.execute(select, [var])
        result = cursor.fetchall()
        totalnet = []
        for i in result:
            totalnet.append(self.computeAPVNetAmount(i[0], i[1]))
        return self.validateAmount2(sum(totalnet))
    
    def returnTotalAPVGross(self, var):
        select = "SELECT amount FROM tblpayables WHERE apvNumber = %s"
        db.commit()
        cursor.execute(select, [var])
        result = cursor.fetchall()
        totalnet = []
        for i in result:
            totalnet.append(i[0])
        return self.validateAmount2(sum(totalnet))
        
    def returnTotalAPVBalance(self, var):
        select = "SELECT amount, taxType, isVoid FROM tblpayables WHERE apvNumber = %s "
        db.commit()
        cursor.execute(select, [var])
        result = cursor.fetchall()
        totalnet = []
        if result[0][2] == "No":
            for i in result:
                totalnet.append(self.computeAPVNetAmount(i[0], i[1]))
            return self.validateAmount2(sum(totalnet)-float(self.returnDVNetAmount2(str(var).zfill(8))))
        else:
            return self.validateAmount2(0)

    def returnDVStatus(self, var1, var2):
        if var1 == "No":
            return "For approval"
        else:
            if var2 == "No":
                return "Approved"
            else:
                return "Void"

    def searchAPV(self, var, *args):
        db.commit()
        find = "SELECT apvNumber, apvDate, glDate, supplierCode, particulars, amount, isApproved, encoder, DATE(encoded), approver, DATE(approved) FROM tblpayables WHERE (apvNumber LIKE %s OR supplierCode LIKE %s OR particulars LIKE %s OR reference LIKE %s) AND apvDate BETWEEN %s AND %s ORDER BY apvNumber DESC"
        cursor.execute(find, [f"%{var}%", f"%{var}%", f"%{var}%", f"%{var}%", CALENDAR_START.get_date(), CALENDAR_END.get_date()])
        result = cursor.fetchall()
        for i in TREE_PAYABLE.get_children():
            TREE_PAYABLE.delete(i)
        if result:
            PROGRESS_BAR.grid(column = 7, row = 0, padx = TOP_PADX + 10, sticky = E)
            count = 0
            apvnumbers, skipped = [], []
            for i in result:
                if i[0] not in apvnumbers:
                    if count % 2 == 0:
                        TREE_PAYABLE.insert("", "end", values = (str(i[0]).zfill(8),i[1],i[2],self.returnSupplierName(i[3]),i[4],self.returnTotalAPVNet(int(i[0])),self.returnTotalAPVBalance(int(i[0])),i[6],self.returnUserName(i[7], 0),i[8],self.returnUserName(i[9], 0),i[10]), tags = "evenrow")
                    else:
                        TREE_PAYABLE.insert("", "end", values = (str(i[0]).zfill(8),i[1],i[2],self.returnSupplierName(i[3]),i[4],self.returnTotalAPVNet(int(i[0])),self.returnTotalAPVBalance(int(i[0])),i[6],self.returnUserName(i[7], 0),i[8],self.returnUserName(i[9], 0),i[10]), tags = "oddrow")
                    count += 1
                    apvnumbers.append(i[0])
                    PROGRESS_BAR["value"] = round((count/(len(result)-len(skipped)))*100, 0)
                    SUB1_PAYABLE.update()
                else:
                    skipped.append(i[0])
            PROGRESS_BAR.grid_remove()
        else:
            messagebox.showerror("Payables", "No match found!")

    def refreshAPV(self, *args):
        for i in TREE_PAYABLE.get_children():
            TREE_PAYABLE.delete(i)
        db.commit()
        cursor.execute(f"SELECT apvNumber, apvDate, glDate, supplierCode, particulars, amount, isApproved, encoder, DATE(encoded), approver, DATE(approved) FROM tblpayables WHERE apvDate BETWEEN '{CALENDAR_START.get_date()}' AND '{CALENDAR_END.get_date()}' ORDER BY apvNumber DESC")
        result = cursor.fetchall()
        PROGRESS_BAR.grid(column = 7, row = 0, padx = TOP_PADX + 10, sticky = E)
        count = 0
        apvnumbers, skipped = [], []
        for i in result:
            if i[0] not in apvnumbers:
                if count % 2 == 0:
                    TREE_PAYABLE.insert("", "end", values = (str(i[0]).zfill(8),i[1],i[2],self.returnSupplierName(i[3]),i[4],self.returnTotalAPVNet(int(i[0])),self.returnTotalAPVBalance(int(i[0])),i[6],self.returnUserName(i[7], 0),i[8],self.returnUserName(i[9], 0),i[10]), tags = "evenrow")
                else:
                    TREE_PAYABLE.insert("", "end", values = (str(i[0]).zfill(8),i[1],i[2],self.returnSupplierName(i[3]),i[4],self.returnTotalAPVNet(int(i[0])),self.returnTotalAPVBalance(int(i[0])),i[6],self.returnUserName(i[7], 0),i[8],self.returnUserName(i[9], 0),i[10]), tags = "oddrow")
                count += 1
                apvnumbers.append(i[0])
                PROGRESS_BAR["value"] = round((count/(len(result)-len(skipped)))*100, 0)
                SUB1_PAYABLE.update()
            else:
                skipped.append(i[0])
        PROGRESS_BAR.grid_remove()

    def modifyAPVPopupMenu(self, tree, pop, e, *args):
        selection = tree.identify_row(e.y)
        if selection:
            tree.selection_set(selection)
        self.copySelection(tree)
        if self.returnUserName(USER, 3) == "ASD":
            pop.entryconfigure("Edit", state = NORMAL)
        if float(self.returnDVNetAmount2(str(content[0]).zfill(8))) > 0:
            pop.entryconfigure("Create Disbursement", state = DISABLED)
            pop.entryconfigure("View Disbursement", state = NORMAL)
        else:
            if content[7] == "Yes": #approved
                pop.entryconfigure("Create Disbursement", state = NORMAL)
                pop.entryconfigure("View Disbursement", state = DISABLED)
            else:
                pop.entryconfigure("Create Disbursement", state = DISABLED)
                pop.entryconfigure("View Disbursement", state = DISABLED)
        self.popupMenu(tree, pop, e)

### MENU_FINANCE_DISBURSEMENT ###
    def showAddEditDisbursement(self, *args):
        global TOP_DISBURSEMENTS
        TOP_DISBURSEMENTS = Toplevel()
        TOP_DISBURSEMENTS.title("Create - Disbursement")
        TOP_DISBURSEMENTS.iconbitmap(PATH_ICON + "icon.ico")
        TOP_DISBURSEMENTS.geometry("985x645+100+20")
        TOP_DISBURSEMENTS.resizable(height = False, width = False)
        TOP_DISBURSEMENTS.grab_set()
        TOP_DISBURSEMENTS.focus()
        # TOP_DISBURSEMENTS.protocol("WM_DELETE_WINDOW", self.showDisbursements)

        global required
        required = []

        SUB_FRAME1 = Frame(TOP_DISBURSEMENTS) #details
        SUB_FRAME1.grid(column = 0, row = 0, sticky = W)

        SUB_SUB1FRAME1 = Frame(SUB_FRAME1) #details-1
        SUB_SUB1FRAME1.grid(column = 0, row = 0, sticky = W)

        LABEL_PAYEE = Label(SUB_SUB1FRAME1, text = "Payee", font = APP_FONT)
        LABEL_PAYEE.grid(column = 0, row = 0, pady = TOP_PADY, sticky = W)

        LABEL_PARTICULARS = Label(SUB_SUB1FRAME1, text = "Particulars", font = APP_FONT)
        LABEL_PARTICULARS.grid(column = 0, row = 1, pady = TOP_PADY, sticky = W)

        LABEL_REFERENCE = Label(SUB_SUB1FRAME1, text = "Reference", font = APP_FONT)
        LABEL_REFERENCE.grid(column = 0, row = 2, pady = TOP_PADY, sticky = W)

        LABEL_SIGN1 = Label(SUB_SUB1FRAME1, text = "Signatory 1", font = APP_FONT)
        LABEL_SIGN1.grid(column = 0, row = 3, pady = TOP_PADY, sticky = W)

        LABEL_SIGN2 = Label(SUB_SUB1FRAME1, text = "Signatory 2", font = APP_FONT)
        LABEL_SIGN2.grid(column = 0, row = 4, pady = TOP_PADY, sticky = W)

        SUB_1_1 = Frame(SUB_SUB1FRAME1)
        SUB_1_1.grid(column = 1, row = 0, sticky = W)

        global TEXTVAR_PAYEECODE, ENTRY_PAYEE
        TEXTVAR_PAYEECODE = StringVar()
        ENTRY_PAYEE = Entry(SUB_1_1, textvariable = TEXTVAR_PAYEECODE, font = APP_FONT, width = 15)
        ENTRY_PAYEE.grid(column = 0, row = 0, sticky = W)
        ENTRY_PAYEE.bind("<FocusOut>", lambda e: self.populateAPVFields(TEXTVAR_PAYEECODE.get()))
        
        global BUTTON_PAYEE
        BUTTON_PAYEE = Button(SUB_1_1, text = "...", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, cursor = "hand2", command = self.showSupplierSelection)
        BUTTON_PAYEE.grid(column = 0, row = 0, sticky = E)
        
        global TEXTVAR_PAYEENAME, ENTRY_PAYEENAME
        TEXTVAR_PAYEENAME = StringVar()
        ENTRY_PAYEENAME = Entry(SUB_1_1, textvariable = TEXTVAR_PAYEENAME, font = APP_FONT, width = 55, state = "readonly")
        ENTRY_PAYEENAME.grid(column = 1, row = 0, sticky = W)
        ENTRY_PAYEENAME.bind("<FocusOut>", lambda e: self.capitalLetters(TEXTVAR_PAYEENAME))
        required.append(TEXTVAR_PAYEENAME)

        global ENTRY_PARTICULARS
        ENTRY_PARTICULARS = Text(SUB_SUB1FRAME1, font = APP_FONT, width = 70, height = 2)
        ENTRY_PARTICULARS.grid(column = 1, row = 1, sticky = W)
        required.append(ENTRY_PARTICULARS)

        global TEXTVAR_REFERENCE, ENTRY_REFERENCE
        TEXTVAR_REFERENCE = StringVar()
        ENTRY_REFERENCE = Entry(SUB_SUB1FRAME1, textvariable = TEXTVAR_REFERENCE, font = APP_FONT, width = 70)
        ENTRY_REFERENCE.grid(column = 1, row = 2, sticky = W)

        global TEXTVAR_SIGN1, COMBO_SIGN1
        TEXTVAR_SIGN1 = StringVar()
        COMBO_SIGN1 = tk.Combobox(SUB_SUB1FRAME1, values = self.returnSignatories(), textvariable = TEXTVAR_SIGN1, font = APP_FONT, width = 35, state = "readonly")
        COMBO_SIGN1.grid(column = 1, row = 3, sticky = W)
        required.append(TEXTVAR_SIGN1)

        global TEXTVAR_SIGN2, COMBO_SIGN2
        TEXTVAR_SIGN2 = StringVar()
        COMBO_SIGN2 = tk.Combobox(SUB_SUB1FRAME1, values = self.returnSignatories(), textvariable = TEXTVAR_SIGN2, font = APP_FONT, width = 35, state = "readonly")
        COMBO_SIGN2.grid(column = 1, row = 4, sticky = W)
        required.append(TEXTVAR_SIGN2)

        SUB_SUB1FRAME2 = Frame(SUB_FRAME1) #details-2
        SUB_SUB1FRAME2.grid(column = 1, row = 0, sticky = NW, padx = TOP_PADX + 10)

        LABEL_DVNUMBER = Label(SUB_SUB1FRAME2, text = "DV #", font = APP_FONT)
        LABEL_DVNUMBER.grid(column = 0, row = 0, pady = TOP_PADY, sticky = W)

        LABEL_DVDATE = Label(SUB_SUB1FRAME2, text = "DV Date", font = APP_FONT)
        LABEL_DVDATE.grid(column = 0, row = 1, pady = TOP_PADY, sticky = W)

        LABEL_GLDATE = Label(SUB_SUB1FRAME2, text = "GL Date", font = APP_FONT)
        LABEL_GLDATE.grid(column = 0, row = 2, pady = TOP_PADY, sticky = W)
        
        LABEL_TYPE = Label(SUB_SUB1FRAME2, text = "Type", font = APP_FONT)
        LABEL_TYPE.grid(column = 0, row = 3, pady = TOP_PADY, sticky = W)

        global TEXTVAR_DVNUMBER
        TEXTVAR_DVNUMBER = StringVar()
        ENTRY_DVNUMBER = Entry(SUB_SUB1FRAME2, textvariable = TEXTVAR_DVNUMBER, font = APP_FONT, width = 13, state = "readonly")
        ENTRY_DVNUMBER.grid(column = 1, row = 0, sticky = W)

        global CALENDAR_DVDATE
        CALENDAR_DVDATE = DateEntry(SUB_SUB1FRAME2, firstweekday = "sunday", date_pattern = "yyyy-mm-dd")
        CALENDAR_DVDATE.grid(column = 1, row = 1, sticky = W)
        CALENDAR_DVDATE.bind("<FocusOut>", lambda e: CALENDAR_GLDATE.set_date(self.returnLastDayOfMonth(str(CALENDAR_DVDATE.get_date()))))
        required.append(CALENDAR_DVDATE)
        
        global CALENDAR_GLDATE
        CALENDAR_GLDATE = DateEntry(SUB_SUB1FRAME2, firstweekday = "sunday", date_pattern = "yyyy-mm-dd")
        CALENDAR_GLDATE.grid(column = 1, row = 2, sticky = W)
        CALENDAR_GLDATE.set_date(self.returnLastDayOfMonth(str(datetime.datetime.today())))
        required.append(CALENDAR_GLDATE)
        
        global TEXTVAR_TYPE, COMBO_TYPE
        TEXTVAR_TYPE = StringVar()
        COMBO_TYPE = tk.Combobox(SUB_SUB1FRAME2, values = self.returnDisbursementTypes(), textvariable = TEXTVAR_TYPE, font = APP_FONT, width = 20, state = "readonly")
        COMBO_TYPE.grid(column = 1, row = 3, sticky = W)
        COMBO_TYPE.bind("<<ComboboxSelected>>", self.updateDVEntries)
        required.append(TEXTVAR_TYPE)

        SUB_SUB1FRAME3 = Frame(SUB_FRAME1) #details-2 totals
        SUB_SUB1FRAME3.grid(column = 2, row = 0, sticky = NW, padx = TOP_PADX + 10)

        LABEL_GROSS = Label(SUB_SUB1FRAME3, text = "Gross", font = APP_FONT)
        LABEL_GROSS.grid(column = 0, row = 0, sticky = W)
        
        LABEL_VAT = Label(SUB_SUB1FRAME3, text = "VAT", font = APP_FONT)
        LABEL_VAT.grid(column = 0, row = 1, sticky = W)
        
        LABEL_EWT = Label(SUB_SUB1FRAME3, text = "EWT", font = APP_FONT)
        LABEL_EWT.grid(column = 0, row = 2, sticky = W)
        
        LABEL_NET = Label(SUB_SUB1FRAME3, text = "Net", font = APP_FONT)
        LABEL_NET.grid(column = 0, row = 3, sticky = W)
        
        global TEXTVAR_TGROSS
        TEXTVAR_TGROSS = StringVar()
        ENTRY_GROSS = Entry(SUB_SUB1FRAME3, textvariable = TEXTVAR_TGROSS, font = APP_FONT, width = 12, state = "readonly", justify = RIGHT)
        ENTRY_GROSS.grid(column = 1, row = 0, sticky = W)
        
        global TEXTVAR_TVAT
        TEXTVAR_TVAT = StringVar()
        ENTRY_VAT = Entry(SUB_SUB1FRAME3, textvariable = TEXTVAR_TVAT, font = APP_FONT, width = 12, state = "readonly", justify = RIGHT)
        ENTRY_VAT.grid(column = 1, row = 1, sticky = W)
        
        global TEXTVAR_TEWT
        TEXTVAR_TEWT = StringVar()
        ENTRY_EWT = Entry(SUB_SUB1FRAME3, textvariable = TEXTVAR_TEWT, font = APP_FONT, width = 12, state = "readonly", justify = RIGHT)
        ENTRY_EWT.grid(column = 1, row = 2, sticky = W)
        
        global TEXTVAR_TNET
        TEXTVAR_TNET = StringVar()
        ENTRY_NET = Entry(SUB_SUB1FRAME3, textvariable = TEXTVAR_TNET, font = APP_FONT, width = 12, state = "readonly", justify = RIGHT)
        ENTRY_NET.grid(column = 1, row = 3, sticky = W)

        SUB_FRAME2 = Frame(TOP_DISBURSEMENTS) #line headers
        SUB_FRAME2.grid(column = 0, row = 1, sticky = W)

        LABEL_ITEM = Label(SUB_FRAME2, text = "Item", width = 40, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_ITEM.grid(column = 1, row = 0)

        LABEL_BANK = Label(SUB_FRAME2, text = "Bank", width = 27, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_BANK.grid(column = 2, row = 0)

        LABEL_MODE = Label(SUB_FRAME2, text = "Mode", width = 23, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_MODE.grid(column = 3, row = 0)

        LABEL_GROSS = Label(SUB_FRAME2, text = "Gross", width = 15, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_GROSS.grid(column = 4, row = 0)

        LABEL_TAX = Label(SUB_FRAME2, text = "Tax", width = 8, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_TAX.grid(column = 5, row = 0)

        LABEL_NET = Label(SUB_FRAME2, text = "Net", width = 15, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_NET.grid(column = 6, row = 0)

        LABEL_CLEAR = Label(SUB_FRAME2, text = "[x]", width = 2, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_CLEAR.grid(column = 7, row = 0)

        SUB_FRAME3 = Frame(TOP_DISBURSEMENTS) #scroller
        SUB_FRAME3.grid(column = 0, row = 2, sticky = W)

        global dvFrame, dvItem, dvBank, dvMode, dvGross, dvTax, dvNet, dvClear
        dvFrame, dvItem, dvBank, dvMode, dvGross, dvTax, dvNet, dvClear = [], [], [], [], [], [], [], []

        self.createScrollFrame(SUB_FRAME3, 265, 950, 20, 0, 0)
        
        for i in range(10):
            self.showDisbursementLines(SCROLLABLE_FRAME, i)

        SUB_FRAME4 = Frame(TOP_DISBURSEMENTS) #entry
        SUB_FRAME4.grid(column = 0, row = 3, sticky = W)
        
        SUB_SUB4FRAME1 = Frame(SUB_FRAME4)
        SUB_SUB4FRAME1.grid(column = 0, row = 0, sticky = W)

        LABEL_CODE = Label(SUB_SUB4FRAME1, text = "Code", width = 12, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_CODE.grid(column = 0, row = 0)

        LABEL_TITLE = Label(SUB_SUB4FRAME1, text = "Title", width = 30, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_TITLE.grid(column = 1, row = 0)

        LABEL_CENTER = Label(SUB_SUB4FRAME1, text = "Amount", width = 15, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_CENTER.grid(column = 2, row = 0)

        LABEL_GROSS = Label(SUB_SUB4FRAME1, text = "Dr/Cr", width = 10, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_GROSS.grid(column = 3, row = 0)

        LABEL_EXPENSE = Label(SUB_SUB4FRAME1, text = "Remarks", width = 31, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_EXPENSE.grid(column = 4, row = 0)

        LABEL_VAT = Label(SUB_SUB4FRAME1, text = "[x]", width = 2, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_VAT.grid(column = 5, row = 0)

        SUB_SUB4FRAME2 = Frame(SUB_FRAME4)
        SUB_SUB4FRAME2.grid(column = 0, row = 1, sticky = W)
        
        self.createScrollFrame(SUB_SUB4FRAME2, 150, 735, 20, 0, 0)
        
        global entryFrame, entryCode, entryTitle, entryAmount, entryDrCr, entryRemarks, entryClear
        entryFrame, entryCode, entryTitle, entryAmount, entryDrCr, entryRemarks, entryClear = [], [], [], [], [], [], []
        
        for i in range(50):
            self.showSOAEntryLines(SCROLLABLE_FRAME, i)

        SUB_FRAME7 = Frame(SUB_SUB4FRAME2) #entry total frame
        SUB_FRAME7.grid(column = 1, row = 0, sticky = NW)
        
        LABEL_DEBIT = Label(SUB_FRAME7, text = "Debit", width = 12, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_DEBIT.grid(column = 0, row = 0)

        LABEL_CREDIT = Label(SUB_FRAME7, text = "Credit", width = 12, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_CREDIT.grid(column = 0, row = 1)

        LABEL_VARIANCE = Label(SUB_FRAME7, text = "Variance", width = 12, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_VARIANCE.grid(column = 0, row = 2)
        
        global TEXTVAR_DEBIT
        TEXTVAR_DEBIT = StringVar()
        ENTRY_DEBIT = Entry(SUB_FRAME7, textvariable = TEXTVAR_DEBIT, font = APP_FONT, width = 12, justify = RIGHT, state = "readonly")
        ENTRY_DEBIT.grid(column = 1, row = 0, sticky = W)
        
        global TEXTVAR_CREDIT
        TEXTVAR_CREDIT = StringVar()
        ENTRY_CREDIT = Entry(SUB_FRAME7, textvariable = TEXTVAR_CREDIT, font = APP_FONT, width = 12, justify = RIGHT, state = "readonly")
        ENTRY_CREDIT.grid(column = 1, row = 1, sticky = W)
        
        global TEXTVAR_VARIANCE
        TEXTVAR_VARIANCE = StringVar()
        ENTRY_VARIANCE = Entry(SUB_FRAME7, textvariable = TEXTVAR_VARIANCE, font = APP_FONT, width = 12, justify = RIGHT, state = "readonly")
        ENTRY_VARIANCE.grid(column = 1, row = 2, sticky = W)

        SUB_FRAME5 = Frame(TOP_DISBURSEMENTS) #buttons
        SUB_FRAME5.grid(column = 0, row = 4, sticky = NE, pady = TOP_PADY + 5)

        global BUTTON_SAVE
        BUTTON_SAVE = Button(SUB_FRAME5, text = "UPDATE ENTRY", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH + 3, cursor = "arrow", state = DISABLED, command = self.saveDVEntries)
        BUTTON_SAVE.grid(column = 0, row = 0, padx = TOP_PADX)

        global BUTTON_SUBMIT
        BUTTON_SUBMIT = Button(SUB_FRAME5, text = "SUBMIT", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "hand2", state = NORMAL, command = self.saveDV)
        BUTTON_SUBMIT.grid(column = 1, row = 0, padx = TOP_PADX)

        global BUTTON_APPROVE
        BUTTON_APPROVE = Button(SUB_FRAME5, text = "APPROVE", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG_GRN, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "arrow", state = DISABLED, command = self.approveDV)
        BUTTON_APPROVE.grid(column = 2, row = 0, padx = TOP_PADX)

        global BUTTON_VOID
        BUTTON_VOID = Button(SUB_FRAME5, text = "VOID", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG_RED, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "arrow", state = DISABLED, command = self.voidDV)
        BUTTON_VOID.grid(column = 3, row = 0, padx = TOP_PADX)

        global BUTTON_PRINT
        BUTTON_PRINT = Button(SUB_FRAME5, text = "PRINT", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "arrow", state = DISABLED, command = self.printDV)
        BUTTON_PRINT.grid(column = 4, row = 0, padx = TOP_PADX)

        BUTTON_CLOSE = Button(SUB_FRAME5, text = "CLOSE", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "hand2", command = lambda: self.closeTopLevel(TOP_DISBURSEMENTS))
        BUTTON_CLOSE.grid(column = 5, row = 0, padx = TOP_PADX)
        
        if self.returnUserName(USER, 3) != "ASD":
            ENTRY_PAYEENAME.config(state = NORMAL)
            SUB_SUB4FRAME1.grid_remove()
            SUB_SUB4FRAME2.grid_remove()
            SUB_FRAME7.grid_remove()
            BUTTON_SAVE.grid_remove()
            TOP_DISBURSEMENTS.geometry("995x485+100+20")
            
    def showDisbursementLines(self, frame, rown):
        global FRAME_LINE
        FRAME_LINE = Frame(frame)
        FRAME_LINE.grid(column = 0, row = rown, sticky = W)
        dvFrame.append(FRAME_LINE)

        global TEXTVAR_ITEM
        TEXTVAR_ITEM = StringVar()
        ENTRY_ITEM = Entry(FRAME_LINE, textvariable = TEXTVAR_ITEM, font = APP_FONT, width = 40)
        ENTRY_ITEM.grid(column = 1, row = 0, sticky = W)
        dvItem.append(TEXTVAR_ITEM)
        
        global TEXTVAR_BANK
        TEXTVAR_BANK = StringVar()
        COMBO_BANK = tk.Combobox(FRAME_LINE, values = self.returnBanks(), textvariable = TEXTVAR_BANK, font = APP_FONT, width = 25, state = "readonly")
        COMBO_BANK.grid(column = 2, row = 0, sticky = W)
        COMBO_BANK.bind("<<ComboboxSelected>>", lambda e: self.updateDVLine(dvGross[rown], dvTax[rown], rown))
        dvBank.append(TEXTVAR_BANK)
        
        global TEXTVAR_MODE
        TEXTVAR_MODE = StringVar()
        COMBO_MODE = tk.Combobox(FRAME_LINE, values = self.returnBankModes(), textvariable = TEXTVAR_MODE, font = APP_FONT, width = 20)
        COMBO_MODE.grid(column = 3, row = 0, sticky = W)
        dvMode.append(TEXTVAR_MODE)

        global TEXTVAR_GROSS
        TEXTVAR_GROSS = StringVar()
        ENTRY_GROSS = Entry(FRAME_LINE, textvariable = TEXTVAR_GROSS, font = APP_FONT, width = 15, justify = RIGHT)
        ENTRY_GROSS.grid(column = 4, row = 0, sticky = W)
        dvGross.append(TEXTVAR_GROSS)
        ENTRY_GROSS.bind("<FocusOut>", lambda e: self.updateDVLine(dvGross[rown], dvTax[rown], rown))

        global TEXTVAR_TAX, COMBO_TAX
        TEXTVAR_TAX = StringVar()
        COMBO_TAX = tk.Combobox(FRAME_LINE, values = ["NV-00"], textvariable = TEXTVAR_TAX, font = APP_FONT, width = 5, state = DISABLED)
        COMBO_TAX.grid(column = 5, row = 0, sticky = W)
        TEXTVAR_TAX.set("NV-00")
        dvTax.append(TEXTVAR_TAX)

        global TEXTVAR_NET
        TEXTVAR_NET = StringVar()
        ENTRY_NET = Entry(FRAME_LINE, textvariable = TEXTVAR_NET, font = APP_FONT, width = 15, justify = RIGHT, state = "readonly")
        ENTRY_NET.grid(column = 6, row = 0, sticky = W)
        dvNet.append(TEXTVAR_NET)

        global BUTTON_CLEAR
        BUTTON_CLEAR = Button(FRAME_LINE, text = "[X]", font = BUTTON_FONT2, bg = BUTTON_BG, fg = BUTTON_FG_RED, activebackground = CLICK_BG, cursor = "hand2", command = lambda: self.clearDVLine(dvGross[rown], dvTax[rown], rown))
        BUTTON_CLEAR.grid(column = 7, row = 0, sticky = W)
        dvClear.append(BUTTON_CLEAR)

    def saveDV(self):
        insertDV = """INSERT INTO tbldisbursements (
            dvNumber, dvDate, glDate, payeeCode, payeeName,
            particulars, reference, signatory1, signatory2, disbType,
            disbItem, disbBank, disbMode, disbAmount, disbNet, 
            disbTax, isApproved, isVoid, encoder, encoded
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
            
        deleteDV = "DELETE FROM tbldisbursements WHERE dvNumber = %s"
        
        deleteBook = "DELETE FROM tblcashdisbursementbook WHERE dvNumber = %s"
        insertBook = """INSERT INTO tblcashdisbursementbook (
                glDate, dvNumber, chartCode, amount, side,
                remarks, isPosted, isVoid, source, poster, 
                posted
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""

        wrong = []
        for i in required:
            try:
                if i.get() == "":
                    wrong.append("Please check required details!")
            except:
                if i.get("1.0", END) == "":
                    wrong.append("Particulars field is empty!")
        if TEXTVAR_TNET.get() == "" or self.returnFloatAmount(TEXTVAR_TNET.get()) < 0:
            wrong.append("Total Amount is zero!")
        if self.returnPeriodStatus(CALENDAR_GLDATE.get_date()) == "locked":
            wrong.append("GL date is locked")
        if len(wrong) != 0:
            messagebox.showerror("Disbursement Voucher", wrong)
            TOP_DISBURSEMENTS.focus()
        else:
            ask = messagebox.askyesno("Disbursement Voucher", "Are you sure?")
            if ask:
                if TEXTVAR_DVNUMBER.get() == "":
                    DVnumber = int(self.generateDVNumber())
                    message = "DV has been saved!"
                else:
                    cursor.execute(deleteDV, [int(TEXTVAR_DVNUMBER.get())])
                    db.commit()
                    cursor.execute(deleteBook, [int(TEXTVAR_DVNUMBER.get())])
                    db.commit()
                    DVnumber = int(TEXTVAR_DVNUMBER.get())
                    message = "DV has been updated!"
                validlines = []
                for i in range(len(dvFrame)):
                    if self.returnFloatAmount(dvNet[i].get()) != 0 and dvBank[i].get() != "" and dvMode[i].get() != "":
                        validlines.append([
                            DVnumber, CALENDAR_DVDATE.get_date(), CALENDAR_GLDATE.get_date(), TEXTVAR_PAYEECODE.get(), TEXTVAR_PAYEENAME.get(), 
                            ENTRY_PARTICULARS.get("1.0", END).replace("\n", ""), TEXTVAR_REFERENCE.get(), TEXTVAR_SIGN1.get(), TEXTVAR_SIGN2.get(), TEXTVAR_TYPE.get(),
                            dvItem[i].get(), dvBank[i].get(), dvMode[i].get(), self.returnFloatAmount(dvGross[i].get()), self.returnFloatAmount(dvNet[i].get()), 
                            dvTax[i].get(), "No", "No", USER, datetime.datetime.now()
                        ])
                validbook = []
                for i in range(len(entryFrame)):
                    if entryCode[i].get() != "" and self.returnFloatAmount(entryAmount[i].get()) > 0:
                        validbook.append([
                            CALENDAR_GLDATE.get_date(), DVnumber, int(entryCode[i].get()), self.returnFloatAmount(entryAmount[i].get()), entryDrCr[i].get(),
                            entryRemarks[i].get(), "No", "No", "DV", USER,
                            datetime.datetime.now()
                        ])
                
                if len(validlines) > 0 and len(validbook) > 0:
                    for i in validlines:
                        cursor.execute(insertDV, i)
                    db.commit()

                    for i in validbook:
                        cursor.execute(insertBook, i)
                    db.commit()

                    messagebox.showinfo("Disbursement Voucher", message)
                    TEXTVAR_DVNUMBER.set(str(DVnumber).zfill(8))
                    TOP_DISBURSEMENTS.focus()
                    self.disableDVWidgets()
                    BUTTON_SUBMIT.config(state = DISABLED, cursor = "arrow")
                    BUTTON_APPROVE.config(state = NORMAL, cursor = "hand2")
                    BUTTON_VOID.config(state = DISABLED, cursor = "arrow")
                    BUTTON_PRINT.config(state = NORMAL, cursor = "hand2")
                    
                    if self.returnAccess(USER, 10) == 0:
                        BUTTON_APPROVE.config(state = DISABLED, cursor = "arrow")
                else:
                    messagebox.showerror("Disbursement Voucher", "Please check BANK and MODE fields!")
            else:
                TOP_DISBURSEMENTS.focus()

    def editDV(self, *args):
        self.copySelection(TREE_DISBURSEMENT)
        self.showAddEditDisbursement()
        TOP_DISBURSEMENTS.title("Edit - DV")
        
        select = """SELECT 
            dvNumber, dvDate, payeeCode, payeeName, particulars,
            reference, signatory1, signatory2, disbType, disbItem,
            disbBank, disbMode, disbAmount, disbTax, disbNet,
            isApproved, isVoid, encoder, encoded, glDate
                FROM tbldisbursements WHERE dvNumber = %s"""
        
        selectBook = "SELECT chartCode, amount, side, remarks FROM tblcashdisbursementbook WHERE dvNumber = %s"
        db.commit()
        cursor.execute(selectBook, [int(content[0])])
        result2 = cursor.fetchall()
        db.commit()
        cursor.execute(select, [int(content[0])])
        result = cursor.fetchall()
        if result:
            TEXTVAR_DVNUMBER.set(str(result[0][0]).zfill(8))
            TEXTVAR_PAYEECODE.set(result[0][2])
            TEXTVAR_PAYEENAME.set(result[0][3])
            ENTRY_PARTICULARS.insert("1.0", result[0][4])
            TEXTVAR_REFERENCE.set(result[0][5])
            TEXTVAR_SIGN1.set(result[0][6])
            TEXTVAR_SIGN2.set(result[0][7])
            CALENDAR_DVDATE.set_date(result[0][1])
            CALENDAR_GLDATE.set_date(result[0][19])
            TEXTVAR_TYPE.set(result[0][8])
            count = 0
            for i in result:
                dvItem[count].set(i[9])
                dvBank[count].set(i[10])
                dvMode[count].set(i[11])
                dvGross[count].set(self.validateAmount2(i[12]))
                dvTax[count].set(i[13])
                dvNet[count].set(self.validateAmount2(i[14]))
                self.updateDVLine(dvGross[count], dvTax[count], count)
                count += 1
                
            if result2:
                count = 0
                self.clearSOAEntryLines()
                for i in result2:
                    entryCode[count].set(i[0])
                    self.populateChartFields(entryCode[count], entryTitle[count])
                    entryAmount[count].set(self.validateAmount2(i[1]))
                    entryDrCr[count].set(i[2])
                    entryRemarks[count].set(i[3])
                    count += 1
                self.updateSOAEntriesTotals()
                
            if result[0][15] == "No": #approve
                BUTTON_SUBMIT.config(state = NORMAL, cursor = "hand2")
                BUTTON_APPROVE.config(state = NORMAL, cursor = "hand2")
                BUTTON_VOID.config(state = DISABLED, cursor = "arrow")
            else:
                self.disableDVWidgets()
                if result[0][16] == "Yes": #void
                    BUTTON_SUBMIT.config(state = DISABLED, cursor = "arrow")
                    BUTTON_APPROVE.config(state = DISABLED, cursor = "arrow")
                else:
                    BUTTON_SUBMIT.config(state = DISABLED, cursor = "arrow")
                    BUTTON_APPROVE.config(state = DISABLED, cursor = "arrow")
                    BUTTON_VOID.config(state = NORMAL, cursor = "hand2")
            BUTTON_PRINT.config(state = NORMAL, cursor = "hand2")
                    
            if self.returnAccess(USER, 11) == 0:
                BUTTON_VOID.config(state = DISABLED, cursor = "arrow")
                
            if self.returnAccess(USER, 10) == 0:
                BUTTON_APPROVE.config(state = DISABLED, cursor = "arrow")

    def viewDV(self, *args):
        self.copySelection(TREE_PAYABLE)
        self.showAddEditDisbursement()
        TOP_DISBURSEMENTS.title("Edit - DV")
        
        select = """SELECT 
            dvNumber, dvDate, payeeCode, payeeName, particulars,
            reference, signatory1, signatory2, disbType, disbItem,
            disbBank, disbMode, disbAmount, disbTax, disbNet,
            isApproved, isVoid, encoder, encoded, glDate
                FROM tbldisbursements WHERE reference = %s AND isVoid = 'No'"""
        db.commit()
        cursor.execute(select, [content[0]])
        result = cursor.fetchall()
        
        selectBook = "SELECT chartCode, amount, side, remarks FROM tblcashdisbursementbook WHERE dvNumber = %s"
        db.commit()
        cursor.execute(selectBook, [result[0][0]])
        result2 = cursor.fetchall()
        
        if result:
            TEXTVAR_DVNUMBER.set(str(result[0][0]).zfill(8))
            TEXTVAR_PAYEECODE.set(result[0][2])
            TEXTVAR_PAYEENAME.set(result[0][3])
            ENTRY_PARTICULARS.insert("1.0", result[0][4])
            TEXTVAR_REFERENCE.set(result[0][5])
            TEXTVAR_SIGN1.set(result[0][6])
            TEXTVAR_SIGN2.set(result[0][7])
            CALENDAR_DVDATE.set_date(result[0][1])
            CALENDAR_GLDATE.set_date(result[0][19])
            TEXTVAR_TYPE.set(result[0][8])
            count = 0
            for i in result:
                dvItem[count].set(i[9])
                dvBank[count].set(i[10])
                dvMode[count].set(i[11])
                dvGross[count].set(self.validateAmount2(i[12]))
                dvTax[count].set(i[13])
                dvNet[count].set(self.validateAmount2(i[14]))
                self.updateDVLine(dvGross[count], dvTax[count], count)
                count += 1
                
            count = 0
            self.clearSOAEntryLines()
            for i in result2:
                entryCode[count].set(i[0])
                self.populateChartFields(entryCode[count], entryTitle[count])
                entryAmount[count].set(self.validateAmount2(i[1]))
                entryDrCr[count].set(i[2])
                entryRemarks[count].set(i[3])
                count += 1
            self.updateSOAEntriesTotals()
                
            if result[0][15] == "No":
                BUTTON_SUBMIT.config(state = NORMAL, cursor = "hand2")
                BUTTON_APPROVE.config(state = NORMAL, cursor = "hand2")
                BUTTON_VOID.config(state = DISABLED, cursor = "arrow")
                BUTTON_PRINT.config(state = NORMAL, cursor = "hand2")
            else:
                self.disableDVWidgets()
                if result[0][16] == "Yes":
                    BUTTON_SUBMIT.config(state = DISABLED, cursor = "arrow")
                    BUTTON_APPROVE.config(state = DISABLED, cursor = "arrow")
                    BUTTON_VOID.config(state = DISABLED, cursor = "arrow")
                    BUTTON_PRINT.config(state = NORMAL, cursor = "hand2")
                else:
                    BUTTON_SUBMIT.config(state = DISABLED, cursor = "arrow")
                    BUTTON_APPROVE.config(state = DISABLED, cursor = "arrow")
                    BUTTON_VOID.config(state = NORMAL, cursor = "hand2")
                    BUTTON_PRINT.config(state = NORMAL, cursor = "hand2")
                
                    
            if self.returnAccess(USER, 11) == 0:
                BUTTON_VOID.config(state = DISABLED, cursor = "arrow")
                
            if self.returnAccess(USER, 10) == 0:
                BUTTON_APPROVE.config(state = DISABLED, cursor = "arrow")

    def approveDV(self):
        approvedv = "UPDATE tbldisbursements SET approver = %s, approved = %s, isApproved = %s WHERE dvNumber = %s"
        ask = messagebox.askyesno("APPROVE Disbursement Voucher", "Are you sure?")
        if ask:
            cursor.execute(approvedv, [USER, datetime.datetime.now(), "Yes", int(TEXTVAR_DVNUMBER.get())])
            db.commit()
            messagebox.showinfo("APPROVE Disbursement Voucher", "DV has been approved!")
            TOP_DISBURSEMENTS.focus()
            self.disableDVWidgets()
            BUTTON_SUBMIT.config(state = DISABLED, cursor = "arrow")
            BUTTON_APPROVE.config(state = DISABLED, cursor = "arrow")
            BUTTON_VOID.config(state = NORMAL, cursor = "hand2")
            BUTTON_PRINT.config(state = NORMAL, cursor = "hand2")
            
            if self.returnAccess(USER, 11) == 0:
                BUTTON_VOID.config(state = DISABLED, cursor = "arrow")

    def voidDV(self):
        insertbook = """INSERT INTO tblgeneraljournal (
            gjNumber, docDate, glDate, particulars, reference,
            source, chartcode, amount, side, remarks,
            isPosted, isVoid, encoder, encoded) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
        
        voiddv = "UPDATE tbldisbursements SET isVoid = %s WHERE dvNumber = %s"
        
        validbook = []
        for i in range(len(entryFrame)):
            if entryCode[i].get() != "" and self.returnFloatAmount(entryAmount[i].get()) > 0:
                if entryDrCr[i].get() == "Debit":
                    side = "Credit"
                else:
                    side = "Debit"
                validbook.append([
                    int(self.generateGJNumber()), datetime.date.today(), self.returnLastDayOfMonth(str(datetime.date.today())), "REVERSAL OF DV#" + TEXTVAR_DVNUMBER.get() + " " + ENTRY_PARTICULARS.get("1.0", END).replace("\n",""), TEXTVAR_DVNUMBER.get(),
                    "DV", int(entryCode[i].get()), self.returnFloatAmount(entryAmount[i].get()), side, entryRemarks[i].get(),
                    "No", "No", USER, datetime.datetime.now()
                ])
        
        ask = messagebox.askyesno("VOID Disbursement Voucher", "Are you sure?")
        if ask:
            cursor.execute(voiddv, ["Yes", int(TEXTVAR_DVNUMBER.get())])
            db.commit()
            
            if len(validbook) != 0:
                for i in validbook:
                    cursor.execute(insertbook, i)
                db.commit()
            messagebox.showinfo("VOID Disbursement Voucher", "DV has been voided!")
            TOP_DISBURSEMENTS.focus()
            self.disableDVWidgets()
            BUTTON_SUBMIT.config(state = DISABLED, cursor = "arrow")
            BUTTON_APPROVE.config(state = DISABLED, cursor = "arrow")
            BUTTON_VOID.config(state = DISABLED, cursor = "arrow")
            BUTTON_PRINT.config(state = NORMAL, cursor = "hand2")

    def printDV(self):
        wb = load_workbook(PATH_TEMPLATE + "DV.xlsx")
        sheet = wb.active
        sheet["A1"] = f"Printed Date: {datetime.datetime.now()} {self.returnUserName(USER, 0)}"
        sheet["C3"] = f"DV{TEXTVAR_DVNUMBER.get()}"
        sheet["A10"] = TEXTVAR_PAYEENAME.get()
        sheet["C10"] = datetime.datetime.strptime(str(CALENDAR_DVDATE.get_date()), '%Y-%m-%d').strftime('%B %d, %Y')
        sheet["A12"] = self.convertNumberToWords(TEXTVAR_TNET.get())
        sheet["C12"] = self.returnFloatAmount(TEXTVAR_TNET.get())
        sheet["C14"] = TEXTVAR_TYPE.get()
        sheet["A14"] = ENTRY_PARTICULARS.get("1.0", END).replace("\n", "")
        sheet["A41"] = f"{self.returnUserName(USER, 1).upper()} {self.returnUserName(USER, 2).upper()}"
        sheet["A42"] = self.returnUserName(USER, 3)
        sheet["B41"] = TEXTVAR_SIGN1.get()
        sheet["B42"] = self.returnSignatoryPosition(TEXTVAR_SIGN1.get())
        sheet["C41"] = TEXTVAR_SIGN2.get()
        sheet["C42"] = self.returnSignatoryPosition(TEXTVAR_SIGN2.get())

        count = 16
        for i in range(len(dvFrame)):
            if self.returnFloatAmount(dvNet[i].get()) != 0:
                sheet["A" + str(count)] = dvItem[i].get()
                sheet["B" + str(count)] = dvBank[i].get()
                sheet["C" + str(count)] = dvMode[i].get()
                sheet["D" + str(count)] = dvNet[i].get()
                count += 1

        if self.returnUserName(USER, 3) == "ASD":
            count = 30
            for i in range(len(entryFrame)):
                if entryCode[i].get() != "":
                    sheet["A" + str(count)] = entryCode[i].get()
                    sheet["B" + str(count)] = entryTitle[i].get()
                    if entryDrCr[i].get() == "Debit":
                        sheet["C" + str(count)] = self.returnFloatAmount(entryAmount[i].get())
                    else:
                        sheet["D" + str(count)] = self.returnFloatAmount(entryAmount[i].get())
                    count += 1
        else:
            if self.returnAPVExists(TEXTVAR_REFERENCE.get()) == 1:
                count = 30
                for i in range(len(entryFrame)):
                    if entryCode[i].get() != "":
                        sheet["A" + str(count)] = entryCode[i].get()
                        sheet["B" + str(count)] = entryTitle[i].get()
                        if entryDrCr[i].get() == "Debit":
                            sheet["C" + str(count)] = self.returnFloatAmount(entryAmount[i].get())
                        else:
                            sheet["D" + str(count)] = self.returnFloatAmount(entryAmount[i].get())
                        count += 1
            
        for i in range(16,25):
            cell = sheet["D" + str(i)].value
            if cell == None:
                sheet.row_dimensions[i].hidden = True
                
        if self.returnUserName(USER, 3) == "FMD":
            if self.returnAPVExists(TEXTVAR_REFERENCE.get()) == 0:
                for i in range(28,40):
                    sheet.row_dimensions[i].hidden = True
            else:
                for i in range(30,38):
                    cell = sheet["A" + str(i)].value
                    if cell == None:
                        sheet.row_dimensions[i].hidden = True
        else:
            for i in range(30,38):
                cell = sheet["A" + str(i)].value
                if cell == None:
                    sheet.row_dimensions[i].hidden = True
                    
        wb.save(PATH_SAVE + "dv.xlsx")
        startfile(PATH_SAVE + "dv.xlsx", "open")

    def createDV(self):
        self.copySelection(TREE_PAYABLE)
        self.showAddEditDisbursement()
        
        select = """SELECT 
            apvNumber, apvDate, glDate, dueDate, supplierCode,
            particulars, reference, rrNumber, chartCode, centerCode,
            amount, taxType, remarks, encoder, encoded
                FROM tblpayables WHERE apvNumber = %s"""
        db.commit()
        cursor.execute(select, [int(content[0])])
        result = cursor.fetchall()
        if result:
            TEXTVAR_PAYEECODE.set(result[0][4])
            TEXTVAR_PAYEENAME.set(self.returnSupplierName(result[0][4]))
            if result[0][7] == 0:
                ENTRY_PARTICULARS.insert("1.0", result[0][5])
            else:
                ENTRY_PARTICULARS.insert("1.0", result[0][5]+f" RR#{str(result[0][7]).zfill(8)} PO#{str(self.returnPONumber(result[0][7])).zfill(8)}")
            TEXTVAR_REFERENCE.set(str(result[0][0]).zfill(8))
            ENTRY_REFERENCE.config(state = "readonly")
            CALENDAR_DVDATE.set_date(datetime.date.today())
            CALENDAR_GLDATE.set_date(self.returnLastDayOfMonth(str(datetime.date.today())))
            CALENDAR_GLDATE.config(state = "readonly")
            count = 0
            for i in result:
                dvItem[count].set(self.returnChartTitle(i[8])[0])
                dvBank[count].set(self.returnBanks()[30])
                dvMode[count].set("Check")
                dvGross[count].set(i[10])
                dvTax[count].set(i[11])
                self.updateDVLine(dvGross[count], dvTax[count], count)
                count += 1

    def saveDVEntries(self):
        deleteBook = "DELETE FROM tblcashdisbursementbook WHERE dvNumber = %s"
        insertBook = """INSERT INTO tblcashdisbursementbook (
                glDate, dvNumber, chartCode, amount, side,
                remarks, isPosted, isVoid, source, poster, 
                posted
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
        ask = messagebox.askyesno("Disbursement Voucher", "Are you sure?")
        if ask:
            cursor.execute(deleteBook, [int(TEXTVAR_DVNUMBER.get())])
            db.commit()
            DVnumber = int(TEXTVAR_DVNUMBER.get())
        
            validbook = []
            for i in range(len(entryFrame)):
                if entryCode[i].get() != "" and self.returnFloatAmount(entryAmount[i].get()) > 0:
                    validbook.append([
                        CALENDAR_GLDATE.get_date(), DVnumber, int(entryCode[i].get()), self.returnFloatAmount(entryAmount[i].get()), entryDrCr[i].get(),
                        entryRemarks[i].get(), "No", "No", "DV", USER,
                        datetime.datetime.now()
                    ])
            if len(validbook) > 0:
                for i in validbook:
                    cursor.execute(insertBook, i)
                db.commit()
                messagebox.showinfo("Disbursement Voucher", "Entries has been updated!")
        
    def generateDVNumber(self):
        db.commit()
        cursor.execute("SELECT MAX(dvNumber) FROM tbldisbursements")
        result = cursor.fetchone()
        if result[0] == None:
            return str(2035).zfill(8)
        else:
            return str(int(result[0])+1).zfill(8)

    def updateDVLine(self, gross, tax, rown):
        self.validateAmount(gross)
        dvNet[rown].set(self.validateAmount2(self.computeAPVNetAmount(self.returnFloatAmount(gross.get()), tax.get())))
        self.updateDVTotals()
    
    def updateDVTotals(self):
        tgross, tnet = [], []
        for i in range(len(dvFrame)):
            try:
                if self.returnFloatAmount(dvNet[i].get()) != 0:
                    tgross.append(self.returnFloatAmount(dvGross[i].get()))
                    tnet.append(self.returnFloatAmount(dvNet[i].get()))
            except:
                pass
            TEXTVAR_TGROSS.set(self.validateAmount2(sum(tgross)))
            TEXTVAR_TVAT.set(self.validateAmount2(0))
            TEXTVAR_TEWT.set(self.validateAmount2(0))
            TEXTVAR_TNET.set(self.validateAmount2(sum(tnet)))
        self.updateDVEntries()
        self.updateSOAEntriesTotals()
        
    def updateDVEntries(self, *args):
        self.clearSOAEntryLines()
        if self.returnAPVExists(TEXTVAR_REFERENCE.get()) == 1:
            entryCode[0].set("211010")
            self.populateChartFields(entryCode[0], entryTitle[0])
            entryAmount[0].set(self.validateAmount2(self.returnFloatAmount(TEXTVAR_TNET.get())))
            entryDrCr[0].set("Debit")
        else:
            entryCode[0].set(self.returnDisbursementTypeCode(TEXTVAR_TYPE.get()))
            self.populateChartFields(entryCode[0], entryTitle[0])
            entryAmount[0].set(self.validateAmount2(self.returnFloatAmount(TEXTVAR_TNET.get())))
            entryDrCr[0].set("Debit")

        self.declareBankSelected()
        count = 1
        for i in selected:
            entryCode[count].set(self.returnChartCodeUsingTitle(i[0]))
            self.populateChartFields(entryCode[count], entryTitle[count])
            entryAmount[count].set(self.validateAmount2(i[1]))
            entryDrCr[count].set("Credit")
            entryRemarks[count].set(i[2])
            count += 1

    def returnDisbursementTypeCode(self, var):
        cursor.execute(f"SELECT code FROM tbltypes WHERE type = '{var}' LIMIT 1")
        result = cursor.fetchone()
        try:
            return result[0]
        except:
            pass
        
    def returnAPVExists(self, var):
        try:
            cursor.execute(f"SELECT apvNumber FROM tblpayables WHERE apvNumber = {int(var)} LIMIT 1")
            result = cursor.fetchone()
            if result != None:
                return 1
            else:
                return 0
        except:
            return 0
    
    def declareBankSelected(self):
        global selected
        selected = []
        for i in range(len(dvFrame)):
            if self.returnFloatAmount(dvNet[i].get()) != 0:
                selected.append([dvBank[i].get(), self.returnFloatAmount(dvNet[i].get()), dvMode[i].get()])

    def clearDVLine(self, gross, tax, i):
        dvItem[i].set("")
        dvBank[i].set("")
        dvMode[i].set("")
        dvGross[i].set("")
        dvNet[i].set("")
        self.updateDVLine(gross, tax, i)

    def returnDisbursementTypes(self):
        db.commit()
        cursor.execute(f"SELECT type FROM tbltypes ORDER BY type")
        raw = cursor.fetchall()
        result = []
        for i in raw:
            result.append(i[0])
        return result

    def returnBanks(self):
        db.commit()
        cursor.execute("SELECT title FROM tblchart WHERE CAST(code AS CHAR) LIKE '111%' ORDER by title")
        raw = cursor.fetchall()
        result = []
        for i in raw:
            if i[0] != "CASH IN BANK":
                result.append(i[0].split(" - ")[1])
        result.append("CA Fund")
        return result
    
    def returnBankModes(self):
        return ["Cash in Bank", "Check", "CA Fund", "Debit Memo", "E-Banking", "E-Payment", "GCash", "Manager's Check"]

    def returnSignatories(self):
        db.commit()
        cursor.execute("SELECT name FROM tblsignatories ORDER BY name")
        raw = cursor.fetchall()
        result = []
        for i in raw:
            result.append(i[0])
        return result
    
    def returnSignatoryPosition(self, var):
        cursor.execute(f"SELECT position FROM tblsignatories WHERE name = '{var}' LIMIT 1")
        result = cursor.fetchone()
        return result[0]

    def returnChartCodeUsingTitle(self, var):
        if var != "CA Fund":
            select = "SELECT code FROM tblchart WHERE code LIKE '111%' AND title LIKE %s LIMIT 1"
            cursor.execute(select, [f"%{var}%"])
            result = cursor.fetchone()
            try:
                return result[0]
            except:
                pass
        else:
            return 116010

    def returnPayeeName(self, var1, var2):
        try:
            if self.returnSupplierName(var1) != None:
                return self.returnSupplierName(var1)
            else:
                return var2
        except:
            return var2

    def returnDVNetAmount(self, var):
        select = "SELECT sum(disbNet) FROM tbldisbursements WHERE dvNumber = %s LIMIT 1"
        cursor.execute(select, [var])
        result = cursor.fetchone()
        return result[0]
    
    def returnDVNetAmount2(self, var):
        select = "SELECT sum(disbNet) FROM tbldisbursements WHERE reference = %s AND isVoid = 'No' LIMIT 1"
        cursor.execute(select, [var])
        result = cursor.fetchone()
        if result[0] != None:
            return result[0]
        else:
            return 0

    def disableDVWidgets(self):
        ENTRY_PAYEE.config(state = DISABLED)
        BUTTON_PAYEE.config(state = DISABLED, cursor = "arrow")
        ENTRY_PAYEENAME.config(state = DISABLED)
        ENTRY_PARTICULARS.config(state = DISABLED)
        ENTRY_REFERENCE.config(state = DISABLED)
        COMBO_SIGN1.config(state = DISABLED)
        COMBO_SIGN2.config(state = DISABLED)
        CALENDAR_DVDATE.config(state = DISABLED)
        CALENDAR_GLDATE.config(state = DISABLED)
        COMBO_TYPE.config(state = DISABLED)
        
        for i in dvFrame:
            for x in i.winfo_children():
                x.config(state = DISABLED)
        
        if self.returnPostingStatus(int(TEXTVAR_DVNUMBER.get()), 'CDB') == "Unposted":
            if self.returnUserName(USER, 3) == "ASD" and self.returnUserName(USER, 4) == "administrator":
                CALENDAR_GLDATE.config(state = NORMAL)
                BUTTON_SAVE.config(state = NORMAL, cursor = "hand2")
        else:
            for i in entryFrame:
                for x in i.winfo_children():
                    x.config(state = DISABLED)
            BUTTON_SAVE.config(state = DISABLED, cursor = "arrow")

    def searchDV(self, var, *args):
        find = "SELECT dvNumber, dvDate, glDate, payeeCode, payeeName, particulars, disbNet, isApproved, isVoid, encoder, DATE(encoded), approver, DATE(approved) FROM tbldisbursements WHERE (dvNumber LIKE %s OR payeeName LIKE %s OR particulars LIKE %s OR reference LIKE %s) AND dvDate BETWEEN %s AND %s ORDER BY dvNumber DESC"
        db.commit()
        cursor.execute(find, [f"%{var}%", f"%{var}%", f"%{var}%", f"%{var}%", CALENDAR_START.get_date(), CALENDAR_END.get_date()])
        result = cursor.fetchall()
        for i in TREE_DISBURSEMENT.get_children():
            TREE_DISBURSEMENT.delete(i)
        if result:
            PROGRESS_BAR.grid(column = 7, row = 0, padx = TOP_PADX + 10, sticky = E)
            count = 0
            dvnumbers, skipped = [], []
            for i in result:
                if i[0] not in dvnumbers:
                    if i[8] != "Yes":
                        if count % 2 == 0:
                            TREE_DISBURSEMENT.insert("", "end", values = (str(i[0]).zfill(8),i[1],i[2],self.returnPayeeName(i[3],i[4]),i[5],self.validateAmount2(self.returnDVNetAmount(i[0])),self.returnDVStatus(i[7],i[8]),self.returnUserName(i[9], 0),i[10],self.returnUserName(i[11], 0),i[12]), tags = "evenrow")
                        else:
                            TREE_DISBURSEMENT.insert("", "end", values = (str(i[0]).zfill(8),i[1],i[2],self.returnPayeeName(i[3],i[4]),i[5],self.validateAmount2(self.returnDVNetAmount(i[0])),self.returnDVStatus(i[7],i[8]),self.returnUserName(i[9], 0),i[10],self.returnUserName(i[11], 0),i[12]), tags = "oddrow")
                    else:
                        TREE_DISBURSEMENT.insert("", "end", values = (str(i[0]).zfill(8),i[1],i[2],self.returnPayeeName(i[3],i[4]),i[5],self.validateAmount2(self.returnDVNetAmount(i[0])),self.returnDVStatus(i[7],i[8]),self.returnUserName(i[9], 0),i[10],self.returnUserName(i[11], 0),i[12]), tags = "void")
                    count += 1
                    dvnumbers.append(i[0])
                    PROGRESS_BAR["value"] = round((count/(len(result)-len(skipped)))*100, 0)
                    SUB1_DISBURSEMENT.update()
                else:
                    skipped.append(i[0])
            PROGRESS_BAR.grid_remove()
        else:
            messagebox.showerror("Disbursements", "No match found!")
            # db.commit()
            # cursor.execute(f"SELECT dvNumber, dvDate, glDate, payeeCode, payeeName, particulars, disbNet, isApproved, isVoid, encoder, DATE(encoded), approver, DATE(approved) FROM tbldisbursements WHERE dvDate BETWEEN '{CALENDAR_START.get_date()}' AND '{CALENDAR_END.get_date()}' ORDER BY dvNumber DESC")
            # result = cursor.fetchall()

    def refreshDV(self, *args):
        for i in TREE_DISBURSEMENT.get_children():
            TREE_DISBURSEMENT.delete(i)
        db.commit()
        cursor.execute(f"SELECT dvNumber, dvDate, glDate, payeeCode, payeeName, particulars, disbNet, isApproved, isVoid, encoder, DATE(encoded), approver, DATE(approved) FROM tbldisbursements WHERE dvDate BETWEEN '{CALENDAR_START.get_date()}' AND '{CALENDAR_END.get_date()}' ORDER BY dvNumber DESC")
        result = cursor.fetchall()

        count = 0
        dvnumbers, skipped = [], []
        PROGRESS_BAR.grid(column = 7, row = 0, padx = TOP_PADX + 10, sticky = E)
        for i in result:
            if i[0] not in dvnumbers:
                if i[8] != "Yes":
                    if count % 2 == 0:
                        TREE_DISBURSEMENT.insert("", "end", values = (str(i[0]).zfill(8),i[1],i[2],self.returnPayeeName(i[3],i[4]),i[5],self.validateAmount2(self.returnDVNetAmount(i[0])),self.returnDVStatus(i[7],i[8]),self.returnUserName(i[9], 0),i[10],self.returnUserName(i[11], 0),i[12]), tags = "evenrow")
                    else:
                        TREE_DISBURSEMENT.insert("", "end", values = (str(i[0]).zfill(8),i[1],i[2],self.returnPayeeName(i[3],i[4]),i[5],self.validateAmount2(self.returnDVNetAmount(i[0])),self.returnDVStatus(i[7],i[8]),self.returnUserName(i[9], 0),i[10],self.returnUserName(i[11], 0),i[12]), tags = "oddrow")
                else:
                    TREE_DISBURSEMENT.insert("", "end", values = (str(i[0]).zfill(8),i[1],i[2],self.returnPayeeName(i[3],i[4]),i[5],self.validateAmount2(self.returnDVNetAmount(i[0])),self.returnDVStatus(i[7],i[8]),self.returnUserName(i[9], 0),i[10],self.returnUserName(i[11], 0),i[12]), tags = "void")
                count += 1
                dvnumbers.append(i[0])
                PROGRESS_BAR["value"] = round((count/(len(result)-len(skipped)))*100, 0)
                SUB1_DISBURSEMENT.update()
            else:
                skipped.append(i[0])
        PROGRESS_BAR.grid_remove()

    def convertNumberToWords(self, var):
        try:
            if float(var.split(".")[1]) > 0:
                decimal = " AND " + str(int(var.split(".")[1])) +"/100"
            else:
                decimal = " ONLY "
            whole = num2words(TEXTVAR_TNET.get().replace(",","").split(".")[0]).upper()
            return whole.replace(" AND", "") + " PESOS" + decimal
        except:
            return "ZERO"

### MENU_FINANCE_REPORTS ###
    def showSelectedFinanceReport(self, *args):
        for i in SUB_FRAME1.winfo_children():
            i.destroy()
            
        try:
            for i in TREE_DISBURSEMENT.get_children():
                TREE_DISBURSEMENT.delete(i)
        except:
            pass
            
        LABEL_FROM = Label(SUB_FRAME1, text = "From", width = 10, font = APP_FONT, anchor = W)
        LABEL_FROM.grid(column = 0, row = 0, pady = SUBMENU_PADY, sticky = W)
        
        LABEL_TO = Label(SUB_FRAME1, text = "To", width = 10, font = APP_FONT, anchor = W)
        LABEL_TO.grid(column = 0, row = 1, pady = SUBMENU_PADY, sticky = W)
        
        global CALENDAR_FROM
        CALENDAR_FROM = DateEntry(SUB_FRAME1, firstweekday = "sunday", date_pattern = "yyyy-mm-dd", state = "readonly")
        CALENDAR_FROM.grid(column = 1, row = 0, sticky = W, pady = TOP_PADY)
        CALENDAR_FROM.set_date(self.returnFirstDayOfMonth(str(CALENDAR_FROM.get_date())))
        
        global CALENDAR_TO
        CALENDAR_TO = DateEntry(SUB_FRAME1, firstweekday = "sunday", date_pattern = "yyyy-mm-dd", state = "readonly")
        CALENDAR_TO.grid(column = 1, row = 1, sticky = W, pady = TOP_PADY)
        
        global TEXTVAR_TYPE, COMBO_TYPE
        TEXTVAR_TYPE = StringVar()
        COMBO_TYPE = tk.Combobox(SUB_FRAME1, values = self.returnDisbursementTypes(), textvariable = TEXTVAR_TYPE, font = APP_FONT, width = 30, state = "readonly")
        
        global BUTTON_VIEW
        BUTTON_VIEW = Button(SUB_FRAME1, text = "VIEW", width = 5, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2")
        BUTTON_VIEW.grid(column = 1, row = 3, pady = MENU_PADY)
        
        global BUTTON_EXPORT
        BUTTON_EXPORT = Button(SUB_FRAME1, text = "EXPORT", width = 5, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "arrow", state = DISABLED, command = lambda: self.exportDisbursementSummary(TREE_DISBURSEMENT))
        BUTTON_EXPORT.grid(column = 2, row = 3, pady = MENU_PADY)
        if TEXTVAR_REPORTS.get() == "Disbursement Summary":
            BUTTON_VIEW.config(command = lambda: self.viewDisbursementSummary(SUB_FRAME2))
            COMBO_TYPE.grid_remove()
            
        elif TEXTVAR_REPORTS.get() == "Transaction Type Summary":
            COMBO_TYPE.grid(column = 1, row = 2, sticky = W, ipadx = 2, pady = MENU_PADY)
            
            BUTTON_VIEW.config(command = lambda: self.viewTransactionSummary(SUB_FRAME2))
        elif TEXTVAR_REPORTS.get() == "ISO Report":
            COMBO_TYPE.grid_remove()
            BUTTON_VIEW.grid_remove()
            BUTTON_EXPORT.config(state = NORMAL, command = lambda: self.viewISOReport(SUB_FRAME2))

    def viewDisbursementSummary(self, frame):
        for i in frame.winfo_children():
            i.destroy()
        self.showFinanceReportTreeView(frame)
        db.commit()
        cursor.execute(f"SELECT dvNumber, dvDate, payeeCode, payeeName, disbType, particulars, disbItem, disbBank, disbMode, disbNet, isApproved, isVoid, encoder, DATE(encoded), approver, DATE(approved) FROM tbldisbursements WHERE dvDate BETWEEN '{CALENDAR_FROM.get_date()}' AND'{CALENDAR_TO.get_date()}' ORDER BY dvNumber DESC")
        result = cursor.fetchall()
        if result:
            count = 0
            total = []
            for i in result:
                if self.returnDVStatus(i[10],i[11]) == "Void":
                    amt = 0
                else:
                    amt = self.validateAmount2(i[9])
                    total.append(i[9])
                if count % 2 == 0:
                    TREE_DISBURSEMENT.insert("", "end", values = (str(i[0]).zfill(8),i[1],self.returnPayeeName(i[2],i[3]),i[4],i[5],i[6],i[7],i[8],amt,self.returnDVStatus(i[10],i[11]),self.returnUserName(i[12], 0),i[13],self.returnUserName(i[14], 0),i[15]), tags = "evenrow")
                else:
                    TREE_DISBURSEMENT.insert("", "end", values = (str(i[0]).zfill(8),i[1],self.returnPayeeName(i[2],i[3]),i[4],i[5],i[6],i[7],i[8],amt,self.returnDVStatus(i[10],i[11]),self.returnUserName(i[12], 0),i[13],self.returnUserName(i[14], 0),i[15]), tags = "oddrow")
                count += 1
            BUTTON_EXPORT.config(state = NORMAL, cursor = "hand2")
        else:
            BUTTON_EXPORT.config(state = DISABLED, cursor = "arrow")
            
        TREE_DISBURSEMENT.config(height = 24)

    def viewTransactionSummary(self, frame):
        for i in frame.winfo_children():
            i.destroy()
        self.showFinanceReportTreeView(frame)
        db.commit()
        cursor.execute(f"SELECT dvNumber, dvDate, payeeCode, payeeName, disbType, particulars, disbItem, disbBank, disbMode, disbNet, isApproved, isVoid, encoder, DATE(encoded), approver, DATE(approved) FROM tbldisbursements WHERE disbType = '{TEXTVAR_TYPE.get()}' AND dvDate BETWEEN '{CALENDAR_FROM.get_date()}' AND'{CALENDAR_TO.get_date()}' ORDER BY dvNumber DESC")
        result = cursor.fetchall()
        if result:
            count = 0
            total = []
            for i in result:
                if self.returnDVStatus(i[10],i[11]) == "Void":
                    amt = 0
                else:
                    amt = self.validateAmount2(i[9])
                    total.append(i[9])
                if count % 2 == 0:
                    TREE_DISBURSEMENT.insert("", "end", values = (str(i[0]).zfill(8),i[1],self.returnPayeeName(i[2],i[3]),i[4],i[5],i[6],i[7],i[8],amt,self.returnDVStatus(i[10],i[11]),self.returnUserName(i[12], 0),i[13],self.returnUserName(i[14], 0),i[15]), tags = "evenrow")
                else:
                    TREE_DISBURSEMENT.insert("", "end", values = (str(i[0]).zfill(8),i[1],self.returnPayeeName(i[2],i[3]),i[4],i[5],i[6],i[7],i[8],amt,self.returnDVStatus(i[10],i[11]),self.returnUserName(i[12], 0),i[13],self.returnUserName(i[14], 0),i[15]), tags = "oddrow")
                count += 1
            BUTTON_EXPORT.config(state = NORMAL, cursor = "hand2")
        else:
            BUTTON_EXPORT.config(state = DISABLED, cursor = "arrow")
            
        TREE_DISBURSEMENT.config(height = 22)

    def viewISOReport(self, *args):
        db.commit()
        cursor.execute(f"SELECT disbType, SUM(disbNet) FROM tbldisbursements WHERE dvDate BETWEEN '{CALENDAR_FROM.get_date()}' AND'{CALENDAR_TO.get_date()}' GROUP BY disbType ORDER BY disbType")
        result = cursor.fetchall()
        if result:
            wb = load_workbook(PATH_TEMPLATE + "ISOREPORT.xlsx")
            sheet = wb.active
            sheet["B5"] = f"{CALENDAR_FROM.get_date()} to {CALENDAR_TO.get_date()}"
            
            count = 8
            total = []
            for i in result:
                sheet["A" + str(count)] = i[0]
                sheet["B" + str(count)] = i[1]
                total.append(i[1])
                count += 1
            sheet["B" + str(count)] = sum(total)
            
            wb.save(PATH_SAVE + "ISOREPORT.xlsx")
            startfile(PATH_SAVE + "ISOREPORT.xlsx", "open")

    def showFinanceReportTreeView(self, frame):
        global TREE_DISBURSEMENT
        TREE_DISBURSEMENT = tk.Treeview(frame, height = 25, selectmode = "browse")
        TREE_DISBURSEMENT["columns"] = ("DV No.", "Doc Date", "Payee", "Type", "Description", "Item", "Bank", "Mode", "Amount", "Status", "Encoder", "Encoded", "Approver", "Approved")
        TREE_DISBURSEMENT.column("#0", width = 0, minwidth = 0, stretch = True)
        TREE_DISBURSEMENT.column("DV No.", anchor = W, minwidth = 75, width = 75)
        TREE_DISBURSEMENT.column("Doc Date", anchor = W, minwidth = 75, width = 75)
        TREE_DISBURSEMENT.column("Payee", anchor = W, minwidth = 100, width = 135)
        TREE_DISBURSEMENT.column("Type", anchor = W, minwidth = 75, width = 75)
        TREE_DISBURSEMENT.column("Description", anchor = W, minwidth = 100, width = 125)
        TREE_DISBURSEMENT.column("Item", anchor = W, minwidth = 75, width = 75)
        TREE_DISBURSEMENT.column("Bank", anchor = W, minwidth = 75, width = 75)
        TREE_DISBURSEMENT.column("Mode", anchor = W, minwidth = 75, width = 75)
        TREE_DISBURSEMENT.column("Amount", anchor = E, minwidth = 100, width = 75)
        TREE_DISBURSEMENT.column("Status", anchor = W, minwidth = 100, width = 50)
        TREE_DISBURSEMENT.column("Encoder", anchor = W, minwidth = 100, width = 75)
        TREE_DISBURSEMENT.column("Encoded", anchor = W, minwidth = 75, width = 50)
        TREE_DISBURSEMENT.column("Approver", anchor = W, minwidth = 75, width = 75)
        TREE_DISBURSEMENT.column("Approved", anchor = W, minwidth = 75, width = 75)
        
        TREE_DISBURSEMENT.heading("#0", text = "", anchor = W)
        TREE_DISBURSEMENT.heading("DV No.", text = "DV No.", anchor = N)
        TREE_DISBURSEMENT.heading("Doc Date", text = "Doc Date", anchor = N)
        TREE_DISBURSEMENT.heading("Payee", text = "Payee", anchor = N)
        TREE_DISBURSEMENT.heading("Type", text = "Type", anchor = N)
        TREE_DISBURSEMENT.heading("Description", text = "Description", anchor = N)
        TREE_DISBURSEMENT.heading("Item", text = "Item", anchor = N)
        TREE_DISBURSEMENT.heading("Bank", text = "Bank", anchor = N)
        TREE_DISBURSEMENT.heading("Mode", text = "Mode", anchor = N)
        TREE_DISBURSEMENT.heading("Amount", text = "Amount", anchor = N)
        TREE_DISBURSEMENT.heading("Status", text = "Status", anchor = N)
        TREE_DISBURSEMENT.heading("Encoder", text = "Encoder", anchor = N)
        TREE_DISBURSEMENT.heading("Encoded", text = "Encoded", anchor = N)
        TREE_DISBURSEMENT.heading("Approver", text = "Approver", anchor = N)
        TREE_DISBURSEMENT.heading("Approved", text = "Approved", anchor = N)

        POPUP_DISBURSEMENT = Menu(TREE_DISBURSEMENT, tearoff = 0)
        POPUP_DISBURSEMENT.add_command(command = self.editDV, label = "Edit")
        TREE_DISBURSEMENT.bind("<Button-3>", lambda e: self.popupMenu(TREE_DISBURSEMENT, POPUP_DISBURSEMENT, e))
        TREE_DISBURSEMENT.bind("<Double-1>", self.editDV)

        global STYLE_DISBURSEMENT
        STYLE_DISBURSEMENT = tk.Style()
        STYLE_DISBURSEMENT.map("Treeview", foreground = self.fixedMap("foreground", STYLE_DISBURSEMENT), background = self.fixedMap("background", STYLE_DISBURSEMENT))

        TREE_DISBURSEMENT.tag_configure("oddrow", background = None)
        TREE_DISBURSEMENT.tag_configure("evenrow", background = TREE_TAG_EVENROW)

        YSCROLLBAR = tk.Scrollbar(frame, orient = "vertical", command = TREE_DISBURSEMENT.yview)
        YSCROLLBAR.pack(side = "right", fill = "y")

        XSCROLLBAR = tk.Scrollbar(frame, orient = "horizontal", command = TREE_DISBURSEMENT.xview)
        
        TREE_DISBURSEMENT.configure(yscrollcommand = YSCROLLBAR.set, xscrollcommand = XSCROLLBAR.set) 
        TREE_DISBURSEMENT.pack()
        XSCROLLBAR.pack(fill ="x")

    def exportDisbursementSummary(self, tree):
        wb = load_workbook(PATH_TEMPLATE + "DVSUMMARY.xlsx")
        sheet = wb.active
        sheet["B5"] = f"{CALENDAR_FROM.get_date()} to {CALENDAR_TO.get_date()}"
        count = 8
        for i in tree.get_children():
            sheet["A" + str(count)] = str(tree.item(i)['values'][0]).zfill(8)
            sheet["B" + str(count)] = tree.item(i)['values'][1]
            sheet["C" + str(count)] = tree.item(i)['values'][2]
            sheet["D" + str(count)] = tree.item(i)['values'][3]
            sheet["E" + str(count)] = tree.item(i)['values'][4]
            sheet["F" + str(count)] = tree.item(i)['values'][5]
            sheet["G" + str(count)] = tree.item(i)['values'][6]
            sheet["H" + str(count)] = tree.item(i)['values'][7]
            sheet["I" + str(count)] = self.returnFloatAmount(tree.item(i)['values'][8])
            sheet["J" + str(count)] = tree.item(i)['values'][9]
            sheet["K" + str(count)] = tree.item(i)['values'][10]
            sheet["L" + str(count)] = tree.item(i)['values'][11]
            sheet["M" + str(count)] = tree.item(i)['values'][12]
            sheet["N" + str(count)] = tree.item(i)['values'][13]
            count += 1
        sheet["A" + str(count)] == "/" + self.returnUserName(USER, 0) + str(datetime.date.today())
        wb.save(PATH_SAVE + "DVSUMMARY.xlsx")
        startfile(PATH_SAVE + "DVSUMMARY.xlsx", "open")

### MENU_BILLING ###
    def showReceivables(self, *args):
        self.clearWorkspace()
        FRAME_RECEIVABLES = LabelFrame(FRAME_4, text = "Receivables", font = APP_FONT)
        FRAME_RECEIVABLES.grid(column = 1, row = 0)

        global SUB1_RECEIVABLES
        SUB1_RECEIVABLES = Frame(FRAME_RECEIVABLES)
        SUB1_RECEIVABLES.grid(column = 0, row = 0, sticky = W, pady = SUBMENU_PADY)

        SUB2_RECEIVABLES = Frame(FRAME_RECEIVABLES)
        SUB2_RECEIVABLES.grid(column = 0, row = 1, sticky = W)

        LABEL_SEARCH = Label(SUB1_RECEIVABLES, text = "Search", width = 5, font = APP_FONT, anchor = W)
        LABEL_SEARCH.grid(column = 0, row = 0, pady = SUBMENU_PADY, sticky = W)

        TEXTVAR_SEARCH = StringVar()
        ENTRY_SEARCH = Entry(SUB1_RECEIVABLES, textvariable = TEXTVAR_SEARCH, width = 25, font = APP_FONT)
        ENTRY_SEARCH.grid(column = 1, row = 0, padx = MENU_PADX, sticky = W)
        ENTRY_SEARCH.bind("<Return>", lambda e: self.searchReceivable(TEXTVAR_SEARCH.get()))

        BUTTON_GO = Button(SUB1_RECEIVABLES, text = "GO", width = 5, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = lambda: self.searchReceivable(TEXTVAR_SEARCH.get()))
        BUTTON_GO.grid(column = 2, row = 0, padx = MENU_PADX)

        BUTTON_REFRESH = Button(SUB1_RECEIVABLES, text = "REFRESH", width = 10, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = self.refreshReceivable)
        # BUTTON_REFRESH.grid(column = 3, row = 0, padx = MENU_PADX)

        BUTTON_ADD = Button(SUB1_RECEIVABLES, text = "ADD", width = 5, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = self.showAddEditReceivable)
        BUTTON_ADD.grid(column = 4, row = 0, padx = MENU_PADX)
        
        BUTTON_IMPORT = Button(SUB1_RECEIVABLES, text = "IMPORT", width = 10, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, state = DISABLED, cursor = "arrow", command = self.showImportSOA)
        BUTTON_IMPORT.grid(column = 5, row = 0, padx = MENU_PADX)
        
        global CALENDAR_START
        CALENDAR_START = DateEntry(SUB1_RECEIVABLES, firstweekday = "sunday", date_pattern = "yyyy-mm-dd", state = "readonly")
        CALENDAR_START.grid(column = 6, row = 0, sticky = W, padx = TOP_PADX + 10)
        
        global CALENDAR_END
        CALENDAR_END = DateEntry(SUB1_RECEIVABLES, firstweekday = "sunday", date_pattern = "yyyy-mm-dd", state = "readonly")
        CALENDAR_END.grid(column = 7, row = 0, sticky = W, padx = TOP_PADX + 10)

        global PROGRESS_BAR
        PROGRESS_BAR = tk.Progressbar(SUB1_RECEIVABLES, orient = HORIZONTAL, length = 200, mode = "determinate")
        PROGRESS_BAR.grid(column = 8, row = 0, padx = TOP_PADX + 10, sticky = E)

        BUTTON_PREVIOUS = Button(SUB1_RECEIVABLES, text = "<", width = 3, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, state = DISABLED, cursor = "arrow", command = lambda: None)
        # BUTTON_PREVIOUS.grid(column = 9, row = 0, padx = MENU_PADX)

        BUTTON_NEXT = Button(SUB1_RECEIVABLES, text = ">", width = 3, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, state = DISABLED, cursor = "arrow", command = lambda: None)
        # BUTTON_NEXT.grid(column = 10, row = 0, padx = MENU_PADX)

        global TREE_RECEIVABLES
        TREE_RECEIVABLES = tk.Treeview(SUB2_RECEIVABLES, height = 28, selectmode = "browse")
        TREE_RECEIVABLES["columns"] = ("SOA #", "SOA Date", "GL Date", "Client Name", "Particulars", "Amount", "Balance", "isApproved", "isVoid", "Entity", "Encoder", "Encoded", "Approver", "Approved")
        TREE_RECEIVABLES.column("#0", width = 0, minwidth = 0, stretch = True)
        TREE_RECEIVABLES.column("SOA #", anchor = W, minwidth = 50, width = 75)
        TREE_RECEIVABLES.column("SOA Date", anchor = W, minwidth = 50, width = 75)
        TREE_RECEIVABLES.column("GL Date", anchor = W, minwidth = 50, width = 75)
        TREE_RECEIVABLES.column("Client Name", anchor = W, minwidth = 50, width = 200)
        TREE_RECEIVABLES.column("Particulars", anchor = W, minwidth = 50, width = 150)
        TREE_RECEIVABLES.column("Amount", anchor = E, minwidth = 50, width = 85)
        TREE_RECEIVABLES.column("Balance", anchor = E, minwidth = 50, width = 85)
        TREE_RECEIVABLES.column("isApproved", anchor = W, minwidth = 50, width = 75)
        TREE_RECEIVABLES.column("isVoid", anchor = W, minwidth = 50, width = 75)
        TREE_RECEIVABLES.column("Entity", anchor = W, minwidth = 50, width = 75)
        TREE_RECEIVABLES.column("Encoder", anchor = W, minwidth = 50, width = 75)
        TREE_RECEIVABLES.column("Encoded", anchor = W, minwidth = 50, width = 75)
        TREE_RECEIVABLES.column("Approver", anchor = W, minwidth = 50, width = 75)
        TREE_RECEIVABLES.column("Approved", anchor = W, minwidth = 50, width = 75)
        
        TREE_RECEIVABLES.heading("#0", text = "", anchor = W)
        TREE_RECEIVABLES.heading("SOA #", text = "SOA #", anchor = N)
        TREE_RECEIVABLES.heading("SOA Date", text = "SOA Date", anchor = N)
        TREE_RECEIVABLES.heading("GL Date", text = "GL Date", anchor = N)
        TREE_RECEIVABLES.heading("Client Name", text = "Client Name", anchor = N)
        TREE_RECEIVABLES.heading("Particulars", text = "Particulars", anchor = N)
        TREE_RECEIVABLES.heading("Amount", text = "Amount", anchor = N)
        TREE_RECEIVABLES.heading("Balance", text = "Balance", anchor = N)
        TREE_RECEIVABLES.heading("isApproved", text = "isApproved", anchor = N)
        TREE_RECEIVABLES.heading("isVoid", text = "isVoid", anchor = N)
        TREE_RECEIVABLES.heading("Entity", text = "Entity", anchor = N)
        TREE_RECEIVABLES.heading("Encoder", text = "Encoder", anchor = N)
        TREE_RECEIVABLES.heading("Encoded", text = "Encoded", anchor = N)
        TREE_RECEIVABLES.heading("Approver", text = "Approver", anchor = N)
        TREE_RECEIVABLES.heading("Approved", text = "Approved", anchor = N)

        POPUP_RECEIVABLES = Menu(TREE_RECEIVABLES, tearoff = 0)
        POPUP_RECEIVABLES.add_command(command = self.editSOA, label = "Edit")
        # POPUP_RECEIVABLES.add_command(command = None, label = "Delete")
        TREE_RECEIVABLES.bind("<Button-3>", lambda e: self.popupMenu(TREE_RECEIVABLES, POPUP_RECEIVABLES, e))
        TREE_RECEIVABLES.bind("<Double-1>", self.editSOA)

        global STYLE_RECEIVABLES
        STYLE_RECEIVABLES = tk.Style()
        STYLE_RECEIVABLES.map("Treeview", foreground = self.fixedMap("foreground", STYLE_RECEIVABLES), background = self.fixedMap("background", STYLE_RECEIVABLES))

        TREE_RECEIVABLES.tag_configure("oddrow", background = None)
        TREE_RECEIVABLES.tag_configure("evenrow", background = TREE_TAG_EVENROW)
        TREE_RECEIVABLES.tag_configure("void", background = TREE_TAG_VOID)

        db.commit()
        cursor.execute(f"SELECT soaNumber, soaDate, glDate, clientCode, particulars, amount, isApproved, void, encoder, DATE(encoded), approver, DATE(approved) FROM tblreceivables WHERE soaDate BETWEEN '{CALENDAR_START.get_date()}' AND '{CALENDAR_END.get_date()}' ORDER BY soaNumber DESC")
        result = cursor.fetchall()
        
        YSCROLLBAR = tk.Scrollbar(SUB2_RECEIVABLES, orient = "vertical", command = TREE_RECEIVABLES.yview)
        YSCROLLBAR.pack(side = "right", fill = "y")

        XSCROLLBAR = tk.Scrollbar(SUB2_RECEIVABLES, orient = "horizontal", command = TREE_RECEIVABLES.xview)
        
        TREE_RECEIVABLES.configure(yscrollcommand = YSCROLLBAR.set, xscrollcommand = XSCROLLBAR.set) 
        TREE_RECEIVABLES.pack()
        XSCROLLBAR.pack(fill ="x")

        count = 0
        soanumbers, skipped = [], []
        for i in result:
            if i[0] not in soanumbers:
                if i[7] != "Yes":
                    if count % 2 == 0:
                        TREE_RECEIVABLES.insert("", "end", values = (str(i[0]).zfill(8),i[1],i[2],self.returnClientName(i[3]),i[4],self.returnTotalSOAAmount(i[0]),self.returnTotalSOABalance(i[0], self.returnTotalSOAAmount(i[0])),i[6],i[7],self.returnClientEntityType(i[3]),self.returnUserName(i[8], 0),i[9],self.returnUserName(i[10], 0),i[11]), tags = "evenrow")
                    else:
                        TREE_RECEIVABLES.insert("", "end", values = (str(i[0]).zfill(8),i[1],i[2],self.returnClientName(i[3]),i[4],self.returnTotalSOAAmount(i[0]),self.returnTotalSOABalance(i[0], self.returnTotalSOAAmount(i[0])),i[6],i[7],self.returnClientEntityType(i[3]),self.returnUserName(i[8], 0),i[9],self.returnUserName(i[10], 0),i[11]), tags = "oddrow")
                else:
                    TREE_RECEIVABLES.insert("", "end", values = (str(i[0]).zfill(8),i[1],i[2],self.returnClientName(i[3]),i[4],self.returnTotalSOAAmount(i[0]),self.returnTotalSOABalance(i[0], self.returnTotalSOAAmount(i[0])),i[6],i[7],self.returnClientEntityType(i[3]),self.returnUserName(i[8], 0),i[9],self.returnUserName(i[10], 0),i[11]), tags = "void")
                count += 1
                soanumbers.append(i[0])
                PROGRESS_BAR["value"] = round((count/(len(result)-len(skipped)))*100, 0)
                SUB1_RECEIVABLES.update()
            else:
                skipped.append(i[0])
        PROGRESS_BAR.grid_remove()
        
        if self.returnAccess(USER, 12) == 0:
            BUTTON_ADD.config(state = DISABLED, cursor = "arrow")
            BUTTON_IMPORT.config(state = DISABLED, cursor = "arrow")
            
        if self.returnUserName(USER, 3) == "ASD":
            BUTTON_IMPORT.config(state = NORMAL, cursor = "hand2")

    def showCollections(self, *args):
        self.clearWorkspace()
        FRAME_COLLECTIONS = LabelFrame(FRAME_4, text = "Collections", font = APP_FONT)
        FRAME_COLLECTIONS.grid(column = 1, row = 0)

        global SUB1_COLLECTIONS
        SUB1_COLLECTIONS = Frame(FRAME_COLLECTIONS)
        SUB1_COLLECTIONS.grid(column = 0, row = 0, sticky = W, pady = SUBMENU_PADY)

        SUB2_COLLECTIONS = Frame(FRAME_COLLECTIONS)
        SUB2_COLLECTIONS.grid(column = 0, row = 1, sticky = W)

        LABEL_SEARCH = Label(SUB1_COLLECTIONS, text = "Search", width = 5, font = APP_FONT, anchor = W)
        LABEL_SEARCH.grid(column = 0, row = 0, pady = SUBMENU_PADY, sticky = W)

        TEXTVAR_SEARCH = StringVar()
        ENTRY_SEARCH = Entry(SUB1_COLLECTIONS, textvariable = TEXTVAR_SEARCH, width = 25, font = APP_FONT)
        ENTRY_SEARCH.grid(column = 1, row = 0, padx = MENU_PADX, sticky = W)
        ENTRY_SEARCH.bind("<Return>", lambda e: self.searchCollection(TEXTVAR_SEARCH.get()))

        BUTTON_GO = Button(SUB1_COLLECTIONS, text = "GO", width = 5, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = lambda: self.searchCollection(TEXTVAR_SEARCH.get()))
        BUTTON_GO.grid(column = 2, row = 0, padx = MENU_PADX)

        BUTTON_REFRESH = Button(SUB1_COLLECTIONS, text = "REFRESH", width = 10, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = self.refreshCollection)
        # BUTTON_REFRESH.grid(column = 3, row = 0, padx = MENU_PADX)

        BUTTON_ADD = Button(SUB1_COLLECTIONS, text = "ADD", width = 5, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = self.showAddEditCollection)
        BUTTON_ADD.grid(column = 4, row = 0, padx = MENU_PADX)
        
        BUTTON_IMPORT = Button(SUB1_COLLECTIONS, text = "IMPORT", width = 10, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = self.showImportOR)
        BUTTON_IMPORT.grid(column = 5, row = 0, padx = MENU_PADX)

        global CALENDAR_START
        CALENDAR_START = DateEntry(SUB1_COLLECTIONS, firstweekday = "sunday", date_pattern = "yyyy-mm-dd", state = "readonly")
        CALENDAR_START.grid(column = 6, row = 0, sticky = W, padx = TOP_PADX + 10)
        
        global CALENDAR_END
        CALENDAR_END = DateEntry(SUB1_COLLECTIONS, firstweekday = "sunday", date_pattern = "yyyy-mm-dd", state = "readonly")
        CALENDAR_END.grid(column = 7, row = 0, sticky = W, padx = TOP_PADX + 10)

        global PROGRESS_BAR
        PROGRESS_BAR = tk.Progressbar(SUB1_COLLECTIONS, orient = HORIZONTAL, length = 200, mode = "determinate")
        PROGRESS_BAR.grid(column = 8, row = 0, padx = TOP_PADX + 10, sticky = E)

        global TREE_COLLECTIONS
        TREE_COLLECTIONS = tk.Treeview(SUB2_COLLECTIONS, height = 28, selectmode = "browse")
        TREE_COLLECTIONS["columns"] = ("OR #", "OR Date", "GL Date", "Client Name", "Particulars", "Amount", "isApproved", "Void", "Encoder", "Encoded", "Approver", "Approved")
        TREE_COLLECTIONS.column("#0", width = 0, minwidth = 0, stretch = True)
        TREE_COLLECTIONS.column("OR #", anchor = W, minwidth = 50, width = 75)
        TREE_COLLECTIONS.column("OR Date", anchor = W, minwidth = 50, width = 75)
        TREE_COLLECTIONS.column("GL Date", anchor = W, minwidth = 50, width = 75)
        TREE_COLLECTIONS.column("Client Name", anchor = W, minwidth = 50, width = 200)
        TREE_COLLECTIONS.column("Particulars", anchor = W, minwidth = 50, width = 150)
        TREE_COLLECTIONS.column("Amount", anchor = E, minwidth = 50, width = 85)
        TREE_COLLECTIONS.column("isApproved", anchor = E, minwidth = 50, width = 75)
        TREE_COLLECTIONS.column("Void", anchor = E, minwidth = 50, width = 75)
        TREE_COLLECTIONS.column("Encoder", anchor = W, minwidth = 50, width = 75)
        TREE_COLLECTIONS.column("Encoded", anchor = W, minwidth = 50, width = 75)
        TREE_COLLECTIONS.column("Approver", anchor = W, minwidth = 50, width = 75)
        TREE_COLLECTIONS.column("Approved", anchor = W, minwidth = 50, width = 75)
        
        TREE_COLLECTIONS.heading("#0", text = "", anchor = W)
        TREE_COLLECTIONS.heading("OR #", text = "OR #", anchor = N)
        TREE_COLLECTIONS.heading("OR Date", text = "OR Date", anchor = N)
        TREE_COLLECTIONS.heading("GL Date", text = "GL Date", anchor = N)
        TREE_COLLECTIONS.heading("Client Name", text = "Client Name", anchor = N)
        TREE_COLLECTIONS.heading("Particulars", text = "Particulars", anchor = N)
        TREE_COLLECTIONS.heading("Amount", text = "Amount", anchor = N)
        TREE_COLLECTIONS.heading("isApproved", text = "isApproved", anchor = N)
        TREE_COLLECTIONS.heading("Void", text = "Void", anchor = N)
        TREE_COLLECTIONS.heading("Encoder", text = "Encoder", anchor = N)
        TREE_COLLECTIONS.heading("Encoded", text = "Encoded", anchor = N)
        TREE_COLLECTIONS.heading("Approver", text = "Approver", anchor = N)
        TREE_COLLECTIONS.heading("Approved", text = "Approved", anchor = N)

        POPUP_COLLECTIONS = Menu(TREE_COLLECTIONS, tearoff = 0)
        POPUP_COLLECTIONS.add_command(command = self.editCollection, label = "Edit")
        # POPUP_COLLECTIONS.add_command(command = None, label = "Delete")
        TREE_COLLECTIONS.bind("<Button-3>", lambda e: self.popupMenu(TREE_COLLECTIONS, POPUP_COLLECTIONS, e))
        TREE_COLLECTIONS.bind("<Double-1>", self.editCollection)

        global STYLE_COLLECTIONS
        STYLE_COLLECTIONS = tk.Style()
        STYLE_COLLECTIONS.map("Treeview", foreground = self.fixedMap("foreground", STYLE_COLLECTIONS), background = self.fixedMap("background", STYLE_COLLECTIONS))

        TREE_COLLECTIONS.tag_configure("oddrow", background = None)
        TREE_COLLECTIONS.tag_configure("evenrow", background = TREE_TAG_EVENROW)
        TREE_COLLECTIONS.tag_configure("void", background = TREE_TAG_VOID)

        db.commit()
        cursor.execute(f"SELECT orNumber, orDate, glDate, clientCode, particulars, amount, isApproved, isVoid, encoder, DATE(encoded), approver, DATE(approved) FROM tblcollections WHERE orDate BETWEEN '{CALENDAR_START.get_date()}' AND '{CALENDAR_END.get_date()}' ORDER BY orNumber DESC")
        result = cursor.fetchall()
        
        YSCROLLBAR = tk.Scrollbar(SUB2_COLLECTIONS, orient = "vertical", command = TREE_COLLECTIONS.yview)
        YSCROLLBAR.pack(side = "right", fill = "y")

        XSCROLLBAR = tk.Scrollbar(SUB2_COLLECTIONS, orient = "horizontal", command = TREE_COLLECTIONS.xview)
        
        TREE_COLLECTIONS.configure(yscrollcommand = YSCROLLBAR.set, xscrollcommand = XSCROLLBAR.set) 
        TREE_COLLECTIONS.pack()
        XSCROLLBAR.pack(fill ="x")

        count = 0
        ornumbers, skipped = [], []
        for i in result:
            if i[0] not in ornumbers:
                if i[7] != "Yes":
                    if count % 2 == 0:
                        TREE_COLLECTIONS.insert("", "end", values = (str(i[0]).zfill(8),i[1],i[2],self.returnClientName(i[3]),i[4],self.returnTotalORAmount(i[0]),i[6],i[7],self.returnUserName(i[8], 0),i[9],self.returnUserName(i[10], 0),i[11]), tags = "evenrow")
                    else:
                        TREE_COLLECTIONS.insert("", "end", values = (str(i[0]).zfill(8),i[1],i[2],self.returnClientName(i[3]),i[4],self.returnTotalORAmount(i[0]),i[6],i[7],self.returnUserName(i[8], 0),i[9],self.returnUserName(i[10], 0),i[11]), tags = "oddrow")
                else:
                    TREE_COLLECTIONS.insert("", "end", values = (str(i[0]).zfill(8),i[1],i[2],self.returnClientName(i[3]),i[4],self.returnTotalORAmount(i[0]),i[6],i[7],self.returnUserName(i[8], 0),i[9],self.returnUserName(i[10], 0),i[11]), tags = "void")
                count += 1
                ornumbers.append(i[0])
                PROGRESS_BAR["value"] = round((count/(len(result)-len(skipped)))*100, 0)
                SUB1_COLLECTIONS.update()
            else:
                skipped.append(i[0])
        PROGRESS_BAR.grid_remove()

        if self.returnAccess(USER, 15) == 0:
            BUTTON_ADD.config(state = DISABLED, cursor = "arrow")
            BUTTON_IMPORT.config(state = DISABLED, cursor = "arrow")

### MENU_BILLING_RECEIVABLES ###
    def showAddEditReceivable(self):
        global TOP_RECEIVABLES
        TOP_RECEIVABLES = Toplevel()
        TOP_RECEIVABLES.title("Add - Receivables")
        TOP_RECEIVABLES.iconbitmap(PATH_ICON + "icon.ico")
        TOP_RECEIVABLES.geometry("1209x595+100+20")
        TOP_RECEIVABLES.resizable(height = False, width = False)
        TOP_RECEIVABLES.grab_set()
        TOP_RECEIVABLES.focus_force()
        
        global required
        required = []
        
        global SUB_FRAME1
        SUB_FRAME1 = Frame(TOP_RECEIVABLES) #details frame
        SUB_FRAME1.grid(column = 0, row = 0, sticky = W)
        
        SUB_SUBFRAME1 = Frame(SUB_FRAME1)
        SUB_SUBFRAME1.grid(column = 0, row = 0, padx = TOP_PADX + 10, sticky = W)
        
        SUB_SUBSUBFRAME1 = Frame(SUB_SUBFRAME1)
        SUB_SUBSUBFRAME1.grid(column = 1, row = 0, sticky = W)
        
        LABEL_CLIENT = Label(SUB_SUBFRAME1, text = "Client", font = APP_FONT)
        LABEL_CLIENT.grid(column = 0, row = 0, sticky = W)
        
        LABEL_PARTICULARS = Label(SUB_SUBFRAME1, text = "Particulars", font = APP_FONT)
        LABEL_PARTICULARS.grid(column = 0, row = 1, sticky = W)
        
        LABEL_REFERENCE = Label(SUB_SUBFRAME1, text = "Reference", font = APP_FONT)
        LABEL_REFERENCE.grid(column = 0, row = 2, sticky = W)
        
        global TEXTVAR_CLIENTCODE, ENTRY_CLIENTCODE
        TEXTVAR_CLIENTCODE = StringVar()
        ENTRY_CLIENTCODE = Entry(SUB_SUBSUBFRAME1, textvariable = TEXTVAR_CLIENTCODE, font = APP_FONT, width = 15)
        ENTRY_CLIENTCODE.grid(column = 0, row = 0, sticky = W)
        ENTRY_CLIENTCODE.bind("<FocusOut>", lambda e: self.populateSOAFields(TEXTVAR_CLIENTCODE.get()))
        required.append(TEXTVAR_CLIENTCODE)

        global BUTTON_CLIENT
        BUTTON_CLIENT = Button(SUB_SUBSUBFRAME1, text = "...", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, cursor = "hand2", command = self.showClientSelection)
        BUTTON_CLIENT.grid(column = 0, row = 0, sticky = E)
        
        global TEXTVAR_CLIENTNAME
        TEXTVAR_CLIENTNAME = StringVar()
        ENTRY_CLIENTNAME = Entry(SUB_SUBSUBFRAME1, textvariable = TEXTVAR_CLIENTNAME, font = APP_FONT, width = 55, state = "readonly")
        ENTRY_CLIENTNAME.grid(column = 1, row = 0, sticky = W)
        required.append(TEXTVAR_CLIENTNAME)

        global ENTRY_PARTICULARS
        ENTRY_PARTICULARS = Text(SUB_SUBFRAME1, font = APP_FONT, width = 70, height = 2)
        ENTRY_PARTICULARS.grid(column = 1, row = 1, sticky = W)
        required.append(ENTRY_PARTICULARS)
        
        global TEXTVAR_REFERENCE, ENTRY_REFERENCE
        TEXTVAR_REFERENCE = StringVar()
        ENTRY_REFERENCE = Entry(SUB_SUBFRAME1, textvariable = TEXTVAR_REFERENCE, font = APP_FONT, width = 15)
        ENTRY_REFERENCE.grid(column = 1, row = 2, sticky = W)
        # required.append(ENTRY_REFERENCE)
        
        SUB_SUBFRAME2 = Frame(SUB_FRAME1) #soa dates and numbers
        SUB_SUBFRAME2.grid(column = 1, row = 0, padx = TOP_PADX + 10, sticky = W)
        
        LABEL_SOANUMBER = Label(SUB_SUBFRAME2, text = "SOA #", font = APP_FONT)
        LABEL_SOANUMBER.grid(column = 0, row = 0, sticky = W)
        
        LABEL_SOADATE = Label(SUB_SUBFRAME2, text = "SOA Date", font = APP_FONT)
        LABEL_SOADATE.grid(column = 0, row = 1, sticky = W)
        
        LABEL_GLDATE = Label(SUB_SUBFRAME2, text = "GL Date", font = APP_FONT)
        LABEL_GLDATE.grid(column = 0, row = 2, sticky = W)
        
        LABEL_PAYDATE = Label(SUB_SUBFRAME2, text = "Pay Date", font = APP_FONT)
        LABEL_PAYDATE.grid(column = 0, row = 3, sticky = W)
        
        global TEXTVAR_SOANUMBER, ENTRY_SOANUMBER
        TEXTVAR_SOANUMBER = StringVar()
        ENTRY_SOANUMBER = Entry(SUB_SUBFRAME2, textvariable = TEXTVAR_SOANUMBER, font = APP_FONT, width = 13)
        ENTRY_SOANUMBER.grid(column = 1, row = 0, sticky = W)
        ENTRY_SOANUMBER.bind("<FocusOut>", lambda e: self.validateSOANumber(TEXTVAR_SOANUMBER))
        required.append(TEXTVAR_SOANUMBER)
        
        global CALENDAR_SOADATE
        CALENDAR_SOADATE = DateEntry(SUB_SUBFRAME2, firstweekday = "sunday", date_pattern = "yyyy-mm-dd")
        CALENDAR_SOADATE.grid(column = 1, row = 1, sticky = W)
        CALENDAR_SOADATE.bind("<FocusOut>", lambda e: CALENDAR_GLDATE.set_date(self.returnLastDayOfMonth(str(CALENDAR_SOADATE.get_date()))))
        required.append(CALENDAR_SOADATE)
        
        global CALENDAR_GLDATE
        CALENDAR_GLDATE = DateEntry(SUB_SUBFRAME2, firstweekday = "sunday", date_pattern = "yyyy-mm-dd")
        CALENDAR_GLDATE.grid(column = 1, row = 2, sticky = W)
        CALENDAR_GLDATE.set_date(self.returnLastDayOfMonth(str(CALENDAR_SOADATE.get_date())))
        required.append(CALENDAR_GLDATE)
        
        global CALENDAR_PAYDATE
        CALENDAR_PAYDATE = DateEntry(SUB_SUBFRAME2, firstweekday = "sunday", date_pattern = "yyyy-mm-dd")
        CALENDAR_PAYDATE.grid(column = 1, row = 3, sticky = W)
        required.append(CALENDAR_PAYDATE)
        CALENDAR_GLDATE.set_date(self.returnLastDayOfMonth(str(CALENDAR_PAYDATE.get_date())))
        
        SUB_SUBFRAME3 = Frame(SUB_FRAME1)
        SUB_SUBFRAME3.grid(column = 2, row = 0, padx = TOP_PADX, sticky = W)
        
        LABEL_GROSS = Label(SUB_SUBFRAME3, text = "Gross", font = APP_FONT)
        LABEL_GROSS.grid(column = 0, row = 0, sticky = W)
        
        LABEL_VAT = Label(SUB_SUBFRAME3, text = "VAT", font = APP_FONT)
        LABEL_VAT.grid(column = 0, row = 1, sticky = W)
        
        LABEL_EWT = Label(SUB_SUBFRAME3, text = "EWT", font = APP_FONT)
        LABEL_EWT.grid(column = 0, row = 2, sticky = W)
        
        LABEL_CVAT = Label(SUB_SUBFRAME3, text = "CVAT", font = APP_FONT)
        LABEL_CVAT.grid(column = 0, row = 3, sticky = W)
        
        LABEL_NET = Label(SUB_SUBFRAME3, text = "Net", font = APP_FONT)
        LABEL_NET.grid(column = 0, row = 4, sticky = W)
        
        global TEXTVAR_TGROSS
        TEXTVAR_TGROSS = StringVar()
        ENTRY_GROSS = Entry(SUB_SUBFRAME3, textvariable = TEXTVAR_TGROSS, font = APP_FONT, width = 12, state = "readonly", justify = RIGHT)
        ENTRY_GROSS.grid(column = 1, row = 0, sticky = W)
        
        global TEXTVAR_TVAT
        TEXTVAR_TVAT = StringVar()
        ENTRY_VAT = Entry(SUB_SUBFRAME3, textvariable = TEXTVAR_TVAT, font = APP_FONT, width = 12, state = "readonly", justify = RIGHT)
        ENTRY_VAT.grid(column = 1, row = 1, sticky = W)
        
        global TEXTVAR_TEWT
        TEXTVAR_TEWT = StringVar()
        ENTRY_EWT = Entry(SUB_SUBFRAME3, textvariable = TEXTVAR_TEWT, font = APP_FONT, width = 12, state = "readonly", justify = RIGHT)
        ENTRY_EWT.grid(column = 1, row = 2, sticky = W)
        
        global TEXTVAR_TCVAT
        TEXTVAR_TCVAT = StringVar()
        ENTRY_CVAT = Entry(SUB_SUBFRAME3, textvariable = TEXTVAR_TCVAT, font = APP_FONT, width = 12, state = "readonly", justify = RIGHT)
        ENTRY_CVAT.grid(column = 1, row = 3, sticky = W)
        
        global TEXTVAR_TNET
        TEXTVAR_TNET = StringVar()
        ENTRY_NET = Entry(SUB_SUBFRAME3, textvariable = TEXTVAR_TNET, font = APP_FONT, width = 12, state = "readonly", justify = RIGHT)
        ENTRY_NET.grid(column = 1, row = 4, sticky = W)
        
        SUB_SUBFRAME4 = Frame(SUB_FRAME1)
        SUB_SUBFRAME4.grid(column = 3, row = 0, padx = TOP_PADX + 10, sticky = NW)
        
        LABEL_BILLTYPE = Label(SUB_SUBFRAME4, text = "Bill Type", font = APP_FONT)
        LABEL_BILLTYPE.grid(column = 0, row = 0, sticky = W)
        
        LABEL_ENTITYTYPE = Label(SUB_SUBFRAME4, text = "Entity Type", font = APP_FONT)
        LABEL_ENTITYTYPE.grid(column = 0, row = 1, sticky = W)
        
        global TEXTVAR_BILLTYPE
        TEXTVAR_BILLTYPE = StringVar()
        ENTRY_BILLTYPE = Entry(SUB_SUBFRAME4, textvariable = TEXTVAR_BILLTYPE, font = APP_FONT, width = 12, state = "readonly")
        ENTRY_BILLTYPE.grid(column = 1, row = 0, sticky = W)
        TEXTVAR_BILLTYPE.set("Manual")
        required.append(TEXTVAR_BILLTYPE)
        
        global TEXTVAR_ENTITYTYPE
        TEXTVAR_ENTITYTYPE = StringVar()
        ENTRY_ENTITYTYPE = Entry(SUB_SUBFRAME4, textvariable = TEXTVAR_ENTITYTYPE, font = APP_FONT, width = 12, state = "readonly")
        ENTRY_ENTITYTYPE.grid(column = 1, row = 1, sticky = W)
        required.append(TEXTVAR_ENTITYTYPE)

        SUB_FRAME2 = Frame(TOP_RECEIVABLES) #headers frame
        SUB_FRAME2.grid(column = 0, row = 1, sticky = W)
        
        LABEL_CODE = Label(SUB_FRAME2, text = "Transaction Type", width = 30, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_CODE.grid(column = 0, row = 0)

        LABEL_TITLE = Label(SUB_FRAME2, text = "Amount", width = 15, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_TITLE.grid(column = 1, row = 0)

        LABEL_CENTER = Label(SUB_FRAME2, text = "Tax", width = 10, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_CENTER.grid(column = 2, row = 0)

        LABEL_GROSS = Label(SUB_FRAME2, text = "Category", width = 26, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_GROSS.grid(column = 3, row = 0)

        LABEL_EXPENSE = Label(SUB_FRAME2, text = "VAT", width = 12, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_EXPENSE.grid(column = 4, row = 0)

        LABEL_VAT = Label(SUB_FRAME2, text = "EWT", width = 12, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_VAT.grid(column = 5, row = 0)

        LABEL_EWT = Label(SUB_FRAME2, text = "CVAT", width = 12, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_EWT.grid(column = 6, row = 0)

        LABEL_NET = Label(SUB_FRAME2, text = "Net", width = 12, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_NET.grid(column = 7, row = 0)

        LABEL_DESC = Label(SUB_FRAME2, text = "Remarks", width = 30, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_DESC.grid(column = 8, row = 0)

        LABEL_CANCEL = Label(SUB_FRAME2, text = "[x]", width = 2, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_CANCEL.grid(column = 9, row = 0)
        
        SUB_FRAME3 = Frame(TOP_RECEIVABLES) #scroll frame
        SUB_FRAME3.grid(column = 0, row = 2, sticky = W)
        
        self.createScrollFrame(SUB_FRAME3, 280, 1185, 20, 0, 0)
        
        global soaFrame, soaTrans, soaGross, soaTax, soaComboTax, soaCategory, soaRemarks, soaClear
        global soaVAT, soaEWT, soaCVAT, soaNET
        soaFrame, soaTrans, soaGross, soaTax, soaComboTax, soaCategory, soaRemarks, soaClear = [], [], [], [], [], [], [], []
        soaVAT, soaEWT, soaCVAT, soaNET = [], [], [], []
        
        for i in range(11):
            self.showReceivableLines(SCROLLABLE_FRAME, i)
        
        global SUB_FRAME6
        SUB_FRAME6 = Frame(TOP_RECEIVABLES) #entry label frame
        SUB_FRAME6.grid(column = 0, row = 3, sticky = NW)
        
        global SUB_FRAME5
        SUB_FRAME5 = Frame(TOP_RECEIVABLES) #entry lines frame
        SUB_FRAME5.grid(column = 0, row = 4, sticky = NW)
        
        LABEL_CODE = Label(SUB_FRAME6, text = "Code", width = 12, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_CODE.grid(column = 0, row = 0)

        LABEL_TITLE = Label(SUB_FRAME6, text = "Title", width = 30, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_TITLE.grid(column = 1, row = 0)

        LABEL_CENTER = Label(SUB_FRAME6, text = "Amount", width = 15, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_CENTER.grid(column = 2, row = 0)

        LABEL_GROSS = Label(SUB_FRAME6, text = "Dr/Cr", width = 10, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_GROSS.grid(column = 3, row = 0)

        LABEL_EXPENSE = Label(SUB_FRAME6, text = "Remarks", width = 31, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_EXPENSE.grid(column = 4, row = 0)

        LABEL_X = Label(SUB_FRAME6, text = "[x]", width = 2, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_X.grid(column = 5, row = 0)
        
        self.createScrollFrame(SUB_FRAME5, 150, 735, 20, 0, 0)
        
        global entryFrame, entryCode, entryTitle, entryAmount, entryDrCr, entryRemarks, entryClear
        entryFrame, entryCode, entryTitle, entryAmount, entryDrCr, entryRemarks, entryClear = [], [], [], [], [], [], []
        
        for i in range(11):
            self.showSOAEntryLines(SCROLLABLE_FRAME, i)

        SUB_FRAME7 = Frame(SUB_FRAME5) #entry total frame
        SUB_FRAME7.grid(column = 1, row = 0, sticky = NW)
        
        LABEL_DEBIT = Label(SUB_FRAME7, text = "Debit", width = 12, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_DEBIT.grid(column = 0, row = 0)

        LABEL_CREDIT = Label(SUB_FRAME7, text = "Credit", width = 12, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_CREDIT.grid(column = 0, row = 1)

        LABEL_VARIANCE = Label(SUB_FRAME7, text = "Variance", width = 12, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_VARIANCE.grid(column = 0, row = 2)
        
        global TEXTVAR_DEBIT
        TEXTVAR_DEBIT = StringVar()
        ENTRY_DEBIT = Entry(SUB_FRAME7, textvariable = TEXTVAR_DEBIT, font = APP_FONT, width = 12, justify = RIGHT, state = "readonly")
        ENTRY_DEBIT.grid(column = 1, row = 0, sticky = W)
        
        global TEXTVAR_CREDIT
        TEXTVAR_CREDIT = StringVar()
        ENTRY_CREDIT = Entry(SUB_FRAME7, textvariable = TEXTVAR_CREDIT, font = APP_FONT, width = 12, justify = RIGHT, state = "readonly")
        ENTRY_CREDIT.grid(column = 1, row = 1, sticky = W)
        
        global TEXTVAR_VARIANCE
        TEXTVAR_VARIANCE = StringVar()
        ENTRY_VARIANCE = Entry(SUB_FRAME7, textvariable = TEXTVAR_VARIANCE, font = APP_FONT, width = 12, justify = RIGHT, state = "readonly")
        ENTRY_VARIANCE.grid(column = 1, row = 2, sticky = W)
        
        SUB_FRAME4 = Frame(TOP_RECEIVABLES) #button frame
        SUB_FRAME4.grid(column = 0, row = 4, sticky = SE)
        
        global BUTTON_SAVE
        BUTTON_SAVE = Button(SUB_FRAME4, text = "UPDATE ENTRY", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH + 3, cursor = "arrow", state = DISABLED, command = self.saveSOAEntries)
        BUTTON_SAVE.grid(column = 0, row = 0, padx = TOP_PADX)
        
        global BUTTON_SUBMIT
        BUTTON_SUBMIT = Button(SUB_FRAME4, text = "SUBMIT", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "hand2", state = NORMAL, command = self.saveSOA)
        BUTTON_SUBMIT.grid(column = 1, row = 0, padx = TOP_PADX)

        global BUTTON_APPROVE
        BUTTON_APPROVE = Button(SUB_FRAME4, text = "APPROVE", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG_GRN, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "arrow", state = DISABLED, command = self.approveSOA)
        BUTTON_APPROVE.grid(column = 2, row = 0, padx = TOP_PADX)

        global BUTTON_VOID
        BUTTON_VOID = Button(SUB_FRAME4, text = "VOID", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG_RED, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "arrow", state = DISABLED, command = self.voidSOA)
        BUTTON_VOID.grid(column = 3, row = 0, padx = TOP_PADX)

        global BUTTON_PRINT
        BUTTON_PRINT = Button(SUB_FRAME4, text = "PRINT", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "arrow", state = DISABLED, command = self.printSOA)
        BUTTON_PRINT.grid(column = 4, row = 0, padx = TOP_PADX)

        BUTTON_CLOSE = Button(SUB_FRAME4, text = "CLOSE", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "hand2", command = lambda: self.closeTopLevel(TOP_RECEIVABLES))
        BUTTON_CLOSE.grid(column = 5, row = 0, padx = TOP_PADX)

        if self.returnUserName(USER, 3) != "ASD":
            BUTTON_SAVE.grid_remove()
            for i in entryFrame:
                for x in i.winfo_children():
                    x.config(state = DISABLED)
            
    def showReceivableLines(self, frame, rown):
        FRAME_LINE = Frame(frame)
        FRAME_LINE.grid(column = 0, row = rown, sticky = W)
        soaFrame.append(FRAME_LINE)

        global TEXTVAR_TRANS
        TEXTVAR_TRANS = StringVar()
        COMBO_TRANS = tk.Combobox(FRAME_LINE, textvariable = TEXTVAR_TRANS, values = self.listTransactions(), font = APP_FONT, width = 27, state = "readonly")
        COMBO_TRANS.grid(column = 0, row = 0, sticky = W)
        COMBO_TRANS.bind("<<ComboboxSelected>>", lambda e: self.updateSOALine(soaGross[rown], TEXTVAR_ENTITYTYPE, soaTax[rown], rown))
        soaTrans.append(TEXTVAR_TRANS)

        global TEXTVAR_GROSS
        TEXTVAR_GROSS = StringVar()
        ENTRY_GROSS = Entry(FRAME_LINE, textvariable = TEXTVAR_GROSS, font = APP_FONT, width = 15, justify = RIGHT)
        ENTRY_GROSS.grid(column = 1, row = 0, sticky = W, ipadx = 1)
        ENTRY_GROSS.bind("<FocusOut>", lambda e: self.updateSOALine(soaGross[rown], TEXTVAR_ENTITYTYPE, soaTax[rown], rown))
        ENTRY_GROSS.bind("<FocusIn>", lambda e: soaGross[rown].set(soaGross[rown].get().replace(",", "")))
        soaGross.append(TEXTVAR_GROSS)
        
        global TEXTVAR_TAX, COMBO_TAX
        TEXTVAR_TAX = StringVar()
        COMBO_TAX = tk.Combobox(FRAME_LINE, values = ["NV-00", "NV-02", "WV-02"], textvariable = TEXTVAR_TAX, font = APP_FONT, width = 7, state = "readonly")
        COMBO_TAX.grid(column = 2, row = 0, sticky = W, ipadx = 2)
        soaTax.append(TEXTVAR_TAX)
        soaComboTax.append(COMBO_TAX)
        COMBO_TAX.bind("<<ComboboxSelected>>", lambda e: self.computeSOALine(soaGross[rown], TEXTVAR_ENTITYTYPE, soaTax[rown], rown))
        
        global TEXTVAR_CATEGORY
        TEXTVAR_CATEGORY = StringVar()
        COMBO_CATEGORY = tk.Combobox(FRAME_LINE, values = ["Administrative", "Maintenance"], textvariable = TEXTVAR_CATEGORY, font = APP_FONT, width = 22, state = "readonly")
        COMBO_CATEGORY.grid(column = 3, row = 0, sticky = W, ipadx = 2)
        soaCategory.append(TEXTVAR_CATEGORY)
        COMBO_CATEGORY.bind("<<ComboboxSelected>>", lambda e: self.computeSOALine(soaGross[rown], TEXTVAR_ENTITYTYPE, soaTax[rown], rown))

        global TEXTVAR_VAT
        TEXTVAR_VAT = StringVar()
        ENTRY_VAT = Entry(FRAME_LINE, textvariable = TEXTVAR_VAT, font = APP_FONT, width = 12, justify = RIGHT, state = "readonly")
        ENTRY_VAT.grid(column = 4, row = 0, sticky = W, ipadx = 1)
        soaVAT.append(TEXTVAR_VAT)

        global TEXTVAR_EWT
        TEXTVAR_EWT = StringVar()
        ENTRY_EWT = Entry(FRAME_LINE, textvariable = TEXTVAR_EWT, font = APP_FONT, width = 12, justify = RIGHT, state = "readonly")
        ENTRY_EWT.grid(column = 5, row = 0, sticky = W, ipadx = 1)
        soaEWT.append(TEXTVAR_EWT)

        global TEXTVAR_CVAT
        TEXTVAR_CVAT = StringVar()
        ENTRY_CVAT = Entry(FRAME_LINE, textvariable = TEXTVAR_CVAT, font = APP_FONT, width = 12, justify = RIGHT, state = "readonly")
        ENTRY_CVAT.grid(column = 6, row = 0, sticky = W, ipadx = 1)
        soaCVAT.append(TEXTVAR_CVAT)

        global TEXTVAR_NET
        TEXTVAR_NET = StringVar()
        ENTRY_NET = Entry(FRAME_LINE, textvariable = TEXTVAR_NET, font = APP_FONT, width = 12, justify = RIGHT, state = "readonly")
        ENTRY_NET.grid(column = 7, row = 0, sticky = W, ipadx = 1)
        soaNET.append(TEXTVAR_NET)

        global TEXTVAR_REMARKS
        TEXTVAR_REMARKS = StringVar()
        ENTRY_REMARKS = Entry(FRAME_LINE, textvariable = TEXTVAR_REMARKS, font = APP_FONT, width = 30)
        ENTRY_REMARKS.grid(column = 8, row = 0, sticky = W, ipadx = 1)
        soaRemarks.append(TEXTVAR_REMARKS)

        global BUTTON_CLEAR
        BUTTON_CLEAR = Button(FRAME_LINE, text = "[X]", font = BUTTON_FONT2, bg = BUTTON_BG, fg = BUTTON_FG_RED, activebackground = CLICK_BG, cursor = "hand2", command = lambda: self.clearSOALine(soaGross[rown], TEXTVAR_ENTITYTYPE, soaTax[rown], rown))
        BUTTON_CLEAR.grid(column = 9, row = 0, sticky = W)
        soaClear.append(BUTTON_CLEAR)

    def clearSOALine(self, gross, entity, tax, i):
        soaTrans[i].set("")
        soaGross[i].set("")
        soaTax[i].set("")
        soaCategory[i].set("")
        soaRemarks[i].set("")
        self.updateSOALine(gross, entity, tax, i)
    
    def saveSOA(self):
        if ENTRY_SOANUMBER["state"] == "readonly":
            switch = "update"
            message = "Update SOA"
        else:
            switch = "insert"
            message = "Add SOA"
        wrong = []
        for i in required:
            try:
                if i.get() == "":
                    wrong.append(i)
            except:
                try:
                    if i.get_date() == "":
                        wrong.append(i)
                except:
                    if i.get("1.0", END) == "":
                        wrong.append(i)
        
        if self.returnFloatAmount(TEXTVAR_VARIANCE.get()) != 0:
            if self.returnUserName(USER, 3) == "ASD":
                wrong.append("Check entry variance!")
        if self.returnPeriodStatus(CALENDAR_GLDATE.get_date()) == "locked":
            wrong.append("GL date is locked")
                        
        if len(wrong) > 0:
            messagebox.showerror(message, "Fill up all required fields!" + str(wrong))
            TOP_RECEIVABLES.focus()
        else:
            if self.returnFloatAmount(TEXTVAR_TNET.get()) == 0:
                messagebox.showerror(message, "Check the amounts!")
                TOP_RECEIVABLES.focus()
            else:
                ask = messagebox.askyesno(message, "Are you sure?")
                if ask:
                    insert = """INSERT INTO tblreceivables (soaDate, soaNumber, payDate, coveredPeriod, particulars,
                                                            reference, clientCode, billType, transaction, amount,
                                                            remarks, taxType, category, encoder, encoded,
                                                            modifier, modified, glDate, void, isApproved) VALUES 
                                                            (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
                    select = "SELECT encoder, encoded FROM tblreceivables WHERE soaNumber = %s"
                    delete = "DELETE FROM tblreceivables WHERE soaNumber = %s"
                    
                    insertbook = """INSERT INTO tblsalesbook (glDate, clientCode, soaNumber, chartCode, amount,
                                                                side, remarks, isPosted, void, poster,
                                                                posted, modifier, modified, source) VALUES 
                                                                (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
                    deletebook = "DELETE FROM tblsalesbook WHERE soaNumber = %s"
                    
                    validlines = []
                    validlinesbook = []
                    if switch == "update":
                        db.commit()
                        cursor.execute(select, [int(TEXTVAR_SOANUMBER.get())])
                        result = cursor.fetchall()
                        cursor.execute(delete, [int(TEXTVAR_SOANUMBER.get())])
                        cursor.execute(deletebook, [int(TEXTVAR_SOANUMBER.get())])
                        db.commit()
                        
                    for i in range(len(soaGross)):
                        if self.returnFloatAmount(soaGross[i].get()) != 0:
                            if switch == "insert":
                                validlines.append([CALENDAR_SOADATE.get_date(), TEXTVAR_SOANUMBER.get(), CALENDAR_PAYDATE.get_date(), self.returnCoveredPeriod(str(CALENDAR_PAYDATE.get_date())), ENTRY_PARTICULARS.get("1.0", END).replace("\n", ""),
                                                    TEXTVAR_REFERENCE.get(), TEXTVAR_CLIENTCODE.get(), TEXTVAR_BILLTYPE.get(), soaTrans[i].get(), self.returnFloatAmount(soaGross[i].get()),
                                                    soaRemarks[i].get(), soaTax[i].get(), soaCategory[i].get(), USER, datetime.datetime.now(),
                                                    USER, datetime.datetime.now(),CALENDAR_GLDATE.get_date(), "No", "No"])
                                
                                
                            elif switch == "update":
                                validlines.append([CALENDAR_SOADATE.get_date(), TEXTVAR_SOANUMBER.get(), CALENDAR_PAYDATE.get_date(), self.returnCoveredPeriod(str(CALENDAR_PAYDATE.get_date())), ENTRY_PARTICULARS.get("1.0", END).replace("\n", ""),
                                                    TEXTVAR_REFERENCE.get(), TEXTVAR_CLIENTCODE.get(), TEXTVAR_BILLTYPE.get(), soaTrans[i].get(), self.returnFloatAmount(soaGross[i].get()),
                                                    soaRemarks[i].get(), soaTax[i].get(), soaCategory[i].get(), result[0][0], result[0][1],
                                                    USER, datetime.datetime.now(),CALENDAR_GLDATE.get_date(), "No", "No"])
                    for i in range(len(entryAmount)):
                        if self.returnFloatAmount(entryAmount[i].get()) != 0:
                            validlinesbook.append([CALENDAR_GLDATE.get_date(), TEXTVAR_CLIENTCODE.get(), TEXTVAR_SOANUMBER.get(), entryCode[i].get(), self.returnFloatAmount(entryAmount[i].get()),
                                                    entryDrCr[i].get(), entryRemarks[i].get(), "No", "No", USER,
                                                    datetime.datetime.now(), USER, datetime.datetime.now(), "SB"])
                    
                    if len(validlines) > 0 and len(validlinesbook) > 0:
                        for i in validlines:
                            cursor.execute(insert, i)
                        for i in validlinesbook:
                            cursor.execute(insertbook, i)

                        db.commit()
                        ENTRY_SOANUMBER.config(state = "readonly")
                        messagebox.showinfo(message, "SOA successfully submitted!")
                        BUTTON_SUBMIT.config(state = DISABLED, cursor = "arrow")
                        BUTTON_APPROVE.config(state = NORMAL, cursor = "hand2")
                        BUTTON_PRINT.config(state = NORMAL, cursor = "hand2")
                        TOP_RECEIVABLES.focus()
                        
                        if self.returnAccess(USER, 13) == 0:
                            BUTTON_APPROVE.config(state = DISABLED, cursor = "arrow")
                    else:
                        messagebox.showerror(message, "Length of valid lines is zero!")
                        TOP_RECEIVABLES.focus()

    def voidSOA(self):
        insertbook = """INSERT INTO tblgeneraljournal (
            gjNumber, docDate, glDate, particulars, reference,
            source, chartcode, amount, side, remarks,
            isPosted, isVoid, encoder, encoded) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
        
        voidsoa = "UPDATE tblreceivables SET void = %s WHERE soaNumber = %s"
        
        validbook = []
        for i in range(len(entryFrame)):
            if entryCode[i].get() != "" and self.returnFloatAmount(entryAmount[i].get()) > 0:
                if entryDrCr[i].get() == "Debit":
                    side = "Credit"
                else:
                    side = "Debit"
                validbook.append([
                    int(self.generateGJNumber()), datetime.date.today(), self.returnLastDayOfMonth(str(datetime.date.today())), "REVERSAL OF SOA#" + TEXTVAR_SOANUMBER.get() + " " + ENTRY_PARTICULARS.get("1.0", END).replace("\n",""), TEXTVAR_SOANUMBER.get(),
                    "SOA", int(entryCode[i].get()), self.returnFloatAmount(entryAmount[i].get()), side, entryRemarks[i].get(),
                    "No", "No", USER, datetime.datetime.now()
                ])
        
        ask = messagebox.askyesno("VOID SOA", "Are you sure?")
        if ask:
            cursor.execute(voidsoa, ["Yes", int(TEXTVAR_SOANUMBER.get())])
            db.commit()
            
            if len(validbook) != 0:
                for i in validbook:
                    cursor.execute(insertbook, i)
                db.commit()
            messagebox.showinfo("VOID soa", "SOA has been voided!")
            TOP_RECEIVABLES.focus()
            self.disableSOAWidgets()
            BUTTON_SUBMIT.config(state = DISABLED, cursor = "arrow")
            BUTTON_APPROVE.config(state = DISABLED, cursor = "arrow")
            BUTTON_VOID.config(state = DISABLED, cursor = "arrow")
            BUTTON_PRINT.config(state = NORMAL, cursor = "hand2")

    def approveSOA(self):
        wrong = []
        for i in required:
            try:
                if i.get() == "":
                    wrong.append(i)
            except:
                try:
                    if i.get_date() == "":
                        wrong.append(i)
                except:
                    if i.get("1.0", END) == "":
                        wrong.append(i)
        if self.returnFloatAmount(TEXTVAR_VARIANCE.get()) != 0:
            wrong.append("Check entry variance!")
                        
        if len(wrong) > 0:
            messagebox.showerror("Approve SOA", "Fill up all required fields!" + str(wrong))
            TOP_RECEIVABLES.focus()
        else:
            if self.returnFloatAmount(TEXTVAR_TNET.get()) == 0:
                messagebox.showerror("Approve SOA", "Check the amounts!")
                TOP_RECEIVABLES.focus()
            else:
                ask = messagebox.askyesno("Approve SOA", "Are you sure?")
                if ask:
                    insert = "UPDATE tblreceivables SET isApproved = 'Yes', approver = %s, approved = %s WHERE soaNumber = %s"
                    cursor.execute(insert, [USER, datetime.datetime.now(), TEXTVAR_SOANUMBER.get()])
                    db.commit()
                    self.disableSOAWidgets()
                    BUTTON_SUBMIT.config(state = DISABLED, cursor = "arrow")
                    BUTTON_APPROVE.config(state = DISABLED, cursor = "arrow")
                    BUTTON_VOID.config(state = NORMAL, cursor = "hand2")
                    
                    messagebox.showinfo("Approve SOA", "SOA has been approved!")
                    
                    if self.returnAccess(USER, 14) == 0:
                        BUTTON_VOID.config(state = DISABLED, cursor = "arrow")

    def showImportSOA(self):
        global TOP_IMPORT
        TOP_IMPORT = Toplevel()
        TOP_IMPORT.title("Import - SOA")
        TOP_IMPORT.iconbitmap(PATH_ICON + "icon.ico")
        TOP_IMPORT.geometry("350x325+550+100")
        TOP_IMPORT.resizable(height = False, width = False)
        TOP_IMPORT.grab_set()
        
        LABEL_FOLDER = Label(TOP_IMPORT, text = "Folder", font = APP_FONT)
        LABEL_FOLDER.grid(column = 0, row = 0, pady = TOP_PADY)
        
        global TEXTVAR_FOLDER
        TEXTVAR_FOLDER = StringVar()
        ENTRY_FOLDER = Entry(TOP_IMPORT, textvariable = TEXTVAR_FOLDER, font = APP_FONT, width = 40, state = "readonly")
        ENTRY_FOLDER.grid(column = 1, row = 0, pady = TOP_PADY)
        
        BUTTON_FOLDER = Button(TOP_IMPORT, text = "...", font = APP_FONT, command = lambda: self.getFolderAddress(TEXTVAR_FOLDER))
        BUTTON_FOLDER.grid(column = 1, row = 0, sticky = E, pady = TOP_PADY)
        
        BUTTON_IMPORT = Button(TOP_IMPORT, text = "Import", font = APP_FONT, command = self.importSOASummary)
        BUTTON_IMPORT.grid(column = 1, row = 1, sticky = E, pady = TOP_PADY)

    def importSOASummary(self):
        files = listdir(TEXTVAR_FOLDER.get())
        inserter = """INSERT INTO tblreceivables (soaDate, soaNumber, payDate, coveredPeriod, particulars,
                                                reference, clientCode, billType, chartCode, transaction, 
                                                amount, taxType, category, void, isApproved,
                                                encoder, encoded, glDate) VALUES 
                                                (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
        insertbook = """INSERT INTO tblsalesbook (glDate, clientCode, soaNumber, chartCode, amount,
                                                    side, remarks, isPosted, void, poster,
                                                    posted, modifier, modified, source) VALUES 
                                                    (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
        soaItems, tempItems, duplicates = [], [], []
        excelerrors = []
        for x in files:
            wb = load_workbook(TEXTVAR_FOLDER.get() + "/" + x)
            st = wb.active
            for i in range(st.max_row-1):
                try:
                    if st["A" + str(i+9)].value != None and str(st["D" + str(i+9)].value).replace("-","")[0] == "4":
                        if self.returnSOADuplicate(st["H" + str(i+9)].value) == False:
                            clientCode = st["A" + str(i+9)].value
                            chartCode = str(st["D" + str(i+9)].value).replace("-","")[5:]
                            if chartCode[0] == "1" or chartCode[0] == "3":
                                category = "Administrative"
                            else:
                                category = "Maintenance"
                            amount = float(st["E" + str(i+9)].value)*self.returnVATMultiplier(clientCode)
                            payDate = str(st["G" + str(i+9)].value)
                            soaNumber = st["H" + str(i+9)].value
                            soaDate = st["K" + str(i+9)].value
                            soaItems.append([soaDate, soaNumber, payDate, self.returnCoveredPeriod(payDate), "Billing for services rendered " + self.returnCoveredPeriod(payDate),
                                            "EIS", clientCode, "Auto", chartCode, "Standard Services",
                                            amount, self.returnClientTaxType(clientCode), category, "No", "No", 
                                            USER, datetime.datetime.now(), self.returnLastDayOfMonth(payDate)])
                            tempItems.append([self.returnLastDayOfMonth(payDate), clientCode, soaNumber, amount, category, self.returnClientTaxType2(clientCode)])
                        else:
                            duplicates.append([st["H" + str(i+9)].value, x])
                except Exception as e:
                    excelerrors.append(e)
                    
        if len(excelerrors) != 0:
            messagebox.showerror("Excel Error", excelerrors)
        else:
            bookItems, tempGross, tempVAT, tempIncomeAdmin, tempIncomeMaint = [], [], [], [], []
            for i in range(len(tempItems)):
                try:
                    if tempItems[i][2] != soaNum:
                        bookItems.append([tempItems[i-1][0], tempItems[i-1][1], tempItems[i-1][2], "115010", sum(tempGross),"Debit", "", "No", "No", USER, datetime.datetime.now(), USER, datetime.datetime.now(), "SB"])
                        if sum(tempVAT) > 0:
                            bookItems.append([tempItems[i-1][0], tempItems[i-1][1], tempItems[i-1][2], "213140", sum(tempVAT),"Credit", "", "No", "No", USER, datetime.datetime.now(), USER, datetime.datetime.now(), "SB"])
                        if sum(tempIncomeAdmin) > 0:
                            bookItems.append([tempItems[i-1][0], tempItems[i-1][1], tempItems[i-1][2], "511010", sum(tempIncomeAdmin),"Credit", "", "No", "No", USER, datetime.datetime.now(), USER, datetime.datetime.now(), "SB"])
                        if sum(tempIncomeMaint) > 0:
                            bookItems.append([tempItems[i-1][0], tempItems[i-1][1], tempItems[i-1][2], "512010", sum(tempIncomeMaint),"Credit", "", "No", "No", USER, datetime.datetime.now(), USER, datetime.datetime.now(), "SB"])
                        tempGross, tempVAT, tempIncomeAdmin, tempIncomeMaint = [], [], [], []
                        soaNum = tempItems[i][2]
                        tempGross.append(tempItems[i][3])
                        if tempItems[i][5] == "VAT":
                            tempVAT.append((tempItems[i][3])/1.12*.12)
                            if tempItems[i][4] == "Administrative":
                                tempIncomeAdmin.append(tempItems[i][3]-(tempItems[i][3])/1.12*.12)
                            else:
                                tempIncomeMaint.append(tempItems[i][3]-(tempItems[i][3])/1.12*.12)
                        else:
                            tempVAT.append(0)
                            if tempItems[i][4] == "Administrative":
                                tempIncomeAdmin.append(tempItems[i][3])
                            else:
                                tempIncomeMaint.append(tempItems[i][3])
                    else:
                        tempGross.append(tempItems[i][3])
                        if tempItems[i][5] == "VAT":
                            tempVAT.append((tempItems[i][3])/1.12*.12)
                            if tempItems[i][4] == "Administrative":
                                tempIncomeAdmin.append(tempItems[i][3]-(tempItems[i][3])/1.12*.12)
                            else:
                                tempIncomeMaint.append(tempItems[i][3]-(tempItems[i][3])/1.12*.12)
                        else:
                            tempVAT.append(0)
                            if tempItems[i][4] == "Administrative":
                                tempIncomeAdmin.append(tempItems[i][3])
                            else:
                                tempIncomeMaint.append(tempItems[i][3])
                        if i+1 == len(tempItems):
                            bookItems.append([tempItems[i-1][0], tempItems[i-1][1], tempItems[i-1][2], "115010", sum(tempGross),"Debit", "", "No", "No", USER, datetime.datetime.now(), USER, datetime.datetime.now(), "SB"])
                            if sum(tempVAT) > 0:
                                bookItems.append([tempItems[i-1][0], tempItems[i-1][1], tempItems[i-1][2], "213140", sum(tempVAT),"Credit", "", "No", "No", USER, datetime.datetime.now(), USER, datetime.datetime.now(), "SB"])
                            if sum(tempIncomeAdmin) > 0:
                                bookItems.append([tempItems[i-1][0], tempItems[i-1][1], tempItems[i-1][2], "511010", sum(tempIncomeAdmin),"Credit", "", "No", "No", USER, datetime.datetime.now(), USER, datetime.datetime.now(), "SB"])
                            if sum(tempIncomeMaint) > 0:
                                bookItems.append([tempItems[i-1][0], tempItems[i-1][1], tempItems[i-1][2], "512010", sum(tempIncomeMaint),"Credit", "", "No", "No", USER, datetime.datetime.now(), USER, datetime.datetime.now(), "SB"])
                            
                except:
                    soaNum = tempItems[i][2]
                    tempGross.append(tempItems[i][3])
                    if tempItems[i][5] == "VAT":
                        tempVAT.append((tempItems[i][3])/1.12*.12)
                        if tempItems[i][4] == "Administrative":
                            tempIncomeAdmin.append(tempItems[i][3]-(tempItems[i][3])/1.12*.12)
                        else:
                            tempIncomeMaint.append(tempItems[i][3]-(tempItems[i][3])/1.12*.12)
                    else:
                        tempVAT.append(0)
                        if tempItems[i][4] == "Administrative":
                            tempIncomeAdmin.append(tempItems[i][3])
                        else:
                            tempIncomeMaint.append(tempItems[i][3])

            if len(soaItems) > 0:
                for i in soaItems:
                    cursor.execute(inserter, i)
            if len(bookItems) > 0:
                for i in bookItems:
                    cursor.execute(insertbook, i)
            ask = messagebox.askyesno("Import", "Are you sure?")
            if ask:
                db.commit()
                if len(duplicates) > 0:
                    messagebox.showinfo("Import", "SOA import completed! \n Duplicate SOAs detected" + str(duplicates))
                else:
                    messagebox.showinfo("Import", "SOA import completed! No duplicates detected")
                TOP_IMPORT.grab_release()
                TOP_IMPORT.destroy()
                self.showReceivables()

    def editSOA(self, *args):
        self.copySelection(TREE_RECEIVABLES)
        self.showAddEditReceivable()
        TOP_RECEIVABLES.title("Edit - SOA")
        
        find = """SELECT soaDate, soaNumber, payDate, glDate, coveredPeriod,
                        particulars, reference, clientCode, billType, transaction,
                        chartCode, amount, remarks, taxType, category,
                        isApproved, encoder
                        FROM tblreceivables WHERE soaNumber = %s"""
                        
        findinbook = "SELECT chartCode, amount, side, remarks FROM tblsalesbook WHERE soaNumber = %s"
        db.commit()
        cursor.execute(find, [int(content[0])])
        result = cursor.fetchall()
        cursor.execute(findinbook, [int(content[0])])
        resultbook = cursor.fetchall()
        if result:
            CALENDAR_SOADATE.set_date(result[0][0])
            TEXTVAR_SOANUMBER.set(str(result[0][1]).zfill(8))
            ENTRY_SOANUMBER.config(state = "readonly")
            ENTRY_SOANUMBER.unbind("<FocusOut>")
            CALENDAR_PAYDATE.set_date(result[0][2])
            CALENDAR_GLDATE.set_date(result[0][3])
            ENTRY_PARTICULARS.insert(1.0, result[0][5])
            TEXTVAR_REFERENCE.set(result[0][6])
            TEXTVAR_CLIENTCODE.set(result[0][7])
            self.populateSOAFields(result[0][7])
            TEXTVAR_BILLTYPE.set(result[0][8])
            TEXTVAR_ENTITYTYPE.set(self.returnClientEntityType(result[0][7])[0])
            for i in range(len(result)):
                soaTrans[i].set(result[i][9])
                soaGross[i].set(self.validateAmount2(result[i][11]))
                soaTax[i].set(result[i][13])
                soaCategory[i].set(result[i][14])
                soaRemarks[i].set(result[i][12])
                self.computeSOALine(soaGross[i], TEXTVAR_ENTITYTYPE, soaTax[i], i)
            self.updateSOATotals()
        
            self.clearSOAEntryLines()
            for i in range(len(resultbook)):
                if resultbook[i][1] != 0:
                    entryCode[i].set(resultbook[i][0])
                    self.populateChartFields(entryCode[i], entryTitle[i])
                    entryAmount[i].set(self.validateAmount2(resultbook[i][1]))
                    entryDrCr[i].set(resultbook[i][2])
                    entryRemarks[i].set(resultbook[i][3])
            
            if result[0][15] == "Yes": #isApproved
                self.disableSOAWidgets()
                BUTTON_SUBMIT.config(text = "UPDATE", state = DISABLED, cursor = "arrow")
                BUTTON_APPROVE.config(state = DISABLED, cursor = "arrow")
                if self.returnIfSOAVoid(TEXTVAR_SOANUMBER) == True:
                    BUTTON_VOID.config(state = DISABLED, cursor = "arrow")
                else:
                    BUTTON_VOID.config(state = NORMAL, cursor = "hand2")
            else:
                BUTTON_SUBMIT.config(text = "UPDATE", state = NORMAL, cursor = "hand2")
                BUTTON_APPROVE.config(state = NORMAL, cursor = "hand2")
            BUTTON_PRINT.config(state = NORMAL, cursor = "hand2")
            
            if self.returnUserName(USER, 3) == "ASD":
                BUTTON_SAVE.config(state = NORMAL, cursor = "hand2")
            if self.returnAccess(USER, 13) == 0 and self.returnAccess(USER, 14) == 0:
                BUTTON_APPROVE.config(state = DISABLED, cursor = "arrow")
            if result[0][16] != USER:
                BUTTON_SUBMIT.config(state = DISABLED, cursor = "arrow")

    def printSOA(self):
        wb = load_workbook(PATH_TEMPLATE + "SOA.xlsx")
        sheet = wb.active
        sheet["AG12"] = f"No.         {TEXTVAR_SOANUMBER.get()}"
        sheet["L14"] = TEXTVAR_CLIENTNAME.get()
        sheet["AN15"] = CALENDAR_SOADATE.get()
        sheet["G29"] = f"SA NO: I{TEXTVAR_SOANUMBER.get()}"
        sheet["H33"] = ENTRY_PARTICULARS.get("1.0", END).replace("\n","")
        sheet["AJ35"] = TEXTVAR_TGROSS.get()
        sheet["AK45"] = TEXTVAR_TNET.get()
        sheet["H38"] = TEXTVAR_ENTITYTYPE.get()
        sheet["Y40"] = TEXTVAR_TEWT.get()
        sheet["Y42"] = TEXTVAR_TCVAT.get()
        sheet["F54"] = self.returnUserName(USER, 1).upper() + " " + self.returnUserName(USER, 2).upper()
        wb.save(PATH_SAVE + "SOA.xlsx")
        startfile(PATH_SAVE + "SOA.xlsx", "open")

    def saveSOAEntries(self):
        insertbook = """INSERT INTO tblsalesbook (glDate, clientCode, soaNumber, chartCode, amount,
                                                side, remarks, isPosted, void, poster,
                                                posted, modifier, modified, source) VALUES 
                                                (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
        deletebook = "DELETE FROM tblsalesbook WHERE soaNumber = %s"
        ask = messagebox.askyesno("Receivables", "Are you sure?")
        if ask:
            cursor.execute(deletebook, [int(TEXTVAR_SOANUMBER.get())])
            db.commit()
            validlinesbook = []
            for i in range(len(entryAmount)):
                if self.returnFloatAmount(entryAmount[i].get()) != 0:
                    validlinesbook.append([CALENDAR_GLDATE.get_date(), TEXTVAR_CLIENTCODE.get(), TEXTVAR_SOANUMBER.get(), entryCode[i].get(), self.returnFloatAmount(entryAmount[i].get()),
                                            entryDrCr[i].get(), entryRemarks[i].get(), "No", "No", USER,
                                            datetime.datetime.now(), USER, datetime.datetime.now(), "SB"])
            if len(validlinesbook) > 0:
                for i in validlinesbook:
                    cursor.execute(insertbook, i)
                db.commit()
                messagebox.showinfo("Receivables", "Entries has been updated!")

    def updateSOATotals(self):
        tgross, tvat, tewt, tcvat, tnet = [], [], [], [], []
        for i in range(len(soaGross)):
            try:
                tgross.append(self.returnFloatAmount(soaGross[i].get()))
                tvat.append(self.returnFloatAmount(soaVAT[i].get()))
                tewt.append(self.returnFloatAmount(soaEWT[i].get()))
                tcvat.append(self.returnFloatAmount(soaCVAT[i].get()))
                tnet.append(self.returnFloatAmount(soaNET[i].get()))
            except:
                pass
        TEXTVAR_TGROSS.set(self.validateAmount2(sum(tgross)))
        TEXTVAR_TVAT.set(self.validateAmount2(sum(tvat)))
        TEXTVAR_TEWT.set(self.validateAmount2(sum(tewt)))
        TEXTVAR_TCVAT.set(self.validateAmount2(sum(tcvat)))
        TEXTVAR_TNET.set(self.validateAmount2(sum(tnet)))
        
        self.updateSOAEntries()
        self.updateSOAEntriesTotals()
        
    def computeSOALine(self, gross, entity, tax, i):
        if entity.get() == "Government":
            if tax.get()[:2] == "WV":
                soaCVAT[i].set(self.validateAmount2((self.returnFloatAmount(gross.get())/1.12)*.05))
            else:
                soaCVAT[i].set(self.validateAmount2(0))
        else:
            soaCVAT[i].set(self.validateAmount2(0))
            
        if tax.get() == "NV-00":
            soaVAT[i].set(self.validateAmount2(0))
            soaEWT[i].set(self.validateAmount2(0))
            soaNET[i].set(self.validateAmount2(self.returnFloatAmount(gross.get())-self.returnFloatAmount(soaCVAT[i].get())))
        elif tax.get() == "NV-02":
            soaVAT[i].set(self.validateAmount2(0))
            soaEWT[i].set(self.validateAmount2(self.returnFloatAmount(gross.get())*.02))
            soaNET[i].set(self.validateAmount2(self.returnFloatAmount(gross.get())-self.returnFloatAmount(soaEWT[i].get())-self.returnFloatAmount(soaCVAT[i].get())))
        elif tax.get() == "WV-02":
            soaVAT[i].set(self.validateAmount2((self.returnFloatAmount(gross.get())/1.12)*.12))
            soaEWT[i].set(self.validateAmount2((self.returnFloatAmount(gross.get())/1.12)*.02))
            soaNET[i].set(self.validateAmount2(self.returnFloatAmount(gross.get())-self.returnFloatAmount(soaEWT[i].get())-self.returnFloatAmount(soaCVAT[i].get())))
        else:
            soaVAT[i].set(self.validateAmount2(0))
            soaEWT[i].set(self.validateAmount2(0))
            soaNET[i].set(self.validateAmount2(0))
        self.updateSOATotals()

    def updateSOALine(self, gross, entity, tax, i):
        self.validateAmount(gross)
        self.computeSOALine(gross, entity, tax, i)

    def returnIfSOAVoid(self, var):
        find = "SELECT void FROM tblreceivables WHERE soaNumber = %s LIMIT 1"
        try:
            cursor.execute(find, [int(var.get())])
        except:
            cursor.execute(find, [var])
        result = cursor.fetchone()
        if result[0] == "Yes":
            return True
        else:
            return False

    def returnIfSOAApproved(self, var):
        find = "SELECT isApproved FROM tblreceivables WHERE soaNumber = %s LIMIT 1"
        cursor.execute(find, [int(var.get())])
        result = cursor.fetchone()
        if result[0] == "Yes":
            return True
        else:
            return False
        
    def disableSOAWidgets(self):
        ENTRY_CLIENTCODE.config(state = DISABLED)
        BUTTON_CLIENT.config(state = DISABLED, cursor = "arrow")
        ENTRY_PARTICULARS.config(state = DISABLED)
        ENTRY_REFERENCE.config(state = DISABLED)
        CALENDAR_SOADATE.config(state = DISABLED)
        CALENDAR_GLDATE.config(state = DISABLED)
        CALENDAR_PAYDATE.config(state = DISABLED)
        for i in soaFrame:
            for x in i.winfo_children():
                x.config(state = DISABLED)
        for i in entryFrame:
            for x in i.winfo_children():
                x.config(state = DISABLED)
        if self.returnPostingStatus(int(TEXTVAR_SOANUMBER.get()), 'SB') == "Unposted":
            if self.returnUserName(USER, 3) == "ASD" and self.returnUserName(USER, 4) == "administrator":
                CALENDAR_GLDATE.config(state = NORMAL)
        else:
            for i in entryFrame:
                for x in i.winfo_children():
                    x.config(state = DISABLED)
            BUTTON_SAVE.config(state = DISABLED, cursor = "arrow")

    def returnVATMultiplier(self, var):
        find = "SELECT taxType FROM tblclients WHERE clientCode = %s LIMIT 1"
        cursor.execute(find, [var])
        result = cursor.fetchone()
        if result:
            if result[0] == "VAT":
                return 1.12
            else:
                return 1
        else:
            return 1

    def returnTotalSOAAmount(self, var):
        db.commit()
        find = "SELECT soaNumber, clientCode, amount, taxType FROM tblreceivables WHERE soaNumber = %s"
        cursor.execute(find, [var])
        result = cursor.fetchall()
        NET = []
        for i in result:
            GROSS = float(i[2])
            if self.returnClientEntityType(i[1])[0] == "Government":
                if i[3][:2] == "WV":
                    CVAT = (GROSS/1.12)*.05
                else:
                    CVAT = 0
            else:
                CVAT = 0
    
            if i[3] == "NV-00":
                EWT = 0
                NET.append(GROSS-CVAT)
            elif i[3] == "NV-02":
                EWT = float(i[2])*.02
                NET.append(GROSS-EWT-CVAT)
            elif i[3] == "WV-02":
                EWT = (GROSS/1.12)*.02
                NET.append(GROSS-EWT-CVAT)
            else:
                EWT = 0
                NET.append(0)
        return format(sum(NET), ",.2f")
    
    def returnClientEntityType(self, var):
        find = "SELECT entityType FROM tblclients WHERE clientCode = %s LIMIT 1"
        cursor.execute(find, [var])
        result = cursor.fetchone()
        if result:
            return result
    
    def populateSOAFields(self, var):
        self.capitalLetters(var)
        try:
            if TOP_COLLECTIONS.winfo_exists() == 0:
                select = "SELECT clientCode, clientName, entityType FROM tblclients WHERE clientCode = %s"
            else:
                select = "SELECT parentCode, clientName, entityType FROM tblclients WHERE clientCode = %s"
        except:
            select = "SELECT clientCode, clientName, entityType FROM tblclients WHERE clientCode = %s"
        db.commit()
        cursor.execute(select, [var])
        result = cursor.fetchall()
        if result:
            TEXTVAR_CLIENTCODE.set(result[0][0])
            TEXTVAR_CLIENTNAME.set(result[0][1])
            TEXTVAR_ENTITYTYPE.set(result[0][2])
            try:
                if TOP_COLLECTIONS.winfo_exists() == 0:
                    for i in soaComboTax:
                        if self.returnClientTaxType2(result[0][0]) == "NONVAT":
                            i.config(values = ["NV-00", "NV-02"])
                        else:
                            i.config(values = ["NV-00", "NV-02", "WV-02"])
                else:
                    self.disableCollectionSOANumberFields(DISABLED)
            except:
                for i in soaComboTax:
                    if self.returnClientTaxType2(result[0][0]) == "NONVAT":
                        i.config(values = ["NV-00", "NV-02"])
                    else:
                        i.config(values = ["NV-00", "NV-02", "WV-02"])
        else:
            try:
                if TOP_COLLECTIONS.winfo_exists() == 1:
                    self.disableCollectionSOANumberFields(DISABLED)
            except:
                pass
            TEXTVAR_CLIENTCODE.set("")
            TEXTVAR_CLIENTNAME.set("")
            TEXTVAR_ENTITYTYPE.set("")
        try:
            TOP_RECEIVABLES.focus()
        except:
            TOP_COLLECTIONS.focus()
    
    def updateSOAEntries(self):
        self.clearSOAEntryLines()
        entryGross, entryVAT, entryIncomeAdmin, entryIncomeMaint, entryCommunication, entryFamily, entryTravel, entryOthers, entryDueFrom, entryFinancialAssistance, entryGratuity, entryHazard, entrySoloParent, entryOvertime = [], [], [], [], [], [], [], [], [], [], [], [], [], []
        for i in range(len(soaGross)):
            if soaTrans[i].get() == "Standard Services":
                try:
                    if self.returnFloatAmount(soaGross[i].get()) != 0 and soaCategory[i].get() != "":
                        entryGross.append(self.returnFloatAmount(soaGross[i].get()))
                        entryVAT.append(self.returnFloatAmount(soaVAT[i].get()))
                    if soaCategory[i].get() == "Administrative":
                        entryIncomeAdmin.append(self.returnFloatAmount(soaGross[i].get())-self.returnFloatAmount(soaVAT[i].get()))
                    elif soaCategory[i].get() == "Maintenance":
                        entryIncomeMaint.append(self.returnFloatAmount(soaGross[i].get())-self.returnFloatAmount(soaVAT[i].get()))
                except Exception as e:
                    print(e)
            if soaTrans[i].get() == "Semi-Annual Incentive (SAI)":
                try:
                    if self.returnFloatAmount(soaGross[i].get()) != 0 and soaCategory[i].get() != "":
                        entryGross.append(self.returnFloatAmount(soaGross[i].get()))
                        entryVAT.append(self.returnFloatAmount(soaVAT[i].get()))
                    if soaCategory[i].get() == "Administrative":
                        entryIncomeAdmin.append(self.returnFloatAmount(soaGross[i].get())-self.returnFloatAmount(soaVAT[i].get()))
                    elif soaCategory[i].get() == "Maintenance":
                        entryIncomeMaint.append(self.returnFloatAmount(soaGross[i].get())-self.returnFloatAmount(soaVAT[i].get()))
                except Exception as e:
                    print(e)
            if soaTrans[i].get().split(" - ")[0] == "Allowance":
                try:
                    if self.returnFloatAmount(soaGross[i].get()) != 0 and soaCategory[i].get() != "":
                        entryGross.append(self.returnFloatAmount(soaGross[i].get()))
                        entryVAT.append(self.returnFloatAmount(soaVAT[i].get()))
                    if soaTrans[i].get().split(" - ")[1] == "Communication":
                        entryCommunication.append(self.returnFloatAmount(soaGross[i].get())-self.returnFloatAmount(soaVAT[i].get()))
                    elif soaTrans[i].get().split(" - ")[1] == "Family":
                        entryFamily.append(self.returnFloatAmount(soaGross[i].get())-self.returnFloatAmount(soaVAT[i].get()))
                    elif soaTrans[i].get().split(" - ")[1] == "Internet":
                        entryCommunication.append(self.returnFloatAmount(soaGross[i].get())-self.returnFloatAmount(soaVAT[i].get()))
                    elif soaTrans[i].get().split(" - ")[1] == "Load":
                        entryCommunication.append(self.returnFloatAmount(soaGross[i].get())-self.returnFloatAmount(soaVAT[i].get()))
                    elif soaTrans[i].get().split(" - ")[1] == "Gas":
                        entryTravel.append(self.returnFloatAmount(soaGross[i].get())-self.returnFloatAmount(soaVAT[i].get()))
                    elif soaTrans[i].get().split(" - ")[1] == "Travel":
                        entryTravel.append(self.returnFloatAmount(soaGross[i].get())-self.returnFloatAmount(soaVAT[i].get()))
                    elif soaTrans[i].get().split(" - ")[1] == "Others":
                        entryOthers.append(self.returnFloatAmount(soaGross[i].get())-self.returnFloatAmount(soaVAT[i].get()))
                except Exception as e:
                    print(e)
            if soaTrans[i].get().split(" - ")[0] == "Reimbursement":
                try:
                    if self.returnFloatAmount(soaGross[i].get()) != 0 and soaCategory[i].get() != "":
                        entryGross.append(self.returnFloatAmount(soaGross[i].get()))
                        entryVAT.append(self.returnFloatAmount(soaVAT[i].get()))
                    if soaTrans[i].get().split(" - ")[1] == "Financial Assistance":
                        entryFinancialAssistance.append(self.returnFloatAmount(soaGross[i].get())-self.returnFloatAmount(soaVAT[i].get()))
                    elif soaTrans[i].get().split(" - ")[1] == "Gratuity Pay":
                        entryGratuity.append(self.returnFloatAmount(soaGross[i].get())-self.returnFloatAmount(soaVAT[i].get()))
                    elif soaTrans[i].get().split(" - ")[1] == "Hazard Pay":
                        entryHazard.append(self.returnFloatAmount(soaGross[i].get())-self.returnFloatAmount(soaVAT[i].get()))
                    elif soaTrans[i].get().split(" - ")[1] == "Solo Parent Leave":
                        entrySoloParent.append(self.returnFloatAmount(soaGross[i].get())-self.returnFloatAmount(soaVAT[i].get()))
                    elif soaTrans[i].get().split(" - ")[1] == "Travel":
                        entryTravel.append(self.returnFloatAmount(soaGross[i].get())-self.returnFloatAmount(soaVAT[i].get()))
                except Exception as e:
                    print(e)
            if soaTrans[i].get() == "Overtime - PPA":
                try:
                    if self.returnFloatAmount(soaGross[i].get()) != 0 and soaCategory[i].get() != "":
                        entryDueFrom.append(self.returnFloatAmount(soaGross[i].get()))
                        entryVAT.append(self.returnFloatAmount(soaVAT[i].get()))
                        entryOvertime.append(self.returnFloatAmount(soaGross[i].get())-self.returnFloatAmount(soaVAT[i].get()))
                except Exception as e:
                    print(e)
            if soaTrans[i].get() == "Travel - Admin Fee":
                try:
                    if self.returnFloatAmount(soaGross[i].get()) != 0 and soaCategory[i].get() != "":
                        entryGross.append(self.returnFloatAmount(soaGross[i].get()))
                        entryVAT.append(self.returnFloatAmount(soaVAT[i].get()))
                    if soaCategory[i].get() == "Administrative":
                        entryIncomeAdmin.append(self.returnFloatAmount(soaGross[i].get())-self.returnFloatAmount(soaVAT[i].get()))
                        entryDueFrom.append(self.returnFloatAmount(soaGross[i].get()))
                    elif soaCategory[i].get() == "Maintenance":
                        entryIncomeMaint.append(self.returnFloatAmount(soaGross[i].get())-self.returnFloatAmount(soaVAT[i].get()))
                        entryDueFrom.append(self.returnFloatAmount(soaGross[i].get()))
                except Exception as e:
                    print(e)
        count = 0
        occurence = 0
        if sum(entryGross) > 0:
            entryCode[count].set("115010")
            self.populateChartFields(entryCode[count], entryTitle[count])
            entryAmount[count].set(self.validateAmount2(sum(entryGross)))
            entryDrCr[count].set("Debit")
            count += 1
        if sum(entryDueFrom) > 0:
            entryCode[count].set("116070")
            self.populateChartFields(entryCode[count], entryTitle[count])
            entryAmount[count].set(self.validateAmount2(sum(entryDueFrom)))
            entryDrCr[count].set("Debit")
            count += 1
            occurence += 1
        if sum(entryVAT) > 0:
            entryCode[count].set("213140")
            self.populateChartFields(entryCode[count], entryTitle[count])
            entryAmount[count].set(self.validateAmount2(sum(entryVAT)))
            entryDrCr[count].set("Credit")
            count += 1
        if sum(entryIncomeAdmin) > 0:
            entryCode[count].set("511010")
            self.populateChartFields(entryCode[count], entryTitle[count])
            entryAmount[count].set(self.validateAmount2(sum(entryIncomeAdmin)))
            entryDrCr[count].set("Credit")
            count += 1
        if sum(entryIncomeMaint) > 0:
            entryCode[count].set("512010")
            self.populateChartFields(entryCode[count], entryTitle[count])
            entryAmount[count].set(self.validateAmount2(sum(entryIncomeMaint)))
            entryDrCr[count].set("Credit")
            count += 1
        if sum(entryCommunication) > 0:
            entryCode[count].set("636040")
            self.populateChartFields(entryCode[count], entryTitle[count])
            entryAmount[count].set(self.validateAmount2(sum(entryCommunication)))
            entryDrCr[count].set("Credit")
            count += 1
        if sum(entryFamily) > 0:
            entryCode[count].set("116070")
            self.populateChartFields(entryCode[count], entryTitle[count])
            entryAmount[count].set(self.validateAmount2(sum(entryFamily)))
            entryDrCr[count].set("Credit")
            count += 1
        if sum(entryTravel) > 0:
            entryCode[count].set("636070")
            self.populateChartFields(entryCode[count], entryTitle[count])
            entryAmount[count].set(self.validateAmount2(sum(entryTravel)))
            entryDrCr[count].set("Credit")
            count += 1
        if sum(entryOthers) > 0:
            entryCode[count].set("116070")
            self.populateChartFields(entryCode[count], entryTitle[count])
            entryAmount[count].set(self.validateAmount2(sum(entryOthers)))
            entryDrCr[count].set("Credit")
            count += 1
        if sum(entryDueFrom) > 0:
            if occurence == 0:
                entryCode[count].set("116070")
                self.populateChartFields(entryCode[count], entryTitle[count])
                entryAmount[count].set(self.validateAmount2(sum(entryDueFrom)))
                entryDrCr[count].set("Credit")
                count += 1
        if sum(entryFinancialAssistance) > 0:
            entryCode[count].set("641090")
            self.populateChartFields(entryCode[count], entryTitle[count])
            entryAmount[count].set(self.validateAmount2(sum(entryFinancialAssistance)))
            entryDrCr[count].set("Credit")
            count += 1
        if sum(entryGratuity) > 0:
            entryCode[count].set("116030")
            self.populateChartFields(entryCode[count], entryTitle[count])
            entryAmount[count].set(self.validateAmount2(sum(entryGratuity)))
            entryDrCr[count].set("Credit")
            count += 1
        if sum(entryHazard) > 0:
            entryCode[count].set("116070")
            self.populateChartFields(entryCode[count], entryTitle[count])
            entryAmount[count].set(self.validateAmount2(sum(entryHazard)))
            entryDrCr[count].set("Credit")
            count += 1
        if sum(entrySoloParent) > 0:
            entryCode[count].set("641090")
            self.populateChartFields(entryCode[count], entryTitle[count])
            entryAmount[count].set(self.validateAmount2(sum(entrySoloParent)))
            entryDrCr[count].set("Credit")
            count += 1
        if sum(entryOvertime) > 0:
            entryCode[count].set("611050")
            self.populateChartFields(entryCode[count], entryTitle[count])
            entryAmount[count].set(self.validateAmount2(sum(entryOvertime)))
            entryDrCr[count].set("Credit")

    def updateSOAEntriesTotals(self, *args):
        debit, credit = [], []
        for i in range(len(entryCode)):
            if entryDrCr[i].get() == "Debit":
                if self.returnFloatAmount(entryAmount[i].get()) > 0:
                    debit.append(self.returnFloatAmount(entryAmount[i].get()))
            else:
                if self.returnFloatAmount(entryAmount[i].get()) > 0:
                    credit.append(self.returnFloatAmount(entryAmount[i].get()))
        TEXTVAR_DEBIT.set(self.validateAmount2(sum(debit)))
        TEXTVAR_CREDIT.set(self.validateAmount2(sum(credit)))
        TEXTVAR_VARIANCE.set(self.validateAmount2(sum(debit)-sum(credit)))

    def showSOAEntryLines(self, frame, rown):
        FRAME_LINE = Frame(frame)
        FRAME_LINE.grid(column = 0, row = rown, sticky = W)
        entryFrame.append(FRAME_LINE)

        global TEXTVAR_CHARTCODE
        TEXTVAR_CHARTCODE = StringVar()
        ENTRY_CHARTCODE = Entry(FRAME_LINE, textvariable = TEXTVAR_CHARTCODE, font = APP_FONT, width = 10)
        ENTRY_CHARTCODE.grid(column = 0, row = 0, sticky = W)
        ENTRY_CHARTCODE.bind("<FocusOut>", lambda e: self.populateChartFields(entryCode[rown], entryTitle[rown]))
        entryCode.append(TEXTVAR_CHARTCODE)
        
        BUTTON_CHARTCODE = Button(FRAME_LINE, text = "...", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, cursor = "hand2", command = lambda: self.showChartSelection(rown, entryCode, entryTitle))
        BUTTON_CHARTCODE.grid(column = 1, row = 0, sticky = E)

        global TEXTVAR_CHARTTITLE
        TEXTVAR_CHARTTITLE = StringVar()
        ENTRY_CHARTTITLE = Entry(FRAME_LINE, textvariable = TEXTVAR_CHARTTITLE, font = APP_FONT, width = 30, state = "readonly")
        ENTRY_CHARTTITLE.grid(column = 2, row = 0, sticky = W, ipadx = 1)
        entryTitle.append(TEXTVAR_CHARTTITLE)
        
        global TEXTVAR_SJAMOUNT
        TEXTVAR_SJAMOUNT = StringVar()
        ENTRY_AMOUNT = Entry(FRAME_LINE, textvariable = TEXTVAR_SJAMOUNT, font = APP_FONT, width = 15, justify = RIGHT)
        ENTRY_AMOUNT.grid(column = 3, row = 0, sticky = W, ipadx = 1)
        ENTRY_AMOUNT.bind("<FocusOut>", lambda e: self.validateAmount(entryAmount[rown]))
        ENTRY_AMOUNT.bind("<Tab>", self.updateSOAEntriesTotals)
        ENTRY_AMOUNT.bind("<FocusIn>", lambda e: entryAmount[rown].set(entryAmount[rown].get().replace(",", "")))
        entryAmount.append(TEXTVAR_SJAMOUNT)
        
        global TEXTVAR_DRCR
        TEXTVAR_DRCR = StringVar()
        COMBO_DRCR = tk.Combobox(FRAME_LINE, values = ["Debit", "Credit"], textvariable = TEXTVAR_DRCR, font = APP_FONT, width = 7, state = "readonly")
        COMBO_DRCR.grid(column = 4, row = 0, sticky = W, ipadx = 2)
        COMBO_DRCR.bind("<<ComboboxSelected>>", self.updateSOAEntriesTotals)
        entryDrCr.append(TEXTVAR_DRCR)

        global TEXTVAR_SJREMARKS
        TEXTVAR_SJREMARKS = StringVar()
        ENTRY_SJREMARKS = Entry(FRAME_LINE, textvariable = TEXTVAR_SJREMARKS, font = APP_FONT, width = 30)
        ENTRY_SJREMARKS.grid(column = 5, row = 0, sticky = W, ipadx = 1)
        entryRemarks.append(TEXTVAR_SJREMARKS)
        
        global BUTTON_CLEAR
        BUTTON_CLEAR = Button(FRAME_LINE, text = "[X]", font = BUTTON_FONT2, bg = BUTTON_BG, fg = BUTTON_FG_RED, activebackground = CLICK_BG, cursor = "hand2", command = lambda: self.clearSOAEntryLine(entryCode[rown], entryTitle[rown], entryAmount[rown], entryDrCr[rown], entryRemarks[rown], rown))
        BUTTON_CLEAR.grid(column = 6, row = 0, sticky = W)
        entryClear.append(BUTTON_CLEAR)

    def clearSOAEntryLines(self):
        for i in entryCode:
            i.set("")
        for i in entryTitle:
            i.set("")
        for i in entryAmount:
            i.set("")
        for i in entryDrCr:
            i.set("") 
        for i in entryRemarks:
            i.set("")

    def clearSOAEntryLine(self, code, title, amt, drcr, rem, i):
        code.set("")
        title.set("")
        amt.set("")
        drcr.set("")
        rem.set("")
        self.updateSOAEntriesTotals()
        
    def validateSOANumber(self, var, *args):
        try:
            find = "SELECT soaNumber FROM tblreceivables WHERE soaNumber = %s"
            db.commit()
            cursor.execute(find, [int(var.get())])
            result = cursor.fetchall()
            if result:
                var.set("")
                messagebox.showerror("Add SOA", "SOA number already used!")
                TOP_RECEIVABLES.focus()
            else:
                try:
                    var.set(str(int(var.get())).zfill(8))
                except:
                    var.set("")
                    messagebox.showerror("Add SOA", "Please enter a valid SOA number")
                    TOP_RECEIVABLES.focus()
        except:
            var.set("")
            
    def returnSOADuplicate(self, var):
        find = "SELECT soaNumber FROM tblreceivables WHERE soaNumber = %s LIMIT 1"
        cursor.execute(find, [int(var)])
        result = cursor.fetchone()
        if result:
            return True
        else:
            return False

    def returnCoveredPeriod(self, var):
        if var.split("-")[2][:2] == "05":
            if var.split("-")[1] == "01":
                year1 = str(int(var.split("-")[0])-1)
                year2 = year1
                month1 = "12"
                month2 = month1
            else:
                year1 = str(int(var.split("-")[0])-1)
                year2 = year1
                month1 = str(int(var.split("-")[1])-1).zfill(2)
                month2 = month1
            day1 = "11"
            day2 = "25"
        elif var.split("-")[2][:2] == "10":
            if var.split("-")[1] == "01":
                year1 = str(int(var.split("-")[0])-1)
                year2 = year1
                month1 = "12"
                month2 = month1
            else:
                year1 = str(int(var.split("-")[0])-1)
                year2 = year1
                month1 = str(int(var.split("-")[1])-1).zfill(2)
                month2 = month1
            day1 = "16"
            day2 = str(monthrange(int(var.split("-")[0]), int(month2))[1])
        elif var.split("-")[2][:2] == "15":
            if var.split("-")[1] == "01":
                year1 = str(int(var.split("-")[0])-1)
                year2 = var.split("-")[0]
                month1 = "12"
                month2 = "01"
            else:
                year1 = var.split("-")[0]
                year2 = year1
                month1 = str(int(var.split("-")[1])-1).zfill(2)
                month2 = var.split("-")[1]
            day1 = "21"
            day2 = "05"
        elif var.split("-")[2][:2] == "20":
            if var.split("-")[1] == "01":
                year1 = str(int(var.split("-")[0])-1)
                year2 = var.split("-")[0]
                month1 = "12"
                month2 = "01"
            else:
                year1 = var.split("-")[0]
                year2 = year1
                month1 = str(int(var.split("-")[1])-1).zfill(2)
                month2 = var.split("-")[1]
            day1 = "26"
            day2 = "10"
        elif var.split("-")[2][:2] == "25":
            year1 = var.split("-")[0]
            year2 = year1
            month1 = var.split("-")[1]
            month2 = var.split("-")[1]
            day1 = "01"
            day2 = "15"
        elif var.split("-")[2][:2] == str(int(var.split("-")[1])-1).zfill(2):
            year1 = var.split("-")[0]
            year2 = year1
            month1 = var.split("-")[1]
            month2 = month1
            day1 = "06"
            day2 = "25"
        try:
            return year1 + "-" + month1 + "-" + day1 + " to " + year2 + "-" + month2 + "-" + day2
        except:
            return var
    
    def returnClientTaxType(self, var):
        find = "SELECT taxType FROM tblclients WHERE clientCode = %s LIMIT 1"
        cursor.execute(find, [var])
        result = cursor.fetchone()
        if result:
            if result[0] == "VAT":
                return "WV-02"
            else:
                return "NV-02"
        else:
            return "WV-02"
        
    def returnClientTaxType2(self, var):
        find = "SELECT taxType FROM tblclients WHERE clientCode = %s LIMIT 1"
        cursor.execute(find, [var])
        result = cursor.fetchone()
        if result:
            return result[0]
        else:
            return "VAT"
    
    def returnLastDayOfMonth(self, var):
        return var.split("-")[0] + "-" + var.split("-")[1] + "-" + str(monthrange(int(var.split("-")[0]), int(var.split("-")[1]))[1])
    
    def returnFirstDayOfMonth(self, var):
        return var.split("-")[0] + "-" + var.split("-")[1] + "-" + "01"

    def returnTotalSOABalance(self, var, var2):
        select = "SELECT SUM(amount) FROM tblcollections WHERE soaNumber = %s AND isVoid = 'No' LIMIT 1"
        cursor.execute(select, [var])
        result = cursor.fetchone()
        if self.returnIfSOAVoid(var) == True:
            return self.validateAmount2(0)
        else:
            if result[0] != None:
                return self.validateAmount2(self.returnFloatAmount(var2)-float(result[0]))
            else:
                return self.validateAmount2(self.returnFloatAmount(var2))

    def searchReceivable(self, var, *args):
        find = f"SELECT soaNumber, soaDate, glDate, clientCode, particulars, amount, isApproved, void, encoder, DATE(encoded), approver, DATE(approved) FROM tblreceivables WHERE (soaNumber LIKE %s OR clientCode LIKE %s OR particulars LIKE %s) AND soaDate BETWEEN %s AND %s ORDER BY soaNumber DESC"
        db.commit()
        cursor.execute(find, [f"%{var}%", f"%{var}%", f"%{var}%", CALENDAR_START.get_date(), CALENDAR_END.get_date()])
        result = cursor.fetchall()
        for i in TREE_RECEIVABLES.get_children():
            TREE_RECEIVABLES.delete(i)
        if result:
            PROGRESS_BAR.grid(column = 8, row = 0, padx = TOP_PADX + 10, sticky = E)
            count = 0
            soanumbers, skipped = [], []
            for i in result:
                if i[0] not in soanumbers:
                    if i[7] != "Yes":
                        if count % 2 == 0:
                            TREE_RECEIVABLES.insert("", "end", values = (str(i[0]).zfill(8),i[1],i[2],self.returnClientName(i[3]),i[4],self.returnTotalSOAAmount(i[0]),self.returnTotalSOABalance(i[0], self.returnTotalSOAAmount(i[0])),i[6],i[7],self.returnClientEntityType(i[3]),self.returnUserName(i[8], 0),i[9],self.returnUserName(i[10], 0),i[11]), tags = "evenrow")
                        else:
                            TREE_RECEIVABLES.insert("", "end", values = (str(i[0]).zfill(8),i[1],i[2],self.returnClientName(i[3]),i[4],self.returnTotalSOAAmount(i[0]),self.returnTotalSOABalance(i[0], self.returnTotalSOAAmount(i[0])),i[6],i[7],self.returnClientEntityType(i[3]),self.returnUserName(i[8], 0),i[9],self.returnUserName(i[10], 0),i[11]), tags = "oddrow")
                    else:
                        TREE_RECEIVABLES.insert("", "end", values = (str(i[0]).zfill(8),i[1],i[2],self.returnClientName(i[3]),i[4],self.returnTotalSOAAmount(i[0]),self.returnTotalSOABalance(i[0], self.returnTotalSOAAmount(i[0])),i[6],i[7],self.returnClientEntityType(i[3]),self.returnUserName(i[8], 0),i[9],self.returnUserName(i[10], 0),i[11]), tags = "void")
                    count += 1
                    soanumbers.append(i[0])
                    PROGRESS_BAR["value"] = round((count/(len(result)-len(skipped)))*100, 0)
                    SUB1_RECEIVABLES.update()
                else:
                    skipped.append(i[0])
            PROGRESS_BAR.grid_remove()
        else:
            messagebox.showerror("Receivables", "No match found!")

    def refreshReceivable(self, *args):
        for i in TREE_RECEIVABLES.get_children():
            TREE_RECEIVABLES.delete(i)
        db.commit()
        cursor.execute(f"SELECT soaNumber, soaDate, glDate, clientCode, particulars, amount, isApproved, void, encoder, DATE(encoded), approver, DATE(approved) FROM tblreceivables WHERE soaDate BETWEEN '{CALENDAR_START.get_date()}' AND '{CALENDAR_END.get_date()}' ORDER BY soaNumber DESC")
        result = cursor.fetchall()
        PROGRESS_BAR.grid(column = 8, row = 0, padx = TOP_PADX + 10, sticky = E)
        count = 0
        soanumbers, skipped = [], []
        for i in result:
            if i[0] not in soanumbers:
                if count % 2 == 0:
                    TREE_RECEIVABLES.insert("", "end", values = (str(i[0]).zfill(8),i[1],i[2],self.returnClientName(i[3]),i[4],self.returnTotalSOAAmount(i[0]),self.returnTotalSOABalance(i[0], self.returnTotalSOAAmount(i[0])),i[6],i[7],self.returnClientEntityType(i[3]),self.returnUserName(i[8], 0),i[9],self.returnUserName(i[10], 0),i[11]), tags = "evenrow")
                else:
                    TREE_RECEIVABLES.insert("", "end", values = (str(i[0]).zfill(8),i[1],i[2],self.returnClientName(i[3]),i[4],self.returnTotalSOAAmount(i[0]),self.returnTotalSOABalance(i[0], self.returnTotalSOAAmount(i[0])),i[6],i[7],self.returnClientEntityType(i[3]),self.returnUserName(i[8], 0),i[9],self.returnUserName(i[10], 0),i[11]), tags = "oddrow")
                count += 1
                soanumbers.append(i[0])
                PROGRESS_BAR["value"] = round((count/(len(result)-len(skipped)))*100, 0)
                SUB1_RECEIVABLES.update()
            else:
                skipped.append(i[0])
        PROGRESS_BAR.grid_remove()

### MENU_BILLING_COLLECTIONS ###
    def showAddEditCollection(self):
        global TOP_COLLECTIONS
        TOP_COLLECTIONS = Toplevel()
        TOP_COLLECTIONS.title("Add - Collection")
        TOP_COLLECTIONS.iconbitmap(PATH_ICON + "icon.ico")
        TOP_COLLECTIONS.geometry("1215x595+20+20")
        TOP_COLLECTIONS.resizable(height = False, width = False)
        TOP_COLLECTIONS.grab_set()
        TOP_COLLECTIONS.focus_force()
        
        global required
        required = []
        
        global SUB_FRAME1
        SUB_FRAME1 = Frame(TOP_COLLECTIONS) #details frame
        SUB_FRAME1.grid(column = 0, row = 0, sticky = W, padx = TOP_PADX + 15)
        
        SUB_SUB1FRAME1 = Frame(SUB_FRAME1)
        SUB_SUB1FRAME1.grid(column = 0, row = 0, sticky = W)
        
        LABEL_ORNO = Label(SUB_SUB1FRAME1, text = "Receipt No.", font = APP_FONT)
        LABEL_ORNO.grid(column = 0, row = 0, sticky = W)
        
        LABEL_CLIENT = Label(SUB_SUB1FRAME1, text = "Client Name", font = APP_FONT)
        LABEL_CLIENT.grid(column = 0, row = 1, sticky = W)
        
        LABEL_PARTICULARS = Label(SUB_SUB1FRAME1, text = "Particulars", font = APP_FONT)
        LABEL_PARTICULARS.grid(column = 0, row = 2, sticky = W)
        
        LABEL_ENTITYTYPE = Label(SUB_SUB1FRAME1, text = "Entity Type", font = APP_FONT)
        LABEL_ENTITYTYPE.grid(column = 0, row = 3, sticky = W)
        
        SUB_SUB1SUB1FRAME1 = Frame(SUB_SUB1FRAME1)
        SUB_SUB1SUB1FRAME1.grid(column = 1, row = 0, sticky = W)
        
        SUB_SUB1SUB1FRAME2 = Frame(SUB_SUB1FRAME1)
        SUB_SUB1SUB1FRAME2.grid(column = 1, row = 1, sticky = W)
        
        global TEXTVAR_ORNUMBER, ENTRY_ORNUMBER
        TEXTVAR_ORNUMBER = StringVar()
        ENTRY_ORNUMBER = Entry(SUB_SUB1SUB1FRAME1, textvariable = TEXTVAR_ORNUMBER, font = APP_FONT, width = 13)
        ENTRY_ORNUMBER.grid(column = 0, row = 0, sticky = W)
        ENTRY_ORNUMBER.bind("<FocusOut>", lambda e: self.validateORNumber(TEXTVAR_ORNUMBER))
        required.append(TEXTVAR_ORNUMBER)
        
        global TEXTVAR_TAX, COMBO_TAX
        TEXTVAR_TAX = StringVar()
        COMBO_TAX = tk.Combobox(SUB_SUB1SUB1FRAME1, values = ["VAT", "NONVAT"], textvariable = TEXTVAR_TAX, font = APP_FONT, width = 7, state = "readonly")
        COMBO_TAX.grid(column = 2, row = 0, sticky = W, padx = TOP_PADX + 10)
        required.append(TEXTVAR_TAX)
        # soaTax.append(TEXTVAR_TAX)
        # soaComboTax.append(COMBO_TAX)

        global TEXTVAR_CLIENTCODE, ENTRY_CLIENTCODE
        TEXTVAR_CLIENTCODE = StringVar()
        ENTRY_CLIENTCODE = Entry(SUB_SUB1SUB1FRAME2, textvariable = TEXTVAR_CLIENTCODE, font = APP_FONT, width = 15)
        ENTRY_CLIENTCODE.grid(column = 0, row = 0, sticky = W)
        ENTRY_CLIENTCODE.bind("<FocusOut>", lambda e: self.populateSOAFields(TEXTVAR_CLIENTCODE.get()))
        required.append(TEXTVAR_CLIENTCODE)

        global BUTTON_CLIENT
        BUTTON_CLIENT = Button(SUB_SUB1SUB1FRAME2, text = "...", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, cursor = "hand2", command = self.showClientSelection)
        BUTTON_CLIENT.grid(column = 0, row = 0, sticky = E)
        
        global TEXTVAR_CLIENTNAME
        TEXTVAR_CLIENTNAME = StringVar()
        ENTRY_CLIENTNAME = Entry(SUB_SUB1SUB1FRAME2, textvariable = TEXTVAR_CLIENTNAME, font = APP_FONT, width = 55, state = "readonly")
        ENTRY_CLIENTNAME.grid(column = 1, row = 0, sticky = W)
        required.append(TEXTVAR_CLIENTNAME)
        
        global ENTRY_PARTICULARS
        ENTRY_PARTICULARS = Text(SUB_SUB1FRAME1, font = APP_FONT, width = 70, height = 2)
        ENTRY_PARTICULARS.grid(column = 1, row = 2, sticky = W)
        required.append(ENTRY_PARTICULARS)
        
        global TEXTVAR_ENTITYTYPE
        TEXTVAR_ENTITYTYPE = StringVar()
        ENTRY_ENTITYTYPE = Entry(SUB_SUB1FRAME1, textvariable = TEXTVAR_ENTITYTYPE, font = APP_FONT, width = 12, state = "readonly")
        ENTRY_ENTITYTYPE.grid(column = 1, row = 3, sticky = W)
        
        global BUTTON_SELECTSOA
        BUTTON_SELECTSOA = Button(SUB_SUB1FRAME1, text = "Select SOA", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, cursor = "hand2", command = lambda: self.showSOASelection(TEXTVAR_CLIENTCODE))
        BUTTON_SELECTSOA.grid(column = 1, row = 3)
        
        SUB_SUB1FRAME2 = Frame(SUB_FRAME1) #dates and cash or cheque mode
        SUB_SUB1FRAME2.grid(column = 1, row = 0, sticky = W, padx = TOP_PADX + 15)
        
        LABEL_ORDATE = Label(SUB_SUB1FRAME2, text = "OR Date", font = APP_FONT)
        LABEL_ORDATE.grid(column = 0, row = 0, sticky = W)
        
        LABEL_GLDATE = Label(SUB_SUB1FRAME2, text = "GL Date", font = APP_FONT)
        LABEL_GLDATE.grid(column = 0, row = 1, sticky = W)
        
        LABEL_MODE = Label(SUB_SUB1FRAME2, text = "Mode", font = APP_FONT)
        LABEL_MODE.grid(column = 0, row = 2, sticky = W)
        
        LABEL_CHECK = Label(SUB_SUB1FRAME2, text = "Check No.", font = APP_FONT)
        LABEL_CHECK.grid(column = 0, row = 3, sticky = W)
        
        global CALENDAR_ORDATE
        CALENDAR_ORDATE = DateEntry(SUB_SUB1FRAME2, firstweekday = "sunday", date_pattern = "yyyy-mm-dd")
        CALENDAR_ORDATE.grid(column = 1, row = 0, sticky = W)
        CALENDAR_ORDATE.bind("<FocusOut>", lambda e: CALENDAR_GLDATE.set_date(self.returnLastDayOfMonth(str(CALENDAR_ORDATE.get_date()))))
        required.append(CALENDAR_ORDATE)
        
        global CALENDAR_GLDATE
        CALENDAR_GLDATE = DateEntry(SUB_SUB1FRAME2, firstweekday = "sunday", date_pattern = "yyyy-mm-dd")
        CALENDAR_GLDATE.grid(column = 1, row = 1, sticky = W)
        required.append(CALENDAR_GLDATE)
        CALENDAR_GLDATE.set_date(self.returnLastDayOfMonth(str(CALENDAR_ORDATE.get_date())))
        
        global TEXTVAR_MODE, COMBO_MODE
        TEXTVAR_MODE = StringVar()
        COMBO_MODE = tk.Combobox(SUB_SUB1FRAME2, values = ["Cash", "Check", "Auto Credit"], textvariable = TEXTVAR_MODE, font = APP_FONT, width = 15, state = "readonly")
        COMBO_MODE.grid(column = 1, row = 2, sticky = W)
        required.append(TEXTVAR_MODE)
        # soaTax.append(TEXTVAR_MODE)
        # soaComboTax.append(COMBO_MODE)
        
        global TEXTVAR_CHECKNUMBER
        TEXTVAR_CHECKNUMBER = StringVar()
        ENTRY_CHECKNUMBER = Entry(SUB_SUB1FRAME2, textvariable = TEXTVAR_CHECKNUMBER, font = APP_FONT, width = 15)
        ENTRY_CHECKNUMBER.grid(column = 1, row = 3, sticky = W)
        
        SUB_SUB1FRAME3 = Frame(SUB_FRAME1) #totals frame
        SUB_SUB1FRAME3.grid(column = 2, row = 0, sticky = W)
        
        LABEL_GROSS = Label(SUB_SUB1FRAME3, text = "Collection", font = APP_FONT)
        LABEL_GROSS.grid(column = 0, row = 0, sticky = W)
        
        LABEL_VAT = Label(SUB_SUB1FRAME3, text = "VAT", font = APP_FONT)
        LABEL_VAT.grid(column = 0, row = 1, sticky = W)
        
        LABEL_EWT = Label(SUB_SUB1FRAME3, text = "EWT", font = APP_FONT)
        LABEL_EWT.grid(column = 0, row = 2, sticky = W)
        
        LABEL_CVAT = Label(SUB_SUB1FRAME3, text = "CVAT", font = APP_FONT)
        LABEL_CVAT.grid(column = 0, row = 3, sticky = W)
        
        LABEL_NET = Label(SUB_SUB1FRAME3, text = "Income", font = APP_FONT)
        LABEL_NET.grid(column = 0, row = 4, sticky = W)
        
        global TEXTVAR_TGROSS
        TEXTVAR_TGROSS = StringVar()
        ENTRY_GROSS = Entry(SUB_SUB1FRAME3, textvariable = TEXTVAR_TGROSS, font = APP_FONT, width = 12, state = "readonly", justify = RIGHT)
        ENTRY_GROSS.grid(column = 1, row = 0, sticky = W)
        
        global TEXTVAR_TVAT
        TEXTVAR_TVAT = StringVar()
        ENTRY_VAT = Entry(SUB_SUB1FRAME3, textvariable = TEXTVAR_TVAT, font = APP_FONT, width = 12, state = "readonly", justify = RIGHT)
        ENTRY_VAT.grid(column = 1, row = 1, sticky = W)
        
        global TEXTVAR_TEWT
        TEXTVAR_TEWT = StringVar()
        ENTRY_EWT = Entry(SUB_SUB1FRAME3, textvariable = TEXTVAR_TEWT, font = APP_FONT, width = 12, state = "readonly", justify = RIGHT)
        ENTRY_EWT.grid(column = 1, row = 2, sticky = W)
        
        global TEXTVAR_TCVAT
        TEXTVAR_TCVAT = StringVar()
        ENTRY_CVAT = Entry(SUB_SUB1FRAME3, textvariable = TEXTVAR_TCVAT, font = APP_FONT, width = 12, state = "readonly", justify = RIGHT)
        ENTRY_CVAT.grid(column = 1, row = 3, sticky = W)
        
        global TEXTVAR_TNET
        TEXTVAR_TNET = StringVar()
        ENTRY_NET = Entry(SUB_SUB1FRAME3, textvariable = TEXTVAR_TNET, font = APP_FONT, width = 12, state = "readonly", justify = RIGHT)
        ENTRY_NET.grid(column = 1, row = 4, sticky = W)
        
        SUB_FRAME2 = Frame(TOP_COLLECTIONS) #headers frame
        SUB_FRAME2.grid(column = 0, row = 1, sticky = W)
        
        LABEL_CODE = Label(SUB_FRAME2, text = "SOA No.", width = 10, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_CODE.grid(column = 0, row = 0)

        LABEL_TITLE = Label(SUB_FRAME2, text = "Client Name", width = 15, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_TITLE.grid(column = 1, row = 0)
        
        LABEL_TITLE = Label(SUB_FRAME2, text = "Transaction", width = 12, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_TITLE.grid(column = 2, row = 0)
        
        LABEL_TITLE = Label(SUB_FRAME2, text = "Tax", width = 4, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_TITLE.grid(column = 3, row = 0)

        LABEL_CENTER = Label(SUB_FRAME2, text = "Receivable", width = 15, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_CENTER.grid(column = 4, row = 0)

        LABEL_GROSS = Label(SUB_FRAME2, text = "Collection", width = 15, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_GROSS.grid(column = 5, row = 0)

        LABEL_EXPENSE = Label(SUB_FRAME2, text = "Balance", width = 15, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_EXPENSE.grid(column = 6, row = 0)

        LABEL_VAT = Label(SUB_FRAME2, text = "Payment", width = 10, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_VAT.grid(column = 7, row = 0)

        LABEL_EWT = Label(SUB_FRAME2, text = "VAT", width = 12, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_EWT.grid(column = 8, row = 0)

        LABEL_NET = Label(SUB_FRAME2, text = "EWT", width = 12, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_NET.grid(column = 9, row = 0)

        LABEL_DESC = Label(SUB_FRAME2, text = "CVAT", width = 12, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_DESC.grid(column = 10, row = 0)

        LABEL_CANCEL = Label(SUB_FRAME2, text = "Income", width = 12, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_CANCEL.grid(column = 11, row = 0)
        
        LABEL_CANCEL = Label(SUB_FRAME2, text = "Remarks", width = 12, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_CANCEL.grid(column = 12, row = 0)
        
        LABEL_CANCEL = Label(SUB_FRAME2, text = "[x]", width = 2, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_CANCEL.grid(column = 13, row = 0)
        
        global SUB_FRAME3
        SUB_FRAME3 = Frame(TOP_COLLECTIONS) #transactions scroll frame
        SUB_FRAME3.grid(column = 0, row = 2, sticky = W)
        
        global orFrame, orSOANumber, orClientName, orTransaction, orTax, orReceivable, orCollection, orBalance, orPayment, orComboPayment, orVAT, orEWT, orCVAT, orINC, orRemarks, orClear, orOutput
        orFrame, orSOANumber, orClientName, orTransaction, orTax, orReceivable, orCollection, orBalance, orPayment, orComboPayment, orVAT, orEWT, orCVAT, orINC, orRemarks, orClear, orOutput = [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], []
        outputtax = 0

        self.createScrollFrame(SUB_FRAME3, 280, 1190, 20, 0, 0)
        for i in range(11):
            self.showCollectionLines(SCROLLABLE_FRAME, i)
            orOutput.append(outputtax)

        global SUB_FRAME4
        SUB_FRAME4 = Frame(TOP_COLLECTIONS) #entry
        SUB_FRAME4.grid(column = 0, row = 3, sticky = NW)
        
        global SUB_SUB4FRAME1
        SUB_SUB4FRAME1 = Frame(SUB_FRAME4) #entry headers frame
        SUB_SUB4FRAME1.grid(column = 0, row = 0, sticky = NW)
        
        LABEL_CODE = Label(SUB_SUB4FRAME1, text = "Code", width = 12, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_CODE.grid(column = 0, row = 0)

        LABEL_TITLE = Label(SUB_SUB4FRAME1, text = "Title", width = 30, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_TITLE.grid(column = 1, row = 0)

        LABEL_CENTER = Label(SUB_SUB4FRAME1, text = "Amount", width = 15, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_CENTER.grid(column = 2, row = 0)

        LABEL_GROSS = Label(SUB_SUB4FRAME1, text = "Dr/Cr", width = 10, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_GROSS.grid(column = 3, row = 0)

        LABEL_EXPENSE = Label(SUB_SUB4FRAME1, text = "Remarks", width = 31, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_EXPENSE.grid(column = 4, row = 0)

        LABEL_X = Label(SUB_SUB4FRAME1, text = "[x]", width = 2, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_X.grid(column = 5, row = 0)
        
        global SUB_SUB4FRAME2
        SUB_SUB4FRAME2 = Frame(SUB_FRAME4) #entry lines frame
        SUB_SUB4FRAME2.grid(column = 0, row = 1, sticky = NW)
        
        self.createScrollFrame(SUB_SUB4FRAME2, 150, 735, 20, 0, 0)
        
        global entryFrame, entryCode, entryTitle, entryAmount, entryDrCr, entryRemarks, entryClear
        entryFrame, entryCode, entryTitle, entryAmount, entryDrCr, entryRemarks, entryClear = [], [], [], [], [], [], []
        
        for i in range(35):
            self.showSOAEntryLines(SCROLLABLE_FRAME, i)
            
        self.disableCollectionSOANumberFields(DISABLED)
            
        SUB_SUB4FRAME3 = Frame(SUB_FRAME4) #entry total frame
        SUB_SUB4FRAME3.grid(column = 1, row = 1, sticky = NW)
        
        LABEL_DEBIT = Label(SUB_SUB4FRAME3, text = "Debit", width = 12, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_DEBIT.grid(column = 0, row = 0)

        LABEL_CREDIT = Label(SUB_SUB4FRAME3, text = "Credit", width = 12, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_CREDIT.grid(column = 0, row = 1)

        LABEL_VARIANCE = Label(SUB_SUB4FRAME3, text = "Variance", width = 12, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
        LABEL_VARIANCE.grid(column = 0, row = 2)
        
        global TEXTVAR_DEBIT
        TEXTVAR_DEBIT = StringVar()
        ENTRY_DEBIT = Entry(SUB_SUB4FRAME3, textvariable = TEXTVAR_DEBIT, font = APP_FONT, width = 12, justify = RIGHT, state = "readonly")
        ENTRY_DEBIT.grid(column = 1, row = 0, sticky = W)
        
        global TEXTVAR_CREDIT
        TEXTVAR_CREDIT = StringVar()
        ENTRY_CREDIT = Entry(SUB_SUB4FRAME3, textvariable = TEXTVAR_CREDIT, font = APP_FONT, width = 12, justify = RIGHT, state = "readonly")
        ENTRY_CREDIT.grid(column = 1, row = 1, sticky = W)
        
        global TEXTVAR_VARIANCE
        TEXTVAR_VARIANCE = StringVar()
        ENTRY_VARIANCE = Entry(SUB_SUB4FRAME3, textvariable = TEXTVAR_VARIANCE, font = APP_FONT, width = 12, justify = RIGHT, state = "readonly")
        ENTRY_VARIANCE.grid(column = 1, row = 2, sticky = W)
        
        global SUB_FRAME5
        SUB_FRAME5 = Frame(TOP_COLLECTIONS) #button frame
        SUB_FRAME5.grid(column = 0, row = 3, sticky = SE)
        
        global BUTTON_SUBMIT
        BUTTON_SUBMIT = Button(SUB_FRAME5, text = "SUBMIT", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "hand2", state = NORMAL, command = self.saveCollection)
        BUTTON_SUBMIT.grid(column = 1, row = 0, padx = TOP_PADX)

        global BUTTON_APPROVE
        BUTTON_APPROVE = Button(SUB_FRAME5, text = "APPROVE", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG_GRN, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "arrow", state = DISABLED, command = None)
        BUTTON_APPROVE.grid(column = 2, row = 0, padx = TOP_PADX)

        global BUTTON_VOID
        BUTTON_VOID = Button(SUB_FRAME5, text = "VOID", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG_RED, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "arrow", state = DISABLED, command = None)
        BUTTON_VOID.grid(column = 3, row = 0, padx = TOP_PADX)

        global BUTTON_PRINT
        BUTTON_PRINT = Button(SUB_FRAME5, text = "PRINT", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "arrow", state = DISABLED, command = None)
        BUTTON_PRINT.grid(column = 4, row = 0, padx = TOP_PADX)

        BUTTON_CLOSE = Button(SUB_FRAME5, text = "CLOSE", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "hand2", command = lambda: self.closeTopLevel(TOP_COLLECTIONS))
        BUTTON_CLOSE.grid(column = 5, row = 0, padx = TOP_PADX)

    def showCollectionLines(self, frame, rown):
        FRAME_LINE = Frame(frame)
        FRAME_LINE.grid(column = 0, row = rown, sticky = W)
        orFrame.append(FRAME_LINE)

        global TEXTVAR_SOANUMBER
        TEXTVAR_SOANUMBER = StringVar()
        ENTRY_SOANUMBER = Entry(FRAME_LINE, textvariable = TEXTVAR_SOANUMBER, font = APP_FONT, width = 10, justify = CENTER, state = "readonly")
        ENTRY_SOANUMBER.grid(column = 0, row = 0, sticky = W, ipadx = 1)
        # ENTRY_SOANUMBER.bind("<FocusOut>", lambda e: self.getSOADetails(orSOANumber[rown], rown))
        orSOANumber.append(TEXTVAR_SOANUMBER)

        global TEXTVAR_ORCLIENTNAME
        TEXTVAR_ORCLIENTNAME = StringVar()
        ENTRY_ORCLIENTNAME = Entry(FRAME_LINE, textvariable = TEXTVAR_ORCLIENTNAME, font = APP_FONT_SMALL, width = 17, state = "readonly")
        ENTRY_ORCLIENTNAME.grid(column = 1, row = 0, sticky = W, ipadx = 1)
        orClientName.append(TEXTVAR_ORCLIENTNAME)
        
        global TEXTVAR_ORTRANSACTION
        TEXTVAR_ORTRANSACTION = StringVar()
        ENTRY_ORTRANSACTION = Entry(FRAME_LINE, textvariable = TEXTVAR_ORTRANSACTION, font = APP_FONT, width = 12, state = "readonly")
        ENTRY_ORTRANSACTION.grid(column = 2, row = 0, sticky = W, ipadx = 1)
        orTransaction.append(TEXTVAR_ORTRANSACTION)
        
        global TEXTVAR_ORTAX
        TEXTVAR_ORTAX = StringVar()
        ENTRY_ORTAX = Entry(FRAME_LINE, textvariable = TEXTVAR_ORTAX, font = APP_FONT, width = 4, state = "readonly")
        ENTRY_ORTAX.grid(column = 3, row = 0, sticky = W, ipadx = 1)
        orTax.append(TEXTVAR_ORTAX)

        global TEXTVAR_RECEIVABLE
        TEXTVAR_RECEIVABLE = StringVar()
        ENTRY_RECEIVABLE = Entry(FRAME_LINE, textvariable = TEXTVAR_RECEIVABLE, font = APP_FONT, width = 15, justify = RIGHT, state = "readonly")
        ENTRY_RECEIVABLE.grid(column = 4, row = 0, sticky = W, ipadx = 1)
        orReceivable.append(TEXTVAR_RECEIVABLE)
        
        global TEXTVAR_COLLECTION
        TEXTVAR_COLLECTION = StringVar()
        ENTRY_COLLECTION = Entry(FRAME_LINE, textvariable = TEXTVAR_COLLECTION, font = APP_FONT, width = 15, justify = RIGHT)
        ENTRY_COLLECTION.grid(column = 5, row = 0, sticky = W, ipadx = 1)
        ENTRY_COLLECTION.bind("<FocusOut>", lambda e: self.updateORLine(orCollection[rown], TEXTVAR_ENTITYTYPE, orTax[rown].get(), rown))
        ENTRY_COLLECTION.bind("<FocusIn>", lambda e: orCollection[rown].set(orCollection[rown].get().replace(",", "")))
        orCollection.append(TEXTVAR_COLLECTION)
        
        global TEXTVAR_BALANCE
        TEXTVAR_BALANCE = StringVar()
        ENTRY_BALANCE = Entry(FRAME_LINE, textvariable = TEXTVAR_BALANCE, font = APP_FONT, width = 15, justify = RIGHT, state = "readonly")
        ENTRY_BALANCE.grid(column = 6, row = 0, sticky = W, ipadx = 1)
        # ENTRY_GROSS.bind("<FocusOut>", lambda e: self.updateSOALine(soaGross[rown], TEXTVAR_ENTITYTYPE, soaTax[rown], rown))
        # ENTRY_GROSS.bind("<FocusIn>", lambda e: soaGross[rown].set(soaGross[rown].get().replace(",", "")))
        orBalance.append(TEXTVAR_BALANCE)
        
        global TEXTVAR_PAYMENT, COMBO_PAYMENT
        TEXTVAR_PAYMENT = StringVar()
        COMBO_PAYMENT = tk.Combobox(FRAME_LINE, values = ["Full", "Partial"], textvariable = TEXTVAR_PAYMENT, font = APP_FONT, width = 7, state = "readonly")
        COMBO_PAYMENT.grid(column = 7, row = 0, sticky = W, ipadx = 1)
        orPayment.append(TEXTVAR_PAYMENT)
        orComboPayment.append(COMBO_PAYMENT)
        COMBO_PAYMENT.bind("<<ComboboxSelected>>", lambda e: self.updateORLine(orCollection[rown], TEXTVAR_ENTITYTYPE, orTax[rown].get(), rown))

        global TEXTVAR_VAT
        TEXTVAR_VAT = StringVar()
        ENTRY_VAT = Entry(FRAME_LINE, textvariable = TEXTVAR_VAT, font = APP_FONT, width = 12, justify = RIGHT, state = "readonly")
        ENTRY_VAT.grid(column = 8, row = 0, sticky = W, ipadx = 1)
        orVAT.append(TEXTVAR_VAT)

        global TEXTVAR_EWT
        TEXTVAR_EWT = StringVar()
        ENTRY_EWT = Entry(FRAME_LINE, textvariable = TEXTVAR_EWT, font = APP_FONT, width = 12, justify = RIGHT, state = "readonly")
        ENTRY_EWT.grid(column = 9, row = 0, sticky = W, ipadx = 1)
        orEWT.append(TEXTVAR_EWT)

        global TEXTVAR_CVAT
        TEXTVAR_CVAT = StringVar()
        ENTRY_CVAT = Entry(FRAME_LINE, textvariable = TEXTVAR_CVAT, font = APP_FONT, width = 12, justify = RIGHT, state = "readonly")
        ENTRY_CVAT.grid(column = 10, row = 0, sticky = W, ipadx = 1)
        orCVAT.append(TEXTVAR_CVAT)

        global TEXTVAR_INC
        TEXTVAR_INC = StringVar()
        ENTRY_INC = Entry(FRAME_LINE, textvariable = TEXTVAR_INC, font = APP_FONT, width = 12, justify = RIGHT, state = "readonly")
        ENTRY_INC.grid(column = 11, row = 0, sticky = W, ipadx = 1)
        orINC.append(TEXTVAR_INC)

        global TEXTVAR_REMARKS
        TEXTVAR_REMARKS = StringVar()
        ENTRY_REMARKS = Entry(FRAME_LINE, textvariable = TEXTVAR_REMARKS, font = APP_FONT, width = 12)
        ENTRY_REMARKS.grid(column = 12, row = 0, sticky = W, ipadx = 1)
        orRemarks.append(TEXTVAR_REMARKS)

        global BUTTON_CLEAR
        BUTTON_CLEAR = Button(FRAME_LINE, text = "[X]", font = BUTTON_FONT2, bg = BUTTON_BG, fg = BUTTON_FG_RED, activebackground = CLICK_BG, cursor = "hand2", command = lambda: self.clearORLine(orCollection[rown], TEXTVAR_ENTITYTYPE, orTax[rown].get(), rown))
        BUTTON_CLEAR.grid(column = 13, row = 0, sticky = W)
        orClear.append(BUTTON_CLEAR)

    def saveCollection(self):
        wrong = []
        for i in required:
            try:
                if i.get() == "":
                    wrong.append(i)
            except:
                try:
                    if i.get_date() == "":
                        wrong.append(i)
                except:
                    if i.get("1.0", END) == "":
                        wrong.append(i)
        
        if self.returnFloatAmount(TEXTVAR_VARIANCE.get()) != 0:
            if self.returnUserName(USER, 3) == "ASD":
                wrong.append("Check entry variance!")
        if self.returnPeriodStatus(CALENDAR_GLDATE.get_date()) == "locked":
            wrong.append("GL date is locked")
        if self.returnFloatAmount(TEXTVAR_TGROSS.get()) == 0:
                messagebox.showerror("Collection", "Check the amounts!")

        if len(wrong) > 0:
            messagebox.showerror("Collection", "Fill up all required fields!" + str(wrong))
            TOP_COLLECTIONS.focus()
        else:
            ask = messagebox.askyesno("Collection", "Are you sure?")
            if ask:
                insert = """INSERT INTO tblcollections (
                            orNumber, orDate, glDate, mode, checkNumber,
                            soaNumber, clientCode, particulars, amount, payment,
                            remarks, isApproved, isVoid, encoder, encoded,
                            orType
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
                delete = "DELETE FROM tblcollections WHERE orNumber = %s"
                
                insertbook = """INSERT INTO tblcashreceiptbook (glDate, clientCode, orNumber, chartCode, amount,
                                                            side, remarks, isPosted, isVoid, poster,
                                                            posted, modifier, modified, source) VALUES 
                                                            (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
                deletebook = "DELETE FROM tblcashreceiptbook WHERE orNumber = %s"
                
                if TEXTVAR_ORNUMBER.get() != "":
                    cursor.execute(delete, [int(TEXTVAR_ORNUMBER.get())])
                    cursor.execute(deletebook, [int(TEXTVAR_ORNUMBER.get())])
                    db.commit()
                    message = "Collection has been successfully updated!"
                else:
                    message = "Collection has been successfully submitted!"

                validlines, validlinesbook = [], []
                for i in range(len(orFrame)):
                    if self.returnFloatAmount(orCollection[i].get()) > 0 and orSOANumber[i].get() != "":
                        validlines.append([
                            int(TEXTVAR_ORNUMBER.get()), CALENDAR_ORDATE.get_date(), CALENDAR_GLDATE.get_date(), TEXTVAR_MODE.get(), TEXTVAR_CHECKNUMBER.get(),
                            int(orSOANumber[i].get()), TEXTVAR_CLIENTCODE.get(), ENTRY_PARTICULARS.get("1.0", END).replace("\n",""), self.returnFloatAmount(orCollection[i].get()), orPayment[i].get(),
                            orRemarks[i].get(), "No", "No", USER, datetime.datetime.now(),
                            TEXTVAR_TAX.get()])

                for i in range(len(entryAmount)):
                    if self.returnFloatAmount(entryAmount[i].get()) != 0:
                        validlinesbook.append([CALENDAR_GLDATE.get_date(), TEXTVAR_CLIENTCODE.get(), int(TEXTVAR_ORNUMBER.get()), entryCode[i].get(), self.returnFloatAmount(entryAmount[i].get()),
                                                entryDrCr[i].get(), entryRemarks[i].get(), "No", "No", USER,
                                                datetime.datetime.now(), USER, datetime.datetime.now(), "SB"])
                    
                if len(validlines) > 0 and len(validlinesbook) > 0:
                    for i in validlines:
                        cursor.execute(insert, i)
                    for i in validlinesbook:
                        cursor.execute(insertbook, i)
                    db.commit()

                    ENTRY_ORNUMBER.config(state = "readonly")
                    messagebox.showinfo("Collection", message)
                    BUTTON_SUBMIT.config(state = DISABLED, cursor = "arrow")
                    BUTTON_APPROVE.config(state = NORMAL, cursor = "hand2")
                    BUTTON_PRINT.config(state = NORMAL, cursor = "hand2")
                    TOP_COLLECTIONS.focus()
                    
                    if self.returnAccess(USER, 16) == 0:
                        BUTTON_APPROVE.config(state = DISABLED, cursor = "arrow")
                else:
                    messagebox.showerror("Collection", "Length of valid lines is zero!")
                    TOP_COLLECTIONS.focus()

    def editCollection(self, *args):
        self.copySelection(TREE_COLLECTIONS)
        self.showAddEditCollection()
        TOP_COLLECTIONS.title("Edit - OR")
        
        find = """SELECT orNumber, orDate, glDate, mode, checkNumber,
                        soaNumber, clientCode, particulars, amount, payment,
                        remarks, encoder, isApproved, orType
                        FROM tblcollections WHERE orNumber = %s"""
                        
        findinbook = "SELECT chartCode, amount, side, remarks FROM tblcashreceiptbook WHERE orNumber = %s"
        db.commit()
        cursor.execute(find, [int(content[0])])
        result = cursor.fetchall()
        cursor.execute(findinbook, [int(content[0])])
        resultbook = cursor.fetchall()
        if result:
            CALENDAR_ORDATE.set_date(result[0][1])
            TEXTVAR_ORNUMBER.set(str(result[0][0]).zfill(8))
            TEXTVAR_TAX.set(result[0][13])
            ENTRY_ORNUMBER.config(state = "readonly")
            ENTRY_ORNUMBER.unbind("<FocusOut>")
            CALENDAR_GLDATE.set_date(result[0][2])
            TEXTVAR_MODE.set(result[0][3])
            TEXTVAR_CHECKNUMBER.set(result[0][4])
            ENTRY_PARTICULARS.insert(1.0, result[0][7])
            TEXTVAR_CLIENTCODE.set(result[0][6])
            self.populateSOAFields(result[0][6])
            TEXTVAR_ENTITYTYPE.set(self.returnClientEntityType(result[0][6])[0])
            BUTTON_SELECTSOA.config(state = DISABLED, cursor = "arrow")
            for i in range(len(result)):
                orSOANumber[i].set(result[i][5])
                self.getSOADetails(result[i][5], i)
                orCollection[i].set(self.validateAmount2(result[i][8]))
                orComboPayment[i].set(result[i][9])
                self.updateORLine(orCollection[i], TEXTVAR_ENTITYTYPE, orTax[i].get(), i)
            self.updateORTotals()
        
            self.clearSOAEntryLines()
            for i in range(len(resultbook)):
                if resultbook[i][1] != 0:
                    entryCode[i].set(resultbook[i][0])
                    self.populateChartFields(entryCode[i], entryTitle[i])
                    entryAmount[i].set(self.validateAmount2(resultbook[i][1]))
                    entryDrCr[i].set(resultbook[i][2])
                    entryRemarks[i].set(resultbook[i][3])
                    
            if result[0][12] == "Yes": #isApproved
                # self.disableSOAWidgets()
                BUTTON_SUBMIT.config(text = "UPDATE", state = DISABLED, cursor = "arrow")
                BUTTON_APPROVE.config(state = DISABLED, cursor = "arrow")
                # if self.returnIfSOAVoid(TEXTVAR_SOANUMBER) == True:
                #     BUTTON_VOID.config(state = DISABLED, cursor = "arrow")
                # else:
                #     BUTTON_VOID.config(state = NORMAL, cursor = "hand2")
            else:
                BUTTON_SUBMIT.config(text = "UPDATE", state = NORMAL, cursor = "hand2")
                BUTTON_APPROVE.config(state = NORMAL, cursor = "hand2")
            BUTTON_PRINT.config(state = NORMAL, cursor = "hand2")
            
            if self.returnUserName(USER, 3) == "ASD":
                pass
                # BUTTON_SAVE.config(state = NORMAL, cursor = "hand2")
            if self.returnAccess(USER, 16) == 0 and self.returnAccess(USER, 17) == 0:
                BUTTON_APPROVE.config(state = DISABLED, cursor = "arrow")
            if result[0][11] != USER:
                BUTTON_SUBMIT.config(state = DISABLED, cursor = "arrow")

    def showSOASelection(self, var):
        if TEXTVAR_CLIENTNAME.get() != "":
            global TOP_SOASELECTION
            TOP_SOASELECTION = Toplevel()
            TOP_SOASELECTION.title("Select SOA")
            TOP_SOASELECTION.iconbitmap(PATH_ICON + "icon.ico")
            TOP_SOASELECTION.geometry("600x300+200+200")
            TOP_SOASELECTION.resizable(height = False, width = False)
            TOP_SOASELECTION.grab_set()

            FRAME1 = Frame(TOP_SOASELECTION)
            FRAME1.pack(fill = "x")
            
            FRAME2 = Frame(TOP_SOASELECTION)
            FRAME2.pack(fill = "x")
            
            FRAME3 = Frame(TOP_SOASELECTION)
            FRAME3.pack(fill = "x", side = RIGHT)
            
            select = "SELECT soaNumber, tblclients.clientName, amount, transaction, tblreceivables.taxType, tblclients.parentCode FROM tblreceivables INNER JOIN tblclients ON tblreceivables.clientCode = tblclients.clientCode WHERE tblclients.parentCode = %s AND void = 'No'" #AND isApproved = 'Yes'
            db.commit()
            cursor.execute(select, [var.get()])
            result = cursor.fetchall()
            if result:
                LABEL_CHECK = Label(FRAME1, text = "[ ]", width = 2, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
                LABEL_CHECK.grid(column = 0, row = 0, ipadx = TOP_PADX + 2)
                
                LABEL_CODE = Label(FRAME1, text = "SOA No.", width = 12, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
                LABEL_CODE.grid(column = 1, row = 0)

                LABEL_TITLE = Label(FRAME1, text = "Client Name", width = 25, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
                LABEL_TITLE.grid(column = 2, row = 0)

                LABEL_CENTER = Label(FRAME1, text = "Amount", width = 12, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
                LABEL_CENTER.grid(column = 3, row = 0)

                LABEL_GROSS = Label(FRAME1, text = "Transaction", width = 15, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
                LABEL_GROSS.grid(column = 4, row = 0)
                
                LABEL_GROSS = Label(FRAME1, text = "Tax Type", width = 10, font = APP_FONT, bg = LABEL_BG, relief = RIDGE)
                LABEL_GROSS.grid(column = 5, row = 0)
                
                global selectFrame, selectCheck, selectSOA, selectClient, selectAmount, selectTransaction, selectTax
                selectFrame, selectCheck, selectSOA, selectClient, selectAmount, selectTransaction, selectTax = [], [], [], [], [], [], []
                
                self.createScrollFrame2(FRAME2, 250, 575, 20, 0, 0)
                for i in range(len(result)):
                    self.showSOASelectionLines(SCROLLABLE_FRAME2, i)
                    selectSOA[i].set(result[i][0])
                    selectClient[i].set(result[i][1])
                    selectAmount[i].set(self.validateAmount2(result[i][2]))
                    selectTransaction[i].set(result[i][3])
                    selectTax[i].set(result[i][4])
                
                BUTTON_SELECT = Button(FRAME3, text = "Import Selection", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, cursor = "hand2", command = self.importSOASelection)
                BUTTON_SELECT.grid(column = 0, row = 0, padx = TOP_PADX)
            else:
                messagebox.showerror("Select SOA", "No remaining SOA found for this Client!")
                TOP_SOASELECTION.grab_release()
                TOP_SOASELECTION.destroy()
        else:
            messagebox.showerror("Select SOA", "Client Name should be declared first!")
            TOP_COLLECTIONS.focus()
    
    def importSOASelection(self):
        selected = []
        for i in range(len(selectCheck)):
            if selectCheck[i].get() == 1:
                selected.append(selectSOA[i])
        count = 0
        for i in selected:
            orSOANumber[count].set(selectSOA[selectSOA.index(i)].get())
            orClientName[count].set(selectClient[selectSOA.index(i)].get())
            orReceivable[count].set(selectAmount[selectSOA.index(i)].get())
            orTransaction[count].set(selectTransaction[selectSOA.index(i)].get())
            orTax[count].set(selectTax[selectSOA.index(i)].get())
            orPayment[count].set("Full")
            orCollection[count].set(self.computeFullCollection(selectAmount[selectSOA.index(i)].get(), TEXTVAR_ENTITYTYPE, selectTax[selectSOA.index(i)].get()))
            self.updateORLine(orCollection[count], TEXTVAR_ENTITYTYPE, orTax[count].get(), count)
            count += 1
        TOP_SOASELECTION.grab_release()
        TOP_SOASELECTION.destroy()
        
    def showSOASelectionLines(self, frame, rown):
        FRAME_SELECTLINE = Frame(frame)
        FRAME_SELECTLINE.grid(column = 0, row = rown, sticky = W)
        selectFrame.append(FRAME_SELECTLINE)
        
        CHECKVAR_SELECT = IntVar()
        CHECKBOX_SELECT = Checkbutton(FRAME_SELECTLINE, variable = CHECKVAR_SELECT, font = BUTTON_FONT, bg = BUTTON_BG)
        CHECKBOX_SELECT.grid(column = 0, row = 0)
        selectCheck.append(CHECKVAR_SELECT)
        
        global TEXTVAR_SELECTSOANUMBER
        TEXTVAR_SELECTSOANUMBER = StringVar()
        ENTRY_SELECTSOANUMBER = Entry(FRAME_SELECTLINE, textvariable = TEXTVAR_SELECTSOANUMBER, font = APP_FONT, width = 12, justify = CENTER)
        ENTRY_SELECTSOANUMBER.grid(column = 1, row = 0, sticky = W)
        selectSOA.append(TEXTVAR_SELECTSOANUMBER)
        
        global TEXTVAR_SELECTCLIENTCODE
        TEXTVAR_SELECTCLIENTCODE = StringVar()
        ENTRY_SELECTCLIENTCODE = Entry(FRAME_SELECTLINE, textvariable = TEXTVAR_SELECTCLIENTCODE, font = APP_FONT, width = 26)
        ENTRY_SELECTCLIENTCODE.grid(column = 2, row = 0, sticky = W)
        selectClient.append(TEXTVAR_SELECTCLIENTCODE)
        
        global TEXTVAR_SELECTAMOUNT
        TEXTVAR_SELECTAMOUNT = StringVar()
        ENTRY_SELECTAMOUNT = Entry(FRAME_SELECTLINE, textvariable = TEXTVAR_SELECTAMOUNT, font = APP_FONT, width = 12, justify = RIGHT)
        ENTRY_SELECTAMOUNT.grid(column = 3, row = 0, sticky = W)
        selectAmount.append(TEXTVAR_SELECTAMOUNT)
        
        global TEXTVAR_SELECTTRANSACTION
        TEXTVAR_SELECTTRANSACTION = StringVar()
        ENTRY_SELECTTRANSACTION = Entry(FRAME_SELECTLINE, textvariable = TEXTVAR_SELECTTRANSACTION, font = APP_FONT, width = 15, justify = CENTER)
        ENTRY_SELECTTRANSACTION.grid(column = 4, row = 0, sticky = W)
        selectTransaction.append(TEXTVAR_SELECTTRANSACTION)
        
        global TEXTVAR_SELECTTAXTYPE
        TEXTVAR_SELECTTAXTYPE = StringVar()
        ENTRY_SELECTTAXTYPE = Entry(FRAME_SELECTLINE, textvariable = TEXTVAR_SELECTTAXTYPE, font = APP_FONT, width = 10, justify = CENTER)
        ENTRY_SELECTTAXTYPE.grid(column = 5, row = 0, sticky = W)
        selectTax.append(TEXTVAR_SELECTTAXTYPE)

    def showImportOR(self):
        global TOP_IMPORT
        TOP_IMPORT = Toplevel()
        TOP_IMPORT.title("Import - OR")
        TOP_IMPORT.iconbitmap(PATH_ICON + "icon.ico")
        TOP_IMPORT.geometry("350x325+550+100")
        TOP_IMPORT.resizable(height = False, width = False)
        TOP_IMPORT.grab_set()
        
        LABEL_FOLDER = Label(TOP_IMPORT, text = "File", font = APP_FONT)
        LABEL_FOLDER.grid(column = 0, row = 0, pady = TOP_PADY)
        
        global TEXTVAR_FILE
        TEXTVAR_FILE = StringVar()
        ENTRY_FILE = Entry(TOP_IMPORT, textvariable = TEXTVAR_FILE, font = APP_FONT, width = 40, state = "readonly")
        ENTRY_FILE.grid(column = 1, row = 0, pady = TOP_PADY)

        BUTTON_FOLDER = Button(TOP_IMPORT, text = "...", font = APP_FONT, command = lambda: self.getFileAddress2("Excel files", "*.xlsx"))
        BUTTON_FOLDER.grid(column = 1, row = 0, sticky = E, pady = TOP_PADY)
        
        BUTTON_TEMPLATE = Button(TOP_IMPORT, text = "Get template", font = APP_FONT, command = lambda: self.exportTemplate("tblor.xlsx"))
        BUTTON_TEMPLATE.grid(column = 1, row = 1, sticky = E, pady = TOP_PADY)

        BUTTON_IMPORT = Button(TOP_IMPORT, text = "Import", font = APP_FONT, command = self.importORSummary)
        BUTTON_IMPORT.grid(column = 1, row = 2, sticky = E, pady = TOP_PADY)

        global PROGRESS_TOPBAR
        PROGRESS_TOPBAR = tk.Progressbar(TOP_IMPORT, orient = HORIZONTAL, length = 200, mode = "determinate")
    
    def importORSummary(self):
        insert = """INSERT INTO tblcollections (
                            orNumber, orDate, glDate, mode, checkNumber,
                            soaNumber, clientCode, particulars, amount, payment,
                            remarks, isApproved, isVoid, encoder, encoded,
                            orType
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
        # insertbook = """INSERT INTO tblcashreceiptbook (glDate, clientCode, orNumber, chartCode, amount,
        #                     side, remarks, isPosted, isVoid, poster,
        #                     posted, modifier, modified, source) VALUES 
        #                     (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
        excelerrors, validlines, duplicates = [], [], []
        wb = load_workbook(TEXTVAR_FILE.get())
        st = wb.active
        PROGRESS_TOPBAR.grid(column = 1, row = 3, padx = TOP_PADX + 10, sticky = E)
        for i in range(st.max_row-1):
            # try:
            if st["A" + str(i+2)].value != None and float(st["J" + str(i+2)].value) > 0:
                if self.returnORDuplicate(st["A" + str(i+2)].value) == False:
                    orNumber = int(st["A" + str(i+2)].value)
                    orDate = st["B" + str(i+2)].value
                    orType = st["C" + str(i+2)].value
                    glDate = st["D" + str(i+2)].value
                    mode = st["E" + str(i+2)].value
                    checkNumber = st["F" + str(i+2)].value
                    soaNumber = int(st["G" + str(i+2)].value)
                    clientCode = st["H" + str(i+2)].value
                    particulars = st["I" + str(i+2)].value
                    amount = float(st["J" + str(i+2)].value)
                    payment = st["K" + str(i+2)].value
                    remarks = st["L" + str(i+2)].value
                    validlines.append([
                        orNumber, orDate, glDate, mode, checkNumber,
                        soaNumber, clientCode, particulars, amount, payment,
                        remarks, "Yes", "No", USER, datetime.datetime.now(),
                        orType
                    ])
                else:
                    duplicates.append(st["A" + str(i+2)].value)
            else:
                excelerrors.append("Invalid OR Number and/or invalid amount in row" + str(i+2))
            # except Exception as e:
            #     excelerrors.append(e + str(i+2))
            PROGRESS_TOPBAR["value"] = round((i/st.max_row)*100, 0)
            TOP_IMPORT.update()
        PROGRESS_TOPBAR.grid_remove()
                    
        if len(excelerrors) != 0:
            messagebox.showerror("Excel Error", excelerrors)
        else:
            if len(validlines) != 0:
                ask = messagebox.askyesno("Import", "Are you sure?")
                if ask:
                    count = 0
                    PROGRESS_TOPBAR.grid(column = 1, row = 3, padx = TOP_PADX + 10, sticky = E)
                    for i in validlines:
                        cursor.execute(insert, i)
                        PROGRESS_TOPBAR["value"] = round((count/len(validlines))*100, 0)
                        TOP_IMPORT.update()
                        count += 1
                    PROGRESS_TOPBAR.grid_remove()
                    db.commit()
                    if len(duplicates) != 0:
                        messagebox.showinfo("Import", "OR import completed! \n Duplicate ORs detected" + str(duplicates))
                    else:
                        messagebox.showinfo("Import", "OR import completed! No duplicates detected")
            else:
                if len(duplicates) != 0:
                    messagebox.showerror("Import", "No records to import! \n Duplicate ORs detected" + str(duplicates))
                else:
                    messagebox.showerror("Import", "No records to import!")
                
    def returnORDuplicate(self, var):
        find = "SELECT orNumber FROM tblcollections WHERE orNumber = %s LIMIT 1"
        cursor.execute(find, [int(var)])
        result = cursor.fetchone()
        if result:
            return True
        else:
            return False

    def computeFullCollection(self, gross, entity, tax):
        if entity.get() == "Government":
            if tax[:2] == "WV":
                cvat = self.validateAmount2((self.returnFloatAmount(gross)/1.12)*.05)
            else:
                cvat = self.validateAmount2(0)
        else:
            cvat = self.validateAmount2(0)
        if tax == "NV-00":
            ewt = self.validateAmount2(0)
            return self.validateAmount2(self.returnFloatAmount(gross)-self.returnFloatAmount(cvat))
        elif tax == "NV-02":
            ewt = self.validateAmount2(self.returnFloatAmount(gross)*.02)
            return self.validateAmount2(self.returnFloatAmount(gross)-self.returnFloatAmount(ewt)-self.returnFloatAmount(cvat))
        elif tax == "WV-02":
            ewt = self.validateAmount2((self.returnFloatAmount(gross)/1.12)*.02)
            return self.validateAmount2(self.returnFloatAmount(gross)-self.returnFloatAmount(ewt)-self.returnFloatAmount(cvat))
        else:
            ewt = self.validateAmount2(0)
            return self.validateAmount2(0)

    def validateORNumber(self, var, *args):
        try:
            find = "SELECT orNumber FROM tblcollections WHERE orNumber = %s"
            db.commit()
            cursor.execute(find, [int(var.get())])
            result = cursor.fetchall()
            if result:
                var.set("")
                messagebox.showerror("Add Collection", "OR number already used!")
                TOP_COLLECTIONS.focus()
            else:
                try:
                    var.set(str(int(var.get())).zfill(8))
                except:
                    var.set("")
                    messagebox.showerror("Add Collection", "Please enter a valid OR number")
                    TOP_COLLECTIONS.focus()
        except:
            var.set("")

    def getSOADetails(self, var, i):
        select = "SELECT clientCode, amount, transaction, taxType FROM tblreceivables WHERE soaNumber = %s"
        db.commit()
        cursor.execute(select, [var])
        result = cursor.fetchall()
        if result:
            totalSOAAmount = []
            if self.returnParentCode(result[0][0]) == TEXTVAR_CLIENTCODE.get():
                for x in result:
                    totalSOAAmount.append(x[1])
                orClientName[i].set(self.returnClientName(result[0][0]))
                orReceivable[i].set(self.validateAmount2(sum(totalSOAAmount)))
                orTransaction[i].set(result[0][2])
                orTax[i].set(result[0][3])
        else:
            orClientName[0].set("None")
            orReceivable[0].set(self.validateAmount2(0))
            orTransaction[0].set("None")
            orTax[0].set("NV-00")

    def returnClientName(self, var):
        cursor.execute(f"SELECT clientName FROM tblclients WHERE clientCode = '{var}' LIMIT 1")
        result = cursor.fetchone()
        return result

    def disableCollectionSOANumberFields(self, switch):
        for i in orFrame:
            i.winfo_children()[0].config(state = switch)

    def returnParentCode(self, var):
        cursor.execute(f"SELECT parentCode FROM tblclients WHERE clientCode = '{var}' LIMIT 1")
        result = cursor.fetchone()
        return result[0]

    def returnTotalORAmount(self, var):
        find = "SELECT SUM(amount) FROM tblcollections WHERE orNumber = %s LIMIT 1"
        cursor.execute(find, [var])
        result = cursor.fetchone()
        return format(result[0], ",.2f")

    def updateORLine(self, col, ent, tax, i):
        self.validateAmount(col)
        if orSOANumber[i].get() != "":
            if tax[:2] == "WV":
                if ent.get() == "Private":
                    orVAT[i].set(self.validateAmount2((self.returnFloatAmount(col.get()) / 1.10) * .12))
                    orEWT[i].set(self.validateAmount2((self.returnFloatAmount(col.get()) / 1.10) * .02))
                    orCVAT[i].set(self.validateAmount2(0))
                    orBalance[i].set(self.validateAmount2(self.returnFloatAmount(orReceivable[i].get())-(self.returnFloatAmount(col.get()) / 1.10)*1.12))
                else:
                    orVAT[i].set(self.validateAmount2((self.returnFloatAmount(col.get()) / 1.05) * .12))
                    orEWT[i].set(self.validateAmount2((self.returnFloatAmount(col.get()) / 1.05) * .02))
                    orCVAT[i].set(self.validateAmount2((self.returnFloatAmount(col.get()) / 1.05) * .05))
                    orBalance[i].set(self.validateAmount2(self.returnFloatAmount(orReceivable[i].get())-(self.returnFloatAmount(col.get()) / 1.05)*1.12))
            else:
                orVAT[i].set(self.validateAmount2(0))
                if tax[3:] == "00":
                    orEWT[i].set(self.validateAmount2(0))
                    orCVAT[i].set(self.validateAmount2(0))
                    orBalance[i].set(self.validateAmount2(self.returnFloatAmount(orReceivable[i].get())-self.returnFloatAmount(col.get())))
                else: #NV-02
                    orVAT[i].set(self.validateAmount2(0))
                    if ent.get() == "Private":
                        orEWT[i].set(self.validateAmount2((self.returnFloatAmount(col.get()) / 0.98) * .02))
                        orCVAT[i].set(self.validateAmount2(0))
                        orBalance[i].set(self.validateAmount2(self.returnFloatAmount(orReceivable[i].get())-((self.returnFloatAmount(col.get()) / 0.98))))
                    else:
                        orEWT[i].set(self.validateAmount2((self.returnFloatAmount(col.get()) / 0.93) * .02))
                        orCVAT[i].set(self.validateAmount2(0))
                        orBalance[i].set(self.validateAmount2(self.returnFloatAmount(orReceivable[i].get())-((self.returnFloatAmount(col.get()) / 0.93))))
            
            if orPayment[i].get() == "Full" and self.returnFloatAmount(orBalance[i].get()) > 0:
                addtocollection = self.returnFloatAmount(orReceivable[i].get()) - (self.returnFloatAmount(col.get()) + self.returnFloatAmount(orEWT[i].get()) + self.returnFloatAmount(orCVAT[i].get()))
                if tax[:2] == "WV":
                    orINC[i].set(self.validateAmount2(-addtocollection/1.12))
                    orBalance[i].set(self.validateAmount2(0))
                    orOutput[i] = (addtocollection/1.12)*.12
                else:
                    orINC[i].set(self.validateAmount2(addtocollection))
                    orBalance[i].set(self.validateAmount2(0))
                    orOutput[i] = 0
            else:
                orINC[i].set(self.validateAmount2(0))
                orOutput[i] = 0

        else:
            orVAT[i].set(self.validateAmount2(0))
            orEWT[i].set(self.validateAmount2(0))
            orCVAT[i].set(self.validateAmount2(0))
            orBalance[i].set(self.validateAmount2(0))
            orINC[i].set(self.validateAmount2(0))
            orOutput[i] = 0
            
        self.updateORTotals()

    def updateORTotals(self):
        tcollection, tvat, tewt, tcvat, tinc = [], [], [], [], []
        for i in range(len(orCollection)):
            try:
                tcollection.append(self.returnFloatAmount(orCollection[i].get()))
                tvat.append(self.returnFloatAmount(orVAT[i].get()))
                tewt.append(self.returnFloatAmount(orEWT[i].get()))
                tcvat.append(self.returnFloatAmount(orCVAT[i].get()))
                tinc.append(self.returnFloatAmount(orINC[i].get()))
            except:
                pass
        TEXTVAR_TGROSS.set(self.validateAmount2(sum(tcollection)))
        TEXTVAR_TVAT.set(self.validateAmount2(sum(tvat)))
        TEXTVAR_TEWT.set(self.validateAmount2(sum(tewt)))
        TEXTVAR_TCVAT.set(self.validateAmount2(sum(tcvat)))
        TEXTVAR_TNET.set(self.validateAmount2(sum(tinc)))
        
        self.updateOREntries()
        self.updateSOAEntriesTotals()

    def updateOREntries(self):
        self.clearSOAEntryLines()
        count = 0
        if self.returnFloatAmount(TEXTVAR_TGROSS.get()) > 0:
            entryCode[count].set("115020")
            self.populateChartFields(entryCode[count], entryTitle[count])
            entryAmount[count].set(TEXTVAR_TGROSS.get())
            entryDrCr[count].set("Debit")
            count += 1
        
        if sum(orOutput) > 0:
            entryCode[count].set("213140")
            self.populateChartFields(entryCode[count], entryTitle[count])
            entryAmount[count].set(self.validateAmount2(sum(orOutput)))
            entryDrCr[count].set("Debit")
            count += 1
            
        if self.returnFloatAmount(TEXTVAR_TCVAT.get()) > 0:
            entryCode[count].set("117030")
            self.populateChartFields(entryCode[count], entryTitle[count])
            entryAmount[count].set(TEXTVAR_TCVAT.get())
            entryDrCr[count].set("Debit")
            count += 1
            
        if self.returnFloatAmount(TEXTVAR_TVAT.get())> 0:
            entryCode[count].set("213140")
            self.populateChartFields(entryCode[count], entryTitle[count])
            entryAmount[count].set(self.validateAmount2(self.returnFloatAmount(TEXTVAR_TVAT.get())))
            entryDrCr[count].set("Debit")
            count += 1
        
        if self.returnFloatAmount(TEXTVAR_TNET.get()) < 0:
            total = []
            for i in range(len(orSOANumber)):
                if orSOANumber[i].get() != "":
                    if self.getSOACategory(orSOANumber[i].get())[0] == "Administrative" and self.returnFloatAmount(orINC[i].get()) != 0:
                        account = "511010"
                        total.append(self.returnFloatAmount(orINC[i].get()))
                    if self.getSOACategory(orSOANumber[i].get())[0] == "Maintenance" and self.returnFloatAmount(orINC[i].get()) != 0:
                        account = "512010"
                        total.append(self.returnFloatAmount(orINC[i].get()))
            entryCode[count].set(account)
            self.populateChartFields(entryCode[count], entryTitle[count])
            entryAmount[count].set(self.validateAmount2(-sum(total)))
            entryDrCr[count].set("Debit")
            count += 1
            
        if self.returnFloatAmount(TEXTVAR_TEWT.get()) > 0:
            entryCode[count].set("117010")
            self.populateChartFields(entryCode[count], entryTitle[count])
            entryAmount[count].set(TEXTVAR_TEWT.get())
            entryDrCr[count].set("Debit")
            count += 1
            
        if self.returnFloatAmount(TEXTVAR_TVAT.get()) > 0:
            entryCode[count].set("213040")
            self.populateChartFields(entryCode[count], entryTitle[count])
            entryAmount[count].set(TEXTVAR_TVAT.get())
            entryDrCr[count].set("Credit")
            count += 1
            
        if self.returnFloatAmount(TEXTVAR_TNET.get()) > 0:
            total = []
            for i in range(len(orSOANumber)):
                if orSOANumber[i].get() != "":
                    if self.getSOACategory(orSOANumber[i].get())[0] == "Administrative" and self.returnFloatAmount(orINC[i].get()) != 0:
                        account = "511010"
                        total.append(self.returnFloatAmount(orINC[i].get()))
                    if self.getSOACategory(orSOANumber[i].get())[0] == "Maintenance" and self.returnFloatAmount(orINC[i].get()) != 0:
                        account = "512010"
                        total.append(self.returnFloatAmount(orINC[i].get()))
            entryCode[count].set(account)
            self.populateChartFields(entryCode[count], entryTitle[count])
            entryAmount[count].set(self.validateAmount2(sum(total)))
            entryDrCr[count].set("Credit")
            count += 1
        
        totalar = []
        for i in orReceivable:
            if self.returnFloatAmount(i.get()) > 0:
                totalar.append(self.returnFloatAmount(i.get()))
                
        totalbal = []
        for i in orBalance:
            if self.returnFloatAmount(i.get()) > 0:
                totalbal.append(self.returnFloatAmount(i.get()))
                
        if sum(totalar) + sum(totalbal) > 0:
            entryCode[count].set("115010")
            self.populateChartFields(entryCode[count], entryTitle[count])
            entryAmount[count].set(self.validateAmount2(sum(totalar) - sum(totalbal)))
            entryDrCr[count].set("Credit")

    def getSOATaxType(self, var):
        cursor.execute(f"SELECT taxType FROM tblreceivables WHERE soaNumber = '{var}' LIMIT 1")
        result = cursor.fetchone()
        return result[0]

    def clearORLine(self, col, ent, num, i):
        if orSOANumber[i].get() != "" and self.returnFloatAmount(orCollection[i].get()) > 0:
            orSOANumber[i].set("")
            orClientName[i].set("")
            orReceivable[i].set("")
            orBalance[i].set("")
            orCollection[i].set(0)
            orPayment[i].set("")
            orRemarks[i].set("")
            self.updateORLine(col, ent, num, i)

    def getSOACategory(self, var):
        cursor.execute(f"SELECT category FROM tblreceivables WHERE soaNumber = '{var}' LIMIT 1")
        result = cursor.fetchone()
        if result:
            return result

    def searchCollection(self, var, *args):
        find = "SELECT orNumber, orDate, glDate, clientCode, particulars, amount, isApproved, isVoid, encoder, DATE(encoded), approver, DATE(approved) FROM tblcollections WHERE (orNumber LIKE %s OR soaNumber LIKE %s OR clientCode LIKE %s) AND orDate BETWEEN %s AND %s ORDER BY orNumber DESC"
        db.commit()
        cursor.execute(find, [f"%{var}%", f"%{var}%", f"%{var}%", CALENDAR_START.get_date(), CALENDAR_END.get_date()])
        result = cursor.fetchall()
        for i in TREE_COLLECTIONS.get_children():
            TREE_COLLECTIONS.delete(i)
        if result:
            PROGRESS_BAR.grid(column = 8, row = 0, padx = TOP_PADX + 10, sticky = E)
            count = 0
            ornumbers, skipped = [], []
            for i in result:
                if i[0] not in ornumbers:
                    if i[7] != "Yes":
                        if count % 2 == 0:
                            TREE_COLLECTIONS.insert("", "end", values = (str(i[0]).zfill(8),i[1],i[2],self.returnClientName(i[3]),i[4],self.returnTotalORAmount(i[0]),i[6],i[7],self.returnUserName(i[8], 0),i[9],self.returnUserName(i[10], 0),i[11]), tags = "evenrow")
                        else:
                            TREE_COLLECTIONS.insert("", "end", values = (str(i[0]).zfill(8),i[1],i[2],self.returnClientName(i[3]),i[4],self.returnTotalORAmount(i[0]),i[6],i[7],self.returnUserName(i[8], 0),i[9],self.returnUserName(i[10], 0),i[11]), tags = "oddrow")
                    else:
                        TREE_COLLECTIONS.insert("", "end", values = (str(i[0]).zfill(8),i[1],i[2],self.returnClientName(i[3]),i[4],self.returnTotalORAmount(i[0]),i[6],i[7],self.returnUserName(i[8], 0),i[9],self.returnUserName(i[10], 0),i[11]), tags = "void")
                    count += 1
                    ornumbers.append(i[0])
                    PROGRESS_BAR["value"] = round((count/(len(result)-len(skipped)))*100, 0)
                    SUB1_COLLECTIONS.update()
                else:
                    skipped.append(i[0])
            PROGRESS_BAR.grid_remove()
        else:
            messagebox.showerror("Collections", "No match found!")

    def refreshCollection(self, *args):
        for i in TREE_COLLECTIONS.get_children():
            TREE_COLLECTIONS.delete(i)
        db.commit()
        cursor.execute(f"SELECT orNumber, orDate, glDate, clientCode, particulars, amount, isApproved, isVoid, encoder, DATE(encoded), approver, DATE(approved) FROM tblcollections WHERE orDate BETWEEN '{CALENDAR_START.get_date()}' AND '{CALENDAR_END.get_date()}' ORDER BY orNumber DESC")
        result = cursor.fetchall()
        PROGRESS_BAR.grid(column = 8, row = 0, padx = TOP_PADX + 10, sticky = E)
        count = 0
        ornumbers, skipped = [], []
        for i in result:
            if i[0] not in ornumbers:
                if count % 2 == 0:
                    TREE_COLLECTIONS.insert("", "end", values = (str(i[0]).zfill(8),i[1],i[2],self.returnClientName(i[3]),i[4],self.returnTotalSOAAmount(i[0]),self.returnTotalSOABalance(i[0], self.returnTotalSOAAmount(i[0])),i[6],i[7],self.returnClientEntityType(i[3]),self.returnUserName(i[8], 0),i[9],self.returnUserName(i[10], 0),i[11]), tags = "evenrow")
                else:
                    TREE_COLLECTIONS.insert("", "end", values = (str(i[0]).zfill(8),i[1],i[2],self.returnClientName(i[3]),i[4],self.returnTotalSOAAmount(i[0]),self.returnTotalSOABalance(i[0], self.returnTotalSOAAmount(i[0])),i[6],i[7],self.returnClientEntityType(i[3]),self.returnUserName(i[8], 0),i[9],self.returnUserName(i[10], 0),i[11]), tags = "oddrow")
                count += 1
                ornumbers.append(i[0])
                PROGRESS_BAR["value"] = round((count/(len(result)-len(skipped)))*100, 0)
                SUB1_COLLECTIONS.update()
            else:
                skipped.append(i[0])
        PROGRESS_BAR.grid_remove()

### MENU_GSAD ###
    def showInventory(self, *args):
        self.clearWorkspace()
        FRAME_INVENTORY = LabelFrame(FRAME_4, text = "Inventory", font = APP_FONT)
        FRAME_INVENTORY.grid(column = 1, row = 0)

        SUB1_INVENTORY = Frame(FRAME_INVENTORY)
        SUB1_INVENTORY.grid(column = 0, row = 0, sticky = W, pady = SUBMENU_PADY)

        SUB2_INVENTORY = Frame(FRAME_INVENTORY)
        SUB2_INVENTORY.grid(column = 0, row = 1, sticky = W)

        LABEL_SEARCH = Label(SUB1_INVENTORY, text = "Search", width = 5, font = APP_FONT, anchor = W)
        LABEL_SEARCH.grid(column = 0, row = 0, pady = SUBMENU_PADY, sticky = W)

        TEXTVAR_SEARCH = StringVar()
        ENTRY_SEARCH = Entry(SUB1_INVENTORY, textvariable = TEXTVAR_SEARCH, width = 25, font = APP_FONT)
        ENTRY_SEARCH.grid(column = 1, row = 0, padx = MENU_PADX, sticky = W)
        ENTRY_SEARCH.bind("<Return>", lambda e: self.searchInventory(TEXTVAR_SEARCH.get()))

        BUTTON_GO = Button(SUB1_INVENTORY, text = "GO", width = 5, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = lambda: self.searchInventory(TEXTVAR_SEARCH.get()))
        BUTTON_GO.grid(column = 2, row = 0, padx = MENU_PADX)

        BUTTON_REFRESH = Button(SUB1_INVENTORY, text = "REFRESH", width = 10, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = self.refreshInventory)
        BUTTON_REFRESH.grid(column = 3, row = 0, padx = MENU_PADX)

        BUTTON_ADD = Button(SUB1_INVENTORY, text = "ADD", width = 5, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = self.showAddEditInventory)
        BUTTON_ADD.grid(column = 4, row = 0, padx = MENU_PADX)
        
        global TEXTVAR_CATEGORY
        TEXTVAR_CATEGORY = StringVar()
        COMBO_CATEGORY = tk.Combobox(SUB1_INVENTORY, values = ["All", "Corporate Giveaways", "Equipment - Janitorial", "Equipment - Office", "Furniture - Office", "Service Vehicle", "Supplies - Janitorial", "Supplies - Office", "Supplies - Medicine", "Uniform - Clerical", "Uniform - Janitorial", "Uniform - DBPSC", "Others"], textvariable = TEXTVAR_CATEGORY, font = APP_FONT, width = 25, state = "readonly")
        COMBO_CATEGORY.grid(column = 5, row = 0, sticky = E, padx = TOP_PADX)
        COMBO_CATEGORY.bind("<<ComboboxSelected>>", lambda e: self.searchInventory(TEXTVAR_SEARCH.get()))
        TEXTVAR_CATEGORY.set("All")

        global TREE_INVENTORY
        TREE_INVENTORY = tk.Treeview(SUB2_INVENTORY, height = 28, selectmode = "browse")
        TREE_INVENTORY["columns"] = ("Code", "Name", "Description", "Tax", "Stock", "U.O.M.", "Cost", "Status", "Category", "Asset Code", "Expense Code")
        TREE_INVENTORY.column("#0", width = 0, minwidth = 0, stretch = True)
        TREE_INVENTORY.column("Code", anchor = W, minwidth = 50, width = 100)
        TREE_INVENTORY.column("Name", anchor = W, minwidth = 50, width = 225)
        TREE_INVENTORY.column("Description", anchor = W, minwidth = 50, width = 175)
        TREE_INVENTORY.column("Tax", anchor = W, minwidth = 50, width = 75)
        TREE_INVENTORY.column("Stock", anchor = W, minwidth = 50, width = 75)
        TREE_INVENTORY.column("U.O.M.", anchor = W, minwidth = 50, width = 75)
        TREE_INVENTORY.column("Cost", anchor = E, minwidth = 50, width = 75)
        TREE_INVENTORY.column("Status", anchor = W, minwidth = 50, width = 75)
        TREE_INVENTORY.column("Category", anchor = W, minwidth = 50, width = 75)
        TREE_INVENTORY.column("Asset Code", anchor = W, minwidth = 50, width = 75)
        TREE_INVENTORY.column("Expense Code", anchor = W, minwidth = 50, width = 75)
        
        TREE_INVENTORY.heading("#0", text = "", anchor = W)
        TREE_INVENTORY.heading("Code", text = "Code", anchor = N)
        TREE_INVENTORY.heading("Name", text = "Name", anchor = N)
        TREE_INVENTORY.heading("Description", text = "Description", anchor = N)
        TREE_INVENTORY.heading("Tax", text = "Tax", anchor = N)
        TREE_INVENTORY.heading("Stock", text = "Stock", anchor = N)
        TREE_INVENTORY.heading("U.O.M.", text = "U.O.M.", anchor = N)
        TREE_INVENTORY.heading("Cost", text = "Cost", anchor = N)
        TREE_INVENTORY.heading("Status", text = "Status", anchor = N)
        TREE_INVENTORY.heading("Category", text = "Category", anchor = N)
        TREE_INVENTORY.heading("Asset Code", text = "Asset Code", anchor = N)
        TREE_INVENTORY.heading("Expense Code", text = "Expense Code", anchor = N)

        POPUP_INVENTORY = Menu(TREE_INVENTORY, tearoff = 0)
        POPUP_INVENTORY.add_command(command = self.editInventory, label = "Edit")
        # POPUP_INVENTORY.add_command(command = self.deleteInventory, label = "Delete")
        TREE_INVENTORY.bind("<Button-3>", lambda e: self.popupMenu(TREE_INVENTORY, POPUP_INVENTORY, e))
        TREE_INVENTORY.bind("<Double-1>", self.editInventory)

        global STYLE_INVENTORY
        STYLE_INVENTORY = tk.Style()
        STYLE_INVENTORY.map("Treeview", foreground = self.fixedMap("foreground", STYLE_INVENTORY), background = self.fixedMap("background", STYLE_INVENTORY))

        TREE_INVENTORY.tag_configure("oddrow", background = None)
        TREE_INVENTORY.tag_configure("evenrow", background = TREE_TAG_EVENROW)
        db.commit()
        cursor.execute("SELECT code, name, description, tax, quantity, uom, cost, status, category, assetcode, expensecode FROM tblinventory")
        result = cursor.fetchall()
        
        count = 0
        for i in result:
            if count % 2 == 0:
                TREE_INVENTORY.insert("", "end", values = (i[0],i[1],i[2],i[3],i[4],i[5],self.validateAmount2(i[6]),i[7],i[8],i[9],i[10]), tags = "evenrow")
            else:
                TREE_INVENTORY.insert("", "end", values = (i[0],i[1],i[2],i[3],i[4],i[5],self.validateAmount2(i[6]),i[7],i[8],i[9],i[10]), tags = "oddrow")
            count += 1

        YSCROLLBAR = tk.Scrollbar(SUB2_INVENTORY, orient = "vertical", command = TREE_INVENTORY.yview)
        YSCROLLBAR.pack(side = "right", fill = "y")

        XSCROLLBAR = tk.Scrollbar(SUB2_INVENTORY, orient = "horizontal", command = TREE_INVENTORY.xview)
        
        TREE_INVENTORY.configure(yscrollcommand = YSCROLLBAR.set, xscrollcommand = XSCROLLBAR.set) 
        TREE_INVENTORY.pack()
        XSCROLLBAR.pack(fill ="x")

    def showPurchaseOrder(self, *args):
        self.clearWorkspace()
        FRAME_PO = LabelFrame(FRAME_4, text = "Purchase Order", font = APP_FONT)
        FRAME_PO.grid(column = 1, row = 0)

        SUB1_PO = Frame(FRAME_PO)
        SUB1_PO.grid(column = 0, row = 0, sticky = W, pady = SUBMENU_PADY)

        SUB2_PO = Frame(FRAME_PO)
        SUB2_PO.grid(column = 0, row = 1, sticky = W)

        LABEL_SEARCH = Label(SUB1_PO, text = "Search", width = 5, font = APP_FONT, anchor = W)
        LABEL_SEARCH.grid(column = 0, row = 0, pady = SUBMENU_PADY, sticky = W)

        TEXTVAR_SEARCH = StringVar()
        ENTRY_SEARCH = Entry(SUB1_PO, textvariable = TEXTVAR_SEARCH, width = 25, font = APP_FONT)
        ENTRY_SEARCH.grid(column = 1, row = 0, padx = MENU_PADX, sticky = W)
        ENTRY_SEARCH.bind("<Return>", lambda e: self.searchPurchaseOrder(TEXTVAR_SEARCH.get()))

        BUTTON_GO = Button(SUB1_PO, text = "GO", width = 5, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = lambda: self.searchPurchaseOrder(TEXTVAR_SEARCH.get()))
        BUTTON_GO.grid(column = 2, row = 0, padx = MENU_PADX)

        BUTTON_REFRESH = Button(SUB1_PO, text = "REFRESH", width = 10, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = self.refreshPurchaseOrder)
        BUTTON_REFRESH.grid(column = 3, row = 0, padx = MENU_PADX)

        BUTTON_ADD = Button(SUB1_PO, text = "ADD", width = 5, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = self.showAddEditPurchaseOrder)
        BUTTON_ADD.grid(column = 4, row = 0, padx = MENU_PADX)
        
        if self.returnAccess(USER, 0) == 0:
            BUTTON_ADD.config(state = DISABLED, cursor = "arrow")

        global CALENDAR_START
        CALENDAR_START = DateEntry(SUB1_PO, firstweekday = "sunday", date_pattern = "yyyy-mm-dd", state = "readonly")
        CALENDAR_START.grid(column = 5, row = 0, sticky = W, padx = TOP_PADX + 10)
        
        global CALENDAR_END
        CALENDAR_END = DateEntry(SUB1_PO, firstweekday = "sunday", date_pattern = "yyyy-mm-dd", state = "readonly")
        CALENDAR_END.grid(column = 6, row = 0, sticky = W, padx = TOP_PADX + 10)

        global TREE_PO
        TREE_PO = tk.Treeview(SUB2_PO, height = 28, selectmode = "browse")
        TREE_PO["columns"] = ("PO No.", "Date", "Supplier Name", "Remarks", "Amount", "Delivery", "Status", "Encoder", "Date Encoded", "Approver", "Date Approved")
        TREE_PO.column("#0", width = 0, minwidth = 0, stretch = True)
        TREE_PO.column("PO No.", anchor = W, minwidth = 100, width = 75)
        TREE_PO.column("Date", anchor = W, minwidth = 100, width = 100)
        TREE_PO.column("Supplier Name", anchor = W, minwidth = 200, width = 150)
        TREE_PO.column("Remarks", anchor = W, minwidth = 100, width = 200)
        TREE_PO.column("Amount", anchor = E, minwidth = 100, width = 100)
        TREE_PO.column("Delivery", anchor = W, minwidth = 100, width = 100)
        TREE_PO.column("Status", anchor = W, minwidth = 100, width = 75)
        TREE_PO.column("Encoder", anchor = W, minwidth = 100, width = 75)
        TREE_PO.column("Date Encoded", anchor = W, minwidth = 100, width = 75)
        TREE_PO.column("Approver", anchor = W, minwidth = 100, width = 75)
        TREE_PO.column("Date Approved", anchor = W, minwidth = 100, width = 75)
        
        TREE_PO.heading("#0", text = "", anchor = W)
        TREE_PO.heading("PO No.", text = "PO No.", anchor = N)
        TREE_PO.heading("Date", text = "Date", anchor = N)
        TREE_PO.heading("Supplier Name", text = "Supplier Name", anchor = N)
        TREE_PO.heading("Remarks", text = "Remarks", anchor = N)
        TREE_PO.heading("Amount", text = "Amount", anchor = N)
        TREE_PO.heading("Delivery", text = "Delivery", anchor = N)
        TREE_PO.heading("Status", text = "Status", anchor = N)
        TREE_PO.heading("Encoder", text = "Encoder", anchor = N)
        TREE_PO.heading("Date Encoded", text = "Date Encoded", anchor = N)
        TREE_PO.heading("Approver", text = "Approver", anchor = N)
        TREE_PO.heading("Date Approved", text = "Date Approved", anchor = N)

        POPUP_PO = Menu(TREE_PO, tearoff = 0)
        POPUP_PO.add_command(command = self.editPurchaseOrder, label = "Edit")
        # POPUP_PO.add_command(command = self.deletePurchaseOrder, label = "Delete")
        TREE_PO.bind("<Button-3>", lambda e: self.popupMenu(TREE_PO, POPUP_PO, e))
        TREE_PO.bind("<Double-1>", self.editPurchaseOrder)

        global STYLE_PO
        STYLE_PO = tk.Style()
        STYLE_PO.map("Treeview", foreground = self.fixedMap("foreground", STYLE_PO), background = self.fixedMap("background", STYLE_PO))

        TREE_PO.tag_configure("oddrow", background = None)
        TREE_PO.tag_configure("evenrow", background = TREE_TAG_EVENROW)
        db.commit()
        cursor.execute(f"SELECT poNumber, DATE(date), supplierCode, remarks, inventoryCost*inventoryQuantity, delivery, status, encoder, DATE(encoded), approver, DATE(approved) FROM tblpurchaseorder WHERE date BETWEEN '{CALENDAR_START.get_date()}' AND '{CALENDAR_END.get_date()}' ORDER BY poNumber DESC")
        result = cursor.fetchall()
            
        YSCROLLBAR = tk.Scrollbar(SUB2_PO, orient = "vertical", command = TREE_PO.yview)
        YSCROLLBAR.pack(side = "right", fill = "y")

        XSCROLLBAR = tk.Scrollbar(SUB2_PO, orient = "horizontal", command = TREE_PO.xview)
        
        TREE_PO.configure(yscrollcommand = YSCROLLBAR.set, xscrollcommand = XSCROLLBAR.set) 
        TREE_PO.pack()
        XSCROLLBAR.pack(fill ="x")

        count = 0
        ponumbers = []
        for i in result:
            if i[0] not in ponumbers:
                if count % 2 == 0:
                    TREE_PO.insert("", "end", values = (str(i[0]).zfill(8),i[1],self.returnSupplierName(i[2]),i[3],self.returnTotalPOAmount(i[0]),self.returnPODeliverStatus(i[0]),i[6],self.returnUserName(i[7], 0),i[8],self.returnUserName(i[9], 0),i[10]), tags = "evenrow")
                else:
                    TREE_PO.insert("", "end", values = (str(i[0]).zfill(8),i[1],self.returnSupplierName(i[2]),i[3],self.returnTotalPOAmount(i[0]),self.returnPODeliverStatus(i[0]),i[6],self.returnUserName(i[7], 0),i[8],self.returnUserName(i[9], 0),i[10]), tags = "oddrow")
                count += 1
                ponumbers.append(i[0])
    
    def showReceivingReport(self):
        self.clearWorkspace()
        FRAME_RR = LabelFrame(FRAME_4, text = "Receiving Report", font = APP_FONT)
        FRAME_RR.grid(column = 1, row = 0)

        SUB1_RR = Frame(FRAME_RR)
        SUB1_RR.grid(column = 0, row = 0, sticky = W, pady = SUBMENU_PADY)

        SUB2_RR = Frame(FRAME_RR)
        SUB2_RR.grid(column = 0, row = 1, sticky = W)

        LABEL_SEARCH = Label(SUB1_RR, text = "Search", width = 5, font = APP_FONT, anchor = W)
        LABEL_SEARCH.grid(column = 0, row = 0, pady = SUBMENU_PADY, sticky = W)

        TEXTVAR_SEARCH = StringVar()
        ENTRY_SEARCH = Entry(SUB1_RR, textvariable = TEXTVAR_SEARCH, width = 25, font = APP_FONT)
        ENTRY_SEARCH.grid(column = 1, row = 0, padx = MENU_PADX, sticky = W)
        ENTRY_SEARCH.bind("<Return>", lambda e: self.searchReceivingReport(TEXTVAR_SEARCH.get()))

        BUTTON_GO = Button(SUB1_RR, text = "GO", width = 5, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = lambda: self.searchReceivingReport(TEXTVAR_SEARCH.get()))
        BUTTON_GO.grid(column = 2, row = 0, padx = MENU_PADX)

        BUTTON_REFRESH = Button(SUB1_RR, text = "REFRESH", width = 10, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = self.refreshReceivingReport)
        BUTTON_REFRESH.grid(column = 3, row = 0, padx = MENU_PADX)

        BUTTON_ADD = Button(SUB1_RR, text = "ADD", width = 5, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = self.showAddEditReceivingReport)
        BUTTON_ADD.grid(column = 4, row = 0, padx = MENU_PADX)
        
        if self.returnAccess(USER, 3) == 0:
            BUTTON_ADD.config(state = DISABLED, cursor = "arrow")

        global CALENDAR_START
        CALENDAR_START = DateEntry(SUB1_RR, firstweekday = "sunday", date_pattern = "yyyy-mm-dd", state = "readonly")
        CALENDAR_START.grid(column = 5, row = 0, sticky = W, padx = TOP_PADX + 10)
        
        global CALENDAR_END
        CALENDAR_END = DateEntry(SUB1_RR, firstweekday = "sunday", date_pattern = "yyyy-mm-dd", state = "readonly")
        CALENDAR_END.grid(column = 6, row = 0, sticky = W, padx = TOP_PADX + 10)

        global TREE_RR
        TREE_RR = tk.Treeview(SUB2_RR, height = 28, selectmode = "browse")
        TREE_RR["columns"] = ("RR No.", "PO No.","Date", "Supplier Name", "isApproved", "Encoder", "Encoded", "Approver", "Approved")
        TREE_RR.column("#0", width = 0, minwidth = 0, stretch = True)
        TREE_RR.column("RR No.", anchor = W, minwidth = 100, width = 75)
        TREE_RR.column("PO No.", anchor = W, minwidth = 100, width = 75)
        TREE_RR.column("Date", anchor = W, minwidth = 100, width = 100)
        TREE_RR.column("Supplier Name", anchor = W, minwidth = 150, width = 275)
        TREE_RR.column("isApproved", anchor = W, minwidth = 100, width = 75)
        TREE_RR.column("Encoder", anchor = W, minwidth = 125, width = 125)
        TREE_RR.column("Encoded", anchor = W, minwidth = 125, width = 125)
        TREE_RR.column("Approver", anchor = W, minwidth = 125, width = 125)
        TREE_RR.column("Approved", anchor = W, minwidth = 125, width = 125)
        
        TREE_RR.heading("#0", text = "", anchor = W)
        TREE_RR.heading("RR No.", text = "RR No.", anchor = N)
        TREE_RR.heading("PO No.", text = "PO No.", anchor = N)
        TREE_RR.heading("Date", text = "Date", anchor = N)
        TREE_RR.heading("Supplier Name", text = "Supplier Name", anchor = N)
        TREE_RR.heading("isApproved", text = "isApproved", anchor = N)
        TREE_RR.heading("Encoder", text = "Encoder", anchor = N)
        TREE_RR.heading("Encoded", text = "Encoded", anchor = N)
        TREE_RR.heading("Approver", text = "Approver", anchor = N)
        TREE_RR.heading("Approved", text = "Approved", anchor = N)

        POPUP_RR = Menu(TREE_RR, tearoff = 0)
        POPUP_RR.add_command(command = self.editReceivingReport, label = "Edit")
        TREE_RR.bind("<Button-3>", lambda e: self.popupMenu(TREE_RR, POPUP_RR, e))
        TREE_RR.bind("<Double-1>", self.editReceivingReport)

        global STYLE_RR
        STYLE_RR = tk.Style()
        STYLE_RR.map("Treeview", foreground = self.fixedMap("foreground", STYLE_RR), background = self.fixedMap("background", STYLE_RR))

        TREE_RR.tag_configure("oddrow", background = None)
        TREE_RR.tag_configure("evenrow", background = TREE_TAG_EVENROW)
        db.commit()
        cursor.execute(f"SELECT rrNumber, docDate, tblpurchaseorder.supplierCode, isApproved, tblreceivingreports.encoder, DATE(tblreceivingreports.encoded), tblreceivingreports.approver, DATE(tblreceivingreports.approved), tblreceivingreports.PONumber FROM tblreceivingreports INNER JOIN tblpurchaseorder ON tblreceivingreports.poNumber = tblpurchaseorder.poNumber WHERE docDate BETWEEN '{CALENDAR_START.get_date()}' AND '{CALENDAR_END.get_date()}' ORDER BY rrNumber DESC")
        result = cursor.fetchall()

        YSCROLLBAR = tk.Scrollbar(SUB2_RR, orient = "vertical", command = TREE_RR.yview)
        YSCROLLBAR.pack(side = "right", fill = "y")

        XSCROLLBAR = tk.Scrollbar(SUB2_RR, orient = "horizontal", command = TREE_RR.xview)
        
        TREE_RR.configure(yscrollcommand = YSCROLLBAR.set, xscrollcommand = XSCROLLBAR.set) 
        TREE_RR.pack()
        XSCROLLBAR.pack(fill ="x")

        count = 0
        rrnumbers = []
        for i in result:
            if i[0] not in rrnumbers:
                if count % 2 == 0:
                    TREE_RR.insert("", "end", values = (str(i[0]).zfill(8), str(i[8]).zfill(8),i[1],self.returnSupplierName(i[2]),i[3],self.returnUserName(i[4], 0),i[5],self.returnUserName(i[6], 0),i[7]), tags = "evenrow")
                else:
                    TREE_RR.insert("", "end", values = (str(i[0]).zfill(8), str(i[8]).zfill(8),i[1],self.returnSupplierName(i[2]),i[3],self.returnUserName(i[4], 0),i[5],self.returnUserName(i[6], 0),i[7]), tags = "oddrow")
                count += 1
                rrnumbers.append(i[0])
    
### MENU_GSAD_INVENTORY ###
    def showAddEditInventory(self, *args):
        global TOP_INVENTORY
        TOP_INVENTORY = Toplevel()
        TOP_INVENTORY.title("Add - Inventory")
        TOP_INVENTORY.iconbitmap(PATH_ICON + "icon.ico")
        TOP_INVENTORY.geometry("350x400+550+100")
        TOP_INVENTORY.resizable(height = False, width = False)
        TOP_INVENTORY.grab_set()
        
        global required
        required = []
        
        LABEL_CODE = Label(TOP_INVENTORY, text = "Code", font = APP_FONT)
        LABEL_CODE.grid(column = 0, row = 0, pady = TOP_PADY, sticky = E)

        LABEL_NAME = Label(TOP_INVENTORY, text = "Name", font = APP_FONT)
        LABEL_NAME.grid(column = 0, row = 1, pady = TOP_PADY, sticky = E)

        LABEL_DESC = Label(TOP_INVENTORY, text = "Description", font = APP_FONT)
        LABEL_DESC.grid(column = 0, row = 2, pady = TOP_PADY, sticky = E)
        
        LABEL_TAX = Label(TOP_INVENTORY, text = "Tax", font = APP_FONT)
        LABEL_TAX.grid(column = 0, row = 3, pady = TOP_PADY, sticky = E)

        LABEL_COST = Label(TOP_INVENTORY, text = "Cost", font = APP_FONT)
        LABEL_COST.grid(column = 0, row = 4, pady = TOP_PADY, sticky = E)

        LABEL_QUANTITY = Label(TOP_INVENTORY, text = "Quantity", font = APP_FONT)
        LABEL_QUANTITY.grid(column = 0, row = 5, pady = TOP_PADY, sticky = E)

        LABEL_UOM = Label(TOP_INVENTORY, text = "U.O.M.", font = APP_FONT)
        LABEL_UOM.grid(column = 0, row = 6, pady = TOP_PADY, sticky = E)

        LABEL_STATUS = Label(TOP_INVENTORY, text = "Status", font = APP_FONT)
        LABEL_STATUS.grid(column = 0, row = 7, pady = TOP_PADY, sticky = E)
        
        LABEL_CATEGORY = Label(TOP_INVENTORY, text = "Category", font = APP_FONT)
        LABEL_CATEGORY.grid(column = 0, row = 8, pady = TOP_PADY, sticky = E)
    
        LABEL_ASSCODE = Label(TOP_INVENTORY, text = "Asset Code", font = APP_FONT)
        LABEL_ASSCODE.grid(column = 0, row = 9, pady = TOP_PADY, sticky = E)
        
        LABEL_EXPCODE = Label(TOP_INVENTORY, text = "Expense Code", font = APP_FONT)
        LABEL_EXPCODE.grid(column = 0, row = 10, pady = TOP_PADY, sticky = E)
    
        global TEXTVAR_CODE, ENTRY_CODE
        TEXTVAR_CODE = StringVar()
        ENTRY_CODE = Entry(TOP_INVENTORY, textvariable = TEXTVAR_CODE, font = APP_FONT, width = 35)
        ENTRY_CODE.grid(column = 1, row = 0, sticky = W)
        required.append(TEXTVAR_CODE)

        global TEXTVAR_NAME
        TEXTVAR_NAME = StringVar()
        ENTRY_NAME = Entry(TOP_INVENTORY, textvariable = TEXTVAR_NAME, font = APP_FONT, width = 35)
        ENTRY_NAME.grid(column = 1, row = 1, sticky = W)
        required.append(TEXTVAR_NAME)

        global TEXTVAR_DESC
        TEXTVAR_DESC = StringVar()
        ENTRY_DESC = Entry(TOP_INVENTORY, textvariable = TEXTVAR_DESC, font = APP_FONT, width = 35)
        ENTRY_DESC.grid(column = 1, row = 2, sticky = W)
        
        global TEXTVAR_TAX
        TEXTVAR_TAX = StringVar()
        COMBO_TAX = tk.Combobox(TOP_INVENTORY, values = ["VAT", "Non-VAT"], textvariable = TEXTVAR_TAX, font = APP_FONT, width = 10, state = "readonly")
        COMBO_TAX.grid(column = 1, row = 3, sticky = W)
        TEXTVAR_TAX.set("VAT")
        required.append(TEXTVAR_TAX)

        global TEXTVAR_COST
        TEXTVAR_COST = StringVar()
        ENTRY_COST = Entry(TOP_INVENTORY, textvariable = TEXTVAR_COST, font = APP_FONT, width = 12, justify = RIGHT)
        ENTRY_COST.grid(column = 1, row = 4, sticky = W)
        required.append(TEXTVAR_COST)
        ENTRY_COST.bind("<FocusOut>", lambda e: self.validateAmount(TEXTVAR_COST))
        ENTRY_COST.bind("<FocusIn>", lambda e: TEXTVAR_COST.set(TEXTVAR_COST.get().replace(",", "")))

        global TEXTVAR_QUANTITY, SPINBOX_QUANTITY
        TEXTVAR_QUANTITY = StringVar()
        SPINBOX_QUANTITY = Spinbox(TOP_INVENTORY, textvariable = TEXTVAR_QUANTITY, from_ = 0, to = 9999, font = (APP_FONT[0], 11), width = 12, justify = RIGHT, state = DISABLED)
        SPINBOX_QUANTITY.grid(column = 1, row = 5, sticky = W)
        SPINBOX_QUANTITY.bind("<FocusOut>", lambda e: self.validateInteger(TEXTVAR_QUANTITY))
        TEXTVAR_QUANTITY.set(0)
        
        global TEXTVAR_UOM
        TEXTVAR_UOM = StringVar()
        COMBO_UOM = tk.Combobox(TOP_INVENTORY, values = self.listUnits(), textvariable = TEXTVAR_UOM, font = APP_FONT, width = 10, state = "readonly")
        COMBO_UOM.grid(column = 1, row = 6, sticky = W)
        required.append(TEXTVAR_UOM)
        
        global TEXTVAR_STATUS
        TEXTVAR_STATUS = StringVar()
        COMBO_STATUS = tk.Combobox(TOP_INVENTORY, values = ["active", "inactive"], textvariable = TEXTVAR_STATUS, font = APP_FONT, width = 10, state = "readonly")
        COMBO_STATUS.grid(column = 1, row = 7, sticky = W)
        TEXTVAR_STATUS.set("active")
        required.append(TEXTVAR_STATUS)

        global TEXTVAR_CATEGORY
        TEXTVAR_CATEGORY = StringVar()
        COMBO_CATEGORY = tk.Combobox(TOP_INVENTORY, values = ["Corporate Giveaways", "Equipment - Janitorial", "Equipment - Office", "Furniture - Office", "Service Vehicle", "Supplies - Janitorial", "Supplies - Office", "Supplies - Medicine", "Uniform - Clerical", "Uniform - Janitorial", "Uniform - DBPSC", "Others"], textvariable = TEXTVAR_CATEGORY, font = APP_FONT, width = 25, state = "readonly")
        COMBO_CATEGORY.grid(column = 1, row = 8, sticky = W)
        COMBO_CATEGORY.bind("<<ComboboxSelected>>", lambda e: self.populateItemCodes(e))
        required.append(TEXTVAR_CATEGORY)
        
        global TEXTVAR_ASSCODE
        TEXTVAR_ASSCODE = StringVar()
        ENTRY_ASSCODE = Entry(TOP_INVENTORY, textvariable = TEXTVAR_ASSCODE, font = APP_FONT, width = 35, state = "readonly")
        ENTRY_ASSCODE.grid(column = 1, row = 9, sticky = W)
        
        global TEXTVAR_EXPCODE
        TEXTVAR_EXPCODE = StringVar()
        ENTRY_EXPCODE = Entry(TOP_INVENTORY, textvariable = TEXTVAR_EXPCODE, font = APP_FONT, width = 35, state = "readonly")
        ENTRY_EXPCODE.grid(column = 1, row = 10, sticky = W)
        
        FRAME_BUTTON = Frame(TOP_INVENTORY)
        FRAME_BUTTON.grid(column = 1, row = 11, sticky = E, pady = TOP_PADY + 10)

        BUTTON_SAVE = Button(FRAME_BUTTON, text = "SAVE", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "hand2", command = self.saveInventory)
        BUTTON_SAVE.grid(column = 0, row = 0, padx = TOP_PADX)

        BUTTON_CLOSE = Button(FRAME_BUTTON, text = "CLOSE", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "hand2", command = lambda: self.closeTopLevel(TOP_INVENTORY))
        BUTTON_CLOSE.grid(column = 1, row = 0, padx = TOP_PADX)
        
    def saveInventory(self, *args):
        wrong = []
        for i in required:
            if i.get() == "":
                wrong.append(i)
                break
            
        if len(wrong) > 0:
            messagebox.showerror("Save Item", f"Please provide {str(wrong[0])}!")
        else:
            ask = messagebox.askyesno("Save Item", "Are you sure?")
            if ask == True:
                insert = """INSERT INTO tblinventory (
                            code, name, description, uom, cost, status, category, tax, assetcode, expensecode, encoded)
                            values (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
                update = """UPDATE tblinventory SET code = %s, name = %s, description = %s, uom = %s, cost = %s, status = %s, category = %s, tax = %s, assetcode = %s, expensecode = %s, encoded = %s
                            WHERE code = %s"""
                find = "SELECT code FROM tblinventory WHERE code = %s"
                try:
                    db.commit()
                    cursor.execute(find, [content[0]])
                    result = cursor.fetchall()
                    if result:
                        cursor.execute(update, [TEXTVAR_CODE.get(), TEXTVAR_NAME.get(), TEXTVAR_DESC.get(), TEXTVAR_UOM.get(), TEXTVAR_COST.get().replace(",",""), TEXTVAR_STATUS.get(), TEXTVAR_CATEGORY.get(), TEXTVAR_TAX.get(), TEXTVAR_ASSCODE.get(), TEXTVAR_EXPCODE.get(), datetime.datetime.now(), content[0]])
                        saved = "Item has been updated!"
                    else:
                        cursor.execute(insert, [TEXTVAR_CODE.get(), TEXTVAR_NAME.get(), TEXTVAR_DESC.get(), TEXTVAR_UOM.get(), TEXTVAR_COST.get().replace(",",""), TEXTVAR_STATUS.get(), TEXTVAR_CATEGORY.get(), TEXTVAR_TAX.get(), TEXTVAR_ASSCODE.get(), TEXTVAR_EXPCODE.get(), datetime.datetime.now()])
                        saved = "A new item has been saved!"
                except:
                    cursor.execute(insert, [TEXTVAR_CODE.get(), TEXTVAR_NAME.get(), TEXTVAR_DESC.get(), TEXTVAR_UOM.get(), TEXTVAR_COST.get().replace(",",""), TEXTVAR_STATUS.get(), TEXTVAR_CATEGORY.get(), TEXTVAR_TAX.get(), TEXTVAR_ASSCODE.get(), TEXTVAR_EXPCODE.get(), datetime.datetime.now()])
                    saved = "A new item has been saved!"
                finally:
                    db.commit()
                    messagebox.showinfo("Save Item", saved)
                    TOP_INVENTORY.grab_release()
                    TOP_INVENTORY.destroy()
                    self.showInventory()
                
    def editInventory(self, *args):
        self.copySelection(TREE_INVENTORY)
        self.showAddEditInventory()
        TOP_INVENTORY.title("Edit - Inventory")

        SPINBOX_QUANTITY.config(state = DISABLED)
        
        TEXTVAR_CODE.set(content[0])
        ENTRY_CODE.config(state = "readonly")
        TEXTVAR_NAME.set(content[1])
        TEXTVAR_DESC.set(content[2])
        TEXTVAR_TAX.set(content[3])
        TEXTVAR_COST.set(content[6])
        TEXTVAR_UOM.set(content[5])
        TEXTVAR_STATUS.set(content[7])
        TEXTVAR_CATEGORY.set(content[8])
        TEXTVAR_ASSCODE.set(content[9])
        TEXTVAR_EXPCODE.set(content[10])

    def deleteInventory(self, *args):
        delete = "DELETE FROM tblinventory WHERE code = %s"
        self.copySelection(TREE_INVENTORY)
        ask = messagebox.askyesno("Delete Item", "Are you sure?")
        if ask == True:
            cursor.execute(delete, [content[0]])
            db.commit()
            messagebox.showinfo("Delete Item", "Item has been deleted!")
            self.showInventory()

    def showInventoryTreeview(self, frame):
        global TREE_INVENTORY
        TREE_INVENTORY = tk.Treeview(frame, height = 28, selectmode = "browse")
        TREE_INVENTORY["columns"] = ("Code", "Name", "Description", "Tax", "Stock", "U.O.M.", "Cost", "Status", "Category", "Asset Code", "Expense Code")
        TREE_INVENTORY.column("#0", width = 0, minwidth = 0, stretch = True)
        TREE_INVENTORY.column("Code", anchor = W, minwidth = 50, width = 100)
        TREE_INVENTORY.column("Name", anchor = W, minwidth = 50, width = 225)
        TREE_INVENTORY.column("Description", anchor = W, minwidth = 50, width = 175)
        TREE_INVENTORY.column("Tax", anchor = W, minwidth = 50, width = 75)
        TREE_INVENTORY.column("Stock", anchor = W, minwidth = 50, width = 75)
        TREE_INVENTORY.column("U.O.M.", anchor = W, minwidth = 50, width = 75)
        TREE_INVENTORY.column("Cost", anchor = E, minwidth = 50, width = 75)
        TREE_INVENTORY.column("Status", anchor = W, minwidth = 50, width = 75)
        TREE_INVENTORY.column("Category", anchor = W, minwidth = 50, width = 75)
        TREE_INVENTORY.column("Asset Code", anchor = W, minwidth = 50, width = 75)
        TREE_INVENTORY.column("Expense Code", anchor = W, minwidth = 50, width = 75)
        
        TREE_INVENTORY.heading("#0", text = "", anchor = W)
        TREE_INVENTORY.heading("Code", text = "Code", anchor = N)
        TREE_INVENTORY.heading("Name", text = "Name", anchor = N)
        TREE_INVENTORY.heading("Description", text = "Description", anchor = N)
        TREE_INVENTORY.heading("Tax", text = "Tax", anchor = N)
        TREE_INVENTORY.heading("Stock", text = "Stock", anchor = N)
        TREE_INVENTORY.heading("U.O.M.", text = "U.O.M.", anchor = N)
        TREE_INVENTORY.heading("Cost", text = "Cost", anchor = N)
        TREE_INVENTORY.heading("Status", text = "Status", anchor = N)
        TREE_INVENTORY.heading("Category", text = "Category", anchor = N)
        TREE_INVENTORY.heading("Asset Code", text = "Asset Code", anchor = N)
        TREE_INVENTORY.heading("Expense Code", text = "Expense Code", anchor = N)

        POPUP_INVENTORY = Menu(TREE_INVENTORY, tearoff = 0)
        POPUP_INVENTORY.add_command(command = self.editInventory, label = "Edit")
        POPUP_INVENTORY.add_command(command = self.deleteInventory, label = "Delete")
        TREE_INVENTORY.bind("<Button-3>", lambda e: self.popupMenu(TREE_INVENTORY, POPUP_INVENTORY, e))

        global STYLE_INVENTORY
        STYLE_INVENTORY = tk.Style()
        STYLE_INVENTORY.map("Treeview", foreground = self.fixedMap("foreground", STYLE_INVENTORY), background = self.fixedMap("background", STYLE_INVENTORY))

        TREE_INVENTORY.tag_configure("oddrow", background = None)
        TREE_INVENTORY.tag_configure("evenrow", background = TREE_TAG_EVENROW)
        db.commit()
        cursor.execute("SELECT code, name, description, tax, quantity, uom, cost, status, category, assetcode, expensecode FROM tblinventory")
        result = cursor.fetchall()
        
        count = 0
        for i in result:
            if count % 2 == 0:
                TREE_INVENTORY.insert("", "end", values = (i[0],i[1],i[2],i[3],i[4],i[5],i[6],i[7],i[8],i[9],i[10]), tags = "evenrow")
            else:
                TREE_INVENTORY.insert("", "end", values = (i[0],i[1],i[2],i[3],i[4],i[5],i[6],i[7],i[8],i[9],i[10]), tags = "oddrow")
            count += 1

        YSCROLLBAR = tk.Scrollbar(frame, orient = "vertical", command = TREE_INVENTORY.yview)
        YSCROLLBAR.pack(side = "right", fill = "y")

        XSCROLLBAR = tk.Scrollbar(frame, orient = "horizontal", command = TREE_INVENTORY.xview)
        
        TREE_INVENTORY.configure(yscrollcommand = YSCROLLBAR.set, xscrollcommand = XSCROLLBAR.set) 
        TREE_INVENTORY.pack()
        XSCROLLBAR.pack(fill ="x")

    def showInventorySelection(self, *args):
        global TOP_INVENTORYSELECTION
        TOP_INVENTORYSELECTION = Toplevel()
        TOP_INVENTORYSELECTION.title("Choose Inventory Code")
        TOP_INVENTORYSELECTION.iconbitmap(PATH_ICON + "icon.ico")
        TOP_INVENTORYSELECTION.geometry("600x300+100+100")
        TOP_INVENTORYSELECTION.resizable(height = False, width = False)
        TOP_INVENTORYSELECTION.grab_set()

        FRAME1 = Frame(TOP_INVENTORYSELECTION)
        FRAME1.pack(fill = "x")

        FRAME2 = Frame(TOP_INVENTORYSELECTION)
        FRAME2.pack(fill = "x")

        LABEL_SEARCH = Label(FRAME1, text = "Search", font = APP_FONT)
        LABEL_SEARCH.grid(column = 0, row = 0, padx = TOP_PADX)
        
        ENTRY_SEARCH = Entry(FRAME1, font = APP_FONT, width = 20)
        ENTRY_SEARCH.grid(column = 1, row = 0, padx = TOP_PADX)
        ENTRY_SEARCH.bind("<Return>", lambda e: self.searchInventory(ENTRY_SEARCH.get()))
        
        BUTTON_SEARCH = Button(FRAME1, text = "GO", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, command = lambda: self.searchInventory(ENTRY_SEARCH.get()))
        BUTTON_SEARCH.grid(column = 2, row = 0, padx = TOP_PADX)
        
        global TEXTVAR_CATEGORY
        TEXTVAR_CATEGORY = StringVar()
        COMBO_CATEGORY = tk.Combobox(FRAME1, values = ["Corporate Giveaways", "Equipment - Janitorial", "Equipment - Office", "Furniture - Office", "Service Vehicle", "Supplies - Janitorial", "Supplies - Office", "Supplies - Medicine", "Uniform - Clerical", "Uniform - Janitorial", "Uniform - DBPSC", "Others"], textvariable = TEXTVAR_CATEGORY, font = APP_FONT, width = 25, state = "readonly")
        COMBO_CATEGORY.grid(column = 3, row = 0, sticky = E, padx = TOP_PADX)
        COMBO_CATEGORY.bind("<<ComboboxSelected>>", lambda e: self.searchInventory(ENTRY_SEARCH.get()))
        TEXTVAR_CATEGORY.set("Supplies - Janitorial")

        self.showInventoryTreeview(FRAME2)
        TREE_INVENTORY.config(height = 10)
        TREE_INVENTORY.unbind("<Button-3>")
        TREE_INVENTORY.bind("<Double-1>", self.enterInventorySelection)
        
        self.searchInventory(ENTRY_SEARCH.get())
        ENTRY_SEARCH.focus()
        
    def enterInventorySelection(self, *args):
        self.copySelection(TREE_INVENTORY)
        TEXTVAR_INV_CODE.set(content[0])
        TOP_INVENTORYSELECTION.grab_release()
        TOP_INVENTORYSELECTION.destroy()
        try:
            self.populateInventoryFields(TEXTVAR_INV_CODE)
            TOP_PO.lift()
            SPINBOX_QUANTITY.focus()
        except:
            pass

    def populateInventoryFields(self, var):
        global result
        self.capitalLetters(var)
        select = "SELECT code, name, uom, cost FROM tblinventory WHERE code = %s AND status = 'active'"
        db.commit()
        cursor.execute(select, [var.get()])
        result = cursor.fetchall()
        if result:
            TEXTVAR_INV_NAME.set(result[0][1])
            TEXTVAR_INV_COST.set(result[0][3])
        else:
            var.set("")
            TEXTVAR_INV_NAME.set("")
            TEXTVAR_INV_COST.set(0)
            
    def returnInventoryNameUOM(self, var):
        select = "SELECT name, uom FROM tblinventory WHERE code = %s AND status = 'active'"
        cursor.execute(select, [var])
        return cursor.fetchone()

    def populateItemCodes(self, *args):
        if TEXTVAR_CATEGORY.get() == "Corporate Giveaways":
            asset, expense = 117090, 641080
        elif TEXTVAR_CATEGORY.get() == "Equipment - Janitorial":
            asset, expense = 117120, 633020
        elif TEXTVAR_CATEGORY.get() == "Equipment - Office":
            asset, expense = 122060, 122060
        elif TEXTVAR_CATEGORY.get() == "Furniture - Office":
            asset, expense = 122060, 122060
        elif TEXTVAR_CATEGORY.get() == "Service Vehicle":
            asset, expense = 122050, 122050
        elif TEXTVAR_CATEGORY.get() == "Supplies - Janitorial":
            asset, expense = 117120, 633020
        elif TEXTVAR_CATEGORY.get() == "Supplies - Office":
            asset, expense = 117090, 636050
        elif TEXTVAR_CATEGORY.get() == "Supplies - Medicine":
            asset, expense = 117140, 621160
        elif TEXTVAR_CATEGORY.get() == "Uniform - Clerical":
            asset, expense = 611170, 611170
        elif TEXTVAR_CATEGORY.get() == "Uniform - Janitorial":
            asset, expense = 611180, 611180
        elif TEXTVAR_CATEGORY.get() == "Uniform - DBPSC":
            asset, expense = 621140, 621140
        elif TEXTVAR_CATEGORY.get() == "Others":
            asset, expense = 117090, 636050
        TEXTVAR_ASSCODE.set(asset)
        TEXTVAR_EXPCODE.set(expense)

    def searchInventory(self, var, *args):
        if TEXTVAR_CATEGORY.get() != "All":
            find = "SELECT code, name, description, tax, quantity, uom, cost, status, category, assetcode, expensecode FROM tblinventory WHERE (code LIKE %s OR name LIKE %s OR description LIKE %s) AND category = %s ORDER BY name"
        else:
            find = "SELECT code, name, description, tax, quantity, uom, cost, status, category, assetcode, expensecode FROM tblinventory WHERE (code LIKE %s OR name LIKE %s OR description LIKE %s) ORDER BY name"
        db.commit()
        if TEXTVAR_CATEGORY.get() != "All":
            cursor.execute(find, [f"%{var}%", f"%{var}%", f"%{var}%", TEXTVAR_CATEGORY.get()])
        else:
            cursor.execute(find, [f"%{var}%", f"%{var}%", f"%{var}%"])
        result = cursor.fetchall()
        for i in TREE_INVENTORY.get_children():
                TREE_INVENTORY.delete(i)
        if result:
            count = 0
            for i in result:
                if count % 2 == 0:
                    TREE_INVENTORY.insert("", "end", values = (i[0],i[1],i[2],i[3],i[4],i[5],i[6],i[7],i[8],i[9],i[10]), tags = "evenrow")
                else:
                    TREE_INVENTORY.insert("", "end", values = (i[0],i[1],i[2],i[3],i[4],i[5],i[6],i[7],i[8],i[9],i[10]), tags = "oddrow")
                count += 1
        else:
            messagebox.showerror("Inventory", "No match found!")
            # db.commit()
            # cursor.execute("SELECT code, name, description, tax, quantity, uom, cost, status, category, assetcode, expensecode FROM tblinventory ORDER BY name")
            # result = cursor.fetchall()

    def refreshInventory(self, *args):
        for i in TREE_INVENTORY.get_children():
            TREE_INVENTORY.delete(i)
        db.commit()
        cursor.execute("SELECT code, name, description, tax, quantity, uom, cost, status, category, assetcode, expensecode FROM tblinventory ORDER BY name")
        result = cursor.fetchall()

        count = 0
        for i in result:
            if count % 2 == 0:
                TREE_INVENTORY.insert("", "end", values = (i[0],i[1],i[2],i[3],i[4],i[5],i[6],i[7],i[8],i[9],i[10]), tags = "evenrow")
            else:
                TREE_INVENTORY.insert("", "end", values = (i[0],i[1],i[2],i[3],i[4],i[5],i[6],i[7],i[8],i[9],i[10]), tags = "oddrow")
            count += 1

### MENU_GSAD_PURCHASE ORDER ###
    def showAddEditPurchaseOrder(self, *args):
        global TOP_PO
        TOP_PO = Toplevel()
        TOP_PO.title("Add - Purchase Order")
        TOP_PO.iconbitmap(PATH_ICON + "icon.ico")
        TOP_PO.geometry("725x500+300+100")
        TOP_PO.resizable(height = False, width = False)
        TOP_PO.grab_set()
        # TOP_PO.protocol("WM_DELETE_WINDOW", self.exitAddEditPO)
        
        global required
        required = []
        
        FRAME_1 = Frame(TOP_PO)
        FRAME_1.grid(column = 0, row = 0, pady = TOP_PADY, sticky = W)
        
        FRAME_2 = Frame(TOP_PO)
        FRAME_2.grid(column = 0, row = 1, sticky = W)

        LABEL_PO = Label(FRAME_1, text = "PO No.", font = APP_FONT)
        LABEL_PO.grid(column = 0, row = 0, pady = TOP_PADY, sticky = E)

        LABEL_CODE = Label(FRAME_1, text = "Supplier Code", font = APP_FONT)
        LABEL_CODE.grid(column = 0, row = 1, pady = TOP_PADY, sticky = E)

        LABEL_NAME = Label(FRAME_1, text = "Supplier Name", font = APP_FONT)
        LABEL_NAME.grid(column = 0, row = 2, pady = TOP_PADY, sticky = E)

        LABEL_ADDRESS = Label(FRAME_1, text = "Supplier Address", font = APP_FONT)
        LABEL_ADDRESS.grid(column = 0, row = 3, pady = TOP_PADY, sticky = E)

        LABEL_TIN = Label(FRAME_1, text = "Supplier TIN", font = APP_FONT)
        LABEL_TIN.grid(column = 0, row = 4, pady = TOP_PADY, sticky = E)

        LABEL_PARTICULARS = Label(FRAME_1, text = "Particulars", font = APP_FONT)
        LABEL_PARTICULARS.grid(column = 0, row = 5, pady = TOP_PADY, sticky = E)
        
        LABEL_TAX = Label(FRAME_1, text = "Tax", font = APP_FONT)
        LABEL_TAX.grid(column = 0, row = 6, pady = TOP_PADY, sticky = E)

        LABEL_DIVIDER = Label(FRAME_1, text = "", font = APP_FONT)
        LABEL_DIVIDER.grid(column = 2, row = 0, padx = TOP_PADX + 5, sticky = N)

        LABEL_DATE = Label(FRAME_1, text = "Date", font = APP_FONT)
        LABEL_DATE.grid(column = 3, row = 0, pady = TOP_PADY, sticky = E)

        LABEL_GROSS = Label(FRAME_1, text = "Gross", font = APP_FONT)
        LABEL_GROSS.grid(column = 3, row = 1, pady = TOP_PADY, sticky = E)

        LABEL_EXPENSE = Label(FRAME_1, text = "Expense", font = APP_FONT)
        LABEL_EXPENSE.grid(column = 3, row = 2, pady = TOP_PADY, sticky = E)

        LABEL_VAT = Label(FRAME_1, text = "VAT", font = APP_FONT)
        LABEL_VAT.grid(column = 3, row = 3, pady = TOP_PADY, sticky = E)

        LABEL_EWT = Label(FRAME_1, text = "EWT", font = APP_FONT)
        LABEL_EWT.grid(column = 3, row = 4, pady = TOP_PADY, sticky = E)

        LABEL_NET = Label(FRAME_1, text = "Net", font = APP_FONT)
        LABEL_NET.grid(column = 3, row = 5, pady = TOP_PADY, sticky = NE)

        global TEXTVAR_PO, ENTRY_PO
        TEXTVAR_PO = StringVar()
        ENTRY_PO = Entry(FRAME_1, textvariable = TEXTVAR_PO, font = APP_FONT, width = 20, state = "readonly", justify = RIGHT)
        ENTRY_PO.grid(column = 1, row = 0, sticky = W)

        global TEXTVAR_CODE, ENTRY_CODE
        TEXTVAR_CODE = StringVar()
        ENTRY_CODE = Entry(FRAME_1, textvariable = TEXTVAR_CODE, font = APP_FONT, width = 40)
        ENTRY_CODE.grid(column = 1, row = 1, sticky = W)
        ENTRY_CODE.bind("<FocusOut>", lambda e: self.populatePOFields(TEXTVAR_CODE))
        required.append(TEXTVAR_CODE)

        global BUTTON_CODE
        BUTTON_CODE = Button(FRAME_1, text = "...", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, cursor = "hand2", command = self.showSupplierSelection)
        BUTTON_CODE.grid(column = 1, row = 1, sticky = E)

        global TEXTVAR_NAME
        TEXTVAR_NAME = StringVar()
        ENTRY_NAME = Entry(FRAME_1, textvariable = TEXTVAR_NAME, font = APP_FONT, width = 40, state = "readonly")
        ENTRY_NAME.grid(column = 1, row = 2, sticky = W)
        required.append(TEXTVAR_NAME)

        global TEXTVAR_ADDRESS
        TEXTVAR_ADDRESS = StringVar()
        ENTRY_ADDRESS = Entry(FRAME_1, textvariable = TEXTVAR_ADDRESS, font = APP_FONT, width = 40, state = "readonly")
        ENTRY_ADDRESS.grid(column = 1, row = 3, sticky = W)
        required.append(TEXTVAR_ADDRESS)

        global TEXTVAR_TIN
        TEXTVAR_TIN = StringVar()
        ENTRY_TIN = Entry(FRAME_1, textvariable = TEXTVAR_TIN, font = APP_FONT, width = 20, state = "readonly")
        ENTRY_TIN.grid(column = 1, row = 4, sticky = W)
        required.append(TEXTVAR_TIN)
        
        global ENTRY_REMARKS
        ENTRY_REMARKS = Text(FRAME_1, font = APP_FONT, width = 40, height = 4)
        ENTRY_REMARKS.grid(column = 1, row = 5, sticky = W)
        required.append(ENTRY_REMARKS)
        
        global TEXTVAR_TAX
        TEXTVAR_TAX = StringVar()
        COMBO_TAX = tk.Combobox(FRAME_1, values = ["VAT", "Non-VAT"], textvariable = TEXTVAR_TAX, font = APP_FONT, width = 10, state = DISABLED)
        COMBO_TAX.grid(column = 1, row = 6, pady = TOP_PADY, sticky = W)

        global CALENDAR_DOC
        CALENDAR_DOC = DateEntry(FRAME_1, firstweekday = "sunday", date_pattern = "yyyy-mm-dd")
        CALENDAR_DOC.grid(column = 4, row = 0, sticky = W)
        required.append(CALENDAR_DOC)

        global TEXTVAR_GROSS
        TEXTVAR_GROSS = StringVar()
        ENTRY_GROSS = Entry(FRAME_1, textvariable = TEXTVAR_GROSS, font = APP_FONT, width = 20, state = "readonly", justify = RIGHT)
        ENTRY_GROSS.grid(column = 4, row = 1, sticky = W)

        global TEXTVAR_EXPENSE
        TEXTVAR_EXPENSE = StringVar()
        ENTRY_EXPENSE = Entry(FRAME_1, textvariable = TEXTVAR_EXPENSE, font = APP_FONT, width = 20, state = "readonly", justify = RIGHT)
        ENTRY_EXPENSE.grid(column = 4, row = 2, sticky = W)

        global TEXTVAR_VAT
        TEXTVAR_VAT = StringVar()
        ENTRY_VAT = Entry(FRAME_1, textvariable = TEXTVAR_VAT, font = APP_FONT, width = 20, state = "readonly", justify = RIGHT)
        ENTRY_VAT.grid(column = 4, row = 3, sticky = W)

        global TEXTVAR_EWT
        TEXTVAR_EWT = StringVar()
        ENTRY_EWT = Entry(FRAME_1, textvariable = TEXTVAR_EWT, font = APP_FONT, width = 20, state = "readonly", justify = RIGHT)
        ENTRY_EWT.grid(column = 4, row = 4, sticky = W)

        global TEXTVAR_NET
        TEXTVAR_NET = StringVar()
        ENTRY_NET = Entry(FRAME_1, textvariable = TEXTVAR_NET, font = APP_FONT, width = 20, state = "readonly", justify = RIGHT)
        ENTRY_NET.grid(column = 4, row = 5, sticky = N)
        required.append(TEXTVAR_NET)
        
        FRAME_2_1 = Frame(FRAME_2)
        FRAME_2_1.grid(column = 0, row = 0, pady = TOP_PADY, sticky = W)
        
        FRAME_2_2 = Frame(FRAME_2)
        FRAME_2_2.grid(column = 0, row = 1, pady = TOP_PADY, sticky = W)
        
        FRAME_2_3 = Frame(FRAME_2)
        FRAME_2_3.grid(column = 0, row = 2, pady = TOP_PADY, sticky = W)
        
        FRAME_2_4 = Frame(FRAME_2)
        FRAME_2_4.grid(column = 0, row = 3, pady = TOP_PADY, sticky = E)
        
        global TEXTVAR_INV_CODE, ENTRY_INV_CODE
        TEXTVAR_INV_CODE = StringVar()
        ENTRY_INV_CODE = Entry(FRAME_2_1, textvariable = TEXTVAR_INV_CODE, font = APP_FONT, width = 15)
        ENTRY_INV_CODE.grid(column = 0, row = 0, padx = TOP_PADX, sticky = W)
        ENTRY_INV_CODE.bind("<FocusOut>", lambda e: self.populateInventoryFields(TEXTVAR_INV_CODE))
        
        global BUTTON_INV_CODE
        BUTTON_INV_CODE = Button(FRAME_2_1, text = "...", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, cursor = "hand2", command = self.showInventorySelection)
        BUTTON_INV_CODE.grid(column = 0, row = 0, padx = TOP_PADX, sticky = E)
        BUTTON_INV_CODE.bind("<Return>", self.showInventorySelection)
        
        global TEXTVAR_INV_NAME, ENTRY_INV_NAME
        TEXTVAR_INV_NAME = StringVar()
        ENTRY_INV_NAME = Entry(FRAME_2_1, textvariable = TEXTVAR_INV_NAME, font = APP_FONT, width = 69, state = "readonly")
        ENTRY_INV_NAME.grid(column = 1, row = 0, sticky = W)
        
        global TEXTVAR_QUANTITY, SPINBOX_QUANTITY
        TEXTVAR_QUANTITY = StringVar()
        SPINBOX_QUANTITY = Spinbox(FRAME_2_2, textvariable = TEXTVAR_QUANTITY, font = APP_FONT, width = 6, from_ = 0, to = 99999)
        SPINBOX_QUANTITY.grid(column = 0, row = 0, padx = TOP_PADX, sticky = W)
        SPINBOX_QUANTITY.bind("<FocusOut>", lambda e: self.validateInteger(TEXTVAR_QUANTITY))
        
        global TEXTVAR_INV_COST, ENTRY_INV_COST
        TEXTVAR_INV_COST = StringVar()
        ENTRY_INV_COST = Entry(FRAME_2_2, textvariable = TEXTVAR_INV_COST, font = APP_FONT, width = 15, justify = RIGHT)
        ENTRY_INV_COST.grid(column = 1, row = 0, padx = TOP_PADX, sticky = W)
        ENTRY_INV_COST.bind("<FocusOut>", lambda e: self.validateAmount(TEXTVAR_INV_COST))
        
        global TEXTVAR_INV_STOCK
        TEXTVAR_INV_STOCK = StringVar()
        ENTRY_INV_STOCK = Entry(FRAME_2_2, textvariable = TEXTVAR_INV_STOCK, font = APP_FONT, width = 15, state = "readonly")
        ENTRY_INV_STOCK.grid(column = 2, row = 0, padx = TOP_PADX, sticky = W)
        
        global BUTTON_ADD
        BUTTON_ADD = Button(FRAME_2_2, text = "ADD", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "hand2", command = self.addPOInventoryItem)
        BUTTON_ADD.grid(column = 3, row = 0, padx = TOP_PADX)
        BUTTON_ADD.bind("<Return>", self.addPOInventoryItem)
        
        global BUTTON_UPDATE
        BUTTON_UPDATE = Button(FRAME_2_2, text = "UPDATE", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "arrow", state = DISABLED, command = self.addPOInventoryItem)
        BUTTON_UPDATE.grid(column = 4, row = 0, padx = TOP_PADX)
        BUTTON_UPDATE.bind("<Return>", self.addPOInventoryItem)
        
        global BUTTON_CANCEL
        BUTTON_CANCEL = Button(FRAME_2_2, text = "CANCEL", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "arrow", state = DISABLED, command = self.cancelEditPOInventoryItem)
        BUTTON_CANCEL.grid(column = 5, row = 0, padx = TOP_PADX)
        BUTTON_CANCEL.bind("<Return>", self.cancelEditPOInventoryItem)
        
        global TREE_INV_ITEMS
        TREE_INV_ITEMS = tk.Treeview(FRAME_2_3, height = 6, selectmode = "browse")
        TREE_INV_ITEMS["columns"] = ("Code", "Name", "Cost", "UOM", "Quantity", "Total")
        TREE_INV_ITEMS.column("#0", width = 0, minwidth = 0, stretch = True)
        TREE_INV_ITEMS.column("Code", anchor = W, minwidth = 100, width = 75)
        TREE_INV_ITEMS.column("Name", anchor = W, minwidth = 100, width = 200)
        TREE_INV_ITEMS.column("Cost", anchor = E, minwidth = 100, width = 100)
        TREE_INV_ITEMS.column("UOM", anchor = W, minwidth = 100, width = 50)
        TREE_INV_ITEMS.column("Quantity", anchor = E, minwidth = 100, width = 150)
        TREE_INV_ITEMS.column("Total", anchor = E, minwidth = 100, width = 125)
        
        TREE_INV_ITEMS.heading("#0", text = "", anchor = W)
        TREE_INV_ITEMS.heading("Code", text = "Code", anchor = N, command = lambda: self.sortColumn(TREE_INV_ITEMS, "Cost", False))
        TREE_INV_ITEMS.heading("Name", text = "Name", anchor = N, command = lambda: self.sortColumn(TREE_INV_ITEMS, "Name", False))
        TREE_INV_ITEMS.heading("Cost", text = "Cost", anchor = N, command = lambda: self.sortColumn(TREE_INV_ITEMS, "Cost", False))
        TREE_INV_ITEMS.heading("UOM", text = "UOM", anchor = N, command = lambda: self.sortColumn(TREE_INV_ITEMS, "Quantity", False))
        TREE_INV_ITEMS.heading("Quantity", text = "Quantity", anchor = N, command = lambda: self.sortColumn(TREE_INV_ITEMS, "Quantity", False))
        TREE_INV_ITEMS.heading("Total", text = "Total", anchor = N, command = lambda: self.sortColumn(TREE_INV_ITEMS, "Total", False))

        global POPUP_INV_ITEMS
        POPUP_INV_ITEMS = Menu(TREE_INV_ITEMS, tearoff = 0)
        POPUP_INV_ITEMS.add_command(command = self.editPOInventoryItem, label = "Edit")
        POPUP_INV_ITEMS.add_command(command = lambda: self.deleteTreeItem(TREE_INV_ITEMS), label = "Delete")
        TREE_INV_ITEMS.bind("<Button-3>", lambda e: self.popupMenu(TREE_INV_ITEMS, POPUP_INV_ITEMS, e))
        TREE_INV_ITEMS.bind("<Double-1>", self.editPOInventoryItem)

        global STYLE_INV_ITEMS
        STYLE_INV_ITEMS = tk.Style()
        STYLE_INV_ITEMS.map("Treeview", foreground = self.fixedMap("foreground", STYLE_INV_ITEMS), background = self.fixedMap("background", STYLE_INV_ITEMS))

        TREE_INV_ITEMS.tag_configure("oddrow", background = None)
        TREE_INV_ITEMS.tag_configure("evenrow", background = TREE_TAG_EVENROW)

        YSCROLLBAR = tk.Scrollbar(FRAME_2_3, orient = "vertical", command = TREE_INV_ITEMS.yview)
        YSCROLLBAR.pack(side = "right", fill = "y")

        XSCROLLBAR = tk.Scrollbar(FRAME_2_3, orient = "horizontal", command = TREE_INV_ITEMS.xview)
        
        TREE_INV_ITEMS.configure(yscrollcommand = YSCROLLBAR.set, xscrollcommand = XSCROLLBAR.set) 
        TREE_INV_ITEMS.pack()
        XSCROLLBAR.pack(fill ="x")
        
        global BUTTON_SAVE
        BUTTON_SAVE = Button(FRAME_2_4, text = "SAVE", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "hand2", command = self.savePurchaseOrder)
        BUTTON_SAVE.grid(column = 1, row = 1, padx = TOP_PADX)
        BUTTON_SAVE.bind("<Return>", self.savePurchaseOrder)
        
        global BUTTON_APPROVE
        BUTTON_APPROVE = Button(FRAME_2_4, text = "APPROVE", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "arrow", state = DISABLED, command = self.approvePurchaseOrder)
        BUTTON_APPROVE.grid(column = 2, row = 1, padx = TOP_PADX)
        BUTTON_APPROVE.bind("<Return>", None)
        
        global BUTTON_VOID
        BUTTON_VOID = Button(FRAME_2_4, text = "VOID", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "arrow", state = DISABLED, command = self.voidPurchaseOrder)
        BUTTON_VOID.grid(column = 3, row = 1, padx = TOP_PADX)
        BUTTON_VOID.bind("<Return>", None)
        
        global BUTTON_PRINT
        BUTTON_PRINT = Button(FRAME_2_4, text = "PRINT", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "arrow", state = DISABLED, command = self.printPurchaseOrder)
        BUTTON_PRINT.grid(column = 4, row = 1, padx = TOP_PADX)
        BUTTON_PRINT.bind("<Return>", None)
        
        global BUTTON_CLOSE
        BUTTON_CLOSE = Button(FRAME_2_4, text = "CLOSE", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "hand2", command = lambda: self.closeTopLevel(TOP_PO))
        BUTTON_CLOSE.grid(column = 5, row = 1, padx = TOP_PADX)
        BUTTON_CLOSE.bind("<Return>", None)
        
        if self.returnAccess(USER, 1) == 0:
            BUTTON_APPROVE.config(state = DISABLED, cursor = "arrow")
            
        if self.returnAccess(USER, 2) == 0:
            BUTTON_VOID.config(state = DISABLED, cursor = "arrow")
        
    def populatePOFields(self, var):
        self.capitalLetters(var)
        select = "SELECT name, tin, address, vatable FROM tblsuppliers WHERE code = %s"
        db.commit()
        cursor.execute(select, [var.get()])
        result = cursor.fetchall()
        if result:
            TEXTVAR_NAME.set(result[0][0])
            TEXTVAR_ADDRESS.set(result[0][2])
            TEXTVAR_TIN.set(result[0][1])
            TEXTVAR_TAX.set(result[0][3])
        else:
            var.set("")
            TEXTVAR_NAME.set("")
            TEXTVAR_ADDRESS.set("")
            TEXTVAR_TIN.set("")
            TEXTVAR_TAX.set("")

    def addPOInventoryItem(self):
        if TEXTVAR_INV_CODE.get() != "":
            duplicate = False
            for i in TREE_INV_ITEMS.get_children():
                if TEXTVAR_INV_CODE.get() == TREE_INV_ITEMS.item(i)['values'][0] and BUTTON_ADD["state"] == NORMAL:
                    messagebox.showerror("Purchase Order", "You already added this item!")
                    duplicate = True
                    break
            if duplicate == False:
                if BUTTON_UPDATE["state"] == NORMAL:
                    self.deleteTreeItem(TREE_INV_ITEMS)
                count = 0
                if count % 2 == 0:
                    TREE_INV_ITEMS.insert("", "end", values = (TEXTVAR_INV_CODE.get(),TEXTVAR_INV_NAME.get(),self.returnValidatedAmount(TEXTVAR_INV_COST),self.returnInventoryNameUOM(TEXTVAR_INV_CODE.get())[1],TEXTVAR_QUANTITY.get(),self.validateAmount2(float(TEXTVAR_INV_COST.get().replace(",",""))*int(TEXTVAR_QUANTITY.get()))), tags = "evenrow")
                else:
                    TREE_INV_ITEMS.insert("", "end", values = (TEXTVAR_INV_CODE.get(),TEXTVAR_INV_NAME.get(),self.returnValidatedAmount(TEXTVAR_INV_COST),self.returnInventoryNameUOM(TEXTVAR_INV_CODE.get())[1],TEXTVAR_QUANTITY.get(),self.validateAmount2(float(TEXTVAR_INV_COST.get().replace(",",""))*int(TEXTVAR_QUANTITY.get()))), tags = "oddrow")
                count += 1
                self.updatePOTotals()
                TEXTVAR_INV_CODE.set("")
                TEXTVAR_INV_NAME.set("")
                TEXTVAR_QUANTITY.set(0)
                TEXTVAR_INV_COST.set(0)
                TEXTVAR_INV_STOCK.set(0)
                
                BUTTON_UPDATE.config(state = DISABLED, cursor = "arrow")
                BUTTON_CANCEL.config(state = DISABLED, cursor = "arrow")
                BUTTON_ADD.config(state = NORMAL, cursor = "hand2")
            else:
                TOP_PO.focus()

    def editPOInventoryItem(self, *args):
        self.copySelection(TREE_INV_ITEMS)
        TEXTVAR_INV_CODE.set(content[0])
        TEXTVAR_QUANTITY.set(content[4])
        TEXTVAR_INV_NAME.set(content[1])
        TEXTVAR_INV_COST.set(content[2])
        
        BUTTON_UPDATE.config(state = NORMAL, cursor = "hand2")
        BUTTON_CANCEL.config(state = NORMAL, cursor = "hand2")
        BUTTON_ADD.config(state = DISABLED, cursor = "arrow")
        
        self.updatePOTotals()
        
    def printPurchaseOrder(self):
        wb = load_workbook(PATH_TEMPLATE + "PO.xlsx")
        sheet = wb.active
        sheet["I1"] = f"Printed Date: {datetime.datetime.now()} {self.returnUserName(USER, 0)}"
        sheet["H9"] = TEXTVAR_PO.get()
        sheet["B9"] = TEXTVAR_NAME.get()
        sheet["H10"] = datetime.datetime.strptime(str(CALENDAR_DOC.get_date()), '%Y-%m-%d').strftime('%B %d, %Y')
        sheet["B10"] = TEXTVAR_ADDRESS.get()
        sheet["B11"] = TEXTVAR_TIN.get()
        sheet["A157"] = ENTRY_REMARKS.get("1.0", END).replace("\n", "")
        
        sheet["H157"] = self.returnFloatAmount(TEXTVAR_GROSS.get())
        sheet["H157"] = self.returnFloatAmount(TEXTVAR_EXPENSE.get())
        sheet["H158"] = self.returnFloatAmount(TEXTVAR_VAT.get())
        sheet["H159"] = self.returnFloatAmount(TEXTVAR_EWT.get())
        sheet["H160"] = self.returnFloatAmount(TEXTVAR_NET.get())
        
        sheet["B172"] = f"{self.returnUserName(USER, 1).upper()} {self.returnUserName(USER, 2).upper()}"
        sheet["B173"] = self.returnUserName(USER, 3)

        count = 17
        for i in TREE_INV_ITEMS.get_children():
            if self.returnFloatAmount(TREE_INV_ITEMS.item(i)['values'][5]) != 0:
                sheet["A" + str(count)] = TREE_INV_ITEMS.item(i)['values'][1]
                sheet["E" + str(count)] = str(TREE_INV_ITEMS.item(i)['values'][4]) + " " + TREE_INV_ITEMS.item(i)['values'][3]
                sheet["F" + str(count)] = self.returnFloatAmount(TREE_INV_ITEMS.item(i)['values'][2])
                sheet["H" + str(count)] = self.returnFloatAmount(TREE_INV_ITEMS.item(i)['values'][5])
                count += 1

        for i in range(16,155):
            cell = sheet["H" + str(i)].value
            if cell == None:
                sheet.row_dimensions[i].hidden = True
                    
        wb.save(PATH_SAVE + "PO.xlsx")
        startfile(PATH_SAVE + "PO.xlsx", "open")
        
    def cancelEditPOInventoryItem(self):
        TEXTVAR_INV_CODE.set("")
        TEXTVAR_QUANTITY.set(0)
        TEXTVAR_INV_NAME.set("")
        TEXTVAR_INV_COST.set(0)
        
        BUTTON_UPDATE.config(state = DISABLED, cursor = "arrow")
        BUTTON_CANCEL.config(state = DISABLED, cursor = "arrow")
        BUTTON_ADD.config(state = NORMAL, cursor = "hand2")
    
    def updatePOTotals(self):
        total = []
        for i in TREE_INV_ITEMS.get_children():
            total.append(float(str(TREE_INV_ITEMS.item(i)['values'][5]).replace(",","")))

        nonewt = ["EAFC", "SON-0001"]
        gross = sum(total)
        if TEXTVAR_TAX.get() == "VAT":
            expense = sum(total)/1.12
            vat = expense*.12
            ewt = round(expense*.01, 3)
        else:
            expense = gross
            vat = 0
            if TEXTVAR_CODE.get() in nonewt:
                ewt = 0
            else:
                ewt = gross*.01
        TEXTVAR_GROSS.set(self.validateAmount2(gross))
        TEXTVAR_EXPENSE.set(self.validateAmount2(expense))
        TEXTVAR_VAT.set(self.validateAmount2(vat))
        TEXTVAR_EWT.set(self.validateAmount2(ewt))
        TEXTVAR_NET.set(self.validateAmount2(gross-ewt))
        
    def savePurchaseOrder(self):
        wrong = []
        for i in required:
            try:
                if i.get() == "":
                    wrong.append(i)
            except:
                if i.get("1.0",END).replace("\n","") == "":
                    wrong.append(i)
        if len(wrong) != 0:
            messagebox.showerror("Purchase Order", wrong)
            TOP_PO.focus()
        else:
            ask = messagebox.askyesno("Save Purchase Order", "Are you sure?")
            if ask == True:
                if TEXTVAR_PO.get() == "":
                    self.generatePONumber()
                POitemsToInsert = []
                # POitemsToUpdate = []
                for i in TREE_INV_ITEMS.get_children():
                    POitemsToInsert.append([int(TEXTVAR_PO.get()), CALENDAR_DOC.get(), TEXTVAR_CODE.get(), ENTRY_REMARKS.get("1.0", END).replace("\n", ""), TREE_INV_ITEMS.item(i)['values'][0],
                    float(str(TREE_INV_ITEMS.item(i)['values'][2]).replace(",","")), int(TREE_INV_ITEMS.item(i)['values'][4]),"No Action","for approval", USER,
                    datetime.datetime.now()])
                    # POitemsToUpdate.append([CALENDAR_DOC.get(), TEXTVAR_CODE.get(), ENTRY_REMARKS.get("1.0", END).replace("\n", ""), TREE_INV_ITEMS.item(i)['values'][0], float(str(TREE_INV_ITEMS.item(i)['values'][2]).replace(",","")), int(TREE_INV_ITEMS.item(i)['values'][4]), USER, datetime.datetime.now(), int(TEXTVAR_PO.get())])
                if len(POitemsToInsert) > 0:
                    insert = """INSERT INTO tblpurchaseorder (
                                poNumber, date, supplierCode, remarks, inventoryCode,
                                inventoryCost, inventoryQuantity, delivery, status, encoder,
                                encoded)
                                values (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
                    # update = """UPDATE tblpurchaseorder SET date = %s, supplierCode = %s, remarks = %s, inventoryCode = %s, inventoryCost = %s, inventoryQuantity = %s, encoder = %s, encoded = %s
                    #             WHERE poNumber = %s"""
                    delete = "DELETE FROM tblpurchaseorder WHERE poNumber = %s"
                    find = "SELECT poNumber FROM tblpurchaseorder WHERE poNumber = %s"
                    db.commit()
                    cursor.execute(find, [int(TEXTVAR_PO.get())])
                    result = cursor.fetchall()
                    if result:
                        cursor.execute(delete, [int(TEXTVAR_PO.get())])
                        db.commit()
                        saved = "Purchase Order has been updated!"
                    else:
                        saved = "A new Purchase Order has been created!"
                    for i in POitemsToInsert:
                        cursor.execute(insert, i)
                    db.commit()
                    messagebox.showinfo("Purchase Order", saved)
                    TOP_PO.focus()
                    BUTTON_APPROVE.config(state = NORMAL, cursor = "hand2")
                    BUTTON_PRINT.config(state = NORMAL, cursor = "hand2")
                    
                    if self.returnAccess(USER, 1) == 0:
                        BUTTON_APPROVE.config(state = DISABLED, cursor = "arrow")
    
    def approvePurchaseOrder(self):
        ask = messagebox.askyesno("Approve Purchase Order", "Are you sure?")
        if ask:
            update = "UPDATE tblpurchaseorder SET status = %s, approver = %s, approved = %s WHERE poNumber = %s"
            cursor.execute(update, ["Approved", USER, datetime.datetime.now(), int(TEXTVAR_PO.get())])
            db.commit()
            messagebox.showinfo("Approve Purchase Order", "Purchase Order has been approved!")
            self.disablePurchaseOrderWidgets()
            TOP_PO.focus()
            BUTTON_SAVE.config(state = DISABLED, cursor = "arrow")
            BUTTON_APPROVE.config(state = DISABLED, cursor = "arrow")
            BUTTON_VOID.config(state = NORMAL, cursor = "hand2")
            
            if self.returnAccess(USER, 1) == 0:
                BUTTON_VOID.config(state = DISABLED, cursor = "arrow")
    
    def voidPurchaseOrder(self):
        ask = messagebox.askyesno("Void Purchase Order", "Are you sure?")
        if ask:
            update = "UPDATE tblpurchaseorder SET status = %s WHERE poNumber = %s"
            cursor.execute(update, ["void", int(TEXTVAR_PO.get())])
            db.commit()
            messagebox.showinfo("Void Purchase Order", "Purchase Order has been voided!")
            BUTTON_VOID.config(state = DISABLED, cursor = "arrow")
    
    def generatePONumber(self):
        cursor.execute("SELECT MAX(poNumber) FROM tblpurchaseorder")
        result = cursor.fetchone()
        if result[0] == None:
            TEXTVAR_PO.set(str(155321).zfill(8))
        else:
            TEXTVAR_PO.set(str(int(result[0])+1).zfill(8))

    def returnSupplierName(self, code):
        find = "SELECT name FROM tblsuppliers WHERE code = %s"
        cursor.execute(find, [code])
        result = cursor.fetchone()
        return result[0]
    
    def returnTotalPOAmount(self, code):
        find = "SELECT SUM(inventoryCost*inventoryQuantity) FROM tblpurchaseorder WHERE poNumber = %s"
        cursor.execute(find, [code])
        result = cursor.fetchone()
        return format(result[0], ",.2f")

    def editPurchaseOrder(self, *args):
        self.copySelection(TREE_PO)
        self.showAddEditPurchaseOrder()
        TOP_PO.title("Edit - Purchase Order")

        find = "SELECT poNumber, DATE(date), supplierCode, remarks, inventoryCode, inventoryCost, inventoryQuantity, encoder, encoded, status FROM tblpurchaseorder WHERE poNumber = %s"
        db.commit()
        cursor.execute(find, [int(content[0])])
        result = cursor.fetchall()
        if result:
            TEXTVAR_PO.set(result[0][0].zfill(8))
            CALENDAR_DOC.set_date(result[0][1])
            TEXTVAR_CODE.set(result[0][2])
            self.populatePOFields(TEXTVAR_CODE)
            ENTRY_REMARKS.insert(1.0, result[0][3])
            count = 0
            for i in result:
                if count % 2 == 0:
                    TREE_INV_ITEMS.insert("", "end", values = (i[4],self.returnInventoryNameUOM(i[4])[0],self.validateAmount2(float(i[5])),self.returnInventoryNameUOM(i[4])[1],i[6],self.validateAmount2(float(i[5])*int(i[6]))), tags = "evenrow")
                else:
                    TREE_INV_ITEMS.insert("", "end", values = (i[4],self.returnInventoryNameUOM(i[4])[0],self.validateAmount2(float(i[5])),self.returnInventoryNameUOM(i[4])[1],i[6],self.validateAmount2(float(i[5])*int(i[6]))), tags = "oddrow")
                count += 1
                self.updatePOTotals()
            if result[0][9] == "Approved":
                self.disablePurchaseOrderWidgets()
                BUTTON_SAVE.config(state = DISABLED, cursor = "arrow")
                BUTTON_APPROVE.config(state = DISABLED, cursor = "arrow")
                BUTTON_VOID.config(state = NORMAL, cursor = "hand2")
                BUTTON_PRINT.config(state = NORMAL, cursor = "hand2")
            elif result[0][9] == "void":
                self.disablePurchaseOrderWidgets()
                BUTTON_SAVE.config(state = DISABLED, cursor = "arrow")
                BUTTON_APPROVE.config(state = DISABLED, cursor = "arrow")
                BUTTON_VOID.config(state = DISABLED, cursor = "arrow")
                BUTTON_PRINT.config(state = NORMAL, cursor = "hand2")
            else:
                BUTTON_APPROVE.config(state = NORMAL, cursor = "hand2")
                BUTTON_VOID.config(state = DISABLED, cursor = "arrow")
                BUTTON_PRINT.config(state = NORMAL, cursor = "hand2")
                
            if self.returnAccess(USER, 2) == 0:
                BUTTON_VOID.config(state = DISABLED, cursor = "arrow")
                
            if self.returnAccess(USER, 1) == 0:
                BUTTON_APPROVE.config(state = DISABLED, cursor = "arrow")
    
    def disablePurchaseOrderWidgets(self):
        CALENDAR_DOC.config(state = DISABLED)
        ENTRY_CODE.config(state = DISABLED)
        BUTTON_CODE.config(state = DISABLED, cursor = "arrow")
        ENTRY_REMARKS.config(state = DISABLED)
        ENTRY_INV_CODE.config(state = DISABLED)
        BUTTON_INV_CODE.config(state = DISABLED, cursor = "arrow")
        BUTTON_ADD.config(state = DISABLED, cursor = "arrow")
        SPINBOX_QUANTITY.config(state = DISABLED)
        ENTRY_INV_COST.config(state = DISABLED)
        TREE_INV_ITEMS.unbind("<Button-3>")
        TREE_INV_ITEMS.unbind("<Double-1>")
        
    def deletePurchaseOrder(self):
        pass

    def searchPurchaseOrder(self, var, *args):
        find = "SELECT poNumber, DATE(date), supplierCode, remarks, inventoryCost*inventoryQuantity, delivery, status, encoder, DATE(encoded), approver, DATE(approved) FROM tblpurchaseorder WHERE (poNumber LIKE %s OR supplierCode LIKE %s OR remarks LIKE %s) AND date BETWEEN %s AND %s ORDER BY poNumber DESC"
        db.commit()
        cursor.execute(find, [f"%{var}%", f"%{var}%", f"%{var}%", CALENDAR_START.get_date(), CALENDAR_END.get_date()])
        result = cursor.fetchall()
        for i in TREE_PO.get_children():
            TREE_PO.delete(i)
        if result:
            count = 0
            ponumbers = []
            for i in result:
                if i[0] not in ponumbers:
                    if count % 2 == 0:
                        TREE_PO.insert("", "end", values = (str(i[0]).zfill(8),i[1],self.returnSupplierName(i[2]),i[3],self.returnTotalPOAmount(i[0]),self.returnPODeliverStatus(i[0]),i[6],self.returnUserName(i[7], 0),i[8],self.returnUserName(i[9], 0),i[10]), tags = "evenrow")
                    else:
                        TREE_PO.insert("", "end", values = (str(i[0]).zfill(8),i[1],self.returnSupplierName(i[2]),i[3],self.returnTotalPOAmount(i[0]),self.returnPODeliverStatus(i[0]),i[6],self.returnUserName(i[7], 0),i[8],self.returnUserName(i[9], 0),i[10]), tags = "oddrow")
                    count += 1
                    ponumbers.append(i[0])
        else:
            messagebox.showerror("Purchase Order", "No match found!")
            # db.commit()
            # cursor.execute(f"SELECT poNumber, DATE(date), supplierCode, remarks, inventoryCost*inventoryQuantity, delivery, status, encoder, DATE(encoded), approver, DATE(approved) FROM tblpurchaseorder WHERE date BETWEEN '{CALENDAR_START.get_date()}' AND '{CALENDAR_END.get_date()}' ORDER BY poNumber DESC")
            # result = cursor.fetchall()

    def refreshPurchaseOrder(self, *args):
        for i in TREE_PO.get_children():
            TREE_PO.delete(i)
        db.commit()
        cursor.execute(f"SELECT poNumber, DATE(date), supplierCode, remarks, inventoryCost*inventoryQuantity, delivery, status, encoder, DATE(encoded), approver, DATE(approved) FROM tblpurchaseorder WHERE date BETWEEN '{CALENDAR_START.get_date()}' AND '{CALENDAR_END.get_date()}' ORDER BY poNumber DESC")
        result = cursor.fetchall()

        count = 0
        ponumbers = []
        for i in result:
            if i[0] not in ponumbers:
                if count % 2 == 0:
                    TREE_PO.insert("", "end", values = (str(i[0]).zfill(8),i[1],self.returnSupplierName(i[2]),i[3],self.returnTotalPOAmount(i[0]),self.returnPODeliverStatus(i[0]),i[6],self.returnUserName(i[7], 0),i[8],self.returnUserName(i[9], 0),i[10]), tags = "evenrow")
                else:
                    TREE_PO.insert("", "end", values = (str(i[0]).zfill(8),i[1],self.returnSupplierName(i[2]),i[3],self.returnTotalPOAmount(i[0]),self.returnPODeliverStatus(i[0]),i[6],self.returnUserName(i[7], 0),i[8],self.returnUserName(i[9], 0),i[10]), tags = "oddrow")
                count += 1
                ponumbers.append(i[0])

### MENU_GSAD_RECEIVING REPORT ###
    def showAddEditReceivingReport(self):
        global TOP_RR
        TOP_RR = Toplevel()
        TOP_RR.title("Add - Receiving Report")
        TOP_RR.iconbitmap(PATH_ICON + "icon.ico")
        TOP_RR.geometry("925x500+300+100")
        TOP_RR.resizable(height = False, width = False)
        TOP_RR.grab_set()
        
        required = []
        
        FRAME_1 = Frame(TOP_RR)
        FRAME_1.grid(column = 0, row = 0, pady = TOP_PADY, sticky = W)

        FRAME_1_1 = Frame(FRAME_1)
        FRAME_1_1.grid(column = 0, row = 0, pady = TOP_PADY, sticky = W)

        LABEL_RR = Label(FRAME_1_1, text = "RR No.", font = APP_FONT)
        LABEL_RR.grid(column = 0, row = 0, pady = TOP_PADY, sticky = E)

        LABEL_DATE = Label(FRAME_1_1, text = "Doc Date", font = APP_FONT)
        LABEL_DATE.grid(column = 0, row = 1, pady = TOP_PADY, sticky = E)

        LABEL_PO = Label(FRAME_1_1, text = "PO No.", font = APP_FONT)
        LABEL_PO.grid(column = 0, row = 2, pady = TOP_PADY, sticky = E)

        LABEL_SUPPLIERCODE = Label(FRAME_1_1, text = "Supplier Code", font = APP_FONT)
        LABEL_SUPPLIERCODE.grid(column = 0, row = 3, pady = TOP_PADY, sticky = E)

        LABEL_SUPPLIER = Label(FRAME_1_1, text = "Supplier Name", font = APP_FONT)
        LABEL_SUPPLIER.grid(column = 0, row = 4, pady = TOP_PADY, sticky = E)

        LABEL_PARTICULARS = Label(FRAME_1_1, text = "Particulars", font = APP_FONT)
        LABEL_PARTICULARS.grid(column = 0, row = 5, pady = TOP_PADY, sticky = E)

        LABEL_REFERENCE = Label(FRAME_1_1, text = "Invoice No.", font = APP_FONT)
        LABEL_REFERENCE.grid(column = 0, row = 6, pady = TOP_PADY, sticky = E)

        global TEXTVAR_RR
        TEXTVAR_RR = StringVar()
        ENTRY_RR = Entry(FRAME_1_1, textvariable = TEXTVAR_RR, font = APP_FONT, width = 20, state = "readonly", justify = RIGHT)
        ENTRY_RR.grid(column = 1, row = 0, sticky = W)

        global CALENDAR_DOC
        CALENDAR_DOC = DateEntry(FRAME_1_1, firstweekday = "sunday", date_pattern = "yyyy-mm-dd")
        CALENDAR_DOC.grid(column = 1, row = 1, sticky = W)

        global TEXTVAR_PO, ENTRY_PO
        TEXTVAR_PO = StringVar()
        ENTRY_PO = Entry(FRAME_1_1, textvariable = TEXTVAR_PO, font = APP_FONT, width = 20, justify = CENTER)
        ENTRY_PO.grid(column = 1, row = 2, sticky = W)
        
        global BUTTON_GO
        BUTTON_GO = Button(FRAME_1_1, text = "GO", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "hand2", command = lambda: self.getPOdetails(TEXTVAR_PO))
        BUTTON_GO.grid(column = 1, row = 2, padx = TOP_PADX, sticky = E)
        BUTTON_GO.bind("<Return>", lambda e: self.getPOdetails(TEXTVAR_PO))

        global TEXTVAR_SUPPLIERCODE
        TEXTVAR_SUPPLIERCODE = StringVar()
        ENTRY_SUPPLIERCODE = Entry(FRAME_1_1, textvariable = TEXTVAR_SUPPLIERCODE, font = APP_FONT, width = 10, state = "readonly")
        ENTRY_SUPPLIERCODE.grid(column = 1, row = 3, sticky = W)

        global TEXTVAR_SUPPLIERNAME
        TEXTVAR_SUPPLIERNAME = StringVar()
        ENTRY_SUPPLIERNAME = Entry(FRAME_1_1, textvariable = TEXTVAR_SUPPLIERNAME, font = APP_FONT, width = 30, state = "readonly")
        ENTRY_SUPPLIERNAME.grid(column = 1, row = 4, sticky = W)

        global TEXTVAR_PARTICULARS
        TEXTVAR_PARTICULARS = StringVar()
        ENTRY_PARTICULARS = Entry(FRAME_1_1, textvariable = TEXTVAR_PARTICULARS, font = APP_FONT, width = 30, state = "readonly")
        ENTRY_PARTICULARS.grid(column = 1, row = 5, sticky = W)

        global TEXTVAR_REFERENCE, ENTRY_REFERENCE
        TEXTVAR_REFERENCE = StringVar()
        ENTRY_REFERENCE = Entry(FRAME_1_1, textvariable = TEXTVAR_REFERENCE, font = APP_FONT, width = 30)
        ENTRY_REFERENCE.grid(column = 1, row = 6, sticky = W)
        required.append(TEXTVAR_REFERENCE)

        FRAME_1_2 = Frame(FRAME_1)
        FRAME_1_2.grid(column = 1, row = 0, pady = TOP_PADY, sticky = W)

        LABEL_ENCODER = Label(FRAME_1_2, text = "Encoder", font = APP_FONT)
        LABEL_ENCODER.grid(column = 0, row = 0, pady = TOP_PADY, sticky = E)

        LABEL_ENCODED = Label(FRAME_1_2, text = "Encoded", font = APP_FONT)
        LABEL_ENCODED.grid(column = 0, row = 1, pady = TOP_PADY, sticky = E)

        LABEL_APPROVER = Label(FRAME_1_2, text = "Approver", font = APP_FONT)
        LABEL_APPROVER.grid(column = 0, row = 2, pady = TOP_PADY, sticky = E)

        LABEL_APPROVED = Label(FRAME_1_2, text = "Approved", font = APP_FONT)
        LABEL_APPROVED.grid(column = 0, row = 3, pady = TOP_PADY, sticky = E)

        global TEXTVAR_ENCODER
        TEXTVAR_ENCODER = StringVar()
        ENTRY_ENCODER = Entry(FRAME_1_2, textvariable = TEXTVAR_ENCODER, font = APP_FONT, width = 15, state = "readonly")
        ENTRY_ENCODER.grid(column = 1, row = 0, sticky = W)

        global TEXTVAR_ENCODED
        TEXTVAR_ENCODED = StringVar()
        ENTRY_ENCODED = Entry(FRAME_1_2, textvariable = TEXTVAR_ENCODED, font = APP_FONT, width = 15, state = "readonly")
        ENTRY_ENCODED.grid(column = 1, row = 1, sticky = W)

        global TEXTVAR_APPROVER
        TEXTVAR_APPROVER = StringVar()
        ENTRY_APPROVER = Entry(FRAME_1_2, textvariable = TEXTVAR_APPROVER, font = APP_FONT, width = 15, state = "readonly")
        ENTRY_APPROVER.grid(column = 1, row = 2, sticky = W)

        global TEXTVAR_APPROVED
        TEXTVAR_APPROVED = StringVar()
        ENTRY_APPROVED = Entry(FRAME_1_2, textvariable = TEXTVAR_APPROVED, font = APP_FONT, width = 15, state = "readonly")
        ENTRY_APPROVED.grid(column = 1, row = 3, sticky = W)
        
        LABEL_GROSS = Label(FRAME_1_2, text = "Gross", font = APP_FONT)
        LABEL_GROSS.grid(column = 2, row = 0, pady = TOP_PADY, sticky = E)

        LABEL_EXPENSE = Label(FRAME_1_2, text = "Expense", font = APP_FONT)
        LABEL_EXPENSE.grid(column = 2, row = 1, pady = TOP_PADY, sticky = E)

        LABEL_VAT = Label(FRAME_1_2, text = "VAT", font = APP_FONT)
        LABEL_VAT.grid(column = 2, row = 2, pady = TOP_PADY, sticky = E)

        LABEL_EWT = Label(FRAME_1_2, text = "EWT", font = APP_FONT)
        LABEL_EWT.grid(column = 2, row = 3, pady = TOP_PADY, sticky = E)

        LABEL_NET = Label(FRAME_1_2, text = "Net", font = APP_FONT)
        LABEL_NET.grid(column = 2, row = 4, pady = TOP_PADY, sticky = NE)
        
        LABEL_TAX = Label(FRAME_1_2, text = "Tax", font = APP_FONT)
        LABEL_TAX.grid(column = 2, row = 5, pady = TOP_PADY, sticky = NE)
        
        global TEXTVAR_GROSS
        TEXTVAR_GROSS = StringVar()
        ENTRY_GROSS = Entry(FRAME_1_2, textvariable = TEXTVAR_GROSS, font = APP_FONT, width = 20, state = "readonly", justify = RIGHT)
        ENTRY_GROSS.grid(column = 3, row = 0, sticky = W)

        global TEXTVAR_EXPENSE
        TEXTVAR_EXPENSE = StringVar()
        ENTRY_EXPENSE = Entry(FRAME_1_2, textvariable = TEXTVAR_EXPENSE, font = APP_FONT, width = 20, state = "readonly", justify = RIGHT)
        ENTRY_EXPENSE.grid(column = 3, row = 1, sticky = W)

        global TEXTVAR_VAT
        TEXTVAR_VAT = StringVar()
        ENTRY_VAT = Entry(FRAME_1_2, textvariable = TEXTVAR_VAT, font = APP_FONT, width = 20, state = "readonly", justify = RIGHT)
        ENTRY_VAT.grid(column = 3, row = 2, sticky = W)

        global TEXTVAR_EWT
        TEXTVAR_EWT = StringVar()
        ENTRY_EWT = Entry(FRAME_1_2, textvariable = TEXTVAR_EWT, font = APP_FONT, width = 20, state = "readonly", justify = RIGHT)
        ENTRY_EWT.grid(column = 3, row = 3, sticky = W)

        global TEXTVAR_NET
        TEXTVAR_NET = StringVar()
        ENTRY_NET = Entry(FRAME_1_2, textvariable = TEXTVAR_NET, font = APP_FONT, width = 20, state = "readonly", justify = RIGHT)
        ENTRY_NET.grid(column = 3, row = 4, sticky = N)
        required.append(TEXTVAR_NET)
        
        global TEXTVAR_TAX
        TEXTVAR_TAX = StringVar()
        COMBO_TAX = tk.Combobox(FRAME_1_2, values = ["VAT", "Non-VAT"], textvariable = TEXTVAR_TAX, font = APP_FONT, width = 10, state = DISABLED)
        COMBO_TAX.grid(column = 3, row = 5, pady = TOP_PADY, sticky = W)

        FRAME_2 = Frame(TOP_RR)
        FRAME_2.grid(column = 0, row = 1, sticky = W)

        FRAME_2_1 = Frame(FRAME_2)
        FRAME_2_1.grid(column = 0, row = 0, pady = TOP_PADY, sticky = W)
        
        FRAME_2_2 = Frame(FRAME_2)
        FRAME_2_2.grid(column = 0, row = 1, pady = TOP_PADY, sticky = W)
        
        FRAME_2_3 = Frame(FRAME_2)
        FRAME_2_3.grid(column = 0, row = 2, pady = TOP_PADY, sticky = W)
        
        FRAME_2_4 = Frame(FRAME_2)
        FRAME_2_4.grid(column = 0, row = 3, pady = TOP_PADY, sticky = E)
        
        global TEXTVAR_INV_CODE
        TEXTVAR_INV_CODE = StringVar()
        ENTRY_INV_CODE = Entry(FRAME_2_1, textvariable = TEXTVAR_INV_CODE, font = APP_FONT, width = 15, state = "readonly")
        ENTRY_INV_CODE.grid(column = 0, row = 0, padx = TOP_PADX, sticky = W)
        
        global BUTTON_INV_CODE
        BUTTON_INV_CODE = Button(FRAME_2_1, text = "...", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, cursor = "arrow", state = DISABLED)
        BUTTON_INV_CODE.grid(column = 0, row = 0, padx = TOP_PADX, sticky = E)
        
        global TEXTVAR_INV_NAME
        TEXTVAR_INV_NAME = StringVar()
        ENTRY_INV_NAME = Entry(FRAME_2_1, textvariable = TEXTVAR_INV_NAME, font = APP_FONT, width = 69, state = "readonly")
        ENTRY_INV_NAME.grid(column = 1, row = 0, sticky = W)
        
        global TEXTVAR_QUANTITY
        TEXTVAR_QUANTITY = StringVar()
        SPINBOX_QUANTITY = Spinbox(FRAME_2_2, textvariable = TEXTVAR_QUANTITY, font = APP_FONT, width = 6, from_ = 0, to = 99999)
        SPINBOX_QUANTITY.grid(column = 0, row = 0, padx = TOP_PADX, sticky = W)
        SPINBOX_QUANTITY.bind("<FocusOut>", lambda e: self.validateInteger(TEXTVAR_QUANTITY))
        
        global TEXTVAR_INV_COST
        TEXTVAR_INV_COST = StringVar()
        ENTRY_INV_COST = Entry(FRAME_2_2, textvariable = TEXTVAR_INV_COST, font = APP_FONT, width = 15, justify = RIGHT)
        ENTRY_INV_COST.grid(column = 1, row = 0, padx = TOP_PADX, sticky = W)
        ENTRY_INV_COST.bind("<FocusOut>", lambda e: self.validateAmount(TEXTVAR_INV_COST))
        
        global TEXTVAR_INV_STOCK
        TEXTVAR_INV_STOCK = StringVar()
        ENTRY_INV_STOCK = Entry(FRAME_2_2, textvariable = TEXTVAR_INV_STOCK, font = APP_FONT, width = 15, state = "readonly")
        ENTRY_INV_STOCK.grid(column = 2, row = 0, padx = TOP_PADX, sticky = W)

        global BUTTON_UPDATE
        BUTTON_UPDATE = Button(FRAME_2_2, text = "UPDATE", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "arrow", state = DISABLED, command = self.updateRRInventoryItem)
        BUTTON_UPDATE.grid(column = 4, row = 0, padx = TOP_PADX)
        BUTTON_UPDATE.bind("<Return>", self.updateRRInventoryItem)
        
        global BUTTON_CANCEL
        BUTTON_CANCEL = Button(FRAME_2_2, text = "CANCEL", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "arrow", state = DISABLED, command = self.cancelEditRRInventoryItem)
        BUTTON_CANCEL.grid(column = 5, row = 0, padx = TOP_PADX)
        BUTTON_CANCEL.bind("<Return>", self.cancelEditRRInventoryItem)
        
        global TREE_INV_ITEMS
        TREE_INV_ITEMS = tk.Treeview(FRAME_2_3, height = 6, selectmode = "browse")
        TREE_INV_ITEMS["columns"] = ("Code", "Name", "Cost", "UOM", "Ordered", "Delivered", "Total")
        TREE_INV_ITEMS.column("#0", width = 0, minwidth = 0, stretch = True)
        TREE_INV_ITEMS.column("Code", anchor = W, minwidth = 75, width = 75)
        TREE_INV_ITEMS.column("Name", anchor = W, minwidth = 100, width = 200)
        TREE_INV_ITEMS.column("Cost", anchor = E, minwidth = 100, width = 100)
        TREE_INV_ITEMS.column("UOM", anchor = W, minwidth = 100, width = 50)
        TREE_INV_ITEMS.column("Ordered", anchor = E, minwidth = 100, width = 150)
        TREE_INV_ITEMS.column("Delivered", anchor = E, minwidth = 100, width = 150)
        TREE_INV_ITEMS.column("Total", anchor = E, minwidth = 100, width = 125)
        
        TREE_INV_ITEMS.heading("#0", text = "", anchor = W)
        TREE_INV_ITEMS.heading("Code", text = "Code", anchor = N, command = lambda: self.sortColumn(TREE_INV_ITEMS, "Cost", False))
        TREE_INV_ITEMS.heading("Name", text = "Name", anchor = N, command = lambda: self.sortColumn(TREE_INV_ITEMS, "Name", False))
        TREE_INV_ITEMS.heading("Cost", text = "Cost", anchor = N, command = lambda: self.sortColumn(TREE_INV_ITEMS, "Cost", False))
        TREE_INV_ITEMS.heading("UOM", text = "UOM", anchor = N, command = lambda: self.sortColumn(TREE_INV_ITEMS, "Quantity", False))
        TREE_INV_ITEMS.heading("Ordered", text = "Ordered", anchor = N, command = lambda: self.sortColumn(TREE_INV_ITEMS, "Quantity", False))
        TREE_INV_ITEMS.heading("Delivered", text = "Delivered", anchor = N, command = lambda: self.sortColumn(TREE_INV_ITEMS, "Quantity", False))
        TREE_INV_ITEMS.heading("Total", text = "Total", anchor = N, command = lambda: self.sortColumn(TREE_INV_ITEMS, "Total", False))

        global POPUP_INV_ITEMS
        POPUP_INV_ITEMS = Menu(TREE_INV_ITEMS, tearoff = 0)
        POPUP_INV_ITEMS.add_command(command = self.editRRInventoryItem, label = "Update")
        POPUP_INV_ITEMS.add_command(command = lambda: self.deleteRRTreeItem(TREE_INV_ITEMS), label = "Delete")
        TREE_INV_ITEMS.bind("<Button-3>", lambda e: self.popupMenu(TREE_INV_ITEMS, POPUP_INV_ITEMS, e))
        TREE_INV_ITEMS.bind("<Double-1>", self.editRRInventoryItem)

        global STYLE_INV_ITEMS
        STYLE_INV_ITEMS = tk.Style()
        STYLE_INV_ITEMS.map("Treeview", foreground = self.fixedMap("foreground", STYLE_INV_ITEMS), background = self.fixedMap("background", STYLE_INV_ITEMS))

        TREE_INV_ITEMS.tag_configure("oddrow", background = None)
        TREE_INV_ITEMS.tag_configure("evenrow", background = TREE_TAG_EVENROW)

        YSCROLLBAR = tk.Scrollbar(FRAME_2_3, orient = "vertical", command = TREE_INV_ITEMS.yview)
        YSCROLLBAR.pack(side = "right", fill = "y")

        XSCROLLBAR = tk.Scrollbar(FRAME_2_3, orient = "horizontal", command = TREE_INV_ITEMS.xview)
        
        TREE_INV_ITEMS.configure(yscrollcommand = YSCROLLBAR.set, xscrollcommand = XSCROLLBAR.set) 
        TREE_INV_ITEMS.pack()
        XSCROLLBAR.pack(fill ="x")
        
        global BUTTON_SAVE
        BUTTON_SAVE = Button(FRAME_2_4, text = "SUBMIT", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "hand2", command = self.saveReceivingReport)
        BUTTON_SAVE.grid(column = 1, row = 1, padx = TOP_PADX)
        BUTTON_SAVE.bind("<Return>", self.saveReceivingReport)
        
        global BUTTON_APPROVE
        BUTTON_APPROVE = Button(FRAME_2_4, text = "APPROVE", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "arrow", state = DISABLED, command = self.approveReceivingReport)
        BUTTON_APPROVE.grid(column = 2, row = 1, padx = TOP_PADX)
        BUTTON_APPROVE.bind("<Return>", self.approveReceivingReport)
        
        global BUTTON_VOID
        BUTTON_VOID = Button(FRAME_2_4, text = "VOID", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "arrow", state = DISABLED, command = None)
        BUTTON_VOID.grid(column = 3, row = 1, padx = TOP_PADX)
        BUTTON_VOID.bind("<Return>", self.voidReceivingReport)
        
        global BUTTON_PRINT
        BUTTON_PRINT = Button(FRAME_2_4, text = "PRINT", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "arrow", state = DISABLED, command = self.printReceivingReport)
        BUTTON_PRINT.grid(column = 4, row = 1, padx = TOP_PADX)
        BUTTON_PRINT.bind("<Return>", self.printReceivingReport)
        
        global BUTTON_CLOSE
        BUTTON_CLOSE = Button(FRAME_2_4, text = "CLOSE", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "hand2", command = lambda: self.closeTopLevel(TOP_RR))
        BUTTON_CLOSE.grid(column = 5, row = 1, padx = TOP_PADX)
        BUTTON_CLOSE.bind("<Return>", None)

    def getPOdetails(self, var):
        try:
            if int(var.get()) > 0:
                select = "SELECT supplierCode, remarks, inventoryCode, inventoryCost, inventoryQuantity, encoder, DATE(encoded) FROM tblpurchaseorder WHERE poNumber = %s AND status = %s"
                db.commit()
                cursor.execute(select, [int(var.get()), "Approved"])
                result = cursor.fetchall()
                if result:
                    # if self.returnPODeliverStatus(int(var.get())) == "No Action" or self.returnPODeliverStatus(int(var.get())) == "Partially completed":
                    var.set(str(var.get()).zfill(8))
                    TEXTVAR_SUPPLIERCODE.set(result[0][0])
                    TEXTVAR_SUPPLIERNAME.set(self.returnSupplierName(result[0][0]))
                    TEXTVAR_PARTICULARS.set(result[0][1])
                    TEXTVAR_TAX.set(self.getSupplierTaxType(result[0][0]))
                    count = 0
                    for i in TREE_INV_ITEMS.get_children():
                        TREE_INV_ITEMS.delete(i)
                    for i in result:
                        if count % 2 == 0:
                            TREE_INV_ITEMS.insert("", "end", values = (i[2],self.returnInventoryNameUOM(i[2])[0],self.validateAmount2(float(i[3])),self.returnInventoryNameUOM(i[2])[1],self.returnPOUndeliveredQuantity(i[2], int(var.get())),self.returnPOUndeliveredQuantity(i[2], int(var.get())),self.validateAmount2(float(i[3])*self.returnPOUndeliveredQuantity(i[2], int(var.get())))), tags = "evenrow")
                        else:
                            TREE_INV_ITEMS.insert("", "end", values = (i[2],self.returnInventoryNameUOM(i[2])[0],self.validateAmount2(float(i[3])),self.returnInventoryNameUOM(i[2])[1],self.returnPOUndeliveredQuantity(i[2], int(var.get())),self.returnPOUndeliveredQuantity(i[2], int(var.get())),self.validateAmount2(float(i[3])*self.returnPOUndeliveredQuantity(i[2], int(var.get())))), tags = "oddrow")
                        count += 1
                    self.updateRRTotals()
                    # else:
                    #     # print("h")
                else:
                    var.set("")
                    TEXTVAR_SUPPLIERCODE.set("")
                    TEXTVAR_SUPPLIERNAME.set("")
                    TEXTVAR_PARTICULARS.set("")
                    TEXTVAR_TAX.set("")
                    for i in TREE_INV_ITEMS.get_children():
                        TREE_INV_ITEMS.delete(i)
                    self.updateRRTotals()
        except:
            # print(e)
            var.set("")
            TEXTVAR_SUPPLIERCODE.set("")
            TEXTVAR_SUPPLIERNAME.set("")
            TEXTVAR_PARTICULARS.set("")
            TEXTVAR_TAX.set("")
            for i in TREE_INV_ITEMS.get_children():
                TREE_INV_ITEMS.delete(i)
            self.updateRRTotals()

    def saveReceivingReport(self):
        insert = """INSERT INTO tblreceivingreports (
            rrNUmber, docDate, poNumber, remarks, inventoryCode,
            inventoryCost, inventoryQuantity, isApproved, isVoid, encoder,
            encoded
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
        
        delete = "DELETE FROM tblreceivingreports WHERE rrNumber = %s"
        if self.returnFloatAmount(TEXTVAR_NET.get()) != 0:
            ask = messagebox.askyesno("Receiving Report", "Are you sure?")
            if ask:
                if TEXTVAR_RR.get() != "":
                    cursor.execute(delete, [int(TEXTVAR_RR.get())])
                    db.commit()
                    message = "Receiving Report has been updated!"
                else:
                    self.generateRRNumber()
                    message = "Receiving Report has been saved!"
                rritems = []
                if self.returnFloatAmount(TEXTVAR_NET.get()) > 0:
                    for i in TREE_INV_ITEMS.get_children():
                        if int(TREE_INV_ITEMS.item(i)["values"][4]) != 0:
                            rritems.append([
                                int(TEXTVAR_RR.get()), CALENDAR_DOC.get_date(), TEXTVAR_PO.get(), TEXTVAR_REFERENCE.get(), TREE_INV_ITEMS.item(i)["values"][0],
                                self.returnFloatAmount(TREE_INV_ITEMS.item(i)["values"][2]), int(TREE_INV_ITEMS.item(i)["values"][5]), "No", "No", USER,
                                datetime.datetime.now()
                            ])
                        
                    if len(rritems) != 0:
                        for i in rritems:
                            cursor.execute(insert, i)
                        db.commit()
                        messagebox.showinfo("Receiving Report", message)
                        TOP_RR.grab_release()
                        TOP_RR.destroy()
                        
                        if self.returnAccess(USER, 4) == 0:
                            BUTTON_APPROVE.config(state = DISABLED, cursor = "arrow")
                    else:
                        TEXTVAR_RR.set("")
                        messagebox.showerror("Receiving Report", "This PO has already been fulfilled!")
        else:
            messagebox.showerror("Receiving Report", "Total Net is zero")

    def approveReceivingReport(self):
        insertbook = """INSERT INTO tblinventorybook (
            docDate, inventoryCode, cost, quantity, reference,
            source, encoder, encoded
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)"""
            
        update = """UPDATE tblreceivingreports SET isApproved = %s, approver = %s, approved = %s WHERE rrNumber = %s"""
        
        insertAPV = """INSERT INTO tblpayables (
            apvNumber, apvDate, glDate, dueDate, supplierCode,
            particulars, reference, rrNumber, chartCode, centerCode,
            amount, taxType, encoder, encoded, isApproved, isVoid
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""

        rrbookitems = []
        deletebook = "DELETE FROM tblinventorybook WHERE reference = %s"
        if self.returnFloatAmount(TEXTVAR_NET.get()) > 0:
            for i in TREE_INV_ITEMS.get_children():
                rrbookitems.append([
                    CALENDAR_DOC.get_date(), TREE_INV_ITEMS.item(i)["values"][0], self.returnFloatAmount(TREE_INV_ITEMS.item(i)["values"][2]), int(TREE_INV_ITEMS.item(i)["values"][5]), int(TEXTVAR_RR.get()),
                    "RR", USER, datetime.datetime.now()
                ])
                
            if len(rrbookitems) != 0:
                ask = messagebox.askyesno("APPROVE Receiving Report", "Are you sure?")
                if ask:
                    cursor.execute(deletebook, [int(TEXTVAR_RR.get())])
                    db.commit()
                    cursor.execute(update, ["Yes", USER, datetime.datetime.now(), int(TEXTVAR_RR.get())])
                    db.commit()
                    for x in rrbookitems:
                        cursor.execute(insertbook, x)
                        db.commit()
                    
                    lines = []
                    for i in TREE_INV_ITEMS.get_children():
                        if self.returnFloatAmount(TREE_INV_ITEMS.item(i)["values"][6]) != 0:
                            lines.append([self.returnChartCodeOfItem(TREE_INV_ITEMS.item(i)["values"][0], 0), self.returnFloatAmount(TREE_INV_ITEMS.item(i)["values"][6])])

                    df = pd.DataFrame(lines)
                    sumlines = df.groupby([0], as_index = False).sum()
                    sumlist = sumlines.values.tolist()
                    
                    if TEXTVAR_TAX.get() == "VAT":
                        tax = "WV-01"
                    else:
                        tax = "NV-01"
                    
                    insertapv = []
                    apvnum = int(self.generateAPVNumber())
                    for i in range(len(sumlist)):
                        insertapv.append([
                            apvnum, CALENDAR_DOC.get_date(), self.returnLastDayOfMonth(str(CALENDAR_DOC.get_date())), self.returnLastDayOfMonth(str(CALENDAR_DOC.get_date())), TEXTVAR_SUPPLIERCODE.get(),
                            TEXTVAR_PARTICULARS.get(), TEXTVAR_REFERENCE.get(), int(TEXTVAR_RR.get()), sumlist[i][0], 5200,
                            sumlist[i][1], tax, "0000-A0000", datetime.datetime.now(), "No", "No"
                            ])
                    for i in insertapv:
                        cursor.execute(insertAPV, i)
                    db.commit()
                    TEXTVAR_APPROVER.set(self.returnUserName(USER, 0))
                    TEXTVAR_APPROVED.set(datetime.date.today())
                    messagebox.showinfo("APPROVE Receiving Report", "Receiving Report has been approved")
                    self.disableRRWidgets()
                    BUTTON_SAVE.config(state = DISABLED, cursor = "arrow")
                    BUTTON_APPROVE.config(state = DISABLED, cursor = "arrow")
                    BUTTON_VOID.config(state = NORMAL, cursor = "hand2")
                    BUTTON_PRINT.config(state = NORMAL, cursor = "hand2")
                    
                    if self.returnAccess(USER, 5) == 0:
                        BUTTON_VOID.config(state = DISABLED, cursor = "arrow")

    def returnChartCodeOfItem(self, var, i):
        select = "SELECT assetCode, expenseCode FROM tblinventory WHERE code = %s LIMIT 1"
        cursor.execute(select, [var])
        result = cursor.fetchone()
        if result:
            return result[i]

    def voidReceivingReport(self):
        ask = messagebox.askyesno("VOID Receiving Report", "Are you sure?")
        if ask:
            void = "UPDATE tblreceivingreports SET isVoid = %s WHERE rrNumber = %s"
            cursor.execute(void, ["Yes", int(TEXTVAR_RR.get())])
            db.commit()
            
            delete = "DELETE FROM tblinventorybook WHERE source = %s, reference = %s"
            cursor.execute(delete, ["RR", int(TEXTVAR_RR.get())])
            db.commit()
            
            messagebox.showinfo("VOID Receiving Report", "Are you sure?")
            BUTTON_VOID.config(state = DISABLED, cursor = "arrow")

    def printReceivingReport(self):
        wb = load_workbook(PATH_TEMPLATE + "RR.xlsx")
        sheet = wb.active
        sheet["A1"] = f"Printed Date: {datetime.datetime.now()} {self.returnUserName(USER, 0)}"
        sheet["H3"] = TEXTVAR_RR.get()
        sheet["B9"] = TEXTVAR_SUPPLIERNAME.get()
        sheet["I9"] = datetime.datetime.strptime(str(CALENDAR_DOC.get_date()), '%Y-%m-%d').strftime('%B %d, %Y')
        sheet["B10"] = TEXTVAR_PARTICULARS.get()
        sheet["B11"] = TEXTVAR_PO.get()
        sheet["I11"] = TEXTVAR_REFERENCE.get()
        
        sheet["I104"] = self.returnFloatAmount(TEXTVAR_GROSS.get())
        
        sheet["A106"] = f"{self.returnUserName(USER, 1).upper()} {self.returnUserName(USER, 2).upper()}"
        
        count = 13
        itemNumber = 1
        for i in TREE_INV_ITEMS.get_children():
            if self.returnFloatAmount(TREE_INV_ITEMS.item(i)['values'][6]) != 0:
                sheet["A" + str(count)] = itemNumber
                sheet["B" + str(count)] = TREE_INV_ITEMS.item(i)['values'][1]
                sheet["F" + str(count)] = TREE_INV_ITEMS.item(i)['values'][5]
                sheet["G" + str(count)] = TREE_INV_ITEMS.item(i)['values'][3]
                sheet["H" + str(count)] = self.returnFloatAmount(TREE_INV_ITEMS.item(i)['values'][2])
                sheet["I" + str(count)] = self.returnFloatAmount(TREE_INV_ITEMS.item(i)['values'][6])
                itemNumber += 1
                count += 1

        for i in range(13,103):
            cell = sheet["H" + str(i)].value
            if cell == None:
                sheet.row_dimensions[i].hidden = True
                    
        wb.save(PATH_SAVE + "RR" + TEXTVAR_RR.get() + ".xlsx")
        startfile(PATH_SAVE + "RR" + TEXTVAR_RR.get() + ".xlsx", "open")

    def returnPOUndeliveredQuantity(self, var, num):
        select = "SELECT tblpurchaseorder.inventoryCode, tblpurchaseorder.inventoryQuantity, tblreceivingreports.inventoryQuantity FROM tblpurchaseorder INNER JOIN tblreceivingreports ON tblpurchaseorder.poNumber = tblreceivingreports.poNumber AND tblpurchaseorder.inventoryCode = tblreceivingreports.inventoryCode WHERE tblpurchaseorder.inventoryCode = %s AND tblpurchaseorder.poNumber = %s AND tblreceivingreports.isApproved = 'Yes' AND tblreceivingreports.isVoid = 'No' LIMIT 1"
        cursor.execute(select, [var, num])
        result = cursor.fetchone()
        if result:
            if result[1]-result[2] >= 0:
                return result[1]-result[2]
            else:
                return 0
        else:
            return self.returnPOItemQuantity(var, num)
    
    def returnPODeliverStatus(self, num):
        select = "SELECT inventoryCode, inventoryQuantity FROM tblpurchaseorder WHERE poNumber = %s"
        db.commit()
        cursor.execute(select, [num])
        result = cursor.fetchall()
        undelivered = []
        totalitems = []
        if result:
            for i in result:
                if i[1]-self.returnRRItemQuantity(i[0], num) >= 0:
                    undelivered.append(i[1]-self.returnRRItemQuantity(i[0], num))
                    totalitems.append(i[1])
        if sum(undelivered) > 0 and sum(undelivered) < sum(totalitems):
            return "Partially completed"
        elif sum(undelivered) == 0:
            return "Completed"
        elif sum(undelivered) < 0:
            return "Completed"
        elif sum(undelivered) == sum(totalitems):
            return "No Action"

    def returnPONumber(self, num):
        select = "SELECT poNUmber FROM tblreceivingreports WHERE rrNumber = %s LIMIT 1"
        cursor.execute(select, [num])
        result = cursor.fetchone()
        if result:
            return result[0]
        else:
            return None

    def editReceivingReport(self, *args):
        self.copySelection(TREE_RR)
        self.showAddEditReceivingReport()
        TOP_RR.title("Edit - Receiving Report")
        
        find = """SELECT 
                rrNumber, docDate, poNumber, remarks, inventoryCode, 
                inventoryCost, inventoryQuantity, isApproved, isVoid, encoder,
                DATE(encoded), approver, DATE(approved) 
                    FROM tblreceivingreports WHERE rrNumber = %s"""
        db.commit()
        cursor.execute(find, [int(content[0])])
        result = cursor.fetchall()
        if result:
            TEXTVAR_RR.set(str(result[0][0]).zfill(8))
            TEXTVAR_PO.set(result[0][2])
            ENTRY_PO.config(state = DISABLED)
            self.getPOdetails(TEXTVAR_PO)
            CALENDAR_DOC.set_date(result[0][1])
            TEXTVAR_REFERENCE.set(result[0][3])
            TEXTVAR_ENCODER.set(self.returnUserName(result[0][9], 0))
            TEXTVAR_ENCODED.set(result[0][10])
            TEXTVAR_APPROVER.set(self.returnUserName(result[0][11], 0))
            TEXTVAR_APPROVED.set(result[0][12])
            for i in TREE_INV_ITEMS.get_children():
                TREE_INV_ITEMS.delete(i)
            count = 0
            for i in result:
                if count % 2 == 0:
                    TREE_INV_ITEMS.insert("", "end", values = (i[4],self.returnInventoryNameUOM(i[4])[0],self.validateAmount2(float(i[5])),self.returnInventoryNameUOM(i[4])[1],self.returnPOItemQuantity(i[4], int(content[1])),i[6],self.validateAmount2(float(i[5])*i[6])), tags = "evenrow")
                else:
                    TREE_INV_ITEMS.insert("", "end", values = (i[4],self.returnInventoryNameUOM(i[4])[0],self.validateAmount2(float(i[5])),self.returnInventoryNameUOM(i[4])[1],self.returnPOItemQuantity(i[4], int(content[1])),i[6],self.validateAmount2(float(i[5])*i[6])), tags = "oddrow")
                count += 1
                self.updateRRTotals()
            
            if result[0][7] == "Yes" and result[0][8] == "No":
                self.disableRRWidgets()
                BUTTON_SAVE.config(state = DISABLED, cursor = "arrow")
                BUTTON_APPROVE.config(state = DISABLED, cursor = "arrow")
                BUTTON_VOID.config(state = NORMAL, cursor = "hand2")
                BUTTON_PRINT.config(state = NORMAL, cursor = "hand2")
            elif result[0][7] == "Yes" and result[0][8] == "Yes":
                self.disableRRWidgets()
                BUTTON_SAVE.config(state = DISABLED, cursor = "arrow")
                BUTTON_APPROVE.config(state = DISABLED, cursor = "arrow")
                BUTTON_VOID.config(state = DISABLED, cursor = "arrow")
                BUTTON_PRINT.config(state = NORMAL, cursor = "hand2")
            elif result[0][7] == "No" and result[0][8] == "No":
                BUTTON_SAVE.config(state = NORMAL, cursor = "hand2")
                BUTTON_APPROVE.config(state = NORMAL, cursor = "hand2")
                BUTTON_VOID.config(state = DISABLED, cursor = "arrow")
                BUTTON_PRINT.config(state = NORMAL, cursor = "hand2")
                
            if self.returnAccess(USER, 5) == 0:
                BUTTON_VOID.config(state = DISABLED, cursor = "arrow")
                
            if self.returnAccess(USER, 4) == 0:
                BUTTON_APPROVE.config(state = DISABLED, cursor = "arrow")

    def disableRRWidgets(self):
        ENTRY_PO.config(state = DISABLED)
        BUTTON_GO.config(state = DISABLED, cursor = "arrow")
        CALENDAR_DOC.config(state = DISABLED)
        ENTRY_REFERENCE.config(state = DISABLED)
        TREE_INV_ITEMS.unbind("<Button-3>")
        TREE_INV_ITEMS.unbind("<Double-1>")

    def editRRInventoryItem(self, *args):
        global origpoquantity
        self.copySelection(TREE_INV_ITEMS)
        origpoquantity = content[4]
        TEXTVAR_INV_CODE.set(content[0])
        TEXTVAR_QUANTITY.set(content[5])
        TEXTVAR_INV_NAME.set(content[1])
        TEXTVAR_INV_COST.set(content[2])
        
        BUTTON_UPDATE.config(state = NORMAL, cursor = "hand2")
        BUTTON_CANCEL.config(state = NORMAL, cursor = "hand2")

    def updateRRInventoryItem(self):
        if BUTTON_UPDATE["state"] == NORMAL:
            self.deleteRRTreeItem(TREE_INV_ITEMS)
        count = 0
        if count % 2 == 0:
            TREE_INV_ITEMS.insert("", "end", values = (TEXTVAR_INV_CODE.get(),TEXTVAR_INV_NAME.get(),self.returnValidatedAmount(TEXTVAR_INV_COST),self.returnInventoryNameUOM(TEXTVAR_INV_CODE.get())[1],origpoquantity,TEXTVAR_QUANTITY.get(),self.validateAmount2(float(TEXTVAR_INV_COST.get().replace(",",""))*int(TEXTVAR_QUANTITY.get()))), tags = "evenrow")
        else:
            TREE_INV_ITEMS.insert("", "end", values = (TEXTVAR_INV_CODE.get(),TEXTVAR_INV_NAME.get(),self.returnValidatedAmount(TEXTVAR_INV_COST),self.returnInventoryNameUOM(TEXTVAR_INV_CODE.get())[1],origpoquantity,TEXTVAR_QUANTITY.get(),self.validateAmount2(float(TEXTVAR_INV_COST.get().replace(",",""))*int(TEXTVAR_QUANTITY.get()))), tags = "oddrow")
        count += 1
        TEXTVAR_INV_CODE.set("")
        TEXTVAR_INV_NAME.set("")
        TEXTVAR_QUANTITY.set(0)
        TEXTVAR_INV_COST.set(0)
        TEXTVAR_INV_STOCK.set(0)
        
        self.updateRRTotals()
        
        BUTTON_UPDATE.config(state = DISABLED, cursor = "arrow")
        BUTTON_CANCEL.config(state = DISABLED, cursor = "arrow")

    def cancelEditRRInventoryItem(self):
        TEXTVAR_INV_CODE.set("")
        TEXTVAR_QUANTITY.set(0)
        TEXTVAR_INV_NAME.set("")
        TEXTVAR_INV_COST.set(0)
        
        BUTTON_UPDATE.config(state = DISABLED, cursor = "arrow")
        BUTTON_CANCEL.config(state = DISABLED, cursor = "arrow")

    def generateRRNumber(self):
        cursor.execute("SELECT MAX(rrNumber) FROM tblreceivingreports LIMIT 1")
        result = cursor.fetchone()
        if result[0] == None:
            TEXTVAR_RR.set(str(356).zfill(8))
        else:
            TEXTVAR_RR.set(str(int(result[0])+1).zfill(8))

    def returnPOItemQuantity(self, var, num):
        select = "SELECT inventoryCode, inventoryQuantity FROM tblpurchaseorder WHERE inventoryCode = %s AND poNumber = %s LIMIT 1"
        cursor.execute(select , [var, num])
        result = cursor.fetchone()
        if result:
            return result[1]

    def returnRRItemQuantity(self, var, num):
        select = "SELECT SUM(inventoryQuantity) FROM tblreceivingreports WHERE inventoryCode = %s AND poNumber = %s AND isApproved = 'Yes' AND isVoid = 'No' LIMIT 1"
        cursor.execute(select, [var, num])
        result = cursor.fetchone()
        if result[0] != None:
            return result[0]
        else:
            return 0

    def updateRRTotals(self):
        if TEXTVAR_TAX.get() == "VAT":
            vat = 1.12
            multiplier = 0.12
        else:
            vat = 1
            multiplier = 1
        total = []
        for i in TREE_INV_ITEMS.get_children():
            total.append(float(str(TREE_INV_ITEMS.item(i)['values'][6]).replace(",","")))
        TEXTVAR_GROSS.set(self.validateAmount2(sum(total)))
        TEXTVAR_EXPENSE.set(self.validateAmount2(sum(total)/vat))
        TEXTVAR_VAT.set(self.validateAmount2((sum(total)/vat)*multiplier))
        TEXTVAR_EWT.set(self.validateAmount2(round((sum(total)/vat)*.01, 3)))
        TEXTVAR_NET.set(self.validateAmount2(sum(total)-round((sum(total)/vat)*.01, 3)))

    def getSupplierTaxType(self, var):
        select = "SELECT vatable FROM tblsuppliers WHERE code = %s LIMIT 1"
        cursor.execute(select, [var])
        result = cursor.fetchone()
        if result:
            return result[0]

    def searchReceivingReport(self, var, *args):
        find = "SELECT rrNumber, docDate, tblpurchaseorder.supplierCode, isApproved, tblreceivingreports.encoder, DATE(tblreceivingreports.encoded), tblreceivingreports.approver, DATE(tblreceivingreports.approved), tblreceivingreports.PONumber FROM tblreceivingreports INNER JOIN tblpurchaseorder ON tblreceivingreports.poNumber = tblpurchaseorder.poNumber WHERE (rrNumber LIKE %s OR tblpurchaseorder.supplierCode LIKE %s) AND docDate BETWEEN %s AND %s ORDER BY rrNumber DESC"
        db.commit()
        cursor.execute(find, [f"%{var}%", f"%{var}%", CALENDAR_START.get_date(), CALENDAR_END.get_date()])
        result = cursor.fetchall()
        for i in TREE_RR.get_children():
            TREE_RR.delete(i)
        if result:
            count = 0
            rrnumbers = []
            for i in result:
                if i[0] not in rrnumbers:
                    if count % 2 == 0:
                        TREE_RR.insert("", "end", values = (str(i[0]).zfill(8), str(i[8]).zfill(8),i[1],self.returnSupplierName(i[2]),i[3],self.returnUserName(i[4], 0),i[5],self.returnUserName(i[6], 0),i[7]), tags = "evenrow")
                    else:
                        TREE_RR.insert("", "end", values = (str(i[0]).zfill(8), str(i[8]).zfill(8),i[1],self.returnSupplierName(i[2]),i[3],self.returnUserName(i[4], 0),i[5],self.returnUserName(i[6], 0),i[7]), tags = "oddrow")
                    count += 1
                    rrnumbers.append(i[0])
        else:
            messagebox.showerror("Receiving Report", "No match found!")
            # db.commit()
            # cursor.execute(f"SELECT rrNumber, docDate, tblpurchaseorder.supplierCode, isApproved, tblreceivingreports.encoder, DATE(tblreceivingreports.encoded), tblreceivingreports.approver, DATE(tblreceivingreports.approved), tblreceivingreports.PONumber FROM tblreceivingreports INNER JOIN tblpurchaseorder ON tblreceivingreports.poNumber = tblpurchaseorder.poNumber WHERE docDate BETWEEN '{CALENDAR_START.get_date()}' AND '{CALENDAR_END.get_date()}' ORDER BY rrNumber DESC")
            # result = cursor.fetchall()

    def refreshReceivingReport(self, *args):
        for i in TREE_RR.get_children():
            TREE_RR.delete(i)
        db.commit()
        cursor.execute(f"SELECT rrNumber, docDate, tblpurchaseorder.supplierCode, isApproved, tblreceivingreports.encoder, DATE(tblreceivingreports.encoded), tblreceivingreports.approver, DATE(tblreceivingreports.approved), tblreceivingreports.PONumber FROM tblreceivingreports INNER JOIN tblpurchaseorder ON tblreceivingreports.poNumber = tblpurchaseorder.poNumber WHERE docDate BETWEEN '{CALENDAR_START.get_date()}' AND '{CALENDAR_END.get_date()}' ORDER BY rrNumber DESC")
        result = cursor.fetchall()

        count = 0
        rrnumbers = []
        for i in result:
            if i[0] not in rrnumbers:
                if count % 2 == 0:
                    TREE_RR.insert("", "end", values = (str(i[0]).zfill(8), str(i[8]).zfill(8),i[1],self.returnSupplierName(i[2]),i[3],self.returnUserName(i[4], 0),i[5],self.returnUserName(i[6], 0),i[7]), tags = "evenrow")
                else:
                    TREE_RR.insert("", "end", values = (str(i[0]).zfill(8), str(i[8]).zfill(8),i[1],self.returnSupplierName(i[2]),i[3],self.returnUserName(i[4], 0),i[5],self.returnUserName(i[6], 0),i[7]), tags = "oddrow")
                count += 1
                rrnumbers.append(i[0])

    def deleteRRTreeItem(self, tree):
        self.copySelection(tree)
        tree.delete(item)
        try:
            self.updateRRTotals()
        except:
            pass

### MENU_ADMIN ###
    def showBanks(self, *args):
        self.clearWorkspace()
        FRAME_BANKS = LabelFrame(FRAME_4, text = "Banks", font = APP_FONT)
        FRAME_BANKS.grid(column = 1, row = 0)

        SUB1_BANKS = Frame(FRAME_BANKS)
        SUB1_BANKS.grid(column = 0, row = 0, sticky = W, pady = SUBMENU_PADY)

        global SUB2_BANKS
        SUB2_BANKS = Frame(FRAME_BANKS)
        SUB2_BANKS.grid(column = 0, row = 1, sticky = W)

        LABEL_SEARCH = Label(SUB1_BANKS, text = "Search", width = 5, font = APP_FONT, anchor = W)
        LABEL_SEARCH.grid(column = 0, row = 0, pady = SUBMENU_PADY, sticky = W)

        global TEXTVAR_SEARCH
        TEXTVAR_SEARCH = StringVar()
        ENTRY_SEARCH = Entry(SUB1_BANKS, textvariable = TEXTVAR_SEARCH, width = 25, font = APP_FONT)
        ENTRY_SEARCH.grid(column = 1, row = 0, padx = MENU_PADX, sticky = W)
        ENTRY_SEARCH.bind("<Return>", lambda e: self.searchBank(TEXTVAR_SEARCH.get()))

        BUTTON_GO = Button(SUB1_BANKS, text = "GO", width = 5, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = lambda: self.searchBank(TEXTVAR_SEARCH.get()))
        BUTTON_GO.grid(column = 2, row = 0, padx = MENU_PADX)

        BUTTON_REFRESH = Button(SUB1_BANKS, text = "REFRESH", width = 10, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = self.showBanks)
        BUTTON_REFRESH.grid(column = 3, row = 0, padx = MENU_PADX)

        BUTTON_ADD = Button(SUB1_BANKS, text = "ADD", width = 5, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = self.showAddEditBank)
        BUTTON_ADD.grid(column = 4, row = 0, padx = MENU_PADX)

        global TREE_BANKS
        TREE_BANKS = tk.Treeview(SUB2_BANKS, height = 28, selectmode = "browse")
        TREE_BANKS["columns"] = ("Name", "Account No.", "Code", "Contact Person", "Phone", "Address", "Status")
        TREE_BANKS.column("#0", width = 0, minwidth = 0, stretch = True)
        TREE_BANKS.column("Name", anchor = W, minwidth = 100, width = 175)
        TREE_BANKS.column("Account No.", anchor = W, minwidth = 100, width = 150)
        TREE_BANKS.column("Code", anchor = W, minwidth = 100, width = 100)
        TREE_BANKS.column("Contact Person", anchor = W, minwidth = 100, width = 150)
        TREE_BANKS.column("Phone", anchor = W, minwidth = 100, width = 125)
        TREE_BANKS.column("Address", anchor = W, minwidth = 200, width = 175)
        TREE_BANKS.column("Status", anchor = W, minwidth = 100, width = 75)
        
        TREE_BANKS.heading("#0", text = "", anchor = W)
        TREE_BANKS.heading("Name", text = "Name", anchor = N, command = lambda: self.sortColumn(TREE_BANKS, "Name", False))
        TREE_BANKS.heading("Account No.", text = "Account No.", anchor = N, command = lambda: self.sortColumn(TREE_BANKS, "Account No.", False))
        TREE_BANKS.heading("Code", text = "Code", anchor = N, command = lambda: self.sortColumn(TREE_BANKS, "Code", False))
        TREE_BANKS.heading("Contact Person", text = "Contact Person", anchor = N, command = lambda: self.sortColumn(TREE_BANKS, "Contact Person", False))
        TREE_BANKS.heading("Phone", text = "Phone", anchor = N, command = lambda: self.sortColumn(TREE_BANKS, "Phone", False))
        TREE_BANKS.heading("Address", text = "Address", anchor = N, command = lambda: self.sortColumn(TREE_BANKS, "Address", False))
        TREE_BANKS.heading("Status", text = "Status", anchor = N, command = lambda: self.sortColumn(TREE_BANKS, "Status", False))

        global POPUP_BANKS
        POPUP_BANKS = Menu(TREE_BANKS, tearoff = 0)
        POPUP_BANKS.add_command(command = self.editBank, label = "Edit")
        POPUP_BANKS.add_command(command = self.deleteBank, label = "Delete")
        TREE_BANKS.bind("<Button-3>", lambda e: self.popupMenu(TREE_BANKS, POPUP_BANKS, e))
        TREE_BANKS.bind("<Double-1>", self.editBank)

        global STYLE_BANKS
        STYLE_BANKS = tk.Style()
        STYLE_BANKS.map("Treeview", foreground = self.fixedMap("foreground", STYLE_BANKS), background = self.fixedMap("background", STYLE_BANKS))

        TREE_BANKS.tag_configure("oddrow", background = None)
        TREE_BANKS.tag_configure("evenrow", background = TREE_TAG_EVENROW)
        db.commit()
        cursor.execute("SELECT name, account, code, contact, phone, address, status FROM tblbanks WHERE status = 'active'")
        result = cursor.fetchall()
        count = 0
        for i in result:
            if count % 2 == 0:
                TREE_BANKS.insert("", "end", values = (i[0],i[1],i[2],i[3],i[4],i[5],i[6]), tags = "evenrow")
            else:
                TREE_BANKS.insert("", "end", values = (i[0],i[1],i[2],i[3],i[4],i[5],i[6]), tags = "oddrow")
            count += 1

        YSCROLLBAR = tk.Scrollbar(SUB2_BANKS, orient = "vertical", command = TREE_BANKS.yview)
        YSCROLLBAR.pack(side = "right", fill = "y")

        XSCROLLBAR = tk.Scrollbar(SUB2_BANKS, orient = "horizontal", command = TREE_BANKS.xview)
        
        TREE_BANKS.configure(yscrollcommand = YSCROLLBAR.set, xscrollcommand = XSCROLLBAR.set) 
        TREE_BANKS.pack()
        XSCROLLBAR.pack(fill ="x")

    def showCostCenters(self, *args):
        self.clearWorkspace()
        FRAME_CENTERS = LabelFrame(FRAME_4, text = "Cost Centers", font = APP_FONT)
        FRAME_CENTERS.grid(column = 1, row = 0)

        SUB1_CENTERS = Frame(FRAME_CENTERS)
        SUB1_CENTERS.grid(column = 0, row = 0, sticky = W, pady = SUBMENU_PADY)

        SUB2_CENTERS = Frame(FRAME_CENTERS)
        SUB2_CENTERS.grid(column = 0, row = 1, sticky = W)

        LABEL_SEARCH = Label(SUB1_CENTERS, text = "Search", width = 5, font = APP_FONT, anchor = W)
        LABEL_SEARCH.grid(column = 0, row = 0, pady = SUBMENU_PADY, sticky = W)

        global TEXTVAR_SEARCH
        TEXTVAR_SEARCH = StringVar()
        ENTRY_SEARCH = Entry(SUB1_CENTERS, textvariable = TEXTVAR_SEARCH, width = 25, font = APP_FONT)
        ENTRY_SEARCH.grid(column = 1, row = 0, padx = MENU_PADX, sticky = W)
        ENTRY_SEARCH.bind("<Return>", lambda e: self.searchCenter(TEXTVAR_SEARCH.get()))

        BUTTON_GO = Button(SUB1_CENTERS, text = "GO", width = 5, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = lambda: self.searchCenter(TEXTVAR_SEARCH.get()))
        BUTTON_GO.grid(column = 2, row = 0, padx = MENU_PADX)

        BUTTON_REFRESH = Button(SUB1_CENTERS, text = "REFRESH", width = 10, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = self.showCostCenters)
        BUTTON_REFRESH.grid(column = 3, row = 0, padx = MENU_PADX)

        BUTTON_ADD = Button(SUB1_CENTERS, text = "ADD", width = 5, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = self.showAddEditCenter)
        BUTTON_ADD.grid(column = 4, row = 0, padx = MENU_PADX)

        global TREE_CENTERS
        TREE_CENTERS = tk.Treeview(SUB2_CENTERS, height = 28, selectmode = "browse")
        TREE_CENTERS["columns"] = ("Code", "Name", "")
        TREE_CENTERS.column("#0", width = 0, minwidth = 0, stretch = True)
        TREE_CENTERS.column("Code", anchor = W, minwidth = 100, width = 200)
        TREE_CENTERS.column("Name", anchor = W, minwidth = 100, width = 150)
        TREE_CENTERS.column("#3", width = 600, minwidth = 600)
        
        TREE_CENTERS.heading("#0", text = "", anchor = W, command = lambda: self.sortColumn(TREE_CENTERS, "Code", False))
        TREE_CENTERS.heading("Code", text = "Code", anchor = N, command = lambda: self.sortColumn(TREE_CENTERS, "Code", False))
        TREE_CENTERS.heading("Name", text = "Name", anchor = N, command = lambda: self.sortColumn(TREE_CENTERS, "Code", False))
        TREE_CENTERS.heading("#3", text = "", anchor = W, command = lambda: self.sortColumn(TREE_CENTERS, "Code", False))

        global POPUP_CENTERS
        POPUP_CENTERS = Menu(TREE_CENTERS, tearoff = 0)
        POPUP_CENTERS.add_command(command = self.editCenter, label = "Edit")
        # POPUP_CENTERS.add_command(command = self.deleteCenter, label = "Delete")
        TREE_CENTERS.bind("<Button-3>", lambda e: self.popupMenu(TREE_CENTERS, POPUP_CENTERS, e))
        TREE_CENTERS.bind("<Double-1>", self.editCenter)

        global STYLE_CENTERS
        STYLE_CENTERS = tk.Style()
        STYLE_CENTERS.map("Treeview", foreground = self.fixedMap("foreground", STYLE_CENTERS), background = self.fixedMap("background", STYLE_CENTERS))

        TREE_CENTERS.tag_configure("oddrow", background = None)
        TREE_CENTERS.tag_configure("evenrow", background = TREE_TAG_EVENROW)
        db.commit()
        cursor.execute("SELECT code, name FROM tblcenters ORDER BY code")
        result = cursor.fetchall()

        count = 0
        for i in result:
            if count % 2 == 0:
                TREE_CENTERS.insert("", "end", values = (i[0],i[1]), tags = "evenrow")
            else:
                TREE_CENTERS.insert("", "end", values = (i[0],i[1]), tags = "oddrow")
            count += 1

        YSCROLLBAR = tk.Scrollbar(SUB2_CENTERS, orient = "vertical", command = TREE_CENTERS.yview)
        YSCROLLBAR.pack(side = "right", fill = "y")

        XSCROLLBAR = tk.Scrollbar(SUB2_CENTERS, orient = "horizontal", command = TREE_CENTERS.xview)
        
        TREE_CENTERS.configure(yscrollcommand = YSCROLLBAR.set, xscrollcommand = XSCROLLBAR.set) 
        TREE_CENTERS.pack()
        XSCROLLBAR.pack(fill ="x")

    def showChartofAccounts(self, *args):
        self.clearWorkspace()
        FRAME_CHART = LabelFrame(FRAME_4, text = "Chart of Accounts", font = APP_FONT)
        FRAME_CHART.grid(column = 1, row = 0)

        SUB1_CHART = Frame(FRAME_CHART)
        SUB1_CHART.grid(column = 0, row = 0, sticky = W, pady = SUBMENU_PADY)

        global SUB2_CHART
        SUB2_CHART = Frame(FRAME_CHART)
        SUB2_CHART.grid(column = 0, row = 1, sticky = W)

        LABEL_SEARCH = Label(SUB1_CHART, text = "Search", width = 5, font = APP_FONT, anchor = W)
        LABEL_SEARCH.grid(column = 0, row = 0, pady = SUBMENU_PADY, sticky = W)
        
        global TEXTVAR_SEARCH
        TEXTVAR_SEARCH = StringVar()
        ENTRY_SEARCH = Entry(SUB1_CHART, textvariable = TEXTVAR_SEARCH, width = 25, font = APP_FONT)
        ENTRY_SEARCH.grid(column = 1, row = 0, padx = MENU_PADX, sticky = W)
        ENTRY_SEARCH.bind("<Return>", lambda e: self.searchChartofAccount(TEXTVAR_SEARCH.get()))

        BUTTON_GO = Button(SUB1_CHART, text = "GO", width = 5, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = lambda: self.searchChartofAccount(TEXTVAR_SEARCH.get()))
        BUTTON_GO.grid(column = 2, row = 0, padx = MENU_PADX)

        BUTTON_REFRESH = Button(SUB1_CHART, text = "REFRESH", width = 10, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = self.showChartofAccounts)
        BUTTON_REFRESH.grid(column = 3, row = 0, padx = MENU_PADX)

        BUTTON_ADD = Button(SUB1_CHART, text = "ADD", width = 5, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = self.showAddEditChart)
        BUTTON_ADD.grid(column = 4, row = 0, padx = MENU_PADX)
        
        self.showChartTreeview(SUB2_CHART)

    def showChartTreeview(self, frame):
        global TREE_CHART
        TREE_CHART = tk.Treeview(frame, height = 28, selectmode = "browse")
        TREE_CHART["columns"] = ("Code", "Title", "Parent", "Remarks", "Active")
        TREE_CHART.column("#0", width = 0, minwidth = 0, stretch = True)
        TREE_CHART.column("Code", anchor = W, minwidth = 100, width = 150)
        TREE_CHART.column("Title", anchor = W, minwidth = 100, width = 300)
        TREE_CHART.column("Parent", anchor = W, minwidth = 100, width = 200)
        TREE_CHART.column("Remarks", anchor = W, minwidth = 100, width = 150)
        TREE_CHART.column("Active", anchor = W, minwidth = 100, width = 150)
        
        TREE_CHART.heading("#0", text = "", anchor = W)
        TREE_CHART.heading("Code", text = "Code", anchor = N, command = lambda: self.sortColumn(TREE_CHART, "Code", False))
        TREE_CHART.heading("Title", text = "Title", anchor = N, command = lambda: self.sortColumn(TREE_CHART, "Title", False))
        TREE_CHART.heading("Parent", text = "Parent", anchor = N, command = lambda: self.sortColumn(TREE_CHART, "Parent", False))
        TREE_CHART.heading("Remarks", text = "Remarks", anchor = N, command = lambda: self.sortColumn(TREE_CHART, "Remarks", False))
        TREE_CHART.heading("Active", text = "Active", anchor = N, command = lambda: self.sortColumn(TREE_CHART, "Active", False))

        global POPUP_CHART
        POPUP_CHART = Menu(TREE_CHART, tearoff = 0)
        POPUP_CHART.add_command(command = self.editChart, label = "Edit")
        POPUP_CHART.add_command(command = self.deleteChart, label = "Delete")
        TREE_CHART.bind("<Button-3>", lambda e: self.popupMenu(TREE_CHART, POPUP_CHART, e))

        global STYLE_CHART
        STYLE_CHART = tk.Style()
        STYLE_CHART.map("Treeview", foreground = self.fixedMap("foreground", STYLE_CHART), background = self.fixedMap("background", STYLE_CHART))

        TREE_CHART.tag_configure("oddrow", background = None)
        TREE_CHART.tag_configure("evenrow", background = TREE_TAG_EVENROW)
        db.commit()
        cursor.execute("SELECT code, title, parent, remarks, active FROM tblchart WHERE active = 'Yes' ORDER BY code")
        result = cursor.fetchall()

        count = 0
        for i in result:
            if count % 2 == 0:
                TREE_CHART.insert("", "end", values = (i[0],i[1],i[2],i[3],i[4]), tags = "evenrow")
            else:
                TREE_CHART.insert("", "end", values = (i[0],i[1],i[2],i[3],i[4]), tags = "oddrow")
            count += 1

        YSCROLLBAR = tk.Scrollbar(frame, orient = "vertical", command = TREE_CHART.yview)
        YSCROLLBAR.pack(side = "right", fill = "y")

        XSCROLLBAR = tk.Scrollbar(frame, orient = "horizontal", command = TREE_CHART.xview)
        
        TREE_CHART.configure(yscrollcommand = YSCROLLBAR.set, xscrollcommand = XSCROLLBAR.set) 
        TREE_CHART.pack()
        XSCROLLBAR.pack(fill ="x")
        
    def showClients(self, *args):
        self.clearWorkspace()
        FRAME_CLIENTS = LabelFrame(FRAME_4, text = "Clients", font = APP_FONT)
        FRAME_CLIENTS.grid(column = 1, row = 0)

        SUB1_CLIENTS = Frame(FRAME_CLIENTS)
        SUB1_CLIENTS.grid(column = 0, row = 0, sticky = W, pady = SUBMENU_PADY)

        SUB2_CLIENTS = Frame(FRAME_CLIENTS)
        SUB2_CLIENTS.grid(column = 0, row = 1, sticky = W)

        LABEL_SEARCH = Label(SUB1_CLIENTS, text = "Search", width = 5, font = APP_FONT, anchor = W)
        LABEL_SEARCH.grid(column = 0, row = 0, pady = SUBMENU_PADY, sticky = W)

        global TEXTVAR_SEARCH
        TEXTVAR_SEARCH = StringVar()
        ENTRY_SEARCH = Entry(SUB1_CLIENTS, textvariable = TEXTVAR_SEARCH, width = 25, font = APP_FONT)
        ENTRY_SEARCH.grid(column = 1, row = 0, padx = MENU_PADX, sticky = W)
        ENTRY_SEARCH.bind("<Return>", lambda e: self.searchClient(TEXTVAR_SEARCH.get()))

        BUTTON_GO = Button(SUB1_CLIENTS, text = "GO", width = 5, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = lambda: self.searchClient(TEXTVAR_SEARCH.get()))
        BUTTON_GO.grid(column = 2, row = 0, padx = MENU_PADX)

        BUTTON_REFRESH = Button(SUB1_CLIENTS, text = "REFRESH", width = 10, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = self.showClients)
        BUTTON_REFRESH.grid(column = 3, row = 0, padx = MENU_PADX)

        BUTTON_ADD = Button(SUB1_CLIENTS, text = "ADD", width = 5, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = self.showAddEditClient)
        BUTTON_ADD.grid(column = 4, row = 0, padx = MENU_PADX)

        global TREE_CLIENTS
        TREE_CLIENTS = tk.Treeview(SUB2_CLIENTS, height = 28, selectmode = "browse")
        TREE_CLIENTS["columns"] = ("Code", "Name", "Address", "TIN", "Contact", "Entity", "Tax", "Parent")
        TREE_CLIENTS.column("#0", width = 0, minwidth = 0, stretch = True)
        TREE_CLIENTS.column("Code", anchor = W, minwidth = 100, width = 150)
        TREE_CLIENTS.column("Name", anchor = W, minwidth = 100, width = 200)
        TREE_CLIENTS.column("Address", anchor = W, minwidth = 100, width = 300)
        TREE_CLIENTS.column("TIN", anchor = W, minwidth = 100, width = 75)
        TREE_CLIENTS.column("Contact", anchor = W, minwidth = 100, width = 100)
        TREE_CLIENTS.column("Entity", anchor = W, minwidth = 100, width = 100)
        TREE_CLIENTS.column("Tax", anchor = W, minwidth = 100, width = 75)
        TREE_CLIENTS.column("Parent", anchor = W, minwidth = 100, width = 100)
        
        TREE_CLIENTS.heading("#0", text = "", anchor = W)
        TREE_CLIENTS.heading("Code", text = "Code", anchor = N)
        TREE_CLIENTS.heading("Name", text = "Name", anchor = N)
        TREE_CLIENTS.heading("Address", text = "Address", anchor = N)
        TREE_CLIENTS.heading("TIN", text = "TIN", anchor = N)
        TREE_CLIENTS.heading("Contact", text = "Contact", anchor = N)
        TREE_CLIENTS.heading("Entity", text = "Entity", anchor = N)
        TREE_CLIENTS.heading("Tax", text = "Tax", anchor = N)
        TREE_CLIENTS.heading("Parent", text = "Parent", anchor = N)

        global POPUP_CLIENTS
        POPUP_CLIENTS = Menu(TREE_CLIENTS, tearoff = 0)
        POPUP_CLIENTS.add_command(command = self.editClient, label = "Edit")
        # POPUP_CLIENTS.add_command(command = self.deleteClient, label = "Delete")
        TREE_CLIENTS.bind("<Button-3>", lambda e: self.popupMenu(TREE_CLIENTS, POPUP_CLIENTS, e))
        TREE_CLIENTS.bind("<Double-1>", self.editClient)

        global STYLE_CLIENTS
        STYLE_CLIENTS = tk.Style()
        STYLE_CLIENTS.map("Treeview", foreground = self.fixedMap("foreground", STYLE_CLIENTS), background = self.fixedMap("background", STYLE_CLIENTS))

        TREE_CLIENTS.tag_configure("oddrow", background = None)
        TREE_CLIENTS.tag_configure("evenrow", background = TREE_TAG_EVENROW)
        db.commit()
        cursor.execute("""SELECT clientCode, clientName, address, tin, contact, entityType, taxType, parentCode FROM tblclients ORDER BY clientName""")
        result = cursor.fetchall()
        count = 0
        for i in result:
            if count % 2 == 0:
                TREE_CLIENTS.insert("", "end", values = (i[0],i[1],i[2],i[3],i[4],i[5],i[6],i[7]), tags = "evenrow")
            else:
                TREE_CLIENTS.insert("", "end", values = (i[0],i[1],i[2],i[3],i[4],i[5],i[6],i[7]), tags = "oddrow")
            count += 1

        YSCROLLBAR = tk.Scrollbar(SUB2_CLIENTS, orient = "vertical", command = TREE_CLIENTS.yview)
        YSCROLLBAR.pack(side = "right", fill = "y")

        XSCROLLBAR = tk.Scrollbar(SUB2_CLIENTS, orient = "horizontal", command = TREE_CLIENTS.xview)
        
        TREE_CLIENTS.configure(yscrollcommand = YSCROLLBAR.set, xscrollcommand = XSCROLLBAR.set) 
        TREE_CLIENTS.pack()
        XSCROLLBAR.pack(fill ="x")

    def showSuppliers(self, *args):
        self.clearWorkspace()
        FRAME_SUPPLIERS = LabelFrame(FRAME_4, text = "Suppliers", font = APP_FONT)
        FRAME_SUPPLIERS.grid(column = 1, row = 0)

        SUB1_SUPPLIERS = Frame(FRAME_SUPPLIERS)
        SUB1_SUPPLIERS.grid(column = 0, row = 0, sticky = W, pady = SUBMENU_PADY)

        SUB2_SUPPLIERS = Frame(FRAME_SUPPLIERS)
        SUB2_SUPPLIERS.grid(column = 0, row = 1, sticky = W)

        LABEL_SEARCH = Label(SUB1_SUPPLIERS, text = "Search", width = 5, font = APP_FONT, anchor = W)
        LABEL_SEARCH.grid(column = 0, row = 0, pady = SUBMENU_PADY, sticky = W)

        global TEXTVAR_SEARCH
        TEXTVAR_SEARCH = StringVar()
        ENTRY_SEARCH = Entry(SUB1_SUPPLIERS, textvariable = TEXTVAR_SEARCH, width = 25, font = APP_FONT)
        ENTRY_SEARCH.grid(column = 1, row = 0, padx = MENU_PADX, sticky = W)
        ENTRY_SEARCH.bind("<Return>", lambda e: self.searchSupplier(TEXTVAR_SEARCH.get()))

        BUTTON_GO = Button(SUB1_SUPPLIERS, text = "GO", width = 5, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = lambda: self.searchSupplier(TEXTVAR_SEARCH.get()))
        BUTTON_GO.grid(column = 2, row = 0, padx = MENU_PADX)

        BUTTON_REFRESH = Button(SUB1_SUPPLIERS, text = "REFRESH", width = 10, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = self.showSuppliers)
        BUTTON_REFRESH.grid(column = 3, row = 0, padx = MENU_PADX)

        BUTTON_ADD = Button(SUB1_SUPPLIERS, text = "ADD", width = 5, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = self.showAddEditSupplier)
        BUTTON_ADD.grid(column = 4, row = 0, padx = MENU_PADX)
        
        global TEXTVAR_CATEGORY
        TEXTVAR_CATEGORY = StringVar()
        COMBO_CATEGORY = tk.Combobox(SUB1_SUPPLIERS, values = ["Supplier", "Others"], textvariable = TEXTVAR_CATEGORY, font = APP_FONT, width = 15, state = "readonly")
        COMBO_CATEGORY.grid(column = 5, row = 0, sticky = E, padx = TOP_PADX)
        COMBO_CATEGORY.bind("<<ComboboxSelected>>", lambda e: self.searchSupplier(TEXTVAR_SEARCH.get()))
        TEXTVAR_CATEGORY.set("Supplier")

        global TREE_SUPPLIER
        TREE_SUPPLIER = tk.Treeview(SUB2_SUPPLIERS, height = 28, selectmode = "browse")
        TREE_SUPPLIER["columns"] = ("Code", "Name", "TIN", "Address", "Contact", "Tax", "Category", "Entity")
        TREE_SUPPLIER.column("#0", width = 0, minwidth = 0, stretch = True)
        TREE_SUPPLIER.column("Code", anchor = W, minwidth = 100, width = 150)
        TREE_SUPPLIER.column("Name", anchor = W, minwidth = 100, width = 250)
        TREE_SUPPLIER.column("TIN", anchor = W, minwidth = 100, width = 150)
        TREE_SUPPLIER.column("Address", anchor = W, minwidth = 100, width = 200)
        TREE_SUPPLIER.column("Contact", anchor = W, minwidth = 100, width = 150)
        TREE_SUPPLIER.column("Tax", anchor = W, minwidth = 100, width = 50)
        TREE_SUPPLIER.column("Category", anchor = W, minwidth = 100, width = 100)
        TREE_SUPPLIER.column("Entity", anchor = W, minwidth = 100, width = 100)
        
        TREE_SUPPLIER.heading("#0", text = "", anchor = W)
        TREE_SUPPLIER.heading("Code", text = "Code", anchor = N)
        TREE_SUPPLIER.heading("Name", text = "Name", anchor = N)
        TREE_SUPPLIER.heading("TIN", text = "TIN", anchor = N)
        TREE_SUPPLIER.heading("Address", text = "Address", anchor = N)
        TREE_SUPPLIER.heading("Contact", text = "Contact", anchor = N)
        TREE_SUPPLIER.heading("Tax", text = "Tax", anchor = N)
        TREE_SUPPLIER.heading("Category", text = "Category", anchor = N)
        TREE_SUPPLIER.heading("Entity", text = "Entity", anchor = N)

        global POPUP_SUPPLIER
        POPUP_SUPPLIER = Menu(TREE_SUPPLIER, tearoff = 0)
        POPUP_SUPPLIER.add_command(command = self.editSupplier, label = "Edit")
        # POPUP_SUPPLIER.add_command(command = self.deleteSupplier, label = "Delete")
        TREE_SUPPLIER.bind("<Button-3>", lambda e: self.popupMenu(TREE_SUPPLIER, POPUP_SUPPLIER, e))
        TREE_SUPPLIER.bind("<Double-1>", self.editSupplier)

        global STYLE_SUPPLIERS
        STYLE_SUPPLIERS = tk.Style()
        STYLE_SUPPLIERS.map("Treeview", foreground = self.fixedMap("foreground", STYLE_SUPPLIERS), background = self.fixedMap("background", STYLE_SUPPLIERS))

        TREE_SUPPLIER.tag_configure("oddrow", background = None)
        TREE_SUPPLIER.tag_configure("evenrow", background = TREE_TAG_EVENROW)
        db.commit()
        cursor.execute("""SELECT code, name, tin, address, contact, vatable, category, entity FROM tblsuppliers""")
        result = cursor.fetchall()
        count = 0
        for i in result:
            if count % 2 == 0:
                TREE_SUPPLIER.insert("", "end", values = (i[0],i[1],i[2],i[3],i[4],i[5],i[6],i[7]), tags = "evenrow")
            else:
                TREE_SUPPLIER.insert("", "end", values = (i[0],i[1],i[2],i[3],i[4],i[5],i[6],i[7]), tags = "oddrow")
            count += 1

        YSCROLLBAR = tk.Scrollbar(SUB2_SUPPLIERS, orient = "vertical", command = TREE_SUPPLIER.yview)
        YSCROLLBAR.pack(side = "right", fill = "y")

        XSCROLLBAR = tk.Scrollbar(SUB2_SUPPLIERS, orient = "horizontal", command = TREE_SUPPLIER.xview)
        
        TREE_SUPPLIER.configure(yscrollcommand = YSCROLLBAR.set, xscrollcommand = XSCROLLBAR.set) 
        TREE_SUPPLIER.pack()
        XSCROLLBAR.pack(fill ="x")

    def showTaxCodes(self, *args):
        self.clearWorkspace()
        FRAME_TAX = LabelFrame(FRAME_4, text = "Tax Codes", font = APP_FONT)
        FRAME_TAX.grid(column = 1, row = 0)

        SUB1_TAX = Frame(FRAME_TAX)
        SUB1_TAX.grid(column = 0, row = 0, sticky = W, pady = SUBMENU_PADY)

        SUB2_TAX = Frame(FRAME_TAX)
        SUB2_TAX.grid(column = 0, row = 1, sticky = W)

        LABEL_SEARCH = Label(SUB1_TAX, text = "Search", width = 5, font = APP_FONT, anchor = W)
        LABEL_SEARCH.grid(column = 0, row = 0, pady = SUBMENU_PADY, sticky = W)

        TEXTVAR_SEARCH = StringVar()
        ENTRY_SEARCH = Entry(SUB1_TAX, textvariable = TEXTVAR_SEARCH, width = 25, font = APP_FONT)
        ENTRY_SEARCH.grid(column = 1, row = 0, padx = MENU_PADX, sticky = W)
        ENTRY_SEARCH.bind("<Return>", lambda e: self.searchTaxCode(TEXTVAR_SEARCH.get()))

        BUTTON_GO = Button(SUB1_TAX, text = "GO", width = 5, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = lambda: self.searchTaxCode(TEXTVAR_SEARCH.get()))
        BUTTON_GO.grid(column = 2, row = 0, padx = MENU_PADX)

        BUTTON_REFRESH = Button(SUB1_TAX, text = "REFRESH", width = 10, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = self.showTaxCodes)
        BUTTON_REFRESH.grid(column = 3, row = 0, padx = MENU_PADX)

        BUTTON_ADD = Button(SUB1_TAX, text = "ADD", width = 5, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = self.showAddEditTax)
        BUTTON_ADD.grid(column = 4, row = 0, padx = MENU_PADX)

        global TREE_TAX
        TREE_TAX = tk.Treeview(SUB2_TAX, height = 28, selectmode = "browse")
        TREE_TAX["columns"] = ("Code", "Rate (%)", "Description","")
        TREE_TAX.column("#0", width = 0, minwidth = 0, stretch = True)
        TREE_TAX.column("Code", anchor = W, minwidth = 100, width = 150)
        TREE_TAX.column("Rate (%)", anchor = W, minwidth = 100, width = 150)
        TREE_TAX.column("Description", anchor = W, minwidth = 100, width = 400)
        TREE_TAX.column("#4", anchor = W, minwidth = 100, width = 250)
        
        TREE_TAX.heading("#0", text = "", anchor = W)
        TREE_TAX.heading("Code", text = "Code", anchor = N)
        TREE_TAX.heading("Rate (%)", text = "Rate (%)", anchor = N)
        TREE_TAX.heading("Description", text = "Description", anchor = N)
        TREE_TAX.heading("#4", text = "", anchor = N)

        global POPUP_TAX
        POPUP_TAX = Menu(TREE_TAX, tearoff = 0)
        POPUP_TAX.add_command(command = self.editTax, label = "Edit")
        POPUP_TAX.add_command(command = self.deleteTax, label = "Delete")
        TREE_TAX.bind("<Button-3>", lambda e: self.popupMenu(TREE_TAX, POPUP_TAX, e))

        global STYLE_TAX
        STYLE_TAX = tk.Style()
        STYLE_TAX.map("Treeview", foreground = self.fixedMap("foreground", STYLE_TAX), background = self.fixedMap("background", STYLE_TAX))

        TREE_TAX.tag_configure("oddrow", background = None)
        TREE_TAX.tag_configure("evenrow", background = TREE_TAG_EVENROW)
        db.commit()
        cursor.execute("""SELECT code, rate, description FROM tbltaxes""")
        result = cursor.fetchall()
        count = 0
        for i in result:
            if count % 2 == 0:
                TREE_TAX.insert("", "end", values = (i[0],i[1],i[2]), tags = "evenrow")
            else:
                TREE_TAX.insert("", "end", values = (i[0],i[1],i[2]), tags = "oddrow")
            count += 1

        YSCROLLBAR = tk.Scrollbar(SUB2_TAX, orient = "vertical", command = TREE_TAX.yview)
        YSCROLLBAR.pack(side = "right", fill = "y")

        XSCROLLBAR = tk.Scrollbar(SUB2_TAX, orient = "horizontal", command = TREE_TAX.xview)
        
        TREE_TAX.configure(yscrollcommand = YSCROLLBAR.set, xscrollcommand = XSCROLLBAR.set) 
        TREE_TAX.pack()
        XSCROLLBAR.pack(fill ="x")

    def showTransactionTypes(self, *args):
        self.clearWorkspace()
        FRAME_TYPES = LabelFrame(FRAME_4, text = "Transaction Types", font = APP_FONT)
        FRAME_TYPES.grid(column = 1, row = 0)

        SUB1_TYPES = Frame(FRAME_TYPES)
        SUB1_TYPES.grid(column = 0, row = 0, sticky = W, pady = SUBMENU_PADY)

        SUB2_TYPES = Frame(FRAME_TYPES)
        SUB2_TYPES.grid(column = 0, row = 1, sticky = W)

        LABEL_SEARCH = Label(SUB1_TYPES, text = "Search", width = 5, font = APP_FONT, anchor = W)
        LABEL_SEARCH.grid(column = 0, row = 0, pady = SUBMENU_PADY, sticky = W)

        global TEXTVAR_SEARCH
        TEXTVAR_SEARCH = StringVar()
        ENTRY_SEARCH = Entry(SUB1_TYPES, textvariable = TEXTVAR_SEARCH, width = 25, font = APP_FONT)
        ENTRY_SEARCH.grid(column = 1, row = 0, padx = MENU_PADX, sticky = W)
        ENTRY_SEARCH.bind("<Return>", lambda e: self.searchTransactionType(TEXTVAR_SEARCH.get()))

        BUTTON_GO = Button(SUB1_TYPES, text = "GO", width = 5, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = lambda: self.searchTransactionType(TEXTVAR_SEARCH.get()))
        BUTTON_GO.grid(column = 2, row = 0, padx = MENU_PADX)

        BUTTON_REFRESH = Button(SUB1_TYPES, text = "REFRESH", width = 10, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = self.showTransactionTypes)
        BUTTON_REFRESH.grid(column = 3, row = 0, padx = MENU_PADX)

        BUTTON_ADD = Button(SUB1_TYPES, text = "ADD", width = 5, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = self.showAddEditType)
        BUTTON_ADD.grid(column = 4, row = 0, padx = MENU_PADX)

        global TREE_TYPES
        TREE_TYPES = tk.Treeview(SUB2_TYPES, height = 28, selectmode = "browse")
        TREE_TYPES["columns"] = ("Code", "Type", "Description","")
        TREE_TYPES.column("#0", width = 0, minwidth = 0, stretch = True)
        TREE_TYPES.column("Code", anchor = W, minwidth = 100, width = 150)
        TREE_TYPES.column("Type", anchor = W, minwidth = 100, width = 150)
        TREE_TYPES.column("Description", anchor = W, minwidth = 100, width = 400)
        TREE_TYPES.column("#4", anchor = W, minwidth = 100, width = 250)
        
        TREE_TYPES.heading("#0", text = "", anchor = W)
        TREE_TYPES.heading("Code", text = "Code", anchor = N)
        TREE_TYPES.heading("Type", text = "Type", anchor = N)
        TREE_TYPES.heading("Description", text = "Description", anchor = N)
        TREE_TYPES.heading("#4", text = "", anchor = N)

        global POPUP_TYPES
        POPUP_TYPES = Menu(TREE_TYPES, tearoff = 0)
        POPUP_TYPES.add_command(command = self.editType, label = "Edit")
        POPUP_TYPES.add_command(command = self.deleteType, label = "Delete")
        TREE_TYPES.bind("<Button-3>", lambda e: self.popupMenu(TREE_TYPES, POPUP_TYPES, e))

        global STYLE_TYPES
        STYLE_TYPES = tk.Style()
        STYLE_TYPES.map("Treeview", foreground = self.fixedMap("foreground", STYLE_TYPES), background = self.fixedMap("background", STYLE_TYPES))

        TREE_TYPES.tag_configure("oddrow", background = None)
        TREE_TYPES.tag_configure("evenrow", background = TREE_TAG_EVENROW)
        db.commit()
        cursor.execute("""SELECT code, type, description FROM tbltypes""")
        result = cursor.fetchall()
        count = 0
        for i in result:
            if count % 2 == 0:
                TREE_TYPES.insert("", "end", values = (i[0],i[1],i[2]), tags = "evenrow")
            else:
                TREE_TYPES.insert("", "end", values = (i[0],i[1],i[2]), tags = "oddrow")
            count += 1

        YSCROLLBAR = tk.Scrollbar(SUB2_TYPES, orient = "vertical", command = TREE_TYPES.yview)
        YSCROLLBAR.pack(side = "right", fill = "y")

        XSCROLLBAR = tk.Scrollbar(SUB2_TYPES, orient = "horizontal", command = TREE_TYPES.xview)
        
        TREE_TYPES.configure(yscrollcommand = YSCROLLBAR.set, xscrollcommand = XSCROLLBAR.set) 
        TREE_TYPES.pack()
        XSCROLLBAR.pack(fill ="x")

    def showUsers(self, *args):
        self.clearWorkspace()
        FRAME_USERS = LabelFrame(FRAME_4, text = "Users", font = APP_FONT)
        FRAME_USERS.grid(column = 1, row = 0)

        SUB1_USERS = Frame(FRAME_USERS)
        SUB1_USERS.grid(column = 0, row = 0, sticky = W, pady = SUBMENU_PADY)

        SUB2_USERS = Frame(FRAME_USERS)
        SUB2_USERS.grid(column = 0, row = 1, sticky = W)

        LABEL_SEARCH = Label(SUB1_USERS, text = "Search", width = 5, font = APP_FONT, anchor = W)
        LABEL_SEARCH.grid(column = 0, row = 0, pady = SUBMENU_PADY, sticky = W)

        global TEXTVAR_SEARCH
        TEXTVAR_SEARCH = StringVar()
        ENTRY_SEARCH = Entry(SUB1_USERS, textvariable = TEXTVAR_SEARCH, width = 25, font = APP_FONT)
        ENTRY_SEARCH.grid(column = 1, row = 0, padx = MENU_PADX, sticky = W)
        ENTRY_SEARCH.bind("<Return>", lambda e: self.searchUser(TEXTVAR_SEARCH.get()))

        BUTTON_GO = Button(SUB1_USERS, text = "GO", width = 5, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = lambda: self.searchUser(TEXTVAR_SEARCH.get()))
        BUTTON_GO.grid(column = 2, row = 0, padx = MENU_PADX)

        BUTTON_REFRESH = Button(SUB1_USERS, text = "REFRESH", width = 10, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = self.showUsers)
        BUTTON_REFRESH.grid(column = 3, row = 0, padx = MENU_PADX)

        BUTTON_ADD = Button(SUB1_USERS, text = "ADD", width = 5, font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, bd = 1, activebackground = CLICK_BG, cursor = "hand2", command = self.showAddEditUser)
        BUTTON_ADD.grid(column = 4, row = 0, padx = MENU_PADX)

        global TREE_USERS
        TREE_USERS = tk.Treeview(SUB2_USERS, height = 28, selectmode = "browse")
        TREE_USERS["columns"] = ("ID", "Username", "First Name", "Last Name", "User Type", "Department")
        TREE_USERS.column("#0", width = 0, minwidth = 0, stretch = True)
        TREE_USERS.column("ID", anchor = W, minwidth = 100, width = 50)
        TREE_USERS.column("Username", anchor = W, minwidth = 100, width = 150)
        TREE_USERS.column("First Name", anchor = W, minwidth = 100, width = 200)
        TREE_USERS.column("Last Name", anchor = W, minwidth = 100, width = 200)
        TREE_USERS.column("User Type", anchor = W, minwidth = 100, width = 150)
        TREE_USERS.column("Department", anchor = W, minwidth = 100, width = 200)
        
        TREE_USERS.heading("#0", text = "", anchor = W)
        TREE_USERS.heading("ID", text = "ID", anchor = N, command = lambda: self.sortColumn(TREE_USERS, "ID", False))
        TREE_USERS.heading("Username", text = "Username", anchor = N, command = lambda: self.sortColumn(TREE_USERS, "Username", False))
        TREE_USERS.heading("First Name", text = "First Name", anchor = N, command = lambda: self.sortColumn(TREE_USERS, "First Name", False))
        TREE_USERS.heading("Last Name", text = "Last Name", anchor = N, command = lambda: self.sortColumn(TREE_USERS, "Last Name", False))
        TREE_USERS.heading("User Type", text = "User Type", anchor = N, command = lambda: self.sortColumn(TREE_USERS, "User Type", False))
        TREE_USERS.heading("Department", text = "Department", anchor = N, command = lambda: self.sortColumn(TREE_USERS, "Department", False))

        global POPUP_USERS
        POPUP_USERS = Menu(TREE_USERS, tearoff = 0)
        POPUP_USERS.add_command(command = self.editUser, label = "Edit")
        POPUP_USERS.add_command(command = self.deleteUser, label = "Delete")
        POPUP_USERS.add_command(command = self.editAccess, label = "Edit Access")
        TREE_USERS.bind("<Button-3>", lambda e: self.popupMenu(TREE_USERS, POPUP_USERS, e))
        TREE_USERS.bind("<Double-1>", self.editUser)

        global STYLE_USERS
        STYLE_USERS = tk.Style()
        STYLE_USERS.map("Treeview", foreground = self.fixedMap("foreground", STYLE_USERS), background = self.fixedMap("background", STYLE_USERS))

        TREE_USERS.tag_configure("oddrow", background = None)
        TREE_USERS.tag_configure("evenrow", background = TREE_TAG_EVENROW)

        db.commit()
        cursor.execute("SELECT id, username, firstname, lastname, usertype, department FROM tblusers")
        result = cursor.fetchall()
        count = 0
        for i in result:
            if count % 2 == 0:
                TREE_USERS.insert("", "end", values = (i[0],i[1],i[2],i[3],i[4],i[5]), tags = "evenrow")
            else:
                TREE_USERS.insert("", "end", values = (i[0],i[1],i[2],i[3],i[4],i[5]), tags = "oddrow")
            count += 1

        YSCROLLBAR = tk.Scrollbar(SUB2_USERS, orient = "vertical", command = TREE_USERS.yview)
        YSCROLLBAR.pack(side = "right", fill = "y")

        XSCROLLBAR = tk.Scrollbar(SUB2_USERS, orient = "horizontal", command = TREE_USERS.xview)
        
        TREE_USERS.configure(yscrollcommand = YSCROLLBAR.set, xscrollcommand = XSCROLLBAR.set) 
        TREE_USERS.pack()
        XSCROLLBAR.pack(fill ="x")

### MENU_ADMIN_BANKS ###
    def showAddEditBank(self, *args):
        global TOP_BANK
        TOP_BANK = Toplevel()
        TOP_BANK.title("Add - Bank")
        TOP_BANK.iconbitmap(PATH_ICON + "icon.ico")
        TOP_BANK.geometry("350x325+550+100")
        TOP_BANK.resizable(height = False, width = False)
        TOP_BANK.grab_set()

        global required
        required = []

        LABEL_NAME = Label(TOP_BANK, text = "Name", font = APP_FONT)
        LABEL_NAME.grid(column = 0, row = 0, pady = TOP_PADY, sticky = E)

        LABEL_NUMBER = Label(TOP_BANK, text = "Account No.", font = APP_FONT)
        LABEL_NUMBER.grid(column = 0, row = 1, pady = TOP_PADY, sticky = E)

        LABEL_CONTACT = Label(TOP_BANK, text = "Contact Person", font = APP_FONT)
        LABEL_CONTACT.grid(column = 0, row = 3, pady = TOP_PADY, sticky = E)

        LABEL_PHONE = Label(TOP_BANK, text = "Phone", font = APP_FONT)
        LABEL_PHONE.grid(column = 0, row = 4, pady = TOP_PADY, sticky = E)

        LABEL_ADDRESS = Label(TOP_BANK, text = "Address", font = APP_FONT)
        LABEL_ADDRESS.grid(column = 0, row = 5, pady = TOP_PADY, sticky = E)

        global TEXTVAR_NAME
        TEXTVAR_NAME = StringVar()
        ENTRY_NAME = Entry(TOP_BANK, textvariable = TEXTVAR_NAME, font = APP_FONT, width = 35)
        ENTRY_NAME.grid(column = 1, row = 0, sticky = W)
        required.append(TEXTVAR_NAME)

        global TEXTVAR_ACCOUNT
        TEXTVAR_ACCOUNT = StringVar()
        ENTRY_ACCOUNT = Entry(TOP_BANK, textvariable = TEXTVAR_ACCOUNT, font = APP_FONT, width = 35)
        ENTRY_ACCOUNT.grid(column = 1, row = 1, sticky = W)
        required.append(TEXTVAR_ACCOUNT)

        global TEXTVAR_CONTACT
        TEXTVAR_CONTACT = StringVar()
        ENTRY_CONTACT = Entry(TOP_BANK, textvariable = TEXTVAR_CONTACT, font = APP_FONT, width = 35)
        ENTRY_CONTACT.grid(column = 1, row = 3, sticky = W)
        required.append(TEXTVAR_CONTACT)

        global TEXTVAR_PHONE
        TEXTVAR_PHONE = StringVar()
        ENTRY_PHONE = Entry(TOP_BANK, textvariable = TEXTVAR_PHONE, font = APP_FONT, width = 35)
        ENTRY_PHONE.grid(column = 1, row = 4, sticky = W)
        required.append(TEXTVAR_ACCOUNT)

        global TEXTVAR_ADDRESS
        TEXTVAR_ADDRESS = StringVar()
        ENTRY_ADDRESS = Entry(TOP_BANK, textvariable = TEXTVAR_ADDRESS, font = APP_FONT, width = 35)
        ENTRY_ADDRESS.grid(column = 1, row = 5, sticky = W)
        required.append(TEXTVAR_ADDRESS)

        LABEL_DIVIDER = Label(TOP_BANK, text = "", font = APP_FONT)
        LABEL_DIVIDER.grid(column = 1, row = 6, pady = TOP_PADY + 10, sticky = N)

        LABEL_CODE = Label(TOP_BANK, text = "Code", font = APP_FONT)
        LABEL_CODE.grid(column = 0, row = 7, pady = TOP_PADY, sticky = E)

        global TEXTVAR_CODE
        TEXTVAR_CODE = StringVar()
        ENTRY_CODE = Entry(TOP_BANK, textvariable = TEXTVAR_CODE, font = APP_FONT, width = 35)
        ENTRY_CODE.grid(column = 1, row = 7, sticky = W)
        required.append(TEXTVAR_CODE)

        BUTTON_CODE = Button(TOP_BANK, text = "...", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, cursor = "hand2", command = None)
        BUTTON_CODE.grid(column = 1, row = 7, sticky = E)

        LABEL_TITLE = Label(TOP_BANK, text = "Title", font = APP_FONT)
        LABEL_TITLE.grid(column = 0, row = 8, pady = TOP_PADY, sticky = E)

        LABEL_PARENT = Label(TOP_BANK, text = "Parent", font = APP_FONT)
        LABEL_PARENT.grid(column = 0, row = 9, pady = TOP_PADY, sticky = E)

        global TEXTVAR_TITLE
        TEXTVAR_TITLE = StringVar()
        ENTRY_TITLE = Entry(TOP_BANK, textvariable = TEXTVAR_TITLE, font = APP_FONT, width = 35)
        ENTRY_TITLE.grid(column = 1, row = 8, sticky = W)
        required.append(TEXTVAR_TITLE)

        global TEXTVAR_PARENT
        TEXTVAR_PARENT = StringVar()
        ENTRY_PARENT = Entry(TOP_BANK, textvariable = TEXTVAR_PARENT, font = APP_FONT, width = 35)
        ENTRY_PARENT.grid(column = 1, row = 9, sticky = W)
        required.append(TEXTVAR_PARENT)

        BUTTON_PARENT = Button(TOP_BANK, text = "...", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, cursor = "hand2", command = None)
        BUTTON_PARENT.grid(column = 1, row = 9, sticky = E)

        FRAME_BUTTON = Frame(TOP_BANK)
        FRAME_BUTTON.grid(column = 1, row = 10, sticky = E, pady = TOP_PADY + 10)

        BUTTON_SAVE = Button(FRAME_BUTTON, text = "SAVE", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "hand2", command = self.saveBank)
        BUTTON_SAVE.grid(column = 0, row = 0, padx = TOP_PADX)

        BUTTON_CLOSE = Button(FRAME_BUTTON, text = "CLOSE", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "hand2", command = lambda: self.closeTopLevel(TOP_BANK))
        BUTTON_CLOSE.grid(column = 1, row = 0, padx = TOP_PADX)

    def showAddBankChart(self, *args):
        pass

    def showAddBankParent(self, *args):
        pass

    def saveBank(self, *args):
        wrong = []
        for i in required:
            if i.get() == "":
                wrong.append(i)
                break

        if len(wrong) > 0:
            messagebox.showerror("Save Bank", f"Please provide {str(wrong[0])}!")
        else:
            insert = """INSERT INTO tblbanks (
                        name, account, code, parentCode, contact, phone, address, status)
                        values (%s, %s, %s, %s, %s, %s, %s, %s)"""
            update = """UPDATE tblbanks SET name = %s, account = %s, code = %s, parentCode = %s, contact = %s, phone = %s, address = %s, status = %s
                        WHERE name = %s"""
            find = "SELECT name FROM tblbanks WHERE name = %s"
            try:
                db.commit()
                cursor.execute(find, [content[0]])
                result = cursor.fetchall()
                if result:
                    cursor.execute(update, [TEXTVAR_NAME.get(), TEXTVAR_ACCOUNT.get(), TEXTVAR_CODE.get(), TEXTVAR_PARENT.get(), TEXTVAR_CONTACT.get(), TEXTVAR_PHONE.get(), TEXTVAR_ADDRESS.get(), "active", content[0]])
                    saved = "Bank has been updated!"
                else:
                    cursor.execute(insert, [TEXTVAR_NAME.get(), TEXTVAR_ACCOUNT.get(), TEXTVAR_CODE.get(), TEXTVAR_PARENT.get(), TEXTVAR_CONTACT.get(), TEXTVAR_PHONE.get(), TEXTVAR_ADDRESS.get(), "active"])
                    saved = "A new bank has been saved!"
            except:
                cursor.execute(insert, [TEXTVAR_NAME.get(), TEXTVAR_ACCOUNT.get(), TEXTVAR_CODE.get(), TEXTVAR_PARENT.get(), TEXTVAR_CONTACT.get(), TEXTVAR_PHONE.get(), TEXTVAR_ADDRESS.get(), "active"])
                saved = "A new bank has been saved!"
                
            ask = messagebox.askyesno("Save Bank", "Are you sure?")
            if ask == True:
                db.commit()
                messagebox.showinfo("Save Bank", saved)
                TOP_BANK.grab_release()
                TOP_BANK.destroy()
                self.showBanks()

    def editBank(self, *args):
        self.copySelection(TREE_BANKS)
        self.showAddEditBank()
        TOP_BANK.title("Edit - Bank")

        TEXTVAR_NAME.set(content[0])
        TEXTVAR_ACCOUNT.set(content[1])
        TEXTVAR_CODE.set(content[2])
        TEXTVAR_PARENT.set(content[3])
        TEXTVAR_CONTACT.set(content[4])
        TEXTVAR_PHONE.set(content[4])
        TEXTVAR_ADDRESS.set(content[5])

    def deleteBank(self, *args):
        delete = "DELETE FROM tblbanks WHERE name = %s"
        self.copySelection(TREE_BANKS)
        ask = messagebox.askyesno("Delete Bank", "Are you sure?")
        if ask == True:
            cursor.execute(delete, [content[0]])
            db.commit()
            messagebox.showinfo("Delete Bank", "Bank has been deleted!")
            self.showBanks()

    def searchBank(self, var, *args):
        find = "SELECT name, account, code, parentCode, contact, phone, address, status FROM tblbanks WHERE name LIKE %s OR account LIKE %s OR code LIKE %s ORDER BY name"
        db.commit()
        cursor.execute(find, [f"%{var}%", f"%{var}%", f"%{var}%"])
        result = cursor.fetchall()
        if result:
            for i in TREE_BANKS.get_children():
                TREE_BANKS.delete(i)

            count = 0
            for i in result:
                if count % 2 == 0:
                    TREE_BANKS.insert("", "end", values = (i[0],i[1],i[2],i[3],i[4],i[5],i[6],i[7]), tags = "evenrow")
                else:
                    TREE_BANKS.insert("", "end", values = (i[0],i[1],i[2],i[3],i[4],i[5],i[6],i[7]), tags = "oddrow")
                count += 1
        else:
            messagebox.showerror("Banks", "No match found!")
            for i in TREE_BANKS.get_children():
                TREE_BANKS.delete(i)
            db.commit()
            cursor.execute("SELECT name, account, code, parentCode, contact, phone, address, status FROM tblbanks ORDER BY name")
            result = cursor.fetchall()

            count = 0
            for i in result:
                if count % 2 == 0:
                    TREE_BANKS.insert("", "end", values = (i[0],i[1],i[2],i[3],i[4],i[5],i[6],i[7]), tags = "evenrow")
                else:
                    TREE_BANKS.insert("", "end", values = (i[0],i[1],i[2],i[3],i[4],i[5],i[6],i[7]), tags = "oddrow")
                count += 1

### MENU_ADMIN_CENTERS ###
    def showAddEditCenter(self, *args):
        global TOP_CENTERS
        TOP_CENTERS = Toplevel()
        TOP_CENTERS.title("Add - Cost Center")
        TOP_CENTERS.iconbitmap(PATH_ICON + "icon.ico")
        TOP_CENTERS.geometry('350x325+550+100')
        TOP_CENTERS.resizable(height = False, width = False)
        TOP_CENTERS.grab_set()

        global required
        required = []

        LABEL_CODE = Label(TOP_CENTERS, text = "Code", font = APP_FONT)
        LABEL_CODE.grid(column = 0, row = 0, pady = TOP_PADY, sticky = E)

        LABEL_CENTER = Label(TOP_CENTERS, text = "Cost Center", font = APP_FONT)
        LABEL_CENTER.grid(column = 0, row = 1, pady = TOP_PADY, sticky = E)

        global TEXTVAR_CODE
        TEXTVAR_CODE = StringVar()
        ENTRY_CODE = Entry(TOP_CENTERS, textvariable = TEXTVAR_CODE, font = APP_FONT, width = 35)
        ENTRY_CODE.grid(column = 1, row = 0, sticky = W)
        required.append(TEXTVAR_CODE)

        global TEXTVAR_CENTER
        TEXTVAR_CENTER = StringVar()
        ENTRY_CENTER = Entry(TOP_CENTERS, textvariable = TEXTVAR_CENTER, font = APP_FONT, width = 35)
        ENTRY_CENTER.grid(column = 1, row = 1, sticky = W)
        required.append(TEXTVAR_CENTER)

        FRAME_BUTTON = Frame(TOP_CENTERS)
        FRAME_BUTTON.grid(column = 1, row = 2, sticky = E, pady = TOP_PADY + 10)

        BUTTON_SAVE = Button(FRAME_BUTTON, text = "SAVE", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "hand2", command = self.saveCenter)
        BUTTON_SAVE.grid(column = 0, row = 0, padx = TOP_PADX)

        BUTTON_CLOSE = Button(FRAME_BUTTON, text = "CLOSE", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "hand2", command = lambda: self.closeTopLevel(TOP_CENTERS))
        BUTTON_CLOSE.grid(column = 1, row = 0, padx = TOP_PADX)

    def saveCenter(self, *args):
        wrong = []
        for i in required:
            if i.get() == "":
                wrong.append(i)
                break
            
        if len(wrong) > 0:
            messagebox.showerror("Save Center", f"Please provide {str(wrong[0])}!")
        else:
            insert = "INSERT INTO tblcenters (code, name) values (%s, %s)"
            update = "UPDATE tblcenters SET code = %s, name = %s WHERE code = %s"
            find = "SELECT name FROM tblcenters WHERE code = %s"
            try:
                db.commit()
                cursor.execute(find, [content[0]])
                result = cursor.fetchall()
                if result:
                    cursor.execute(update, [TEXTVAR_CODE.get(), TEXTVAR_CENTER.get(), content[0]])
                    saved = "Center has been updated!"
                else:
                    cursor.execute(insert, [TEXTVAR_CODE.get(), TEXTVAR_CENTER.get()])
                    saved = "A new center has been saved!"
            except:
                cursor.execute(insert, [TEXTVAR_CODE.get(), TEXTVAR_CENTER.get()])
                saved = "A new center has been saved!"
                
            ask = messagebox.askyesno("Save Center", "Are you sure?")
            if ask == True:
                db.commit()
                messagebox.showinfo("Save Center", saved)
                TOP_CENTERS.grab_release()
                TOP_CENTERS.destroy()
                self.showCostCenters()

    def editCenter(self, *args):
        self.copySelection(TREE_CENTERS)
        self.showAddEditCenter()
        TOP_CENTERS.title("Edit - Centers")

        TEXTVAR_CODE.set(content[0])
        TEXTVAR_CENTER.set(content[1])

    def deleteCenter(self, *args):
        delete = "DELETE FROM tblcenters WHERE code = %s"
        self.copySelection(TREE_CENTERS)
        ask = messagebox.askyesno("Delete Center", "Are you sure?")
        if ask == True:
            cursor.execute(delete, [content[0]])
            db.commit()
            messagebox.showinfo("Delete Center", "Center has been deleted!")
            self.showCostCenters()

    def searchCenter(self, var, *args):
        find = "SELECT code, name FROM tblcenters WHERE code LIKE %s OR name LIKE %s ORDER BY name"
        db.commit()
        cursor.execute(find, [f"%{var}%", f"%{var}%"])
        result = cursor.fetchall()
        if result:
            for i in TREE_CENTERS.get_children():
                TREE_CENTERS.delete(i)

            count = 0
            for i in result:
                if count % 2 == 0:
                    TREE_CENTERS.insert("", "end", values = (i[0],i[1]), tags = "evenrow")
                else:
                    TREE_CENTERS.insert("", "end", values = (i[0],i[1]), tags = "oddrow")
                count += 1
        else:
            messagebox.showerror("Centers", "No match found!")
            for i in TREE_CENTERS.get_children():
                TREE_CENTERS.delete(i)
            db.commit()
            cursor.execute("SELECT code, name FROM tblcenters ORDER BY name")
            result = cursor.fetchall()

            count = 0
            for i in result:
                if count % 2 == 0:
                    TREE_CENTERS.insert("", "end", values = (i[0],i[1]), tags = "evenrow")
                else:
                    TREE_CENTERS.insert("", "end", values = (i[0],i[1]), tags = "oddrow")
                count += 1

    def showCenterSelection(self, rown, code, name):
        global TOP_CENTERSELECTION
        TOP_CENTERSELECTION = Toplevel()
        TOP_CENTERSELECTION.title("Choose Cost Center")
        TOP_CENTERSELECTION.iconbitmap(PATH_ICON + "icon.ico")
        TOP_CENTERSELECTION.geometry("400x300+200+100")
        TOP_CENTERSELECTION.resizable(height = False, width = False)
        TOP_CENTERSELECTION.grab_set()

        FRAME1 = Frame(TOP_CENTERSELECTION)
        FRAME1.pack(fill = "x")

        FRAME2 = Frame(TOP_CENTERSELECTION)
        FRAME2.pack(fill = "x")

        LABEL_SEARCH = Label(FRAME1, text = "Search", font = APP_FONT)
        LABEL_SEARCH.grid(column = 0, row = 0, padx = TOP_PADX)
        
        global TEXTVAR_SEARCH
        TEXTVAR_SEARCH = StringVar()
        ENTRY_SEARCH = Entry(FRAME1, font = APP_FONT, textvariable = TEXTVAR_SEARCH, width = 20)
        ENTRY_SEARCH.grid(column = 1, row = 0, padx = TOP_PADX)
        ENTRY_SEARCH.bind("<Return>", lambda e: self.searchCenter(TEXTVAR_SEARCH.get()))
        
        BUTTON_SEARCH = Button(FRAME1, text = "GO", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, command = lambda: self.searchCenter(TEXTVAR_SEARCH.get()))
        BUTTON_SEARCH.grid(column = 2, row = 0, padx = TOP_PADX)

        self.showCenterTreeview(FRAME2)
        TREE_CENTER.config(height = 10)
        TREE_CENTER.unbind("<Button-3>")
        TREE_CENTER.bind("<Double-1>", lambda e: self.enterCenterSelection(rown, code, name))

    def showCenterTreeview(self, frame):
        global TREE_CENTER
        TREE_CENTER = tk.Treeview(frame, height = 28, selectmode = "browse")
        TREE_CENTER["columns"] = ("Code", "Name")
        TREE_CENTER.column("#0", width = 0, minwidth = 0, stretch = True)
        TREE_CENTER.column("Code", anchor = W, minwidth = 100, width = 150)
        TREE_CENTER.column("Name", anchor = W, minwidth = 100, width = 250)
        
        TREE_CENTER.heading("#0", text = "", anchor = W)
        TREE_CENTER.heading("Code", text = "Code", anchor = N)
        TREE_CENTER.heading("Name", text = "Name", anchor = N)
        
        global POPUP_CENTER
        POPUP_CENTER = Menu(TREE_CENTER, tearoff = 0)
        POPUP_CENTER.add_command(command = self.editCenter, label = "Edit")
        POPUP_CENTER.add_command(command = self.deleteCenter, label = "Delete")
        TREE_CENTER.bind("<Button-3>", lambda e: self.popupMenu(TREE_CENTER, POPUP_CENTER, e))

        global STYLE_CENTERS
        STYLE_CENTERS = tk.Style()
        STYLE_CENTERS.map("Treeview", foreground = self.fixedMap("foreground", STYLE_CENTERS), background = self.fixedMap("background", STYLE_CENTERS))

        TREE_CENTER.tag_configure("oddrow", background = None)
        TREE_CENTER.tag_configure("evenrow", background = TREE_TAG_EVENROW)
        db.commit()
        cursor.execute("SELECT code, name FROM tblcenters ORDER BY code")
        result = cursor.fetchall()
        count = 0
        for i in result:
            if count % 2 == 0:
                TREE_CENTER.insert("", "end", values = (i[0],i[1]), tags = "evenrow")
            else:
                TREE_CENTER.insert("", "end", values = (i[0],i[1]), tags = "oddrow")
            count += 1

        YSCROLLBAR = tk.Scrollbar(frame, orient = "vertical", command = TREE_CENTER.yview)
        YSCROLLBAR.pack(side = "right", fill = "y")

        XSCROLLBAR = tk.Scrollbar(frame, orient = "horizontal", command = TREE_CENTER.xview)
        
        TREE_CENTER.configure(yscrollcommand = YSCROLLBAR.set, xscrollcommand = XSCROLLBAR.set) 
        TREE_CENTER.pack()
        XSCROLLBAR.pack(fill ="x")

    def enterCenterSelection(self, rown, code, name):
        self.copySelection(TREE_CENTER)
        TOP_CENTERSELECTION.grab_release()
        TOP_CENTERSELECTION.destroy()
        code[rown].set(str(content[0]))
        name[rown].set(str(content[1]))

### MENU_ADMIN_CHART ###
    def showAddEditChart(self, *args):
        global TOP_CHART
        TOP_CHART = Toplevel()
        TOP_CHART.title("Add - Chart of Accounts")
        TOP_CHART.iconbitmap(PATH_ICON + "icon.ico")
        TOP_CHART.geometry('350x325+550+100')
        TOP_CHART.resizable(height = False, width = False)
        TOP_CHART.grab_set()

        global required
        required = []

        LABEL_CODE = Label(TOP_CHART, text = "Code", font = APP_FONT)
        LABEL_CODE.grid(column = 0, row = 0, pady = TOP_PADY, sticky = E)

        LABEL_TITLE = Label(TOP_CHART, text = "Title", font = APP_FONT)
        LABEL_TITLE.grid(column = 0, row = 1, pady = TOP_PADY, sticky = E)

        LABEL_PARENT = Label(TOP_CHART, text = "Parent", font = APP_FONT)
        LABEL_PARENT.grid(column = 0, row = 2, pady = TOP_PADY, sticky = E)
        
        LABEL_PARENTTITLE = Label(TOP_CHART, text = "Parent Title", font = APP_FONT)
        LABEL_PARENTTITLE.grid(column = 0, row = 3, pady = TOP_PADY, sticky = E)

        LABEL_REMARKS = Label(TOP_CHART, text = "Remarks", font = APP_FONT)
        LABEL_REMARKS.grid(column = 0, row = 4, pady = TOP_PADY, sticky = E)

        LABEL_ACTIVE = Label(TOP_CHART, text = "Active", font = APP_FONT)
        LABEL_ACTIVE.grid(column = 0, row = 5, pady = TOP_PADY, sticky = E)

        global TEXTVAR_CODE
        TEXTVAR_CODE = StringVar()
        ENTRY_CODE = Entry(TOP_CHART, textvariable = TEXTVAR_CODE, font = APP_FONT, width = 35)
        ENTRY_CODE.grid(column = 1, row = 0, sticky = W)
        required.append(TEXTVAR_CODE)

        BUTTON_CODE = Button(TOP_CHART, text = "...", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, cursor = "hand2", command = None)
        BUTTON_CODE.grid(column = 1, row = 0, sticky = E)

        global TEXTVAR_TITLE
        TEXTVAR_TITLE = StringVar()
        ENTRY_TITLE = Entry(TOP_CHART, textvariable = TEXTVAR_TITLE, font = APP_FONT, width = 35)
        ENTRY_TITLE.grid(column = 1, row = 1, sticky = W)
        required.append(TEXTVAR_TITLE)

        global TEXTVAR_PARENT
        TEXTVAR_PARENT = StringVar()
        ENTRY_PARENT = Entry(TOP_CHART, textvariable = TEXTVAR_PARENT, font = APP_FONT, width = 35)
        ENTRY_PARENT.grid(column = 1, row = 2, sticky = W)
        required.append(TEXTVAR_PARENT)
        
        BUTTON_PARENT = Button(TOP_CHART, text = "...", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, cursor = "hand2", command = None)
        BUTTON_PARENT.grid(column = 1, row = 2, sticky = E)
        
        global TEXTVAR_PARENTTITLE
        TEXTVAR_PARENTTITLE = StringVar()
        ENTRY_PARENTTITLE = Entry(TOP_CHART, textvariable = TEXTVAR_PARENTTITLE, font = APP_FONT, width = 35, state = "readonly")
        ENTRY_PARENTTITLE.grid(column = 1, row = 3, sticky = W)

        global TEXTVAR_REMARKS
        TEXTVAR_REMARKS = StringVar()
        ENTRY_REMARKS = Entry(TOP_CHART, textvariable = TEXTVAR_REMARKS, font = APP_FONT, width = 35)
        ENTRY_REMARKS.grid(column = 1, row = 4, sticky = W)

        global TEXTVAR_ACTIVE
        TEXTVAR_ACTIVE = StringVar()
        ENTRY_ACTIVE = tk.Combobox(TOP_CHART, values = ["Yes", "No"], textvariable = TEXTVAR_ACTIVE, font = APP_FONT, width = 35, state = "readonly")
        ENTRY_ACTIVE.grid(column = 1, row = 5, sticky = W)
        required.append(TEXTVAR_ACTIVE)

        FRAME_BUTTON = Frame(TOP_CHART)
        FRAME_BUTTON.grid(column = 1, row = 6, sticky = E, pady = TOP_PADY + 10)

        BUTTON_SAVE = Button(FRAME_BUTTON, text = "SAVE", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "hand2", command = self.saveChart)
        BUTTON_SAVE.grid(column = 0, row = 0, padx = TOP_PADX)

        BUTTON_CLOSE = Button(FRAME_BUTTON, text = "CLOSE", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "hand2", command = lambda: self.closeTopLevel(TOP_CHART))
        BUTTON_CLOSE.grid(column = 1, row = 0, padx = TOP_PADX)

    def saveChart(self, *args):
        wrong = []
        for i in required:
            if i.get() == "":
                wrong.append(i)
                break
        
        if len(wrong) > 0:
            messagebox.showerror("Save Code", f"Please provide {str(wrong[0])}!")
        else:
            insert = """INSERT INTO tblchart (
                        code, title, parent, remarks, active, encoder, encoded)
                        values (%s, %s, %s, %s, %s, %s, %s)"""
            update = """UPDATE tblchart SET code = %s, title = %s, parent = %s, remarks = %s, active = %s, modifier = %s, modified = %s
                        WHERE code = %s"""
            find = "SELECT code FROM tblchart WHERE code = %s"
            try:
                db.commit()
                cursor.execute(find, [content[0]])
                result = cursor.fetchall()
                if result:
                    cursor.execute(update, [TEXTVAR_CODE.get(), TEXTVAR_TITLE.get(), TEXTVAR_PARENT.get(), TEXTVAR_REMARKS.get(), TEXTVAR_ACTIVE.get(), USER, datetime.datetime.now(), content[0]])
                    saved = "Code has been updated!"
                else:
                    cursor.execute(insert, [TEXTVAR_CODE.get(), TEXTVAR_TITLE.get(), TEXTVAR_PARENT.get(), TEXTVAR_REMARKS.get(), TEXTVAR_ACTIVE.get(), USER, datetime.datetime.now()])
                    saved = "A new code has been saved!"
            except:
                cursor.execute(insert, [TEXTVAR_CODE.get(), TEXTVAR_TITLE.get(), TEXTVAR_PARENT.get(), TEXTVAR_REMARKS.get(), TEXTVAR_ACTIVE.get(), USER, datetime.datetime.now()])
                saved = "A new code has been saved!"
                
            ask = messagebox.askyesno("Save Code", "Are you sure?")
            if ask == True:
                db.commit()
                messagebox.showinfo("Save Code", saved)
                TOP_CHART.grab_release()
                TOP_CHART.destroy()
                self.showChartofAccounts()

    def editChart(self, *args):
        self.copySelection(TREE_CHART)
        self.showAddEditChart()
        TOP_CHART.title("Edit - Code")

        TEXTVAR_CODE.set(content[0])
        TEXTVAR_TITLE.set(content[1])
        TEXTVAR_PARENT.set(content[2])
        TEXTVAR_REMARKS.set(content[3])
        TEXTVAR_ACTIVE.set(content[4])

    def deleteChart(self, *args):
        delete = "DELETE FROM tblchart WHERE code = %s"
        self.copySelection(TREE_CHART)
        ask = messagebox.askyesno("Delete Code", "Are you sure?")
        if ask == True:
            cursor.execute(delete, [content[0]])
            db.commit()
            messagebox.showinfo("Delete Code", "Code has been deleted!")
            self.showChartofAccounts()

    def populateChartFields(self, var, var2, *args):
        find = "SELECT title FROM tblchart WHERE code = %s LIMIT 1"
        cursor.execute(find, [var.get()])
        result = cursor.fetchone()
        if result:
            if var.get()[3:] != "000":
                var2.set(result[0])
            else:
                messagebox.showerror("Chart of Accounts", f"Parent Account {var.get()} {result[0]} cannot be used!")
                var.set("")
                var2.set("")
        else:
            var.set("")
            var2.set("")

    def searchChartofAccount(self, var, *args):
        find = "SELECT code, title, parent, remarks, active FROM tblchart WHERE code LIKE %s OR title LIKE %s AND active = 'Yes' ORDER BY code"
        db.commit()
        cursor.execute(find, [f"%{var}%", f"%{var}%"])
        result = cursor.fetchall()
        if result:
            for i in TREE_CHART.get_children():
                TREE_CHART.delete(i)

            count = 0
            for i in result:
                if count % 2 == 0:
                    TREE_CHART.insert("", "end", values = (i[0],i[1],i[2],i[3],i[4]), tags = "evenrow")
                else:
                    TREE_CHART.insert("", "end", values = (i[0],i[1],i[2],i[3],i[4]), tags = "oddrow")
                count += 1
        else:
            messagebox.showerror("Chart of Accounts", "No match found!")
            for i in TREE_CHART.get_children():
                TREE_CHART.delete(i)
            db.commit()
            cursor.execute("SELECT code, title, parent, remarks, active FROM tblchart WHERE active = 'Yes' ORDER BY code")
            result = cursor.fetchall()

            count = 0
            for i in result:
                if count % 2 == 0:
                    TREE_CHART.insert("", "end", values = (i[0],i[1],i[2],i[3],i[4]), tags = "evenrow")
                else:
                    TREE_CHART.insert("", "end", values = (i[0],i[1],i[2],i[3],i[4]), tags = "oddrow")
                count += 1
    
    def showChartSelection(self, line, ccode, ctitle, *args):
        global TOP_CHARTSELECTION
        TOP_CHARTSELECTION = Toplevel()
        TOP_CHARTSELECTION.title("Choose Chart Code")
        TOP_CHARTSELECTION.iconbitmap(PATH_ICON + "icon.ico")
        TOP_CHARTSELECTION.geometry("600x300+100+100")
        TOP_CHARTSELECTION.resizable(height = False, width = False)
        TOP_CHARTSELECTION.grab_set()

        FRAME1 = Frame(TOP_CHARTSELECTION)
        FRAME1.pack(fill = "x")

        FRAME2 = Frame(TOP_CHARTSELECTION)
        FRAME2.pack(fill = "x")

        LABEL_SEARCH = Label(FRAME1, text = "Search", font = APP_FONT)
        LABEL_SEARCH.grid(column = 0, row = 0, padx = TOP_PADX)
        
        global TEXTVAR_SEARCH
        TEXTVAR_SEARCH = StringVar()
        ENTRY_SEARCH = Entry(FRAME1, font = APP_FONT, textvariable = TEXTVAR_SEARCH, width = 20)
        ENTRY_SEARCH.grid(column = 1, row = 0, padx = TOP_PADX)
        ENTRY_SEARCH.bind("<Return>", lambda e: self.searchChartofAccount(TEXTVAR_SEARCH.get()))
        
        BUTTON_SEARCH = Button(FRAME1, text = "GO", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, command = lambda: self.searchChartofAccount(TEXTVAR_SEARCH.get()))
        BUTTON_SEARCH.grid(column = 2, row = 0, padx = TOP_PADX)

        self.showChartTreeview(FRAME2)
        TREE_CHART.config(height = 10)
        TREE_CHART.unbind("<Button-3>")
        TREE_CHART.bind("<Double-1>", lambda e: self.enterChartSelection(line, ccode, ctitle))

    def enterChartSelection(self, item, ccode, ctitle, *args):
        self.copySelection(TREE_CHART)
        TOP_CHARTSELECTION.grab_release()
        TOP_CHARTSELECTION.destroy()
        try:
            ccode[item].set(content[0])
            ctitle[item].set(content[1])
        except:
            try:
                TOP_RECEIVABLES.focus()
            except:
                pass

    def returnChartTitle(self, var):
        find = "SELECT title FROM tblchart WHERE code = %s LIMIT 1"
        cursor.execute(find, [var])
        result = cursor.fetchone()
        return result

### MENU_ADMIN_CLIENTS ###
    def showAddEditClient(self, *args):
        global TOP_CLIENT
        TOP_CLIENT = Toplevel()
        TOP_CLIENT.title("Add - Clients")
        TOP_CLIENT.iconbitmap(PATH_ICON + "icon.ico")
        TOP_CLIENT.geometry('350x325+550+100')
        TOP_CLIENT.resizable(height = False, width = False)
        TOP_CLIENT.grab_set()

        global required
        required = []

        LABEL_CODE = Label(TOP_CLIENT, text = "Code", font = APP_FONT)
        LABEL_CODE.grid(column = 0, row = 0, pady = TOP_PADY, sticky = E)

        LABEL_NAME = Label(TOP_CLIENT, text = "Name", font = APP_FONT)
        LABEL_NAME.grid(column = 0, row = 1, pady = TOP_PADY, sticky = E)

        LABEL_TIN = Label(TOP_CLIENT, text = "Address", font = APP_FONT)
        LABEL_TIN.grid(column = 0, row = 2, pady = TOP_PADY, sticky = E)

        LABEL_ADDRESS = Label(TOP_CLIENT, text = "TIN", font = APP_FONT)
        LABEL_ADDRESS.grid(column = 0, row = 3, pady = TOP_PADY, sticky = E)

        LABEL_PARENT = Label(TOP_CLIENT, text = "Contact", font = APP_FONT)
        LABEL_PARENT.grid(column = 0, row = 4, pady = TOP_PADY, sticky = E)

        LABEL_CONTACT = Label(TOP_CLIENT, text = "Entity", font = APP_FONT)
        LABEL_CONTACT.grid(column = 0, row = 5, pady = TOP_PADY, sticky = E)
        
        LABEL_CONTACT = Label(TOP_CLIENT, text = "Tax", font = APP_FONT)
        LABEL_CONTACT.grid(column = 0, row = 6, pady = TOP_PADY, sticky = E)
        
        LABEL_CONTACT = Label(TOP_CLIENT, text = "Parent", font = APP_FONT)
        LABEL_CONTACT.grid(column = 0, row = 7, pady = TOP_PADY, sticky = E)

        global TEXTVAR_CODE
        TEXTVAR_CODE = StringVar()
        ENTRY_CODE = Entry(TOP_CLIENT, textvariable = TEXTVAR_CODE, font = APP_FONT, width = 35)
        ENTRY_CODE.grid(column = 1, row = 0, sticky = W)
        required.append(TEXTVAR_CODE)

        global TEXTVAR_NAME
        TEXTVAR_NAME = StringVar()
        ENTRY_NAME = Entry(TOP_CLIENT, textvariable = TEXTVAR_NAME, font = APP_FONT, width = 35)
        ENTRY_NAME.grid(column = 1, row = 1, sticky = W)
        required.append(TEXTVAR_NAME)

        global TEXTVAR_ADDRESS
        TEXTVAR_ADDRESS = StringVar()
        ENTRY_ADDRESS = Entry(TOP_CLIENT, textvariable = TEXTVAR_ADDRESS, font = APP_FONT, width = 35)
        ENTRY_ADDRESS.grid(column = 1, row = 2, sticky = W)
        required.append(TEXTVAR_ADDRESS)

        global TEXTVAR_TIN
        TEXTVAR_TIN = StringVar()
        ENTRY_TIN = Entry(TOP_CLIENT, textvariable = TEXTVAR_TIN, font = APP_FONT, width = 35)
        ENTRY_TIN.grid(column = 1, row = 3, sticky = W)
        required.append(TEXTVAR_TIN)
        
        global TEXTVAR_CONTACT
        TEXTVAR_CONTACT = StringVar()
        ENTRY_CONTACT = Entry(TOP_CLIENT, textvariable = TEXTVAR_CONTACT, font = APP_FONT, width = 35)
        ENTRY_CONTACT.grid(column = 1, row = 4, sticky = W)
        required.append(TEXTVAR_CONTACT)
        
        global TEXTVAR_ENTITY
        TEXTVAR_ENTITY = StringVar()
        ENTRY_ENTITY = tk.Combobox(TOP_CLIENT, values = ["Private", "Government"], textvariable = TEXTVAR_ENTITY, font = APP_FONT, width = 10, state = "readonly")
        ENTRY_ENTITY.grid(column = 1, row = 5, sticky = W)
        required.append(TEXTVAR_ENTITY)

        global TEXTVAR_TAX
        TEXTVAR_TAX = StringVar()
        ENTRY_TAX = tk.Combobox(TOP_CLIENT, values = ["VAT", "Non-VAT"], textvariable = TEXTVAR_TAX, font = APP_FONT, width = 10, state = "readonly")
        ENTRY_TAX.grid(column = 1, row = 6, sticky = W)
        required.append(TEXTVAR_TAX)
        
        global TEXTVAR_PARENT
        TEXTVAR_PARENT = StringVar()
        ENTRY_PARENT = Entry(TOP_CLIENT, textvariable = TEXTVAR_PARENT, font = APP_FONT, width = 35)
        ENTRY_PARENT.grid(column = 1, row = 7, sticky = W)
        required.append(TEXTVAR_PARENT)

        FRAME_BUTTON = Frame(TOP_CLIENT)
        FRAME_BUTTON.grid(column = 1, row = 8, sticky = E, pady = TOP_PADY + 10)

        BUTTON_SAVE = Button(FRAME_BUTTON, text = "SAVE", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "hand2", command = self.saveClient)
        BUTTON_SAVE.grid(column = 0, row = 0, padx = TOP_PADX)

        BUTTON_CLOSE = Button(FRAME_BUTTON, text = "CLOSE", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "hand2", command = lambda: self.closeTopLevel(TOP_CLIENT))
        BUTTON_CLOSE.grid(column = 1, row = 0, padx = TOP_PADX)

    def saveClient(self, *args):
        wrong = []
        for i in required:
            if i.get() == "":
                wrong.append(i)
                break
            
        if len(wrong) > 0:
            messagebox.showerror("Save Client", f"Please provide {str(wrong[0])}!")
        else:
            insert = """INSERT INTO tblclients (
                        clientCode, clientName, address, tin, contact, entityType, taxType, parentCode, encoder, encoded, company)
                        values (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
            update = """UPDATE tblclients SET clientCode = %s, clientName = %s, address = %s, tin = %s, contact = %s, entityType = %s, taxType = %s, parentCode = %s, modifier = %s, modified = %s
                        WHERE clientCode = %s"""
            find = "SELECT clientName FROM tblclients WHERE clientCode = %s"
            try:
                db.commit()
                cursor.execute(find, [TEXTVAR_CODE.get()])
                result = cursor.fetchall()
                if result:
                    cursor.execute(update, [TEXTVAR_CODE.get(), TEXTVAR_NAME.get(), TEXTVAR_ADDRESS.get(), TEXTVAR_TIN.get(), TEXTVAR_CONTACT.get(), TEXTVAR_ENTITY.get(), TEXTVAR_TAX.get(), TEXTVAR_PARENT.get(), USER, datetime.datetime.now(), TEXTVAR_CODE.get()])
                    saved = "Client has been updated!"
                else:
                    cursor.execute(insert, [TEXTVAR_CODE.get(), TEXTVAR_NAME.get(), TEXTVAR_ADDRESS.get(), TEXTVAR_TIN.get(), TEXTVAR_CONTACT.get(), TEXTVAR_ENTITY.get(), TEXTVAR_TAX.get(), TEXTVAR_PARENT.get(), USER, datetime.datetime.now(), "DBPSC"])
                    saved = "A new client has been saved!"
            except Exception as e:
                print(e)
                cursor.execute(insert, [TEXTVAR_CODE.get(), TEXTVAR_NAME.get(), TEXTVAR_ADDRESS.get(), TEXTVAR_TIN.get(), TEXTVAR_CONTACT.get(), TEXTVAR_ENTITY.get(), TEXTVAR_TAX.get(), TEXTVAR_PARENT.get(), USER, datetime.datetime.now(), "DBPSC"])
                saved = "A new client has been saved!"
                
            ask = messagebox.askyesno("Save Client", "Are you sure?")
            if ask == True:
                db.commit()
                messagebox.showinfo("Save Client", saved)
                TOP_CLIENT.grab_release()
                TOP_CLIENT.destroy()
                self.showClients()

    def editClient(self, *args):
        self.copySelection(TREE_CLIENTS)
        self.showAddEditClient()
        TOP_CLIENT.title("Edit - Client")

        TEXTVAR_CODE.set(content[0])
        TEXTVAR_NAME.set(content[1])
        TEXTVAR_TIN.set(content[3])
        TEXTVAR_ADDRESS.set(content[2])
        TEXTVAR_PARENT.set(content[7])
        TEXTVAR_CONTACT.set(content[4])
        TEXTVAR_ENTITY.set(content[5])
        TEXTVAR_TAX.set(content[6])

    def deleteClient(self, *args):
        delete = "DELETE FROM tblclients WHERE clientCode = %s"
        self.copySelection(TREE_CLIENTS)
        ask = messagebox.askyesno("Delete Client", "Are you sure?")
        if ask == True:
            cursor.execute(delete, [content[0]])
            db.commit()
            messagebox.showinfo("Delete Client", "Client has been deleted!")
            self.showClients()

    def showClientTreeview(self, frame):
        global TREE_CLIENTS
        TREE_CLIENTS = tk.Treeview(frame, height = 28, selectmode = "browse")
        TREE_CLIENTS["columns"] = ("Code", "Name", "Address", "TIN", "Contact", "Entity", "Tax", "Parent")
        TREE_CLIENTS.column("#0", width = 0, minwidth = 0, stretch = True)
        TREE_CLIENTS.column("Code", anchor = W, minwidth = 100, width = 150)
        TREE_CLIENTS.column("Name", anchor = W, minwidth = 100, width = 250)
        TREE_CLIENTS.column("Address", anchor = W, minwidth = 100, width = 150)
        TREE_CLIENTS.column("TIN", anchor = W, minwidth = 100, width = 150)
        TREE_CLIENTS.column("Contact", anchor = W, minwidth = 100, width = 150)
        TREE_CLIENTS.column("Entity", anchor = W, minwidth = 100, width = 75)
        TREE_CLIENTS.column("Tax", anchor = W, minwidth = 100, width = 75)
        TREE_CLIENTS.column("Parent", anchor = W, minwidth = 100, width = 150)
        
        TREE_CLIENTS.heading("#0", text = "", anchor = W)
        TREE_CLIENTS.heading("Code", text = "Code", anchor = N)
        TREE_CLIENTS.heading("Name", text = "Name", anchor = N)
        TREE_CLIENTS.heading("Address", text = "Address", anchor = N)
        TREE_CLIENTS.heading("TIN", text = "TIN", anchor = N)
        TREE_CLIENTS.heading("Contact", text = "Contact", anchor = N)
        TREE_CLIENTS.heading("Entity", text = "Entity", anchor = N)
        TREE_CLIENTS.heading("Tax", text = "Tax", anchor = N)
        TREE_CLIENTS.heading("Parent", text = "Parent", anchor = N)
        
        global POPUP_CLIENT
        POPUP_CLIENT = Menu(TREE_CLIENTS, tearoff = 0)
        POPUP_CLIENT.add_command(command = self.editClient, label = "Edit")
        POPUP_CLIENT.add_command(command = self.deleteClient, label = "Delete")
        TREE_CLIENTS.bind("<Button-3>", lambda e: self.popupMenu(TREE_CLIENTS, POPUP_CLIENT, e))
        TREE_CLIENTS.bind("<Double-1>", self.editClient)

        global STYLE_CLIENT
        STYLE_CLIENT = tk.Style()
        STYLE_CLIENT.map("Treeview", foreground = self.fixedMap("foreground", STYLE_CLIENT), background = self.fixedMap("background", STYLE_CLIENT))

        TREE_CLIENTS.tag_configure("oddrow", background = None)
        TREE_CLIENTS.tag_configure("evenrow", background = TREE_TAG_EVENROW)
        db.commit()
        try:
            if TOP_COLLECTIONS.winfo_exists() == 0:
                cursor.execute("""SELECT clientCode, clientName, address, tin, contact, entityType, taxType, parentCode FROM tblclients ORDER BY clientName""")
            else:
                cursor.execute("""SELECT parentCode, clientName, address, tin, contact, entityType, taxType, clientCode FROM tblclients ORDER BY clientName""")
        except:
            cursor.execute("""SELECT clientCode, clientName, address, tin, contact, entityType, taxType, parentCode FROM tblclients ORDER BY clientName""")
        result = cursor.fetchall()
        count = 0
        for i in result:
            if count % 2 == 0:
                TREE_CLIENTS.insert("", "end", values = (i[0],i[1],i[2],i[3],i[4],i[5],i[6],i[7]), tags = "evenrow")
            else:
                TREE_CLIENTS.insert("", "end", values = (i[0],i[1],i[2],i[3],i[4],i[5],i[6],i[7]), tags = "oddrow")
            count += 1

        YSCROLLBAR = tk.Scrollbar(frame, orient = "vertical", command = TREE_CLIENTS.yview)
        YSCROLLBAR.pack(side = "right", fill = "y")

        XSCROLLBAR = tk.Scrollbar(frame, orient = "horizontal", command = TREE_CLIENTS.xview)
        
        TREE_CLIENTS.configure(yscrollcommand = YSCROLLBAR.set, xscrollcommand = XSCROLLBAR.set) 
        TREE_CLIENTS.pack()
        XSCROLLBAR.pack(fill ="x")

    def showClientSelection(self, *args):
        global TOP_CLIENTSELECTION
        TOP_CLIENTSELECTION = Toplevel()
        TOP_CLIENTSELECTION.title("Choose Client Code")
        TOP_CLIENTSELECTION.iconbitmap(PATH_ICON + "icon.ico")
        TOP_CLIENTSELECTION.geometry("600x300+100+100")
        TOP_CLIENTSELECTION.resizable(height = False, width = False)
        TOP_CLIENTSELECTION.grab_set()

        FRAME1 = Frame(TOP_CLIENTSELECTION)
        FRAME1.pack(fill = "x")

        FRAME2 = Frame(TOP_CLIENTSELECTION)
        FRAME2.pack(fill = "x")

        LABEL_SEARCH = Label(FRAME1, text = "Search", font = APP_FONT)
        LABEL_SEARCH.grid(column = 0, row = 0, padx = TOP_PADX)
        
        global TEXTVAR_SEARCH
        TEXTVAR_SEARCH = StringVar()
        ENTRY_SEARCH = Entry(FRAME1, textvariable = TEXTVAR_SEARCH, font = APP_FONT, width = 20)
        ENTRY_SEARCH.grid(column = 1, row = 0, padx = TOP_PADX)
        ENTRY_SEARCH.bind("<Return>", lambda e: self.searchClient(TEXTVAR_SEARCH.get()))
        
        BUTTON_SEARCH = Button(FRAME1, text = "GO", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, command = lambda: self.searchClient(TEXTVAR_SEARCH.get()))
        BUTTON_SEARCH.grid(column = 2, row = 0, padx = TOP_PADX)

        self.showClientTreeview(FRAME2)
        TREE_CLIENTS.config(height = 10)
        TREE_CLIENTS.unbind("<Button-3>")
        TREE_CLIENTS.bind("<Double-1>", self.enterClientSelection)
    
    def searchClient(self, var, *args):
        find = "SELECT clientCode, clientName, address, tin, contact, entityType, taxType, parentCode FROM tblclients WHERE clientCode LIKE %s OR clientName LIKE %s ORDER BY clientName"
        db.commit()
        cursor.execute(find, [f"%{var}%", f"%{var}%"])
        result = cursor.fetchall()
        if result:
            for i in TREE_CLIENTS.get_children():
                TREE_CLIENTS.delete(i)

            count = 0
            for i in result:
                if count % 2 == 0:
                    TREE_CLIENTS.insert("", "end", values = (i[0],i[1],i[2],i[3],i[4],i[5],i[6],i[7]), tags = "evenrow")
                else:
                    TREE_CLIENTS.insert("", "end", values = (i[0],i[1],i[2],i[3],i[4],i[5],i[6],i[7]), tags = "oddrow")
                count += 1
        else:
            messagebox.showerror("Clients", "No match found!")
            for i in TREE_CLIENTS.get_children():
                TREE_CLIENTS.delete(i)
            db.commit()
            cursor.execute("""SELECT clientCode, clientName, address, tin, contact, entityType, taxType, parentCode FROM tblclients ORDER BY clientName""")
            result = cursor.fetchall()

            count = 0
            for i in result:
                if count % 2 == 0:
                    TREE_CLIENTS.insert("", "end", values = (i[0],i[1],i[2],i[3],i[4],i[5],i[6],i[7]), tags = "evenrow")
                else:
                    TREE_CLIENTS.insert("", "end", values = (i[0],i[1],i[2],i[3],i[4],i[5],i[6],i[7]), tags = "oddrow")
                count += 1
    
    def enterClientSelection(self, *args):
        self.copySelection(TREE_CLIENTS)
        TEXTVAR_CLIENTCODE.set(content[0])
        TOP_CLIENTSELECTION.destroy()
        try:
            self.populateSOAFields(TEXTVAR_CLIENTCODE.get())
            TOP_RECEIVABLES.lift()
            ENTRY_PARTICULARS.focus()
        except:
            pass
        
### MENU_ADMIN_SUPPLIERS ###
    def showAddEditSupplier(self, *args):
        global TOP_SUPPLIER
        TOP_SUPPLIER = Toplevel()
        TOP_SUPPLIER.title("Add - Supplier")
        TOP_SUPPLIER.iconbitmap(PATH_ICON + "icon.ico")
        TOP_SUPPLIER.geometry('350x325+550+100')
        TOP_SUPPLIER.resizable(height = False, width = False)
        TOP_SUPPLIER.grab_set()

        global required
        required = []

        LABEL_CODE = Label(TOP_SUPPLIER, text = "Code", font = APP_FONT)
        LABEL_CODE.grid(column = 0, row = 0, pady = TOP_PADY, sticky = E)

        LABEL_NAME = Label(TOP_SUPPLIER, text = "Name", font = APP_FONT)
        LABEL_NAME.grid(column = 0, row = 1, pady = TOP_PADY, sticky = E)

        LABEL_TIN = Label(TOP_SUPPLIER, text = "TIN", font = APP_FONT)
        LABEL_TIN.grid(column = 0, row = 2, pady = TOP_PADY, sticky = E)

        LABEL_ADDRESS = Label(TOP_SUPPLIER, text = "Address", font = APP_FONT)
        LABEL_ADDRESS.grid(column = 0, row = 3, pady = TOP_PADY, sticky = E)

        LABEL_CONTACT = Label(TOP_SUPPLIER, text = "Contact", font = APP_FONT)
        LABEL_CONTACT.grid(column = 0, row = 4, pady = TOP_PADY, sticky = E)
        
        LABEL_TAX = Label(TOP_SUPPLIER, text = "Tax", font = APP_FONT)
        LABEL_TAX.grid(column = 0, row = 5, pady = TOP_PADY, sticky = E)
        
        LABEL_CATEGORY = Label(TOP_SUPPLIER, text = "Category", font = APP_FONT)
        LABEL_CATEGORY.grid(column = 0, row = 6, pady = TOP_PADY, sticky = E)
        
        LABEL_ENTITY = Label(TOP_SUPPLIER, text = "Entity", font = APP_FONT)
        LABEL_ENTITY.grid(column = 0, row = 7, pady = TOP_PADY, sticky = E)

        global TEXTVAR_CODE, ENTRY_CODE
        TEXTVAR_CODE = StringVar()
        ENTRY_CODE = Entry(TOP_SUPPLIER, textvariable = TEXTVAR_CODE, font = APP_FONT, width = 35)
        ENTRY_CODE.grid(column = 1, row = 0, sticky = W)
        ENTRY_CODE.bind("<FocusOut>", lambda e: self.capitalLetters(TEXTVAR_CODE))
        required.append(TEXTVAR_CODE)

        global TEXTVAR_NAME
        TEXTVAR_NAME = StringVar()
        ENTRY_NAME = Entry(TOP_SUPPLIER, textvariable = TEXTVAR_NAME, font = APP_FONT, width = 35)
        ENTRY_NAME.grid(column = 1, row = 1, sticky = W)
        required.append(TEXTVAR_NAME)

        global TEXTVAR_TIN
        TEXTVAR_TIN = StringVar()
        ENTRY_TIN = Entry(TOP_SUPPLIER, textvariable = TEXTVAR_TIN, font = APP_FONT, width = 35)
        ENTRY_TIN.grid(column = 1, row = 2, sticky = W)
        required.append(TEXTVAR_TIN)

        global TEXTVAR_ADDRESS
        TEXTVAR_ADDRESS = StringVar()
        ENTRY_ADDRESS = Entry(TOP_SUPPLIER, textvariable = TEXTVAR_ADDRESS, font = APP_FONT, width = 35)
        ENTRY_ADDRESS.grid(column = 1, row = 3, sticky = W)
        required.append(TEXTVAR_ADDRESS)

        global TEXTVAR_CONTACT
        TEXTVAR_CONTACT = StringVar()
        ENTRY_CONTACT = Entry(TOP_SUPPLIER, textvariable = TEXTVAR_CONTACT, font = APP_FONT, width = 35)
        ENTRY_CONTACT.grid(column = 1, row = 4, sticky = W)
        
        global TEXTVAR_TAX, ENTRY_TAX
        TEXTVAR_TAX = StringVar()
        ENTRY_TAX = tk.Combobox(TOP_SUPPLIER, values = ["VAT", "Non-VAT"], textvariable = TEXTVAR_TAX, font = APP_FONT, width = 10, state = "readonly")
        ENTRY_TAX.grid(column = 1, row = 5, sticky = W)
        required.append(TEXTVAR_TAX)
        
        global TEXTVAR_CATEGORY
        TEXTVAR_CATEGORY = StringVar()
        ENTRY_CATEGORY = tk.Combobox(TOP_SUPPLIER, values = ["Supplier", "Others"], textvariable = TEXTVAR_CATEGORY, font = APP_FONT, width = 15, state = "readonly")
        ENTRY_CATEGORY.grid(column = 1, row = 6, sticky = W)
        required.append(TEXTVAR_CATEGORY)
        
        global TEXTVAR_ENTITY
        TEXTVAR_ENTITY = StringVar()
        ENTRY_ENTITY = tk.Combobox(TOP_SUPPLIER, values = ["Corporation", "Individual"], textvariable = TEXTVAR_ENTITY, font = APP_FONT, width = 15, state = "readonly")
        ENTRY_ENTITY.grid(column = 1, row = 7, sticky = W)
        required.append(TEXTVAR_ENTITY)

        FRAME_BUTTON = Frame(TOP_SUPPLIER)
        FRAME_BUTTON.grid(column = 1, row = 8, sticky = E, pady = TOP_PADY + 10)

        BUTTON_SAVE = Button(FRAME_BUTTON, text = "SAVE", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "hand2", command = self.saveSupplier)
        BUTTON_SAVE.grid(column = 0, row = 0, padx = TOP_PADX)

        BUTTON_CLOSE = Button(FRAME_BUTTON, text = "CLOSE", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "hand2", command = lambda: self.closeTopLevel(TOP_SUPPLIER))
        BUTTON_CLOSE.grid(column = 1, row = 0, padx = TOP_PADX)

    def saveSupplier(self, *args):
        wrong = []
        for i in required:
            if i.get() == "":
                wrong.append(i)
                break
            
        if len(wrong) > 0:
            messagebox.showerror("Save Supplier", f"Please provide {str(wrong[0])}!")
        else:
            insert = """INSERT INTO tblsuppliers (
                        code, name, tin, address, contact, encoded, vatable, category, entity)
                        values (%s, %s, %s, %s, %s, %s, %s, %s, %s)"""
            update = """UPDATE tblsuppliers SET name = %s, tin = %s, address = %s, contact = %s, encoded = %s, vatable = %s, category = %s, entity = %s
                        WHERE code = %s"""
            find = "SELECT code FROM tblsuppliers WHERE code = %s"
            db.commit()
            cursor.execute(find, [TEXTVAR_CODE.get()])
            result = cursor.fetchall()
            if result:
                cursor.execute(update, [TEXTVAR_NAME.get(), TEXTVAR_TIN.get(), TEXTVAR_ADDRESS.get(), TEXTVAR_CONTACT.get(), datetime.datetime.now(), TEXTVAR_TAX.get(), TEXTVAR_CATEGORY.get(), TEXTVAR_ENTITY.get(), TEXTVAR_CODE.get()])
                saved = "Supplier has been updated!"
            else:
                cursor.execute(insert, [TEXTVAR_CODE.get(), TEXTVAR_NAME.get(), TEXTVAR_TIN.get(), TEXTVAR_ADDRESS.get(), TEXTVAR_CONTACT.get(), datetime.datetime.now(), TEXTVAR_TAX.get(), TEXTVAR_CATEGORY.get(), TEXTVAR_ENTITY.get()])
                saved = "A new supplier has been saved!"
            ask = messagebox.askyesno("Save Supplier", "Are you sure?")
            if ask == True:
                db.commit()
                messagebox.showinfo("Save Supplier", saved)
                TOP_SUPPLIER.grab_release()
                TOP_SUPPLIER.destroy()
                self.showSuppliers()

    def editSupplier(self, *args):
        self.copySelection(TREE_SUPPLIER)
        self.showAddEditSupplier()
        TOP_SUPPLIER.title("Edit - Supplier")
        TEXTVAR_CODE.set(content[0])
        ENTRY_CODE.config(state = DISABLED)
        TEXTVAR_NAME.set(content[1])
        TEXTVAR_TIN.set(content[2])
        TEXTVAR_ADDRESS.set(content[3])
        TEXTVAR_CONTACT.set(content[4])
        TEXTVAR_TAX.set(content[5])
        ENTRY_TAX.config(state = DISABLED)
        TEXTVAR_CATEGORY.set(content[6])
        TEXTVAR_ENTITY.set(content[7])
        
    def deleteSupplier(self, *args):
        self.copySelection(TREE_SUPPLIER)
        ask = messagebox.askyesno("Delete Supplier", "Are you sure?")
        if ask:
            cursor.execute(f"DELETE FROM tblsuppliers WHERE code = '{content[0]}'")
            db.commit()
            messagebox.showinfo("Delete Supplier", "Supplier has been deleted!")
            self.showSuppliers()

    def showSupplierTreeview(self, frame):
        global TREE_SUPPLIER
        TREE_SUPPLIER = tk.Treeview(frame, height = 28, selectmode = "browse")
        TREE_SUPPLIER["columns"] = ("Code", "Name", "TIN", "Address", "Contact", "Tax")
        TREE_SUPPLIER.column("#0", width = 0, minwidth = 0, stretch = True)
        TREE_SUPPLIER.column("Code", anchor = W, minwidth = 100, width = 150)
        TREE_SUPPLIER.column("Name", anchor = W, minwidth = 100, width = 250)
        TREE_SUPPLIER.column("TIN", anchor = W, minwidth = 100, width = 150)
        TREE_SUPPLIER.column("Address", anchor = W, minwidth = 100, width = 200)
        TREE_SUPPLIER.column("Contact", anchor = W, minwidth = 100, width = 150)
        TREE_SUPPLIER.column("Tax", anchor = W, minwidth = 100, width = 50)
        
        TREE_SUPPLIER.heading("#0", text = "", anchor = W)
        TREE_SUPPLIER.heading("Code", text = "Code", anchor = N)
        TREE_SUPPLIER.heading("Name", text = "Name", anchor = N)
        TREE_SUPPLIER.heading("TIN", text = "TIN", anchor = N)
        TREE_SUPPLIER.heading("Address", text = "Address", anchor = N)
        TREE_SUPPLIER.heading("Contact", text = "Contact", anchor = N)
        TREE_SUPPLIER.heading("Tax", text = "Tax", anchor = N)
        
        global POPUP_SUPPLIER
        POPUP_SUPPLIER = Menu(TREE_SUPPLIER, tearoff = 0)
        POPUP_SUPPLIER.add_command(command = self.editSupplier, label = "Edit")
        POPUP_SUPPLIER.add_command(command = self.deleteSupplier, label = "Delete")
        TREE_SUPPLIER.bind("<Button-3>", lambda e: self.popupMenu(TREE_SUPPLIER, POPUP_SUPPLIER, e))

        global STYLE_SUPPLIERS
        STYLE_SUPPLIERS = tk.Style()
        STYLE_SUPPLIERS.map("Treeview", foreground = self.fixedMap("foreground", STYLE_SUPPLIERS), background = self.fixedMap("background", STYLE_SUPPLIERS))

        TREE_SUPPLIER.tag_configure("oddrow", background = None)
        TREE_SUPPLIER.tag_configure("evenrow", background = TREE_TAG_EVENROW)
        db.commit()
        cursor.execute("""SELECT code, name, tin, address, contact, vatable FROM tblsuppliers WHERE category = 'Supplier' ORDER BY name""")
        result = cursor.fetchall()
        count = 0
        for i in result:
            if count % 2 == 0:
                TREE_SUPPLIER.insert("", "end", values = (i[0],i[1],i[2],i[3],i[4],i[5]), tags = "evenrow")
            else:
                TREE_SUPPLIER.insert("", "end", values = (i[0],i[1],i[2],i[3],i[4],i[5]), tags = "oddrow")
            count += 1

        YSCROLLBAR = tk.Scrollbar(frame, orient = "vertical", command = TREE_SUPPLIER.yview)
        YSCROLLBAR.pack(side = "right", fill = "y")

        XSCROLLBAR = tk.Scrollbar(frame, orient = "horizontal", command = TREE_SUPPLIER.xview)
        
        TREE_SUPPLIER.configure(yscrollcommand = YSCROLLBAR.set, xscrollcommand = XSCROLLBAR.set) 
        TREE_SUPPLIER.pack()
        XSCROLLBAR.pack(fill ="x")

    def showSupplierSelection(self, *args):
        global TOP_SUPPLIERSELECTION
        TOP_SUPPLIERSELECTION = Toplevel()
        TOP_SUPPLIERSELECTION.title("Choose Supplier Code")
        TOP_SUPPLIERSELECTION.iconbitmap(PATH_ICON + "icon.ico")
        TOP_SUPPLIERSELECTION.geometry("600x300+100+100")
        TOP_SUPPLIERSELECTION.resizable(height = False, width = False)
        TOP_SUPPLIERSELECTION.grab_set()

        FRAME1 = Frame(TOP_SUPPLIERSELECTION)
        FRAME1.pack(fill = "x")

        FRAME2 = Frame(TOP_SUPPLIERSELECTION)
        FRAME2.pack(fill = "x")

        LABEL_SEARCH = Label(FRAME1, text = "Search", font = APP_FONT)
        LABEL_SEARCH.grid(column = 0, row = 0, padx = TOP_PADX)
        
        global TEXTVAR_SEARCH
        TEXTVAR_SEARCH = StringVar()
        ENTRY_SEARCH = Entry(FRAME1, textvariable = TEXTVAR_SEARCH, font = APP_FONT, width = 20)
        ENTRY_SEARCH.grid(column = 1, row = 0, padx = TOP_PADX)
        ENTRY_SEARCH.bind("<Return>", lambda e: self.searchSupplier(TEXTVAR_SEARCH.get()))
        
        BUTTON_SEARCH = Button(FRAME1, text = "GO", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, command = lambda: self.searchSupplier(TEXTVAR_SEARCH.get()))
        BUTTON_SEARCH.grid(column = 2, row = 0, sticky = W, padx = TOP_PADX)
        
        global TEXTVAR_CATEGORY
        TEXTVAR_CATEGORY = StringVar()
        COMBO_CATEGORY = tk.Combobox(FRAME1, values = ["Supplier", "Others"], textvariable = TEXTVAR_CATEGORY, font = APP_FONT, width = 15, state = "readonly")
        COMBO_CATEGORY.grid(column = 3, row = 0, sticky = E, padx = TOP_PADX)
        COMBO_CATEGORY.bind("<<ComboboxSelected>>", lambda e: self.searchSupplier(TEXTVAR_SEARCH.get()))
        TEXTVAR_CATEGORY.set("Supplier")

        self.showSupplierTreeview(FRAME2)
        TREE_SUPPLIER.config(height = 10)
        TREE_SUPPLIER.unbind("<Button-3>")
        TREE_SUPPLIER.bind("<Double-1>", self.enterSupplierSelection)
        try:
            if TOP_PAYABLE.winfo_exists() == 1:
                TREE_SUPPLIER.bind("<Double-1>", self.enterAPVSupplierSelection)
            else:
                if TOP_DISBURSEMENTS.winfo_exists() == 1:
                    TREE_SUPPLIER.bind("<Double-1>", self.enterAPVSupplierSelection)
        except:
            try:
                if TOP_DISBURSEMENTS.winfo_exists() == 1:
                    TREE_SUPPLIER.bind("<Double-1>", self.enterAPVSupplierSelection)
            except:
                pass
        
    def enterSupplierSelection(self, *args):
        self.copySelection(TREE_SUPPLIER)
        TEXTVAR_CODE.set(content[0])
        TOP_SUPPLIERSELECTION.destroy()
        try:
            self.populatePOFields(TEXTVAR_CODE)
        except:
            pass
    
    def enterAPVSupplierSelection(self, *args):
        self.copySelection(TREE_SUPPLIER)
        TEXTVAR_PAYEECODE.set(content[0])
        TEXTVAR_PAYEENAME.set(content[1])
        TOP_SUPPLIERSELECTION.destroy()
    
    def searchSupplier(self, var, *args):
        find = "SELECT code, name, tin, address, contact, vatable, category, entity FROM tblsuppliers WHERE (code LIKE %s OR name LIKE %s) AND category = %s ORDER BY name"
        db.commit()
        cursor.execute(find, [f"%{var}%", f"%{var}%", TEXTVAR_CATEGORY.get()])
        result = cursor.fetchall()
        for i in TREE_SUPPLIER.get_children():
            TREE_SUPPLIER.delete(i)
        if result:
            count = 0
            for i in result:
                if count % 2 == 0:
                    TREE_SUPPLIER.insert("", "end", values = (i[0],i[1],i[2],i[3],i[4],i[5],i[6],i[7]), tags = "evenrow")
                else:
                    TREE_SUPPLIER.insert("", "end", values = (i[0],i[1],i[2],i[3],i[4],i[5],i[6],i[7]), tags = "oddrow")
                count += 1
        else:
            messagebox.showerror("Supplier", "No match found!")
     
### MENU_ADMIN_TAXES ###
    def showAddEditTax(self, *args):
        TOP_TAX = Toplevel()
        TOP_TAX.title("Add - Tax")
        TOP_TAX.iconbitmap(PATH_ICON + "icon.ico")
        TOP_TAX.geometry('350x325+550+100')
        TOP_TAX.resizable(height = False, width = False)
        TOP_TAX.grab_set()

        global required
        required = []

        LABEL_CODE = Label(TOP_TAX, text = "Code", font = APP_FONT)
        LABEL_CODE.grid(column = 0, row = 0, pady = TOP_PADY, sticky = E)

        LABEL_RATE = Label(TOP_TAX, text = "Rate (%)", font = APP_FONT)
        LABEL_RATE.grid(column = 0, row = 1, pady = TOP_PADY, sticky = E)

        LABEL_DESC = Label(TOP_TAX, text = "Description", font = APP_FONT)
        LABEL_DESC.grid(column = 0, row = 2, pady = TOP_PADY, sticky = E)

        global TEXTVAR_CODE
        TEXTVAR_CODE = StringVar()
        ENTRY_CODE = Entry(TOP_TAX, textvariable = TEXTVAR_CODE, font = APP_FONT, width = 35, state = "readonly")
        ENTRY_CODE.grid(column = 1, row = 0, sticky = W)
        required.append(TEXTVAR_CODE)

        global TEXTVAR_RATE
        TEXTVAR_RATE = StringVar()
        ENTRY_RATE = Entry(TOP_TAX, textvariable = TEXTVAR_RATE, font = APP_FONT, width = 35)
        ENTRY_RATE.grid(column = 1, row = 1, sticky = W)
        required.append(TEXTVAR_RATE)

        global TEXTVAR_DESC
        TEXTVAR_DESC = StringVar()
        ENTRY_DESC = Entry(TOP_TAX, textvariable = TEXTVAR_DESC, font = APP_FONT, width = 35)
        ENTRY_DESC.grid(column = 1, row = 2, sticky = W)
        required.append(TEXTVAR_DESC)

        FRAME_BUTTON = Frame(TOP_TAX)
        FRAME_BUTTON.grid(column = 1, row = 3, sticky = E, pady = TOP_PADY + 10)

        BUTTON_SAVE = Button(FRAME_BUTTON, text = "SAVE", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "hand2", command = None)
        BUTTON_SAVE.grid(column = 0, row = 0, padx = TOP_PADX)

        BUTTON_CLOSE = Button(FRAME_BUTTON, text = "CLOSE", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "hand2", command = lambda: self.closeTopLevel(TOP_TAX))
        BUTTON_CLOSE.grid(column = 1, row = 0, padx = TOP_PADX)

    def saveTax(self, *args):
        pass

    def editTax(self, *args):
        pass

    def deleteTax(self, *args):
        pass

    def searchTaxCode(self, var, *args):
        find = "SELECT code, rate, description FROM tbltaxes WHERE code LIKE %s OR description LIKE %s ORDER BY code"
        db.commit()
        cursor.execute(find, [f"%{var}%", f"%{var}%"])
        result = cursor.fetchall()
        for i in TREE_TAX.get_children():
            TREE_TAX.delete(i)
        if result:
            count = 0
            for i in result:
                if count % 2 == 0:
                    TREE_TAX.insert("", "end", values = (i[0],i[1],i[2]), tags = "evenrow")
                else:
                    TREE_TAX.insert("", "end", values = (i[0],i[1],i[2]), tags = "oddrow")
                count += 1
        else:
            messagebox.showerror("Tax Code", "No match found!")

### MENU_ADMIN_TYPES ###
    def showAddEditType(self, *args):
        TOP_TYPE = Toplevel()
        TOP_TYPE.title("Add - Transaction Type")
        TOP_TYPE.iconbitmap(PATH_ICON + "icon.ico")
        TOP_TYPE.geometry('350x325+550+100')
        TOP_TYPE.resizable(height = False, width = False)
        TOP_TYPE.grab_set()

        global required
        required = []

        LABEL_CODE = Label(TOP_TYPE, text = "Code", font = APP_FONT)
        LABEL_CODE.grid(column = 0, row = 0, pady = TOP_PADY, sticky = E)

        LABEL_TYPE = Label(TOP_TYPE, text = "Type", font = APP_FONT)
        LABEL_TYPE.grid(column = 0, row = 1, pady = TOP_PADY, sticky = E)

        LABEL_DESC = Label(TOP_TYPE, text = "Description", font = APP_FONT)
        LABEL_DESC.grid(column = 0, row = 2, pady = TOP_PADY, sticky = E)

        global TEXTVAR_CODE
        TEXTVAR_CODE = StringVar()
        ENTRY_CODE = Entry(TOP_TYPE, textvariable = TEXTVAR_CODE, font = APP_FONT, width = 35, state = "readonly")
        ENTRY_CODE.grid(column = 1, row = 0, sticky = W)
        required.append(TEXTVAR_CODE)

        global TEXTVAR_TYPE
        TEXTVAR_TYPE = StringVar()
        ENTRY_TYPE = Entry(TOP_TYPE, textvariable = TEXTVAR_TYPE, font = APP_FONT, width = 35)
        ENTRY_TYPE.grid(column = 1, row = 1, sticky = W)
        required.append(TEXTVAR_TYPE)

        global TEXTVAR_DESC
        TEXTVAR_DESC = StringVar()
        ENTRY_DESC = Entry(TOP_TYPE, textvariable = TEXTVAR_DESC, font = APP_FONT, width = 35)
        ENTRY_DESC.grid(column = 1, row = 2, sticky = W)
        required.append(TEXTVAR_DESC)

        FRAME_BUTTON = Frame(TOP_TYPE)
        FRAME_BUTTON.grid(column = 1, row = 3, sticky = E, pady = TOP_PADY + 10)

        BUTTON_SAVE = Button(FRAME_BUTTON, text = "SAVE", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "hand2", command = None)
        BUTTON_SAVE.grid(column = 0, row = 0, padx = TOP_PADX)

        BUTTON_CLOSE = Button(FRAME_BUTTON, text = "CLOSE", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "hand2", command = lambda: self.closeTopLevel(TOP_TYPE))
        BUTTON_CLOSE.grid(column = 1, row = 0, padx = TOP_PADX)

    def saveType(self, *args):
        pass

    def editType(self, *args):
        pass

    def deleteType(self, *args):
        pass

    def searchTransactionType(self, var, *args):
        find = "SELECT code, type, description FROM tbltypes WHERE code LIKE %s OR type LIKE %s OR description LIKE %s ORDER BY type"
        db.commit()
        cursor.execute(find, [f"%{var}%", f"%{var}%", f"%{var}%"])
        result = cursor.fetchall()
        for i in TREE_TYPES.get_children():
            TREE_TYPES.delete(i)
        if result:
            count = 0
            for i in result:
                if count % 2 == 0:
                    TREE_TYPES.insert("", "end", values = (i[0],i[1],i[2]), tags = "evenrow")
                else:
                    TREE_TYPES.insert("", "end", values = (i[0],i[1],i[2]), tags = "oddrow")
                count += 1
        else:
            messagebox.showerror("Transaction Type", "No match found!")

### MENU_ADMIN_USERS ###
    def showAddEditUser(self, *args):
        global TOP_USER
        TOP_USER = Toplevel()
        TOP_USER.title("Add - User")
        TOP_USER.iconbitmap(PATH_ICON + "icon.ico")
        TOP_USER.geometry('350x325+550+100')
        TOP_USER.resizable(height = False, width = False)
        TOP_USER.grab_set()

        global required
        required = []

        LABEL_ID = Label(TOP_USER, text = "ID", font = APP_FONT)
        LABEL_ID.grid(column = 0, row = 0, pady = TOP_PADY, sticky = E)

        LABEL_USERNAME = Label(TOP_USER, text = "Username", font = APP_FONT)
        LABEL_USERNAME.grid(column = 0, row = 1, pady = TOP_PADY, sticky = E)
        
        LABEL_PASSWORD = Label(TOP_USER, text = "Password", font = APP_FONT)
        LABEL_PASSWORD.grid(column = 0, row = 2, pady = TOP_PADY, sticky = E)
        
        LABEL_CONFIRM = Label(TOP_USER, text = "Confirm PW", font = APP_FONT)
        LABEL_CONFIRM.grid(column = 0, row = 3, pady = TOP_PADY, sticky = E)

        LABEL_FIRST = Label(TOP_USER, text = "First Name", font = APP_FONT)
        LABEL_FIRST.grid(column = 0, row = 4, pady = TOP_PADY, sticky = E)

        LABEL_LAST = Label(TOP_USER, text = "Last Name", font = APP_FONT)
        LABEL_LAST.grid(column = 0, row = 5, pady = TOP_PADY, sticky = E)

        LABEL_TYPE = Label(TOP_USER, text = "User Type", font = APP_FONT)
        LABEL_TYPE.grid(column = 0, row = 6, pady = TOP_PADY, sticky = E)

        LABEL_DEPT = Label(TOP_USER, text = "Department", font = APP_FONT)
        LABEL_DEPT.grid(column = 0, row = 7, pady = TOP_PADY, sticky = E)

        global TEXTVAR_ID, ENTRY_ID
        TEXTVAR_ID = StringVar()
        ENTRY_ID = Entry(TOP_USER, textvariable = TEXTVAR_ID, font = APP_FONT, width = 35)
        ENTRY_ID.grid(column = 1, row = 0, sticky = W)
        required.append(TEXTVAR_ID)

        global TEXTVAR_USERNAME
        TEXTVAR_USERNAME = StringVar()
        ENTRY_USERNAME = Entry(TOP_USER, textvariable = TEXTVAR_USERNAME, font = APP_FONT, width = 35)
        ENTRY_USERNAME.grid(column = 1, row = 1, sticky = W)
        required.append(TEXTVAR_USERNAME)
        
        global TEXTVAR_PASSWORD, ENTRY_PASSWORD
        TEXTVAR_PASSWORD = StringVar()
        ENTRY_PASSWORD = Entry(TOP_USER, textvariable = TEXTVAR_PASSWORD, font = APP_FONT, width = 35, show = "●")
        ENTRY_PASSWORD.grid(column = 1, row = 2, sticky = W)
        required.append(TEXTVAR_PASSWORD)
        
        global TEXTVAR_CONFIRM, ENTRY_CONFIRM
        TEXTVAR_CONFIRM = StringVar()
        ENTRY_CONFIRM = Entry(TOP_USER, textvariable = TEXTVAR_CONFIRM, font = APP_FONT, width = 35, show = "●")
        ENTRY_CONFIRM.grid(column = 1, row = 3, sticky = W)
        required.append(TEXTVAR_CONFIRM)

        global TEXTVAR_FIRST
        TEXTVAR_FIRST = StringVar()
        ENTRY_FIRST = Entry(TOP_USER, textvariable = TEXTVAR_FIRST, font = APP_FONT, width = 35)
        ENTRY_FIRST.grid(column = 1, row = 4, sticky = W)
        required.append(TEXTVAR_FIRST)

        global TEXTVAR_LAST
        TEXTVAR_LAST = StringVar()
        ENTRY_LAST = Entry(TOP_USER, textvariable = TEXTVAR_LAST, font = APP_FONT, width = 35)
        ENTRY_LAST.grid(column = 1, row = 5, sticky = W)
        required.append(TEXTVAR_LAST)

        global TEXTVAR_TYPE, COMBO_TYPE
        TEXTVAR_TYPE = StringVar()
        COMBO_TYPE = tk.Combobox(TOP_USER, values = ["regular", "super", "administrator"], textvariable = TEXTVAR_TYPE, font = APP_FONT, width = 33, state = "readonly")
        COMBO_TYPE.grid(column = 1, row = 6, sticky = W)
        required.append(TEXTVAR_TYPE)

        global TEXTVAR_DEPT, COMBO_DEPT
        TEXTVAR_DEPT = StringVar()
        COMBO_DEPT = tk.Combobox(TOP_USER, values = self.listCenters(), textvariable = TEXTVAR_DEPT, font = APP_FONT, width = 33, state = "readonly")  
        COMBO_DEPT.grid(column = 1, row = 7, sticky = W)
        required.append(TEXTVAR_DEPT)

        FRAME_BUTTON = Frame(TOP_USER)
        FRAME_BUTTON.grid(column = 1, row = 8, sticky = E, pady = TOP_PADY + 10)

        BUTTON_SAVE = Button(FRAME_BUTTON, text = "SAVE", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "hand2", command = self.saveUser)
        BUTTON_SAVE.grid(column = 0, row = 0, padx = TOP_PADX)

        BUTTON_CLOSE = Button(FRAME_BUTTON, text = "CLOSE", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "hand2", command = lambda: self.closeTopLevel(TOP_USER))
        BUTTON_CLOSE.grid(column = 1, row = 0, padx = TOP_PADX)

    def saveUser(self, *args):
        wrong = []
        for i in required:
            if i.get() == "":
                wrong.append(i)
                break
            
        if TEXTVAR_PASSWORD.get() != TEXTVAR_CONFIRM.get():
            wrong.append("matching password")
        db.commit()
        cursor.execute(f"SELECT id FROM tblusers WHERE id = '{TEXTVAR_ID.get()}'")
        result = cursor.fetchall()
        if result:
            message = "Update User"
        else:
            message = "Save User"
            
        if len(wrong) > 0:
            messagebox.showerror(message, f"Please provide {str(wrong[0])}!")
        else:
            insert = """INSERT INTO tblusers (
                        id, username, password, firstname, lastname, usertype, department)
                        values (%s, %s, %s, %s, %s, %s, %s)"""
            update = """UPDATE tblusers SET id = %s, username = %s, firstname = %s, lastname = %s, usertype = %s, department = %s
                        WHERE id = %s"""
            insertaccess = """INSERT INTO tblaccess (
                ID, poAdd, poApprove, poVoid, rrAdd,
                rrApprove, rrVoid, apvAdd, apvApprove, apvVoid,
                dvAdd, dvApprove, dvVoid, soaAdd, soaApprove,
                soaVoid, orAdd, orApprove, orVoid, gjAdd, gjApprove, gjVoid) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
            find = "SELECT id FROM tblusers WHERE id = %s"
            try:
                db.commit()
                cursor.execute(find, [content[0]])
                result = cursor.fetchall()
                if result:
                    cursor.execute(update, [TEXTVAR_ID.get(), TEXTVAR_USERNAME.get(), TEXTVAR_FIRST.get(), TEXTVAR_LAST.get(), TEXTVAR_TYPE.get(), TEXTVAR_DEPT.get(), content[0]])
                    saved = "User has been updated!"
                else:
                    cursor.execute(insert, [TEXTVAR_ID.get(), TEXTVAR_USERNAME.get(), self.hashPassword(TEXTVAR_PASSWORD.get()), TEXTVAR_FIRST.get(), TEXTVAR_LAST.get(), TEXTVAR_TYPE.get(), TEXTVAR_DEPT.get()])
                    saved = "A new user has been saved!"
            except:
                cursor.execute(insert, [TEXTVAR_ID.get(), TEXTVAR_USERNAME.get(), self.hashPassword(TEXTVAR_PASSWORD.get()), TEXTVAR_FIRST.get(), TEXTVAR_LAST.get(), TEXTVAR_TYPE.get(), TEXTVAR_DEPT.get()])
                saved = "A new user has been saved!"
                
            ask = messagebox.askyesno(message, "Are you sure?")
            if ask == True:
                db.commit()
                cursor.execute(insertaccess, [TEXTVAR_ID.get(), 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])
                db.commit()
                messagebox.showinfo(message, saved)
                TOP_USER.grab_release()
                TOP_USER.destroy()
                self.showUsers()

    def editUser(self, *args):
        self.copySelection(TREE_USERS)
        self.showAddEditUser()
        TOP_USER.title("Edit - User")
        
        cursor.execute(f"SELECT password FROM tblusers WHERE id = '{content[0]}' LIMIT 1")
        result = cursor.fetchone()
        if result:
            TEXTVAR_ID.set(content[0])
            TEXTVAR_USERNAME.set(content[1])
            TEXTVAR_PASSWORD.set(result[0])
            TEXTVAR_CONFIRM.set(result[0])
            TEXTVAR_FIRST.set(content[2])
            TEXTVAR_LAST.set(content[3])
            TEXTVAR_TYPE.set(content[4])
            TEXTVAR_DEPT.set(content[5])
            
            ENTRY_ID.config(state = DISABLED)
            ENTRY_PASSWORD.config(state = DISABLED)
            ENTRY_CONFIRM.config(state = DISABLED)

    def deleteUser(self, *args):
        delete = "DELETE FROM tblusers WHERE id = %s"
        self.copySelection(TREE_USERS)
        ask = messagebox.askyesno("Delete User", "Are you sure?")
        if ask == True:
            cursor.execute(delete, [content[0]])
            db.commit()
            messagebox.showinfo("Delete User", "User has been deleted!")
            self.showUsers()

    def hashPassword(self, pw):
        hashed = bcrypt.hashpw(pw.encode("utf-8"), bcrypt.gensalt())
        return hashed

    def searchUser(self, var, *args):
        find = "SELECT id, username, firstname, lastname, usertype, department FROM tblusers WHERE id LIKE %s OR firstname LIKE %s OR lastname LIKE %s ORDER BY id"
        db.commit()
        cursor.execute(find, [f"%{var}%", f"%{var}%", f"%{var}%"])
        result = cursor.fetchall()
        for i in TREE_USERS.get_children():
            TREE_USERS.delete(i)
        if result:
            count = 0
            for i in result:
                if count % 2 == 0:
                    TREE_USERS.insert("", "end", values = (i[0],i[1],i[2],i[3],i[4],i[5]), tags = "evenrow")
                else:
                    TREE_USERS.insert("", "end", values = (i[0],i[1],i[2],i[3],i[4],i[5]), tags = "oddrow")
                count += 1
        else:
            messagebox.showerror("User", "No match found!")

### MENU_ADMIN_ACCESS ###
    def editAccess(self, *args):
        self.copySelection(TREE_USERS)
        global TOP_ACCESS
        TOP_ACCESS = Toplevel()
        TOP_ACCESS.title("Edit - Access of " + self.returnUserName(content[0], 0) + content[0])
        TOP_ACCESS.iconbitmap(PATH_ICON + "icon.ico")
        TOP_ACCESS.geometry('350x325+550+100')
        TOP_ACCESS.resizable(height = False, width = False)
        TOP_ACCESS.grab_set()

        LABEL_ADD = Label(TOP_ACCESS, text = "ADD", font = APP_FONT)
        LABEL_ADD.grid(column = 1, row = 0, pady = TOP_PADY, sticky = E)
        
        LABEL_APPROVE = Label(TOP_ACCESS, text = "APPROVE", font = APP_FONT)
        LABEL_APPROVE.grid(column = 2, row = 0, pady = TOP_PADY, sticky = E)
        
        LABEL_VOID = Label(TOP_ACCESS, text = "VOID", font = APP_FONT)
        LABEL_VOID.grid(column = 3, row = 0, pady = TOP_PADY, sticky = E)
        
        LABEL_PO = Label(TOP_ACCESS, text = "PO", font = APP_FONT)
        LABEL_PO.grid(column = 0, row = 1, pady = TOP_PADY, sticky = E)
        
        LABEL_RR = Label(TOP_ACCESS, text = "RR", font = APP_FONT)
        LABEL_RR.grid(column = 0, row = 2, pady = TOP_PADY, sticky = E)
        
        LABEL_APV = Label(TOP_ACCESS, text = "APV", font = APP_FONT)
        LABEL_APV.grid(column = 0, row = 3, pady = TOP_PADY, sticky = E)
        
        LABEL_DV = Label(TOP_ACCESS, text = "DV", font = APP_FONT)
        LABEL_DV.grid(column = 0, row = 4, pady = TOP_PADY, sticky = E)
        
        LABEL_DV = Label(TOP_ACCESS, text = "SOA", font = APP_FONT)
        LABEL_DV.grid(column = 0, row = 5, pady = TOP_PADY, sticky = E)
        
        LABEL_DV = Label(TOP_ACCESS, text = "OR", font = APP_FONT)
        LABEL_DV.grid(column = 0, row = 6, pady = TOP_PADY, sticky = E)
        
        LABEL_GJ = Label(TOP_ACCESS, text = "GJ", font = APP_FONT)
        LABEL_GJ.grid(column = 0, row = 7, pady = TOP_PADY, sticky = E)
        
        global CHECKVAR_POADD
        CHECKVAR_POADD = IntVar()
        CHECKVAR_ACCESS = Checkbutton(TOP_ACCESS, variable = CHECKVAR_POADD, font = BUTTON_FONT, bg = BUTTON_BG)
        CHECKVAR_ACCESS.grid(column = 1, row = 1)
        global CHECKVAR_POAPPROVE
        CHECKVAR_POAPPROVE = IntVar()
        CHECKVAR_ACCESS = Checkbutton(TOP_ACCESS, variable = CHECKVAR_POAPPROVE, font = BUTTON_FONT, bg = BUTTON_BG)
        CHECKVAR_ACCESS.grid(column = 2, row = 1)
        global CHECKVAR_POVOID
        CHECKVAR_POVOID = IntVar()
        CHECKVAR_ACCESS = Checkbutton(TOP_ACCESS, variable = CHECKVAR_POVOID, font = BUTTON_FONT, bg = BUTTON_BG)
        CHECKVAR_ACCESS.grid(column = 3, row = 1)
        global CHECKVAR_RRADD
        CHECKVAR_RRADD = IntVar()
        CHECKVAR_ACCESS = Checkbutton(TOP_ACCESS, variable = CHECKVAR_RRADD, font = BUTTON_FONT, bg = BUTTON_BG)
        CHECKVAR_ACCESS.grid(column = 1, row = 2)
        global CHECKVAR_RRAPPROVE
        CHECKVAR_RRAPPROVE = IntVar()
        CHECKVAR_ACCESS = Checkbutton(TOP_ACCESS, variable = CHECKVAR_RRAPPROVE, font = BUTTON_FONT, bg = BUTTON_BG)
        CHECKVAR_ACCESS.grid(column = 2, row = 2)
        global CHECKVAR_RRVOID
        CHECKVAR_RRVOID = IntVar()
        CHECKVAR_ACCESS = Checkbutton(TOP_ACCESS, variable = CHECKVAR_RRVOID, font = BUTTON_FONT, bg = BUTTON_BG)
        CHECKVAR_ACCESS.grid(column = 3, row = 2)
        global CHECKVAR_APVADD
        CHECKVAR_APVADD = IntVar()
        CHECKVAR_ACCESS = Checkbutton(TOP_ACCESS, variable = CHECKVAR_APVADD, font = BUTTON_FONT, bg = BUTTON_BG)
        CHECKVAR_ACCESS.grid(column = 1, row = 3)
        global CHECKVAR_APVAPPROVE
        CHECKVAR_APVAPPROVE = IntVar()
        CHECKVAR_ACCESS = Checkbutton(TOP_ACCESS, variable = CHECKVAR_APVAPPROVE, font = BUTTON_FONT, bg = BUTTON_BG)
        CHECKVAR_ACCESS.grid(column = 2, row = 3)
        global CHECKVAR_APVVOID
        CHECKVAR_APVVOID = IntVar()
        CHECKVAR_ACCESS = Checkbutton(TOP_ACCESS, variable = CHECKVAR_APVVOID, font = BUTTON_FONT, bg = BUTTON_BG)
        CHECKVAR_ACCESS.grid(column = 3, row = 3)
        global CHECKVAR_DVADD
        CHECKVAR_DVADD = IntVar()
        CHECKVAR_ACCESS = Checkbutton(TOP_ACCESS, variable = CHECKVAR_DVADD, font = BUTTON_FONT, bg = BUTTON_BG)
        CHECKVAR_ACCESS.grid(column = 1, row = 4)
        global CHECKVAR_DVAPPROVE
        CHECKVAR_DVAPPROVE = IntVar()
        CHECKVAR_ACCESS = Checkbutton(TOP_ACCESS, variable = CHECKVAR_DVAPPROVE, font = BUTTON_FONT, bg = BUTTON_BG)
        CHECKVAR_ACCESS.grid(column = 2, row = 4)
        global CHECKVAR_DVVOID
        CHECKVAR_DVVOID = IntVar()
        CHECKVAR_ACCESS = Checkbutton(TOP_ACCESS, variable = CHECKVAR_DVVOID, font = BUTTON_FONT, bg = BUTTON_BG)
        CHECKVAR_ACCESS.grid(column = 3, row = 4)
        
        global CHECKVAR_SOAADD
        CHECKVAR_SOAADD = IntVar()
        CHECKVAR_ACCESS = Checkbutton(TOP_ACCESS, variable = CHECKVAR_SOAADD, font = BUTTON_FONT, bg = BUTTON_BG)
        CHECKVAR_ACCESS.grid(column = 1, row = 5)
        global CHECKVAR_SOAAPPROVE
        CHECKVAR_SOAAPPROVE = IntVar()
        CHECKVAR_ACCESS = Checkbutton(TOP_ACCESS, variable = CHECKVAR_SOAAPPROVE, font = BUTTON_FONT, bg = BUTTON_BG)
        CHECKVAR_ACCESS.grid(column = 2, row = 5)
        global CHECKVAR_SOAVOID
        CHECKVAR_SOAVOID = IntVar()
        CHECKVAR_ACCESS = Checkbutton(TOP_ACCESS, variable = CHECKVAR_SOAVOID, font = BUTTON_FONT, bg = BUTTON_BG)
        CHECKVAR_ACCESS.grid(column = 3, row = 5)
        
        global CHECKVAR_ORADD
        CHECKVAR_ORADD = IntVar()
        CHECKVAR_ACCESS = Checkbutton(TOP_ACCESS, variable = CHECKVAR_ORADD, font = BUTTON_FONT, bg = BUTTON_BG)
        CHECKVAR_ACCESS.grid(column = 1, row = 6)
        global CHECKVAR_ORAPPROVE
        CHECKVAR_ORAPPROVE = IntVar()
        CHECKVAR_ACCESS = Checkbutton(TOP_ACCESS, variable = CHECKVAR_ORAPPROVE, font = BUTTON_FONT, bg = BUTTON_BG)
        CHECKVAR_ACCESS.grid(column = 2, row = 6)
        global CHECKVAR_ORVOID
        CHECKVAR_ORVOID = IntVar()
        CHECKVAR_ACCESS = Checkbutton(TOP_ACCESS, variable = CHECKVAR_ORVOID, font = BUTTON_FONT, bg = BUTTON_BG)
        CHECKVAR_ACCESS.grid(column = 3, row = 6)
        
        global CHECKVAR_GJADD
        CHECKVAR_GJADD = IntVar()
        CHECKVAR_ACCESS = Checkbutton(TOP_ACCESS, variable = CHECKVAR_GJADD, font = BUTTON_FONT, bg = BUTTON_BG)
        CHECKVAR_ACCESS.grid(column = 1, row = 7)
        global CHECKVAR_GJAPPROVE
        CHECKVAR_GJAPPROVE = IntVar()
        CHECKVAR_ACCESS = Checkbutton(TOP_ACCESS, variable = CHECKVAR_GJAPPROVE, font = BUTTON_FONT, bg = BUTTON_BG)
        CHECKVAR_ACCESS.grid(column = 2, row = 7)
        global CHECKVAR_GJVOID
        CHECKVAR_GJVOID = IntVar()
        CHECKVAR_ACCESS = Checkbutton(TOP_ACCESS, variable = CHECKVAR_GJVOID, font = BUTTON_FONT, bg = BUTTON_BG)
        CHECKVAR_ACCESS.grid(column = 3, row = 7)
        
        BUTTON_SAVE = Button(TOP_ACCESS, text = "SAVE", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "hand2", command = self.saveAccess)
        BUTTON_SAVE.grid(column = 2, row = 8, pady = TOP_PADY)

        BUTTON_CLOSE = Button(TOP_ACCESS, text = "CLOSE", font = BUTTON_FONT, bg = BUTTON_BG, fg = BUTTON_FG, activebackground = CLICK_BG, width = TOP_BUTTON_WIDTH, cursor = "hand2", command = lambda: self.closeTopLevel(TOP_ACCESS))
        BUTTON_CLOSE.grid(column = 2, row = 9, pady = TOP_PADY)
        
        select = "SELECT poAdd, poApprove, poVoid, rrAdd, rrApprove, rrVoid, apvAdd, apvApprove, apvVoid, dvAdd, dvApprove, dvVoid, soaAdd, soaApprove, soaVoid, orAdd, orApprove, orVoid, gjAdd, gjApprove, gjVoid FROM tblaccess WHERE ID = %s LIMIT 1"
        cursor.execute(select, [content[0]])
        result = cursor.fetchone()
        if result:
            CHECKVAR_POADD.set(result[0])
            CHECKVAR_POAPPROVE.set(result[1])
            CHECKVAR_POVOID.set(result[2])
            CHECKVAR_RRADD.set(result[3])
            CHECKVAR_RRAPPROVE.set(result[4])
            CHECKVAR_RRVOID.set(result[5])
            CHECKVAR_APVADD.set(result[6])
            CHECKVAR_APVAPPROVE.set(result[7])
            CHECKVAR_APVVOID.set(result[8])
            CHECKVAR_DVADD.set(result[9])
            CHECKVAR_DVAPPROVE.set(result[10])
            CHECKVAR_DVVOID.set(result[11])
            
            CHECKVAR_SOAADD.set(result[12])
            CHECKVAR_SOAAPPROVE.set(result[13])
            CHECKVAR_SOAVOID.set(result[14])
            
            CHECKVAR_ORADD.set(result[15])
            CHECKVAR_ORAPPROVE.set(result[16])
            CHECKVAR_ORVOID.set(result[17])
            
            CHECKVAR_GJADD.set(result[18])
            CHECKVAR_GJAPPROVE.set(result[19])
            CHECKVAR_GJVOID.set(result[20])

    def saveAccess(self):
        update = """UPDATE tblaccess SET 
        poAdd = %s, poApprove = %s, poVoid = %s,
        rrAdd = %s, rrApprove = %s, rrVoid = %s,
        apvAdd = %s, apvApprove = %s, apvVoid = %s,
        dvAdd = %s, dvApprove = %s, dvVoid = %s,
        soaAdd = %s, soaApprove = %s, soaVoid = %s,
        orAdd = %s, orApprove = %s, orVoid = %s,
        gjAdd = %s, gjApprove = %s, gjVoid = %s
        WHERE ID = %s"""
        
        ask = messagebox.askyesno("User Access", "Are you sure?")
        if ask:
            cursor.execute(update, [
                CHECKVAR_POADD.get(),
                CHECKVAR_POAPPROVE.get(),
                CHECKVAR_POVOID.get(),
                CHECKVAR_RRADD.get(),
                CHECKVAR_RRAPPROVE.get(),
                CHECKVAR_RRVOID.get(),
                CHECKVAR_APVADD.get(),
                CHECKVAR_APVAPPROVE.get(),
                CHECKVAR_APVVOID.get(),
                CHECKVAR_DVADD.get(),
                CHECKVAR_DVAPPROVE.get(),
                CHECKVAR_DVVOID.get(),
                CHECKVAR_SOAADD.get(),
                CHECKVAR_SOAAPPROVE.get(),
                CHECKVAR_SOAVOID.get(),
                CHECKVAR_ORADD.get(),
                CHECKVAR_ORAPPROVE.get(),
                CHECKVAR_ORVOID.get(),
                CHECKVAR_GJADD.get(),
                CHECKVAR_GJAPPROVE.get(),
                CHECKVAR_GJVOID.get(),
                content[0]
            ])
            db.commit()
            messagebox.showinfo("User Access", "User Access has been updated!")
            TOP_ACCESS.focus()

    def returnAccess(self, var, i):
        select = "SELECT poAdd, poApprove, poVoid, rrAdd, rrApprove, rrVoid, apvAdd, apvApprove, apvVoid, dvAdd, dvApprove, dvVoid, soaAdd, soaApprove, soaVoid, orAdd, orApprove, orVoid, gjAdd, gjApprove, gjVoid FROM tblaccess WHERE ID = %s LIMIT 1"
        cursor.execute(select, [var])
        result = cursor.fetchone()
        if result:
            return result[i]

### OTHERS ###
    def morphMenuButton(self, button, *args):
        for i in SUBMENU_FRAMES:
            for x in i.winfo_children():
                x.destroy()
            i.grid_forget()
            i.grid()

        for i in MENU_BUTTONS:
            if i is not button:
                if i["width"] != 107:
                    if MORPH == True:
                        for x in range(31):
                            i.config(width = 77 + x)
                            SUB1_FRAME3.update()
                            SUB1_FRAME3.after(MORPH_SPEED)
                    else:
                        i.config(width = 77 + 30)

        if MORPH == True:
            for i in range(31):
                button.config(width = 107 - i)
                SUB1_FRAME3.update()
                SUB1_FRAME3.after(MORPH_SPEED)
        else:
            button.config(width = 107 - 30)

    def fixedMap(self, option, tree):
        return [elm for elm in tree.map("Treeview", query_opt = option) if elm[:2] != ("!disabled", "!selected")]

    def clearWorkspace(self):
        try:
            FRAME_4.winfo_children()[0].destroy()
        except:
            pass

    def createScrollFrame(self, frame, h, w, w2, coln, rown):
        global SCROLLBOX, SCROLLABLE_FRAME, CANVAS, SCROLLBAR
        SCROLLBOX = Frame(frame)
        CONTAINER = Frame(SCROLLBOX)
        CANVAS = Canvas(CONTAINER)
        SCROLLBAR = Scrollbar(CONTAINER, orient = "vertical", width = w2, command = CANVAS.yview)
        SCROLLABLE_FRAME = Frame(CANVAS)
        SCROLLABLE_FRAME.bind("<Configure>", lambda e: CANVAS.configure(scrollregion = CANVAS.bbox("all")))
        CANVAS.create_window((0,0), window = SCROLLABLE_FRAME, anchor = "nw")
        CANVAS.configure(yscrollcommand = SCROLLBAR.set)

        SCROLLBOX.grid(column = coln, row = rown, sticky = NW)
        CONTAINER.pack()
        CANVAS.pack(side = "left", fill = "both", expand = True)
        CANVAS.config(height = h, width = w)
        SCROLLBAR.pack(side = "right", fill = "y")
    
    def createScrollFrame2(self, frame, h, w, w2, coln, rown):
        global SCROLLBOX2, SCROLLABLE_FRAME2, CANVAS2, SCROLLBAR2
        SCROLLBOX2 = Frame(frame)
        CONTAINER2 = Frame(SCROLLBOX2)
        CANVAS2 = Canvas(CONTAINER2)
        SCROLLBAR2 = Scrollbar(CONTAINER2, orient = "vertical", width = w2, command = CANVAS2.yview)
        SCROLLABLE_FRAME2 = Frame(CANVAS2)
        SCROLLABLE_FRAME2.bind("<Configure>", lambda e: CANVAS2.configure(scrollregion = CANVAS2.bbox("all")))
        CANVAS2.create_window((0,0), window = SCROLLABLE_FRAME2, anchor = "nw")
        CANVAS2.configure(yscrollcommand = SCROLLBAR2.set)

        SCROLLBOX2.grid(column = coln, row = rown, sticky = NW)
        CONTAINER2.pack()
        CANVAS2.pack(side = "left", fill = "both", expand = True)
        CANVAS2.config(height = h, width = w)
        SCROLLBAR2.pack(side = "right", fill = "y")

    def formatDate(self, widget, var, *args):
        try:
            if len(widget.get()) == 10:
                try:
                    FORMATTED = datetime.datetime.strptime(widget.get(), "%Y-%m-%d")
                except:
                    FORMATTED = datetime.datetime.strptime(widget.get(), "%Y/%m/%d")
                finally:
                    var.set(FORMATTED.strftime("%Y-%m-%d"))
            else:
                messagebox.showerror("Date Checker", "Date entered is invalid!")
                var.set(datetime.datetime.now())
        except:
            messagebox.showerror("Date Checker", "Date entered is invalid!")
            var.set(datetime.datetime.now())

    def popupMenu(self, tree, popup, event):
        selection = tree.identify_row(event.y)
        if selection:
            tree.selection_set(selection)
            tree.identify_row(event.y)
            popup.post(event.x_root, event.y_root)

    def copySelection(self, tree):
        try:
            global content, item
            item = tree.selection()[0]
            content = tree.item(item, option = "values")
        except:
            pass

    def sortColumn(self, tree, col, reverse):
        l = [(tree.set(k, col), k) for k in tree.get_children('')]
        l.sort(reverse = reverse)

        # rearrange items in sorted positions
        for index, (val, k) in enumerate(l):
            tree.move(k, '', index)

        # reverse sort next time
        tree.heading(col, command = lambda: self.sortColumn(tree, col, not reverse))

    def closeTopLevel(self, top, *args):
        top.grab_release()
        top.destroy()

    def listCenters(self):
        db.commit()
        cursor.execute("SELECT name FROM tblcenters")
        result = cursor.fetchall()
        return result

    def listUnits(self):
        db.commit()
        cursor.execute("SELECT name FROM tblunits")
        result = cursor.fetchall()
        return result

    def listTransactions(self):
        db.commit()
        cursor.execute("SELECT transaction FROM tbltransactions ORDER BY transaction")
        result = cursor.fetchall()
        new = []
        for i in result:
            new.append(i[0])
        return new

    def capitalLetters(self, var):
        try:
            var.set(var.get().upper())
        except:
            pass

    def validateAmount(self, amt, *args):
        try:
            if amt.get() == "":
                amt.set(format(float(0), ',.2f'))
            else:
                try:
                    comma = format(float(amt.get().replace(",","")), ',.2f')
                    amt.set(comma)
                except:
                    amt.set(format(float(0), ',.2f'))
        except:
            amt.set(format(float(0), ',.2f'))
    
    def validateAmount2(self, amt, *args):
        return format(amt, ',.2f')

    def returnFloatAmount(self, amt):
        try:
            return float(amt.replace(",", ""))
        except:
            return 0

    def returnValidatedAmount(self, amt, *args):
        try:
            if amt.get() == "":
                return format(float(0), ',.2f')
            else:
                try:
                    comma = format(float(amt.get().replace(",","")), ',.2f')
                    return comma
                except:
                    format(float(0), ',.2f')
        except:
            format(float(0), ',.2f')
            
    def validateInteger(self, amt, *args):
        for i in amt.get():
            if i.isdigit() == False:
                amt.set(0)
                break

    def deleteTreeItem(self, tree):
        self.copySelection(tree)
        tree.delete(item)
        try:
            self.updatePOTotals()
        except:
            pass

    def getFolderAddress(self, var):
        filename = filedialog.askdirectory()
        var.set(filename)

    def exportTemplate(self, file):
        wb = load_workbook(PATH_TEMPLATE + file)
        wb.save(PATH_SAVE + file)
        startfile(PATH_SAVE + file, "open")

    def returnPostingStatus(self, var, var2):
        select = "SELECT reference FROM tblgeneralledger WHERE reference = %s AND source = %s LIMIT 1"
        cursor.execute(select, [var, var2])
        result = cursor.fetchone()
        if result:
            return "Posted"
        else:
            return "Unposted"

    def returnClientName(self, var):
        find = "SELECT clientName FROM tblclients WHERE clientCode = %s"
        cursor.execute(find, [var])
        result = cursor.fetchone()
        if result:
            return result[0]
    
    def returnUserName(self, var, i):
        find = "SELECT username, firstname, lastname, department, usertype FROM tblusers WHERE id = %s"
        cursor.execute(find, [var])
        result = cursor.fetchone()
        if result:
            return result[i]

root = Tk()
root.title(APP_NAME)
root.iconbitmap(PATH_ICON + "icon.ico")
root.geometry(f"{root.winfo_screenwidth()}x{root.winfo_screenheight()}+0+0")
root.config(background = APP_BG)
root.resizable(width = True, height = True)

ICON_LOGO = ImageTk.PhotoImage(Image.open(PATH_ICON + "icon.png"))
ICON_PROFILE = ImageTk.PhotoImage(Image.open(PATH_ICON + "user.png"))
ICON_SETTINGS = ImageTk.PhotoImage(Image.open(PATH_ICON + "settings.png"))
ICON_ACCOUNTING = ImageTk.PhotoImage(Image.open(PATH_ICON + "accounting.png"))
ICON_FINANCE = ImageTk.PhotoImage(Image.open(PATH_ICON + "finance.png"))
ICON_BILLING = ImageTk.PhotoImage(Image.open(PATH_ICON + "finance.png"))
ICON_GSAD = ImageTk.PhotoImage(Image.open(PATH_ICON + "gsad.png"))
ICON_ADMIN = ImageTk.PhotoImage(Image.open(PATH_ICON + "admin.png"))

cursor.execute("SELECT vsNumber FROM tblversions ORDER BY id DESC LIMIT 0, 1")
result = cursor.fetchone()
if result[0] != APP_VERSION:
    startfile(PATH_UPDATE + "UPDATER.exe", "open")
else:
    if __name__ == "__main__":
        Main(root)

    root.mainloop()
